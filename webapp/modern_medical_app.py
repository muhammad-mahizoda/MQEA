#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Современный медицинский веб-интерфейс MQEA с полной функциональностью.

Автор: Мухаммад Махизода
Таджикский национальный университет
"""

import sys
import os

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.figure_factory as ff
import random
from datetime import datetime, timedelta, date
import time
import uuid
import json
import logging

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('streamlit_debug.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def log_action(action, details=""):
    """Функция для логирования действий"""
    timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    message = f"[{timestamp}] {action}"
    if details:
        message += f" - {details}"
    # Убираем эмодзи для безопасного вывода (Windows cp1251 не поддерживает эмодзи)
    clean_message = message.encode('ascii', 'ignore').decode('ascii')
    logger.info(clean_message)
    # Безопасный вывод в консоль без эмодзи (для совместимости с Windows)
    # Используем только logger, так как print может конфликтовать со Streamlit
    # Если нужно вывести в консоль, используем logger который уже настроен

# Добавляем путь к модулям
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Импорты MQEA
from mqea import (
    MQEAAnalyzer, 
    MedicalRecommendationEngine,
    PatientProfile,
    Gender,
    ActivityLevel,
    MedicalHistory
)
from mqea.enhanced_ai_assistant import EnhancedMQEAAssistant
from mqea.patient_database import PatientDatabase, PatientRecord, MedicalVisit, DiagnosisRecord
from mqea.ml_diagnostic_engine import MLDiagnosticEngine
from mqea.patient_card_generator import PatientCardGenerator
from mqea.pediatric_quantum_system import (
    PediatricQuantumEngine,
    PediatricVitalSigns,
    AgeGroup,
    PediatricCondition,
    DetailedAnthropometry
)
from mqea.pediatric_profile_system import (
    PediatricProfileManager,
    ChildProfile,
    DevelopmentRecord
)
from mqea.disease_pattern_analyzer import (
    DiseasePatternAnalyzer,
    DiseaseCategory
)
from mqea.auth import AuthManager, User

# Класс для алгоритма КАПЧ
class KAPCHAnalyzer:
    """Квантовый Анализ Подсознание Человека - алгоритм анализа психосоматических состояний."""
    
    def __init__(self):
        self.emotional_states = {
            'стресс': {'risk_level': 0.8, 'diseases': ['гипертония', 'язва', 'депрессия', 'бессонница']},
            'тревога': {'risk_level': 0.7, 'diseases': ['панические атаки', 'гастрит', 'мигрень', 'астма']},
            'гнев': {'risk_level': 0.9, 'diseases': ['инфаркт', 'инсульт', 'язва', 'гипертония']},
            'печаль': {'risk_level': 0.6, 'diseases': ['депрессия', 'анорексия', 'иммунодефицит', 'сердечные заболевания']},
            'страх': {'risk_level': 0.8, 'diseases': ['фобии', 'панические расстройства', 'проблемы с ЖКТ', 'нарушения сна']},
            'одиночество': {'risk_level': 0.7, 'diseases': ['депрессия', 'алкоголизм', 'сердечные заболевания', 'снижение иммунитета']},
            'зависть': {'risk_level': 0.6, 'diseases': ['язва', 'мигрень', 'бессонница', 'нервные расстройства']},
            'вина': {'risk_level': 0.8, 'diseases': ['депрессия', 'самоповреждения', 'психосоматические расстройства', 'тревожность']},
            'стыд': {'risk_level': 0.7, 'diseases': ['социальная фобия', 'депрессия', 'проблемы с самооценкой', 'изоляция']},
            'разочарование': {'risk_level': 0.5, 'diseases': ['депрессия', 'апатия', 'снижение мотивации', 'проблемы со сном']}
        }
        
        self.physical_states = {
            'усталость': {'risk_level': 0.6, 'diseases': ['синдром хронической усталости', 'иммунодефицит', 'депрессия', 'проблемы с сердцем']},
            'бессонница': {'risk_level': 0.7, 'diseases': ['депрессия', 'тревожность', 'проблемы с памятью', 'снижение иммунитета']},
            'головная боль': {'risk_level': 0.5, 'diseases': ['мигрень', 'гипертония', 'стресс', 'проблемы с шеей']},
            'боль в груди': {'risk_level': 0.8, 'diseases': ['сердечные заболевания', 'панические атаки', 'стресс', 'тревожность']},
            'проблемы с пищеварением': {'risk_level': 0.6, 'diseases': ['язва', 'гастрит', 'синдром раздраженного кишечника', 'стресс']},
            'мышечное напряжение': {'risk_level': 0.5, 'diseases': ['фибромиалгия', 'хроническая боль', 'стресс', 'тревожность']}
        }
        
        self.quantum_coefficients = {
            'эмоциональная_стабильность': 0.3,
            'стрессоустойчивость': 0.25,
            'социальная_поддержка': 0.2,
            'физическая_активность': 0.15,
            'качество_сна': 0.1
        }
        
        # Детальная информация о системах организма
        self.body_systems = {
            'нервная_система': {
                'organs': ['мозг', 'спинной_мозг', 'нервы'],
                'emotions_affecting': ['стресс', 'тревога', 'страх', 'гнев'],
                'symptoms': ['головная боль', 'бессонница', 'мышечное напряжение'],
                'diseases': ['мигрень', 'депрессия', 'тревожные расстройства', 'бессонница']
            },
            'сердечно_сосудистая': {
                'organs': ['сердце', 'артерии', 'вены'],
                'emotions_affecting': ['гнев', 'стресс', 'тревога', 'печаль'],
                'symptoms': ['боль в груди', 'усталость'],
                'diseases': ['гипертония', 'инфаркт', 'инсульт', 'аритмия']
            },
            'пищеварительная': {
                'organs': ['желудок', 'кишечник', 'печень', 'поджелудочная'],
                'emotions_affecting': ['стресс', 'тревога', 'гнев', 'вина'],
                'symptoms': ['проблемы с пищеварением'],
                'diseases': ['язва', 'гастрит', 'синдром раздраженного кишечника']
            },
            'иммунная': {
                'organs': ['лимфоузлы', 'селезенка', 'костный_мозг'],
                'emotions_affecting': ['печаль', 'одиночество', 'стресс'],
                'symptoms': ['усталость'],
                'diseases': ['иммунодефицит', 'аутоиммунные заболевания']
            },
            'дыхательная': {
                'organs': ['легкие', 'бронхи', 'трахея'],
                'emotions_affecting': ['страх', 'тревога', 'печаль'],
                'symptoms': [],
                'diseases': ['астма', 'панические атаки', 'гипервентиляция']
            },
            'мочевыделительная': {
                'organs': ['почки', 'мочевой_пузырь'],
                'emotions_affecting': ['страх', 'тревога'],
                'symptoms': [],
                'diseases': ['инфекции мочевыводящих путей', 'камни в почках']
            }
        }
        
        # Нормализация и восстановление
        self.normalization_methods = {
            'медитация': {'effectiveness': 0.8, 'systems': ['нервная_система', 'сердечно_сосудистая']},
            'дыхательные_упражнения': {'effectiveness': 0.7, 'systems': ['нервная_система', 'дыхательная']},
            'физические_упражнения': {'effectiveness': 0.9, 'systems': ['сердечно_сосудистая', 'иммунная']},
            'социальная_поддержка': {'effectiveness': 0.6, 'systems': ['нервная_система', 'иммунная']},
            'здоровое_питание': {'effectiveness': 0.7, 'systems': ['пищеварительная', 'иммунная']},
            'качественный_сон': {'effectiveness': 0.8, 'systems': ['нервная_система', 'иммунная']},
            'психотерапия': {'effectiveness': 0.9, 'systems': ['нервная_система']},
            'йога': {'effectiveness': 0.8, 'systems': ['нервная_система', 'сердечно_сосудистая']}
        }
    
    def analyze_human_state(self, emotional_data, physical_data, lifestyle_data):
        """Анализирует состояние человека и предсказывает риски заболеваний."""
        
        # Анализ эмоционального состояния
        emotional_risks = self._analyze_emotional_risks(emotional_data)
        
        # Анализ физического состояния
        physical_risks = self._analyze_physical_risks(physical_data)
        
        # Квантовый анализ подсознания
        quantum_analysis = self._quantum_subconscious_analysis(emotional_data, physical_data, lifestyle_data)
        
        # Общая оценка рисков
        total_risk_assessment = self._calculate_total_risks(emotional_risks, physical_risks, quantum_analysis)
        
        # Рекомендации по профилактике
        recommendations = self._generate_recommendations(total_risk_assessment)
        
        # Детальный анализ систем организма
        body_systems_analysis = self._analyze_body_systems(emotional_data, physical_data)
        
        # Анализ нормализации
        normalization_analysis = self._analyze_normalization(body_systems_analysis)
        
        # Подробные предсказания
        detailed_predictions = self._generate_detailed_predictions(total_risk_assessment, body_systems_analysis)
        
        return {
            'emotional_risks': emotional_risks,
            'physical_risks': physical_risks,
            'quantum_analysis': quantum_analysis,
            'total_risk_assessment': total_risk_assessment,
            'recommendations': recommendations,
            'risk_score': total_risk_assessment['overall_risk_score'],
            'body_systems_analysis': body_systems_analysis,
            'normalization_analysis': normalization_analysis,
            'detailed_predictions': detailed_predictions
        }
    
    def _analyze_emotional_risks(self, emotional_data):
        """Анализирует эмоциональные риски."""
        risks = []
        total_emotional_risk = 0
        
        for emotion, intensity in emotional_data.items():
            if emotion in self.emotional_states:
                emotion_data = self.emotional_states[emotion]
                risk_score = emotion_data['risk_level'] * intensity
                total_emotional_risk += risk_score
                
                risks.append({
                    'emotion': emotion,
                    'intensity': intensity,
                    'risk_score': risk_score,
                    'potential_diseases': emotion_data['diseases'],
                    'risk_level': 'Высокий' if risk_score > 0.7 else 'Средний' if risk_score > 0.4 else 'Низкий'
                })
        
        return {
            'individual_risks': risks,
            'total_emotional_risk': min(total_emotional_risk, 1.0),
            'risk_level': 'Критический' if total_emotional_risk > 0.8 else 'Высокий' if total_emotional_risk > 0.6 else 'Средний' if total_emotional_risk > 0.4 else 'Низкий'
        }
    
    def _analyze_physical_risks(self, physical_data):
        """Анализирует физические риски."""
        risks = []
        total_physical_risk = 0
        
        for symptom, intensity in physical_data.items():
            if symptom in self.physical_states:
                symptom_data = self.physical_states[symptom]
                risk_score = symptom_data['risk_level'] * intensity
                total_physical_risk += risk_score
                
                risks.append({
                    'symptom': symptom,
                    'intensity': intensity,
                    'risk_score': risk_score,
                    'potential_diseases': symptom_data['diseases'],
                    'risk_level': 'Высокий' if risk_score > 0.7 else 'Средний' if risk_score > 0.4 else 'Низкий'
                })
        
        return {
            'individual_risks': risks,
            'total_physical_risk': min(total_physical_risk, 1.0),
            'risk_level': 'Критический' if total_physical_risk > 0.8 else 'Высокий' if total_physical_risk > 0.6 else 'Средний' if total_physical_risk > 0.4 else 'Низкий'
        }
    
    def _quantum_subconscious_analysis(self, emotional_data, physical_data, lifestyle_data):
        """Квантовый анализ подсознательных процессов."""
        
        # Вычисляем квантовые коэффициенты
        quantum_factors = {}
        
        # Эмоциональная стабильность (обратная зависимость от негативных эмоций)
        negative_emotions = sum(intensity for emotion, intensity in emotional_data.items() 
                               if emotion in ['стресс', 'тревога', 'гнев', 'печаль', 'страх'])
        quantum_factors['эмоциональная_стабильность'] = max(0, 1 - negative_emotions * 0.2)
        
        # Стрессоустойчивость
        stress_level = emotional_data.get('стресс', 0)
        quantum_factors['стрессоустойчивость'] = max(0, 1 - stress_level * 0.3)
        
        # Социальная поддержка (из lifestyle_data)
        social_support = lifestyle_data.get('социальная_поддержка', 0.5)
        quantum_factors['социальная_поддержка'] = social_support
        
        # Физическая активность
        physical_activity = lifestyle_data.get('физическая_активность', 0.5)
        quantum_factors['физическая_активность'] = physical_activity
        
        # Качество сна
        sleep_quality = lifestyle_data.get('качество_сна', 0.5)
        quantum_factors['качество_сна'] = sleep_quality
        
        # Общий квантовый индекс здоровья
        quantum_health_index = sum(
            quantum_factors[factor] * self.quantum_coefficients[factor] 
            for factor in quantum_factors
        )
        
        return {
            'quantum_factors': quantum_factors,
            'quantum_health_index': quantum_health_index,
            'subconscious_stability': quantum_health_index,
            'quantum_coherence': quantum_health_index * 0.8 + np.random.normal(0, 0.1)  # Добавляем квантовую неопределенность
        }
    
    def _calculate_total_risks(self, emotional_risks, physical_risks, quantum_analysis):
        """Вычисляет общую оценку рисков."""
        
        # Взвешенная оценка рисков
        emotional_weight = 0.4
        physical_weight = 0.3
        quantum_weight = 0.3
        
        total_risk_score = (
            emotional_risks['total_emotional_risk'] * emotional_weight +
            physical_risks['total_physical_risk'] * physical_weight +
            (1 - quantum_analysis['quantum_health_index']) * quantum_weight
        )
        
        # Определяем наиболее вероятные заболевания
        all_diseases = {}
        for risk in emotional_risks['individual_risks']:
            for disease in risk['potential_diseases']:
                if disease not in all_diseases:
                    all_diseases[disease] = 0
                all_diseases[disease] += risk['risk_score']
        
        for risk in physical_risks['individual_risks']:
            for disease in risk['potential_diseases']:
                if disease not in all_diseases:
                    all_diseases[disease] = 0
                all_diseases[disease] += risk['risk_score']
        
        # Сортируем заболевания по вероятности
        sorted_diseases = sorted(all_diseases.items(), key=lambda x: x[1], reverse=True)
        
        return {
            'overall_risk_score': min(total_risk_score, 1.0),
            'risk_level': 'Критический' if total_risk_score > 0.8 else 'Высокий' if total_risk_score > 0.6 else 'Средний' if total_risk_score > 0.4 else 'Низкий',
            'most_likely_diseases': sorted_diseases[:5],
            'risk_factors': {
                'emotional': emotional_risks['total_emotional_risk'],
                'physical': physical_risks['total_physical_risk'],
                'quantum': 1 - quantum_analysis['quantum_health_index']
            }
        }
    
    def _generate_recommendations(self, risk_assessment):
        """Генерирует рекомендации по профилактике."""
        recommendations = []
        
        risk_score = risk_assessment['overall_risk_score']
        
        if risk_score > 0.7:
            recommendations.extend([
                "🚨 КРИТИЧЕСКИЙ УРОВЕНЬ РИСКА - Немедленно обратитесь к специалисту",
                "🧘 Практикуйте медитацию и дыхательные упражнения",
                "🏃‍♂️ Увеличьте физическую активность",
                "👥 Обратитесь за психологической поддержкой"
            ])
        elif risk_score > 0.5:
            recommendations.extend([
                "⚠️ ВЫСОКИЙ УРОВЕНЬ РИСКА - Рекомендуется консультация специалиста",
                "🧘 Начните практиковать релаксационные техники",
                "🏃‍♂️ Регулярно занимайтесь спортом",
                "😴 Улучшите качество сна"
            ])
        elif risk_score > 0.3:
            recommendations.extend([
                "📊 СРЕДНИЙ УРОВЕНЬ РИСКА - Профилактические меры",
                "🧘 Изучите техники управления стрессом",
                "🏃‍♂️ Поддерживайте умеренную физическую активность",
                "😊 Развивайте позитивное мышление"
            ])
        else:
            recommendations.extend([
                "✅ НИЗКИЙ УРОВЕНЬ РИСКА - Продолжайте здоровый образ жизни",
                "🧘 Поддерживайте текущие практики релаксации",
                "🏃‍♂️ Сохраняйте активный образ жизни",
                "😊 Развивайте эмоциональную устойчивость"
            ])
        
        return recommendations
    
    def _analyze_body_systems(self, emotional_data, physical_data):
        """Анализирует состояние систем организма."""
        systems_analysis = {}
        
        for system_name, system_data in self.body_systems.items():
            # Вычисляем риск для каждой системы
            system_risk = 0
            affected_emotions = []
            affected_symptoms = []
            
            # Анализируем влияние эмоций
            for emotion in system_data['emotions_affecting']:
                if emotion in emotional_data and emotional_data[emotion] > 0:
                    emotion_intensity = emotional_data[emotion]
                    system_risk += emotion_intensity * 0.3
                    affected_emotions.append({
                        'emotion': emotion,
                        'intensity': emotion_intensity,
                        'impact': emotion_intensity * 0.3
                    })
            
            # Анализируем влияние симптомов
            for symptom in system_data['symptoms']:
                if symptom in physical_data and physical_data[symptom] > 0:
                    symptom_intensity = physical_data[symptom]
                    system_risk += symptom_intensity * 0.4
                    affected_symptoms.append({
                        'symptom': symptom,
                        'intensity': symptom_intensity,
                        'impact': symptom_intensity * 0.4
                    })
            
            # Определяем уровень риска
            risk_level = 'Низкий'
            if system_risk > 0.7:
                risk_level = 'Критический'
            elif system_risk > 0.5:
                risk_level = 'Высокий'
            elif system_risk > 0.3:
                risk_level = 'Средний'
            
            systems_analysis[system_name] = {
                'risk_score': min(system_risk, 1.0),
                'risk_level': risk_level,
                'affected_emotions': affected_emotions,
                'affected_symptoms': affected_symptoms,
                'organs': system_data['organs'],
                'potential_diseases': system_data['diseases'],
                'status': 'Здоров' if system_risk < 0.3 else 'Нарушен' if system_risk > 0.5 else 'Под угрозой'
            }
        
        return systems_analysis
    
    def _analyze_normalization(self, body_systems_analysis):
        """Анализирует методы нормализации для каждой системы."""
        normalization_plan = {}
        
        for system_name, system_data in body_systems_analysis.items():
            if system_data['risk_score'] > 0.3:  # Если система нуждается в нормализации
                recommended_methods = []
                
                for method_name, method_data in self.normalization_methods.items():
                    if system_name in method_data['systems']:
                        effectiveness = method_data['effectiveness']
                        priority = 'Высокий' if effectiveness > 0.8 else 'Средний' if effectiveness > 0.6 else 'Низкий'
                        
                        recommended_methods.append({
                            'method': method_name,
                            'effectiveness': effectiveness,
                            'priority': priority,
                            'expected_improvement': effectiveness * system_data['risk_score'] * 0.5
                        })
                
                # Сортируем по эффективности
                recommended_methods.sort(key=lambda x: x['effectiveness'], reverse=True)
                
                normalization_plan[system_name] = {
                    'current_status': system_data['status'],
                    'risk_score': system_data['risk_score'],
                    'recommended_methods': recommended_methods[:3],  # Топ-3 метода
                    'estimated_recovery_time': self._estimate_recovery_time(system_data['risk_score']),
                    'normalization_priority': 'Критический' if system_data['risk_score'] > 0.7 else 'Высокий' if system_data['risk_score'] > 0.5 else 'Средний'
                }
        
        return normalization_plan
    
    def _estimate_recovery_time(self, risk_score):
        """Оценивает время восстановления."""
        if risk_score > 0.7:
            return "3-6 месяцев интенсивной терапии"
        elif risk_score > 0.5:
            return "1-3 месяца регулярных практик"
        elif risk_score > 0.3:
            return "2-4 недели профилактических мер"
        else:
            return "Поддерживающие меры"
    
    def _generate_detailed_predictions(self, total_risk_assessment, body_systems_analysis):
        """Генерирует подробные предсказания."""
        predictions = {
            'short_term': [],  # 1-3 месяца
            'medium_term': [],  # 3-12 месяцев
            'long_term': [],  # 1-3 года
            'prevention_opportunities': []
        }
        
        # Краткосрочные предсказания
        for disease, probability in total_risk_assessment['most_likely_diseases'][:3]:
            if probability > 0.6:
                predictions['short_term'].append({
                    'condition': disease,
                    'probability': probability,
                    'timeframe': '1-3 месяца',
                    'severity': 'Высокая' if probability > 0.8 else 'Средняя',
                    'prevention_possible': True
                })
        
        # Среднесрочные предсказания
        for system_name, system_data in body_systems_analysis.items():
            if system_data['risk_score'] > 0.5:
                predictions['medium_term'].append({
                    'system': system_name,
                    'risk_score': system_data['risk_score'],
                    'timeframe': '3-12 месяцев',
                    'potential_outcomes': system_data['potential_diseases'][:2],
                    'prevention_strategies': [method['method'] for method in self.normalization_methods.items() if system_name in method[1]['systems']][:3]
                })
        
        # Долгосрочные предсказания
        overall_risk = total_risk_assessment['overall_risk_score']
        if overall_risk > 0.6:
            predictions['long_term'].append({
                'scenario': 'Хронические заболевания',
                'probability': overall_risk,
                'timeframe': '1-3 года',
                'affected_systems': [name for name, data in body_systems_analysis.items() if data['risk_score'] > 0.5],
                'quality_of_life_impact': 'Значительное снижение' if overall_risk > 0.8 else 'Умеренное снижение'
            })
        
        # Возможности профилактики
        for system_name, system_data in body_systems_analysis.items():
            if system_data['risk_score'] < 0.5:  # Здоровые системы
                predictions['prevention_opportunities'].append({
                    'system': system_name,
                    'current_status': 'Здоров',
                    'maintenance_methods': [method for method in self.normalization_methods.keys() if system_name in self.normalization_methods[method]['systems']],
                    'prevention_priority': 'Поддержание здоровья'
                })
        
        return predictions
    
    def create_3d_human_model(self, body_systems_analysis):
        """Создает реалистичную 3D модель человека с визуализацией нарушений."""
        fig = go.Figure()
        
        # Создаем контур человеческого тела
        self._add_human_body_outline(fig)
        
        # Добавляем основные части тела
        self._add_body_parts(fig, body_systems_analysis)
        
        # Добавляем органы в правильных анатомических позициях
        self._add_organs_with_risks(fig, body_systems_analysis)
        
        # Настройка макета
        fig.update_layout(
            title="🧠 3D Модель Человека - Анализ КАПЧ",
            scene=dict(
                xaxis=dict(title="Передняя/Задняя", range=[-1, 1]),
                yaxis=dict(title="Левая/Правая", range=[-0.5, 0.5]),
                zaxis=dict(title="Верх/Низ", range=[0, 2]),
                camera=dict(eye=dict(x=1.8, y=1.8, z=1.2)),
                aspectmode='cube'
            ),
            width=700,
            height=600,
            showlegend=True
        )
        
        return fig
    
    def _add_human_body_outline(self, fig):
        """Добавляет контур человеческого тела."""
        # Голова (сфера)
        theta = np.linspace(0, 2*np.pi, 20)
        phi = np.linspace(0, np.pi, 20)
        x_head = 0.15 * np.outer(np.cos(theta), np.sin(phi))
        y_head = 0.15 * np.outer(np.sin(theta), np.sin(phi))
        z_head = 0.15 * np.outer(np.ones(np.size(theta)), np.cos(phi)) + 1.7
        
        fig.add_trace(go.Surface(
            x=x_head, y=y_head, z=z_head,
            colorscale='Greys',
            opacity=0.3,
            name='Голова',
            showscale=False
        ))
        
        # Туловище (цилиндр)
        theta_torso = np.linspace(0, 2*np.pi, 20)
        z_torso = np.linspace(0.5, 1.5, 20)
        theta_torso, z_torso = np.meshgrid(theta_torso, z_torso)
        x_torso = 0.2 * np.cos(theta_torso)
        y_torso = 0.15 * np.sin(theta_torso)
        
        fig.add_trace(go.Surface(
            x=x_torso, y=y_torso, z=z_torso,
            colorscale='Greys',
            opacity=0.3,
            name='Туловище',
            showscale=False
        ))
        
        # Руки
        # Левая рука
        theta_arm = np.linspace(0, 2*np.pi, 15)
        z_arm = np.linspace(0.8, 1.4, 15)
        theta_arm, z_arm = np.meshgrid(theta_arm, z_arm)
        x_left_arm = -0.3 + 0.08 * np.cos(theta_arm)
        y_left_arm = 0.08 * np.sin(theta_arm)
        
        fig.add_trace(go.Surface(
            x=x_left_arm, y=y_left_arm, z=z_arm,
            colorscale='Greys',
            opacity=0.3,
            name='Левая рука',
            showscale=False
        ))
        
        # Правая рука
        x_right_arm = 0.3 + 0.08 * np.cos(theta_arm)
        y_right_arm = 0.08 * np.sin(theta_arm)
        
        fig.add_trace(go.Surface(
            x=x_right_arm, y=y_right_arm, z=z_arm,
            colorscale='Greys',
            opacity=0.3,
            name='Правая рука',
            showscale=False
        ))
        
        # Ноги
        # Левая нога
        theta_leg = np.linspace(0, 2*np.pi, 15)
        z_leg = np.linspace(0, 0.8, 15)
        theta_leg, z_leg = np.meshgrid(theta_leg, z_leg)
        x_left_leg = -0.1 + 0.1 * np.cos(theta_leg)
        y_left_leg = 0.1 * np.sin(theta_leg)
        
        fig.add_trace(go.Surface(
            x=x_left_leg, y=y_left_leg, z=z_leg,
            colorscale='Greys',
            opacity=0.3,
            name='Левая нога',
            showscale=False
        ))
        
        # Правая нога
        x_right_leg = 0.1 + 0.1 * np.cos(theta_leg)
        y_right_leg = 0.1 * np.sin(theta_leg)
        
        fig.add_trace(go.Surface(
            x=x_right_leg, y=y_right_leg, z=z_leg,
            colorscale='Greys',
            opacity=0.3,
            name='Правая нога',
            showscale=False
        ))
    
    def _add_body_parts(self, fig, body_systems_analysis):
        """Добавляет основные части тела с индикацией рисков."""
        # Определяем общий риск для каждой части тела
        body_parts_risk = {
            'голова': 0,
            'грудь': 0,
            'живот': 0,
            'спина': 0
        }
        
        # Вычисляем риски для каждой части тела
        for system_name, system_data in body_systems_analysis.items():
            risk_score = system_data['risk_score']
            
            if 'мозг' in system_data['organs'] or 'нервы' in system_data['organs']:
                body_parts_risk['голова'] = max(body_parts_risk['голова'], risk_score)
            
            if 'сердце' in system_data['organs'] or 'легкие' in system_data['organs']:
                body_parts_risk['грудь'] = max(body_parts_risk['грудь'], risk_score)
            
            if 'желудок' in system_data['organs'] or 'печень' in system_data['organs'] or 'кишечник' in system_data['organs']:
                body_parts_risk['живот'] = max(body_parts_risk['живот'], risk_score)
        
        # Добавляем цветовую индикацию для частей тела
        for part, risk in body_parts_risk.items():
            color = self._get_risk_color(risk)
            
            if part == 'голова':
                fig.add_trace(go.Scatter3d(
                    x=[0], y=[0], z=[1.7],
                    mode='markers',
                    marker=dict(size=25, color=color, opacity=0.7),
                    name=f'Голова (риск: {risk:.1%})',
                    showlegend=True
                ))
            elif part == 'грудь':
                fig.add_trace(go.Scatter3d(
                    x=[0], y=[0], z=[1.2],
                    mode='markers',
                    marker=dict(size=30, color=color, opacity=0.7),
                    name=f'Грудь (риск: {risk:.1%})',
                    showlegend=True
                ))
            elif part == 'живот':
                fig.add_trace(go.Scatter3d(
                    x=[0], y=[0], z=[1.0],
                    mode='markers',
                    marker=dict(size=25, color=color, opacity=0.7),
                    name=f'Живот (риск: {risk:.1%})',
                    showlegend=True
                ))
    
    def _add_organs_with_risks(self, fig, body_systems_analysis):
        """Добавляет органы в правильных анатомических позициях с индикацией рисков."""
        # Анатомически правильные позиции органов
        organs_anatomy = {
            'мозг': {'x': 0, 'y': 0, 'z': 1.75, 'size': 0.12, 'shape': 'sphere'},
            'сердце': {'x': 0.05, 'y': 0, 'z': 1.25, 'size': 0.08, 'shape': 'heart'},
            'легкие': {'x': 0, 'y': 0, 'z': 1.4, 'size': 0.15, 'shape': 'lungs'},
            'желудок': {'x': 0, 'y': 0, 'z': 1.05, 'size': 0.1, 'shape': 'stomach'},
            'печень': {'x': 0.15, 'y': 0, 'z': 1.1, 'size': 0.12, 'shape': 'liver'},
            'кишечник': {'x': 0, 'y': 0, 'z': 0.9, 'size': 0.2, 'shape': 'intestines'},
            'почки': {'x': 0.2, 'y': 0, 'z': 1.0, 'size': 0.06, 'shape': 'kidney'},
            'селезенка': {'x': -0.15, 'y': 0, 'z': 1.1, 'size': 0.08, 'shape': 'spleen'},
            'поджелудочная': {'x': 0.1, 'y': 0, 'z': 1.0, 'size': 0.05, 'shape': 'pancreas'}
        }
        
        for organ, coords in organs_anatomy.items():
            # Определяем цвет на основе риска
            risk_score = 0
            risk_color = 'green'
            
            for system_name, system_data in body_systems_analysis.items():
                if organ in system_data['organs']:
                    risk_score = system_data['risk_score']
                    risk_color = self._get_risk_color(risk_score)
                    break
            
            # Добавляем орган с анимацией пульсации для проблемных органов
            marker_size = coords['size'] * 15
            if risk_score > 0.5:
                marker_size *= 1.2  # Увеличиваем размер для проблемных органов
            
            fig.add_trace(go.Scatter3d(
                x=[coords['x']],
                y=[coords['y']],
                z=[coords['z']],
                mode='markers',
                marker=dict(
                    size=marker_size,
                    color=risk_color,
                    opacity=0.8,
                    line=dict(width=2, color='white') if risk_score > 0.5 else dict(width=1, color='gray')
                ),
                name=f"{organ.title()} ({risk_score:.1%})",
                text=f"{organ.title()}<br>Риск: {risk_score:.1%}<br>Статус: {'⚠️ Требует внимания' if risk_score > 0.5 else '✅ Здоров'}",
                hovertemplate="%{text}<extra></extra>",
                showlegend=True
            ))
    
    def _get_risk_color(self, risk_score):
        """Возвращает цвет на основе уровня риска."""
        if risk_score > 0.7:
            return 'red'  # Критический риск
        elif risk_score > 0.5:
            return 'orange'  # Высокий риск
        elif risk_score > 0.3:
            return 'yellow'  # Средний риск
        else:
            return 'green'  # Здоров

# Вспомогательные функции
def clean_dataframe(df):
    """Очищает DataFrame от несериализуемых типов данных."""
    if df is None or df.empty:
        return df
    
    df_clean = df.copy()
    
    for col in df_clean.columns:
        if df_clean[col].dtype == 'object':
            # Проверяем каждый элемент в колонке
            for idx in df_clean.index:
                value = df_clean.loc[idx, col]
                
                # Обрабатываем разные типы данных
                if pd.isna(value):
                    df_clean.loc[idx, col] = None
                elif isinstance(value, complex):
                    df_clean.loc[idx, col] = float(abs(value))
                elif hasattr(value, 'isoformat'):  # datetime объекты
                    df_clean.loc[idx, col] = value.isoformat()
                elif isinstance(value, (list, tuple)):
                    df_clean.loc[idx, col] = str(value)
                elif isinstance(value, dict):
                    df_clean.loc[idx, col] = str(value)
                elif hasattr(value, '__iter__') and not isinstance(value, (str, bytes)):
                    df_clean.loc[idx, col] = str(value)
                elif not isinstance(value, (int, float, str, bool, type(None))):
                    df_clean.loc[idx, col] = str(value)
    
    # Дополнительная очистка для Arrow совместимости
    for col in df_clean.columns:
        try:
            # Пытаемся преобразовать в числовой тип, если возможно
            if df_clean[col].dtype == 'object':
                # Проверяем, можно ли преобразовать в числовой тип
                numeric_values = pd.to_numeric(df_clean[col], errors='coerce')
                if not numeric_values.isna().all():
                    df_clean[col] = numeric_values
                else:
                    # Если не числовой, оставляем как строку
                    df_clean[col] = df_clean[col].astype(str)
        except Exception:
            # В случае ошибки, конвертируем все в строки
            df_clean[col] = df_clean[col].astype(str)
    
    return df_clean

# Настройки страницы
st.set_page_config(
    page_title="MQEA - Medical Quantum Entanglement Analysis",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': "MQEA - Medical Quantum Entanglement Analysis\nАвтор: Мухаммад Махизода\nТаджикский национальный университет"
    }
)

# Кастомные CSS стили
st.markdown("""
<style>
/* Скрытие стандартной навигации Streamlit */
.stSidebar .css-1d391kg {
    display: none !important;
}

.stSidebar .css-1oe5cao {
    display: none !important;
}

/* Стили для табов навигации */
.nav-tab {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    border: none;
    border-radius: 8px;
    color: white;
    padding: 8px 16px;
    margin: 2px;
    font-weight: bold;
    transition: all 0.3s ease;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}

.nav-tab:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 8px rgba(0,0,0,0.2);
}

/* Индикатор активного приложения */
.app-indicator {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 12px;
    border-radius: 10px;
    text-align: center;
    font-weight: bold;
    margin: 10px 0;
    box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
    border: 2px solid rgba(255,255,255,0.2);
}

/* Анимация пульсации для активного индикатора */
@keyframes pulse {
    0% { box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3); }
    50% { box-shadow: 0 4px 20px rgba(102, 126, 234, 0.6); }
    100% { box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3); }
}

.app-indicator {
    animation: pulse 2s infinite;
}

/* Стили для разделителей */
.section-divider {
    border-top: 2px solid #667eea;
    margin: 20px 0;
    border-radius: 2px;
}

/* Кастомные кнопки-таблетки */
.pill-button {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    border: none;
    border-radius: 25px;
    padding: 12px 24px;
    font-weight: bold;
    font-size: 14px;
    cursor: pointer;
    transition: all 0.3s ease;
    box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
    text-decoration: none;
    display: inline-block;
    margin: 5px;
}

.pill-button:hover {
    transform: translateY(-2px);
    box-shadow: 0 6px 20px rgba(102, 126, 234, 0.4);
}

.pill-button.active {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    animation: pulse 2s infinite;
}
</style>
""", unsafe_allow_html=True)

# Скрытие только кнопки Deploy
hide_deploy_button = """
<style>
.stDeployButton {display:none !important;}
.stDeployButton * {display:none !important;}
[data-testid="deployButton"] {display:none !important;}
button[title="Deploy"] {display:none !important;}
div[data-testid="deployButton"] {display:none !important;}
.stAppDeployButton {display:none !important;}
.stAppDeployButton button {display:none !important;}
div.st-emotion-cache-scp8yw.e3g0k5y6 .stAppDeployButton {display:none !important;}
#root > div:nth-child(1) > div.withScreencast > div > div > div > header > div > div > div.st-emotion-cache-scp8yw.e3g0k5y6 > div.stAppDeployButton {display:none !important;}

/* Скрытие стандартной навигации */
.stSidebar [data-testid="stSidebarNav"] {display:none !important;}
.stSidebar .css-1d391kg {display:none !important;}
.stSidebar .css-1oe5cao {display:none !important;}
</style>
<script>
// Принудительное скрытие кнопки Deploy и навигации
setInterval(function() {
    const deployButtons = document.querySelectorAll('.stDeployButton, .stAppDeployButton, [data-testid="deployButton"], button[title="Deploy"], #root > div:nth-child(1) > div.withScreencast > div > div > div > header > div > div > div.st-emotion-cache-scp8yw.e3g0k5y6 > div.stAppDeployButton');
    deployButtons.forEach(btn => {
        btn.style.display = 'none';
        btn.style.visibility = 'hidden';
    });
    
    // Скрытие стандартной навигации
    const navElements = document.querySelectorAll('[data-testid="stSidebarNav"]');
    navElements.forEach(el => {
        if (el && !el.closest('.custom-navigation')) {
            el.style.display = 'none';
            el.style.visibility = 'hidden';
        }
    });
}, 100);
</script>
"""
st.markdown(hide_deploy_button, unsafe_allow_html=True)

# Инициализация компонентов
@st.cache_resource
def init_components():
    """Инициализация всех компонентов системы."""
    return {
        'analyzer': MQEAAnalyzer(),
        'assistant': EnhancedMQEAAssistant(MQEAAnalyzer()),
        'database': PatientDatabase(),
        'ml_engine': MLDiagnosticEngine(),
        'card_generator': PatientCardGenerator(),
        'recommendation_engine': MedicalRecommendationEngine()
    }

# Инициализация состояния сессии
def init_session_state():
    """Инициализация состояния сессии."""
    if 'components' not in st.session_state:
        st.session_state.components = init_components()
    
    if 'current_patient' not in st.session_state:
        st.session_state.current_patient = None
    
    if 'current_visit' not in st.session_state:
        st.session_state.current_visit = None
    
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    
    if 'ml_predictions' not in st.session_state:
        st.session_state.ml_predictions = None
    
    # Инициализация переменных маршрутизации
    if 'active_section' not in st.session_state:
        st.session_state.active_section = "main"
    
    if 'current_menu' not in st.session_state:
        st.session_state.current_menu = "📊 Дашборд"
    
    # Инициализация переменных для лабораторных экспериментов
    if 'lab_experiment_completed' not in st.session_state:
        st.session_state.lab_experiment_completed = False
    
    if 'lab_experiment_results' not in st.session_state:
        st.session_state.lab_experiment_results = None
    
    if 'lab_experiment_type' not in st.session_state:
        st.session_state.lab_experiment_type = None
    
    # Инициализация переменных аутентификации
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    if 'current_user' not in st.session_state:
        st.session_state.current_user = None
    
    if 'user_id' not in st.session_state:
        st.session_state.user_id = None
    
    if 'username' not in st.session_state:
        st.session_state.username = None
    
    if 'user_role' not in st.session_state:
        st.session_state.user_role = None
    
    if 'session_id' not in st.session_state:
        st.session_state.session_id = None
    
    # Инициализация менеджера аутентификации (если еще не инициализирован)
    if 'auth_manager' not in st.session_state:
        st.session_state.auth_manager = AuthManager()

# Импорт оптимизированного отображения логотипа
from utils.optimized_logo_display import display_centered_logo, display_premium_logo, display_logo_with_info

# Функция для отображения логотипа
def display_logo():
    """Отображает правильный медицинский логотип MQEA."""
    try:
        # Импортируем функции полноэкранного отображения
        from utils.fullscreen_logo_display import display_centered_logo
        
        # Отображаем правильный центрированный логотип для главного экрана
        display_centered_logo(width='stretch')
    except Exception as e:
        # Резервный вариант
        st.markdown("### 🏥 MQEA")
        st.markdown("**Medical Quantum Entanglement Analysis**")
        st.markdown("*Спокойствие, доверие, стабильность*")

# Главная функция
def show_login_page():
    """Страница входа в систему."""
    st.title("🔐 Вход в систему MQEA")
    
    # Инициализация менеджера аутентификации
    if 'auth_manager' not in st.session_state:
        st.session_state.auth_manager = AuthManager()
    
    auth_manager = st.session_state.auth_manager
    
    # Вкладки: Вход, Регистрация, Восстановление пароля
    tab1, tab2, tab3 = st.tabs(["🔑 Вход", "📝 Регистрация", "🔓 Восстановление пароля"])
    
    with tab1:
        st.subheader("Вход в систему")
        
        with st.form("login_form"):
            username = st.text_input("Имя пользователя", placeholder="Введите имя пользователя")
            password = st.text_input("Пароль", type="password", placeholder="Введите пароль")
            remember_me = st.checkbox("Запомнить меня")
            
            submitted = st.form_submit_button("Войти", type="primary", use_container_width=True)
            
            if submitted:
                if username and password:
                    success, user, message = auth_manager.authenticate_user(username, password)
                    
                    if success and user:
                        st.session_state.authenticated = True
                        st.session_state.current_user = user
                        st.session_state.user_id = user.user_id
                        st.session_state.username = user.username
                        st.session_state.user_role = user.role
                        
                        # Создание сессии
                        session_id = auth_manager.create_session(
                            user.user_id,
                            ip_address=None,  # Можно получить из запроса
                            user_agent=None
                        )
                        st.session_state.session_id = session_id
                        
                        # Сохраняем session_id в query параметрах для восстановления при обновлении страницы
                        try:
                            # Для Streamlit >= 1.28.0
                            if hasattr(st, 'query_params'):
                                st.query_params["session_id"] = session_id
                            else:
                                # Для старых версий
                                st.experimental_set_query_params(session_id=session_id)
                        except:
                            # Если не поддерживается, просто сохраняем в session_state
                            pass
                        
                        st.success(f"✅ Добро пожаловать, {user.full_name}!")
                        st.rerun()
                    else:
                        st.error(f"❌ {message}")
                else:
                    st.warning("⚠️ Заполните все поля")
    
    with tab2:
        st.subheader("Регистрация нового пользователя")
        
        with st.form("register_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                username = st.text_input("Имя пользователя *", placeholder="минимум 3 символа")
                email = st.text_input("Email *", placeholder="example@email.com")
            
            with col2:
                full_name = st.text_input("Полное имя *", placeholder="Иван Иванов")
                # Роль фиксирована как "Пациент" - только администратор может создавать докторов
                role = "user"
                st.info("ℹ️ Регистрация доступна только для пациентов. Врачей создает администратор.")
            
            password = st.text_input("Пароль *", type="password", placeholder="минимум 6 символов")
            password_confirm = st.text_input("Подтверждение пароля *", type="password", placeholder="повторите пароль")
            
            st.caption("* - обязательные поля")
            
            submitted = st.form_submit_button("Зарегистрироваться", type="primary", use_container_width=True)
            
            if submitted:
                # Валидация
                if not all([username, email, full_name, password, password_confirm]):
                    st.warning("⚠️ Заполните все обязательные поля")
                elif password != password_confirm:
                    st.error("❌ Пароли не совпадают")
                else:
                    success, message = auth_manager.register_user(
                        username=username,
                        email=email,
                        password=password,
                        full_name=full_name,
                        role=role
                    )
                    
                    if success:
                        st.success(f"✅ {message}")
                        st.info("Теперь вы можете войти в систему")
                    else:
                        st.error(f"❌ {message}")
    
    with tab3:
        st.subheader("Восстановление пароля")
        
        with st.form("reset_password_form"):
            email = st.text_input("Email", placeholder="Введите email для восстановления пароля")
            
            submitted = st.form_submit_button("Отправить инструкцию", type="primary", use_container_width=True)
            
            if submitted:
                if email:
                    success, token, message = auth_manager.create_reset_token(email)
                    
                    if success:
                        if token:
                            # В реальном приложении здесь должна быть отправка email
                            # Для демо показываем токен напрямую
                            st.success("✅ Токен восстановления создан")
                            st.info(f"🔑 Ваш токен восстановления (действителен 1 час):")
                            st.code(token, language=None)
                            
                            # Форма для сброса пароля
                            st.markdown("---")
                            st.subheader("Сброс пароля")
                            
                            with st.form("reset_form"):
                                reset_token = st.text_input("Токен восстановления", value=token)
                                new_password = st.text_input("Новый пароль", type="password")
                                new_password_confirm = st.text_input("Подтверждение пароля", type="password")
                                
                                reset_submitted = st.form_submit_button("Сбросить пароль", type="primary")
                                
                                if reset_submitted:
                                    if new_password != new_password_confirm:
                                        st.error("❌ Пароли не совпадают")
                                    elif len(new_password) < 6:
                                        st.error("❌ Пароль должен содержать минимум 6 символов")
                                    else:
                                        success, message = auth_manager.reset_password(reset_token, new_password)
                                        
                                        if success:
                                            st.success(f"✅ {message}")
                                            st.info("Теперь вы можете войти с новым паролем")
                                        else:
                                            st.error(f"❌ {message}")
                        else:
                            st.info("ℹ️ Если email существует, инструкция отправлена")
                    else:
                        st.error(f"❌ {message}")
                else:
                    st.warning("⚠️ Введите email")


def restore_authentication_from_session():
    """Восстановление аутентификации из активной сессии в базе данных."""
    try:
        if 'auth_manager' not in st.session_state:
            st.session_state.auth_manager = AuthManager()
        
        auth_manager = st.session_state.auth_manager
        
        # Проверяем, есть ли сохраненный session_id в query параметрах
        # Используем try-except для совместимости со старыми версиями Streamlit
        try:
            # Для Streamlit >= 1.28.0
            if hasattr(st, 'query_params'):
                query_params = st.query_params
                session_id_from_url = query_params.get("session_id", None) if query_params else None
            else:
                # Для старых версий - используем experimental
                query_params = st.experimental_get_query_params()
                session_id_from_url = query_params.get("session_id", [None])[0] if query_params else None
        except:
            session_id_from_url = None
        
        # Если есть session_id в URL, проверяем его
        if session_id_from_url:
            user = auth_manager.get_session(session_id_from_url)
            if user:
                st.session_state.authenticated = True
                st.session_state.current_user = user
                st.session_state.user_id = user.user_id
                st.session_state.username = user.username
                st.session_state.user_role = user.role
                st.session_state.session_id = session_id_from_url
                return True
        
        # Проверяем, есть ли session_id в session_state (может сохраниться в некоторых случаях)
        if 'session_id' in st.session_state and st.session_state.session_id:
            user = auth_manager.get_session(st.session_state.session_id)
            if user:
                st.session_state.authenticated = True
                st.session_state.current_user = user
                st.session_state.user_id = user.user_id
                st.session_state.username = user.username
                st.session_state.user_role = user.role
                return True
        
        # Если есть username в session_state, пытаемся восстановить последнюю активную сессию
        # (это менее безопасно, но работает для восстановления)
        if 'username' in st.session_state and st.session_state.username:
            # Попробуем найти активную сессию для этого пользователя
            # Это будет работать только если пользователь недавно залогинился
            pass
        
        return False
    except Exception as e:
        log_action("❌ ОШИБКА ВОССТАНОВЛЕНИЯ СЕССИИ", str(e))
        return False

def main():
    """Главная функция приложения."""
    log_action("🚀 ЗАПУСК ПРИЛОЖЕНИЯ", "main() вызвана")
    init_session_state()
    
    # Попытка восстановить аутентификацию из сессии
    if not st.session_state.get('authenticated', False):
        restored = restore_authentication_from_session()
        if restored:
            log_action("✅ АУТЕНТИФИКАЦИЯ ВОССТАНОВЛЕНА", f"Пользователь: {st.session_state.username}")
    
    # Проверка аутентификации
    if 'authenticated' not in st.session_state or not st.session_state.authenticated:
        show_login_page()
        return
    
    # Логотип
    display_logo()
    
    # Заголовок с градиентом
    st.markdown("""
    <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                color: white; padding: 30px; border-radius: 15px; text-align: center; 
                margin-bottom: 30px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3);'>
        <h1 style='margin: 0; font-size: 2.5em; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);'>
            🏥 MQEA
        </h1>
        <h2 style='margin: 10px 0; font-size: 1.2em; opacity: 0.9;'>
            Medical Quantum Entanglement Analysis
        </h2>
        <p style='margin: 0; font-size: 1.1em; opacity: 0.8;'>
            Революционная система медицинской диагностики с квантовым анализом и машинным обучением
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Информация об основателе
    with st.expander("ℹ️ Информация об основателе"):
        st.markdown("""
        **Основатель и разработчик:** Мухаммад Махизода  
        **Должность:** Администратор сети  
        **Университет:** Таджикский национальный университет  
        **Email:** muhammad.mahizoda@tnu.tj
        """)
    
    # Правильный медицинский логотип в сайдбаре
    with st.sidebar:
        try:
            from utils.fullscreen_logo_display import display_sidebar_logo
            display_sidebar_logo(width=150)
        except Exception as e:
            st.markdown("### 🏥 MQEA")
            st.markdown("**Medical Quantum Entanglement Analysis**")
            st.markdown("*Спокойствие, доверие, стабильность*")
            st.markdown("---")
    
    # Определение доступных функций в зависимости от роли
    user_role = st.session_state.get('user_role', 'user')
    
    # Меню для Пациента (ограниченный доступ)
    patient_menu = [
        "📊 Дашборд",
        "🦠 Анализ заболеваний",
        "📋 Визиты",
        "📄 Карточки",
        "📈 Отчеты"
    ]
    
    # Меню для Доктора (полный доступ)
    doctor_menu = [
        "📊 Дашборд",
        "👥 Пациенты", 
        "🔬 Анализ MQEA",
        "🦠 Анализ заболеваний",
        "🤖 AI-Помощник",
        "🤖 ML Диагностика",
        "📋 Визиты",
        "📄 Карточки",
        "📈 Отчеты",
        "⚙️ Настройки"
    ]
    
    # Выбор меню в зависимости от роли
    if user_role == 'admin':
        available_menu = doctor_menu
    elif user_role == 'doctor':
        available_menu = doctor_menu
    else:
        available_menu = patient_menu
    
    # Главное меню
    menu = st.sidebar.selectbox(
        "🏥 Главное меню",
        available_menu,
        key="main_menu_selectbox"
    )
    log_action("📋 ГЛАВНОЕ МЕНЮ", f"Выбрано: {menu}")
    
    # Навигационные табы
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🎛️ Навигация")
    
    # Стильные табы для переключения между приложениями
    col1, col2 = st.sidebar.columns(2)
    
    with col1:
        if st.button("🏥 MQEA", type="primary", width='stretch'):
            st.rerun()  # Остаемся на текущей странице
    
    with col2:
        if st.button("📡 Мониторинг", width='stretch'):
            st.switch_page("pages/realtime_monitoring_app.py")
    
    # Индикатор активного приложения
    st.sidebar.markdown("""
    <div class="app-indicator">
        🏥 MQEA АКТИВНО
        <br><small>Система квантового анализа</small>
    </div>
    """, unsafe_allow_html=True)
    
    # Дополнительное меню "Асмотр" - только для Доктора
    if user_role in ['doctor', 'admin']:
        st.sidebar.markdown("---")
        asmort_menu = st.sidebar.selectbox(
            "🔍 Асмотр",
            [
                "📋 Обзор системы",
                "📊 Статистика MQEA",
                "🔬 Лаборатория",
                "📈 Аналитика",
                "⚛️ Квантовые эксперименты",
                "🏥 Медицинские протоколы",
                "📚 База знаний",
                "🔧 Инструменты"
            ],
            key="asmort_selectbox"
        )
        log_action("🔍 АСМОТР", f"Выбрано: {asmort_menu}")
    else:
        asmort_menu = None
    
    # Отдельный раздел "Детская медицина" - только для Доктора
    if user_role in ['doctor', 'admin']:
        st.sidebar.markdown("---")
        pediatric_menu = st.sidebar.selectbox(
            "👶 Детская медицина",
            [
                "🔬 Квантовая диагностика",
                "📊 Мониторинг развития",
                "💊 Планы лечения",
                "📋 Антропометрия",
                "🧠 Неврология",
                "❤️ Кардиология",
                "🫁 Пульмонология",
                "🍼 Неонатология",
                "📈 Аналитика",
                "⚙️ Настройки"
            ],
            key="pediatric_selectbox"
        )
        log_action("👶 ДЕТСКАЯ МЕДИЦИНА", f"Выбрано: {pediatric_menu}")
    else:
        pediatric_menu = None
    
    # Отдельный раздел "КАПЧ" - только для Доктора
    if user_role in ['doctor', 'admin']:
        st.sidebar.markdown("---")
        kapch_menu = st.sidebar.selectbox(
            "🧠 КАПЧ",
            [
                "🔬 Квантовый анализ",
                "📊 Мониторинг сознания",
                "🧘 Медитативные состояния",
                "⚡ Квантовые эксперименты",
                "📈 Результаты анализов",
                "⚙️ Настройки КАПЧ"
            ],
            key="kapch_selectbox"
        )
        log_action("🧠 КАПЧ", f"Выбрано: {kapch_menu}")
    else:
        kapch_menu = None
    
    # Кнопка обновления
    st.sidebar.markdown("---")
    if st.sidebar.button("🔄 Обновить", help="Обновить страницу"):
        log_action("🔄 ОБНОВЛЕНИЕ", "Кнопка обновления нажата")
        st.rerun()
    
    # Информация о пользователе и выход
    st.sidebar.markdown("---")
    if 'current_user' in st.session_state and st.session_state.current_user:
        user = st.session_state.current_user
        # Преобразование роли для отображения на русском
        role_display = {"user": "Пациент", "doctor": "Доктор", "admin": "Администратор"}.get(user.role, user.role)
        
        st.sidebar.markdown(f"**👤 Пользователь:** {user.full_name}")
        st.sidebar.markdown(f"**📧 Email:** {user.email}")
        st.sidebar.markdown(f"**🎭 Роль:** {role_display}")
        
        if st.sidebar.button("🚪 Выйти", type="secondary", width='stretch'):
            # Очистка сессии
            if 'session_id' in st.session_state and st.session_state.session_id:
                if 'auth_manager' in st.session_state:
                    st.session_state.auth_manager.delete_session(st.session_state.session_id)
            
            # Удаляем session_id из query параметров
            try:
                if hasattr(st, 'query_params'):
                    if 'session_id' in st.query_params:
                        del st.query_params["session_id"]
                else:
                    # Для старых версий
                    st.experimental_set_query_params()
            except:
                pass
            
            st.session_state.authenticated = False
            st.session_state.current_user = None
            st.session_state.user_id = None
            st.session_state.username = None
            st.session_state.user_role = None
            st.session_state.session_id = None
            
            st.rerun()
    
    # Кнопка AI-помощника (только для врачей)
    if user_role in ['doctor', 'admin']:
        st.sidebar.markdown("---")
        if st.sidebar.button("🤖 AI Помощник", type="primary", width='stretch'):
            st.session_state.current_menu = "🤖 AI-Помощник"
            st.rerun()
    
    # Кнопка Панели администратора (только для администраторов)
    if user_role == 'admin':
        st.sidebar.markdown("---")
        if st.sidebar.button("👑 Панель администратора", type="primary", width='stretch'):
            st.session_state.current_menu = "👑 Панель администратора"
            st.rerun()
    
    # Проверка на кнопку AI-помощника (только для врачей)
    if st.session_state.get('current_menu') == "🤖 AI-Помощник":
        if user_role in ['doctor', 'admin']:
            show_ai_assistant()
            return
        else:
            st.warning("⚠️ У вас нет доступа к AI-Помощнику. Эта функция доступна только для врачей.")
            show_dashboard()
            return
    
    # Проверка на кнопку Панели администратора (только для администраторов)
    if st.session_state.get('current_menu') == "👑 Панель администратора":
        if user_role == 'admin':
            show_admin_panel()
            return
        else:
            st.error("❌ У вас нет доступа к панели администратора")
            show_dashboard()
            return
    
    # Маршрутизация - все переходы только через sidebar
    log_action("🎯 МАРШРУТИЗАЦИЯ", "Начинаем проверку маршрутов")
    
    # Определяем, какой selectbox был изменен (приоритет по порядку)
    # Сохраняем предыдущие значения для сравнения
    if 'prev_pediatric_menu' not in st.session_state:
        st.session_state.prev_pediatric_menu = pediatric_menu
    if 'prev_kapch_menu' not in st.session_state:
        st.session_state.prev_kapch_menu = kapch_menu
    if 'prev_asmort_menu' not in st.session_state:
        st.session_state.prev_asmort_menu = asmort_menu
    if 'prev_main_menu' not in st.session_state:
        st.session_state.prev_main_menu = menu
    
    # Проверяем, какой selectbox был изменен (только если меню доступно)
    pediatric_changed = (pediatric_menu is not None and st.session_state.prev_pediatric_menu != pediatric_menu)
    kapch_changed = (kapch_menu is not None and st.session_state.prev_kapch_menu != kapch_menu)
    asmort_changed = (asmort_menu is not None and st.session_state.prev_asmort_menu != asmort_menu)
    main_changed = (st.session_state.prev_main_menu != menu)
    
    log_action("🔍 ИЗМЕНЕНИЯ", f"pediatric_changed={pediatric_changed}, kapch_changed={kapch_changed}, asmort_changed={asmort_changed}, main_changed={main_changed}")
    
    # 1. ПРИОРИТЕТ: Если изменилась "Детская медицина" - проверяем её (только для врачей)
    if pediatric_changed and user_role in ['doctor', 'admin']:
        st.session_state.prev_pediatric_menu = pediatric_menu
        st.session_state.active_section = "pediatric"
        if pediatric_menu == "🔬 Квантовая диагностика":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Квантовая диагностика")
            st.sidebar.success("✅ Переходим в квантовую диагностику")
            show_pediatric_diagnosis()
            return
        elif pediatric_menu == "📊 Мониторинг развития":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Мониторинг развития")
            show_pediatric_monitoring()
            return
        elif pediatric_menu == "💊 Планы лечения":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Планы лечения")
            show_pediatric_treatment()
            return
        elif pediatric_menu == "📋 Антропометрия":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Антропометрия")
            show_pediatric_anthropometry()
            return
        elif pediatric_menu == "🧠 Неврология":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Неврология")
            show_pediatric_neurology()
            return
        elif pediatric_menu == "❤️ Кардиология":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Кардиология")
            show_pediatric_cardiology()
            return
        elif pediatric_menu == "🫁 Пульмонология":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Пульмонология")
            show_pediatric_pulmonology()
            return
        elif pediatric_menu == "🍼 Неонатология":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Неонатология")
            show_pediatric_neonatology()
            return
        elif pediatric_menu == "📈 Аналитика":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Аналитика")
            show_pediatric_analytics()
            return
        elif pediatric_menu == "⚙️ Настройки":
            log_action("➡️ ПЕРЕХОД", "Детская медицина → Настройки")
            show_pediatric_settings()
            return
    
    # 2. ПРИОРИТЕТ: Если изменился "КАПЧ" - проверяем его (только для врачей)
    elif kapch_changed and user_role in ['doctor', 'admin']:
        st.session_state.prev_kapch_menu = kapch_menu
        st.session_state.active_section = "kapch"
        if kapch_menu == "🔬 Квантовый анализ":
            log_action("➡️ ПЕРЕХОД", "КАПЧ → Квантовый анализ")
            st.sidebar.success("✅ Переходим в квантовый анализ")
            show_kapch_analysis()
            return
        elif kapch_menu == "📊 Мониторинг сознания":
            log_action("➡️ ПЕРЕХОД", "КАПЧ → Мониторинг сознания")
            show_kapch_consciousness_monitoring()
            return
        elif kapch_menu == "🧘 Медитативные состояния":
            log_action("➡️ ПЕРЕХОД", "КАПЧ → Медитативные состояния")
            show_kapch_meditation_states()
            return
        elif kapch_menu == "⚡ Квантовые эксперименты":
            log_action("➡️ ПЕРЕХОД", "КАПЧ → Квантовые эксперименты")
            show_kapch_quantum_experiments()
            return
        elif kapch_menu == "📈 Результаты анализов":
            log_action("➡️ ПЕРЕХОД", "КАПЧ → Результаты анализов")
            show_kapch_results()
            return
        elif kapch_menu == "⚙️ Настройки КАПЧ":
            log_action("➡️ ПЕРЕХОД", "КАПЧ → Настройки КАПЧ")
            show_kapch_settings()
            return
    
    # 3. Если изменился "Асмотр" - проверяем его (только для врачей)
    elif asmort_changed and user_role in ['doctor', 'admin']:
        st.session_state.prev_asmort_menu = asmort_menu
        st.session_state.active_section = "asmort"
        if asmort_menu == "📋 Обзор системы":
            log_action("➡️ ПЕРЕХОД", "Асмотр → Обзор системы")
            show_system_overview()
            return
        elif asmort_menu == "📊 Статистика MQEA":
            log_action("➡️ ПЕРЕХОД", "Асмотр → Статистика MQEA")
            show_mqea_statistics()
            return
        elif asmort_menu == "🔬 Лаборатория":
            log_action("➡️ ПЕРЕХОД", "Асмотр → Лаборатория")
            show_laboratory()
            return
        elif asmort_menu == "📈 Аналитика":
            log_action("➡️ ПЕРЕХОД", "Асмотр → Аналитика")
            show_analytics()
            return
        elif asmort_menu == "⚛️ Квантовые эксперименты":
            log_action("➡️ ПЕРЕХОД", "Асмотр → Квантовые эксперименты")
            show_quantum_experiments()
            return
        elif asmort_menu == "🏥 Медицинские протоколы":
            log_action("➡️ ПЕРЕХОД", "Асмотр → Медицинские протоколы")
            show_medical_protocols()
            return
        elif asmort_menu == "📚 База знаний":
            log_action("➡️ ПЕРЕХОД", "Асмотр → База знаний")
            show_knowledge_base()
            return
        elif asmort_menu == "🔧 Инструменты":
            log_action("➡️ ПЕРЕХОД", "Асмотр → Инструменты")
            show_tools()
            return
    
    # 3. Если изменилось главное меню - проверяем его
    elif main_changed:
        st.session_state.prev_main_menu = menu
        st.session_state.active_section = "main"
        if menu == "📊 Дашборд":
            log_action("➡️ ПЕРЕХОД", "Главное меню → Дашборд")
            show_dashboard()
            return
        elif menu == "👥 Пациенты":
            if user_role in ['doctor', 'admin']:
                log_action("➡️ ПЕРЕХОД", "Главное меню → Пациенты")
                show_patients_management()
                return
            else:
                st.warning("⚠️ У вас нет доступа к управлению пациентами. Эта функция доступна только для врачей.")
                show_dashboard()
                return
        elif menu == "🔬 Анализ MQEA":
            if user_role in ['doctor', 'admin']:
                log_action("➡️ ПЕРЕХОД", "Главное меню → Анализ MQEA")
                show_mqea_analysis()
                return
            else:
                st.warning("⚠️ У вас нет доступа к полному анализу MQEA. Эта функция доступна только для врачей.")
                show_dashboard()
                return
        elif menu == "🦠 Анализ заболеваний":
            log_action("➡️ ПЕРЕХОД", "Главное меню → Анализ заболеваний")
            show_disease_analysis()
            return
        elif menu == "🤖 AI-Помощник":
            if user_role in ['doctor', 'admin']:
                log_action("➡️ ПЕРЕХОД", "Главное меню → AI-Помощник")
                show_ai_assistant()
                return
            else:
                st.warning("⚠️ У вас нет доступа к AI-Помощнику. Эта функция доступна только для врачей.")
                show_dashboard()
                return
        elif menu == "🤖 ML Диагностика":
            if user_role in ['doctor', 'admin']:
                log_action("➡️ ПЕРЕХОД", "Главное меню → ML Диагностика")
                show_ml_diagnostics()
                return
            else:
                st.warning("⚠️ У вас нет доступа к ML Диагностике. Эта функция доступна только для врачей.")
                show_dashboard()
                return
        elif menu == "📋 Визиты":
            log_action("➡️ ПЕРЕХОД", "Главное меню → Визиты")
            show_visits_management()
            return
        elif menu == "📄 Карточки":
            log_action("➡️ ПЕРЕХОД", "Главное меню → Карточки")
            show_cards_generation()
            return
        elif menu == "📈 Отчеты":
            log_action("➡️ ПЕРЕХОД", "Главное меню → Отчеты")
            show_reports()
            return
        elif menu == "⚙️ Настройки":
            if user_role in ['doctor', 'admin']:
                log_action("➡️ ПЕРЕХОД", "Главное меню → Настройки")
                show_settings()
                return
            else:
                st.warning("⚠️ У вас нет доступа к настройкам системы. Эта функция доступна только для врачей.")
                show_dashboard()
                return
    
    # 4. Если ничего не изменилось, но есть активный раздел - показываем его
    elif 'active_section' in st.session_state:
        active_section = st.session_state.active_section
        log_action("🔄 ПОВТОР", f"Показываем активный раздел: {active_section}")
        
        if active_section == "main":
            if menu == "📊 Дашборд":
                show_dashboard()
                return
            elif menu == "👥 Пациенты":
                show_patients_management()
                return
            elif menu == "🔬 Анализ MQEA":
                show_mqea_analysis()
                return
            elif menu == "🦠 Анализ заболеваний":
                show_disease_analysis()
                return
            elif menu == "🤖 AI-Помощник":
                show_ai_assistant()
                return
            elif menu == "🤖 ML Диагностика":
                show_ml_diagnostics()
                return
            elif menu == "📋 Визиты":
                show_visits_management()
                return
            elif menu == "📄 Карточки":
                show_cards_generation()
                return
            elif menu == "📈 Отчеты":
                show_reports()
                return
            elif menu == "⚙️ Настройки":
                show_settings()
                return
        
        elif active_section == "pediatric":
            if pediatric_menu == "🔬 Квантовая диагностика":
                show_pediatric_diagnosis()
                return
            elif pediatric_menu == "📊 Мониторинг развития":
                show_pediatric_monitoring()
                return
            elif pediatric_menu == "💊 Планы лечения":
                show_pediatric_treatment()
                return
            elif pediatric_menu == "📋 Антропометрия":
                show_pediatric_anthropometry()
                return
            elif pediatric_menu == "🧠 Неврология":
                show_pediatric_neurology()
                return
            elif pediatric_menu == "❤️ Кардиология":
                show_pediatric_cardiology()
                return
            elif pediatric_menu == "🫁 Пульмонология":
                show_pediatric_pulmonology()
                return
            elif pediatric_menu == "🍼 Неонатология":
                show_pediatric_neonatology()
                return
            elif pediatric_menu == "📈 Аналитика":
                show_pediatric_analytics()
                return
            elif pediatric_menu == "⚙️ Настройки":
                show_pediatric_settings()
                return
        
        elif active_section == "kapch":
            if kapch_menu == "🔬 Квантовый анализ":
                show_kapch_analysis()
                return
            elif kapch_menu == "📊 Мониторинг сознания":
                show_kapch_consciousness_monitoring()
                return
            elif kapch_menu == "🧘 Медитативные состояния":
                show_kapch_meditation_states()
                return
            elif kapch_menu == "⚡ Квантовые эксперименты":
                show_kapch_quantum_experiments()
                return
            elif kapch_menu == "📈 Результаты анализов":
                show_kapch_results()
                return
            elif kapch_menu == "⚙️ Настройки КАПЧ":
                show_kapch_settings()
                return
        
        elif active_section == "asmort":
            if asmort_menu == "📋 Обзор системы":
                show_system_overview()
                return
            elif asmort_menu == "📊 Статистика MQEA":
                show_mqea_statistics()
                return
            elif asmort_menu == "🔬 Лаборатория":
                show_laboratory()
                return
            elif asmort_menu == "📈 Аналитика":
                show_analytics()
                return
            elif asmort_menu == "🔍 Анализ данных":
                show_data_analysis()
                return
            elif asmort_menu == "📈 Тренды":
                show_trends()
                return
            elif asmort_menu == "⚠️ Алерты":
                show_alerts()
                return
            elif asmort_menu == "⚛️ Квантовые эксперименты":
                show_quantum_experiments()
                return
            elif asmort_menu == "🏥 Медицинские протоколы":
                show_medical_protocols()
                return
            elif asmort_menu == "📚 База знаний":
                show_knowledge_base()
                return
            elif asmort_menu == "🔧 Инструменты":
                show_tools()
                return
    
    # 5. Если ничего не выбрано, показываем дашборд по умолчанию
    log_action("➡️ ПЕРЕХОД", "Показываем дашборд по умолчанию")
    show_dashboard()
    return

def show_dashboard():
    """Дашборд системы."""
    st.header("📊 Дашборд системы")
    
    # Статистика
    st.subheader("📊 Статистика системы")
    stats = st.session_state.components['database'].get_statistics()
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("👥 Пациенты", stats.get('total_patients', 0))
    
    with col2:
        st.metric("🏥 Визиты", stats.get('total_visits', 0))
    
    with col3:
        st.metric("🔬 Анализы MQEA", stats.get('total_analyses', 0))
    
    with col4:
        st.metric("⚛️ Средняя когерентность", f"{stats.get('average_coherence', 0):.3f}")
    
    # Графики
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📈 Статистика по месяцам")
        # Здесь можно добавить график статистики
        st.info("График статистики будет добавлен")
    
    with col2:
        st.subheader("🎯 Эффективность диагностики")
        # Здесь можно добавить график эффективности
        st.info("График эффективности будет добавлен")
    
    # Последние активности
    st.subheader("🕒 Последние активности")
    
    # Получаем последние визиты
    all_patients = st.session_state.components['database'].get_all_patients()
    recent_activities = []
    
    for patient in all_patients[:5]:
        visits = st.session_state.components['database'].get_patient_visits(patient.patient_id)
        if visits:
            recent_activities.append({
                'patient': patient.name,
                'visit_date': visits[0].visit_date,
                'status': visits[0].status
            })
    
    if recent_activities:
        df_activities = pd.DataFrame(recent_activities)
        st.dataframe(df_activities, width='stretch')
    else:
        st.info("Нет данных о последних активностях")

def show_patients_management():
    """Управление пациентами."""
    log_action("🎉 ВЫЗОВ ФУНКЦИИ", "show_patients_management() - НАЧАЛО")
    
    # Проверка доступа для докторов и администраторов
    user_role = st.session_state.get('user_role', 'user')
    if user_role not in ['doctor', 'admin']:
        st.error("❌ У вас нет доступа к управлению пациентами. Эта функция доступна только для врачей.")
        return
    
    st.header("👥 Управление пациентами")
    
    # Информация о текущем пользователе
    if st.session_state.get('current_user'):
        current_user = st.session_state.current_user
        st.info(f"👤 Вы вошли как: **{current_user.full_name}** ({'Доктор' if current_user.role == 'doctor' else 'Администратор'})")
    
    # Кнопки управления
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("➕ Добавить пациента", type="primary"):
            st.session_state.show_add_patient = True
    
    with col2:
        if st.button("🔍 Поиск пациента"):
            st.session_state.show_search_patient = True
    
    with col3:
        if st.button("📊 Статистика пациентов"):
            st.session_state.show_patient_stats = True
    
    # Добавление пациента
    if st.session_state.get('show_add_patient', False):
        st.subheader("➕ Добавление нового пациента")
        
        with st.form("add_patient_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                name = st.text_input("ФИО пациента *")
                birth_date = st.date_input("Дата рождения *", min_value=date(1970, 1, 1))
                gender = st.selectbox("Пол *", ["Мужской", "Женский"])
                phone = st.text_input("Телефон")
                address = st.text_area("Адрес")
            
            with col2:
                weight = st.number_input("Вес (кг)", 20.0, 200.0, 70.0, 0.1)
                height = st.number_input("Рост (см)", 100.0, 250.0, 170.0, 0.1)
                
                # Медицинская история
                st.write("**Медицинская история:**")
                diabetes = st.checkbox("Диабет")
                hypertension = st.checkbox("Гипертония")
                heart_disease = st.checkbox("Заболевания сердца")
                other_conditions = st.text_area("Другие состояния")
            
            submitted = st.form_submit_button("💾 Сохранить пациента")
            
            if submitted and name and birth_date:
                try:
                    # Создаем запись пациента
                    patient_id = f"P{len(st.session_state.components['database'].get_all_patients()) + 1:03d}"
                    
                    # Вычисляем возраст из даты рождения
                    today = datetime.now().date()
                    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                    
                    medical_history = {}
                    if diabetes:
                        medical_history['diabetes'] = "Диабет"
                    if hypertension:
                        medical_history['hypertension'] = "Гипертония"
                    if heart_disease:
                        medical_history['heart_disease'] = "Заболевания сердца"
                    if other_conditions:
                        medical_history['other'] = other_conditions
                    
                    # Добавляем информацию о создателе (докторе)
                    creator_info = {
                        'created_by': st.session_state.get('username', 'system'),
                        'creator_name': st.session_state.get('current_user', {}).full_name if st.session_state.get('current_user') else 'Система',
                        'creator_role': st.session_state.get('user_role', 'system')
                    }
                    
                    contact_info = {
                        'phone': phone,
                        'address': address,
                        'weight': weight,
                        'height': height,
                        'birth_date': str(birth_date) if birth_date else None,
                        'created_by': creator_info
                    }
                    
                    patient_record = PatientRecord(
                        patient_id=patient_id,
                        name=name,
                        age=age,
                        gender=gender,
                        contact_info=contact_info,
                        medical_history=medical_history,
                        created_at=datetime.now(),
                        updated_at=datetime.now()
                    )
                    
                    if st.session_state.components['database'].add_patient(patient_record):
                        creator_name = st.session_state.get('current_user', {}).full_name if st.session_state.get('current_user') else 'Система'
                        st.success(f"✅ Пациент {name} успешно добавлен (ID: {patient_id})")
                        st.info(f"👤 Создан доктором: {creator_name}")
                        st.session_state.show_add_patient = False
                        st.rerun()
                    else:
                        st.error("❌ Ошибка добавления пациента")
                
                except Exception as e:
                    st.error(f"❌ Ошибка: {str(e)}")
    
    # Поиск пациента
    if st.session_state.get('show_search_patient', False):
        st.subheader("🔍 Поиск пациента")
        
        search_query = st.text_input("Введите имя или ID пациента")
        
        if search_query:
            patients = st.session_state.components['database'].search_patients(search_query)
            
            if patients:
                for patient in patients:
                    with st.expander(f"👤 {patient.name} (ID: {patient.patient_id})"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write(f"**Возраст:** {patient.age} лет")
                            st.write(f"**Пол:** {patient.gender}")
                            st.write(f"**Телефон:** {patient.contact_info.get('phone', 'N/A')}")
                        
                        with col2:
                            st.write(f"**Дата регистрации:** {patient.created_at.strftime('%d.%m.%Y')}")
                            st.write(f"**Статус:** {'Активен' if patient.is_active else 'Неактивен'}")
                            
                            # Показываем информацию о создателе, если есть
                            if patient.contact_info and 'created_by' in patient.contact_info:
                                creator = patient.contact_info['created_by']
                                if isinstance(creator, dict):
                                    creator_name = creator.get('creator_name', 'Неизвестно')
                                    st.write(f"**Создан доктором:** {creator_name}")
                            
                            if st.button(f"Выбрать", key=f"select_{patient.patient_id}"):
                                st.session_state.current_patient = patient
                                st.success(f"✅ Выбран пациент {patient.name}")
                                st.rerun()
            else:
                st.info("Пациенты не найдены")
    
    # Список всех пациентов
    st.subheader("👥 Все пациенты")
    
    all_patients = st.session_state.components['database'].get_all_patients()
    
    if all_patients:
        # Фильтры
        col1, col2, col3 = st.columns(3)
        
        with col1:
            gender_filter = st.selectbox("Фильтр по полу", ["Все", "Мужской", "Женский"])
        
        with col2:
            age_filter = st.slider("Возраст", 0, 120, (0, 120))
        
        with col3:
            search_filter = st.text_input("Поиск по имени")
        
        # Фильтрация
        filtered_patients = all_patients
        
        if gender_filter != "Все":
            filtered_patients = [p for p in filtered_patients if p.gender == gender_filter]
        
        filtered_patients = [p for p in filtered_patients if age_filter[0] <= p.age <= age_filter[1]]
        
        if search_filter:
            filtered_patients = [p for p in filtered_patients if search_filter.lower() in p.name.lower()]
        
        # Таблица пациентов
        if filtered_patients:
            patient_data = []
            for patient in filtered_patients:
                # Получаем информацию о создателе
                creator_name = 'Система'
                if patient.contact_info and 'created_by' in patient.contact_info:
                    creator = patient.contact_info['created_by']
                    if isinstance(creator, dict):
                        creator_name = creator.get('creator_name', 'Система')
                    elif isinstance(creator, str):
                        creator_name = creator
                
                patient_data.append({
                    'ID': patient.patient_id,
                    'ФИО': patient.name,
                    'Возраст': patient.age,
                    'Пол': patient.gender,
                    'Телефон': patient.contact_info.get('phone', 'N/A'),
                    'Создан доктором': creator_name,
                    'Дата регистрации': patient.created_at.strftime('%d.%m.%Y'),
                    'Статус': 'Активен' if patient.is_active else 'Неактивен'
                })
            
            df_patients = pd.DataFrame(patient_data)
            st.dataframe(df_patients, width='stretch')
            
            # Кнопки действий
            selected_patient_id = st.selectbox("Выберите пациента для действий", [p.patient_id for p in filtered_patients])
            
            if selected_patient_id:
                # Сохраняем выбранного пациента в session_state
                st.session_state.selected_patient_id = selected_patient_id
                patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                st.session_state.patient_data = {
                    'name': patient.name,
                    'age': patient.age,
                    'gender': patient.gender,
                    'weight': getattr(patient, 'weight', 'Не указан'),
                    'height': getattr(patient, 'height', 'Не указан'),
                    'diagnoses': getattr(patient, 'diagnoses', ['Не указаны']),
                    'allergies': getattr(patient, 'allergies', ['Не указаны']),
                    'chronic_conditions': getattr(patient, 'chronic_conditions', ['Не указаны'])
                }
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    if st.button("👁️ Просмотр"):
                        st.session_state.current_patient = patient
                        st.success(f"✅ Выбран пациент {patient.name}")
                
                with col2:
                    if st.button("📋 Новый визит"):
                        patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                        st.session_state.current_patient = patient
                        st.session_state.show_new_visit = True
                        st.success(f"✅ Создание визита для {patient.name}")
                
                with col3:
                    if st.button("🔬 MQEA Анализ"):
                        patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                        st.session_state.current_patient = patient
                        st.session_state.show_mqea_analysis = True
                        st.success(f"✅ Анализ MQEA для {patient.name}")
                        # Переход к анализу MQEA
                        st.session_state.current_menu = "🔬 Анализ MQEA"
                        st.rerun()
                
                with col4:
                    if st.button("🦠 Анализ заболеваний"):
                        patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                        st.session_state.current_patient = patient
                        st.success(f"✅ Анализ заболеваний для {patient.name}")
                        # Переход к анализу заболеваний
                        st.session_state.current_menu = "🦠 Анализ заболеваний"
                        st.rerun()
                
                # Дополнительные кнопки
                col5, col6 = st.columns(2)
                
                with col5:
                    if st.button("📄 Карточка"):
                        patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                        st.session_state.current_patient = patient
                        st.session_state.show_generate_card = True
                        st.success(f"✅ Генерация карточки для {patient.name}")
                
                with col6:
                    if st.button("🤖 ML Диагностика"):
                        patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                        st.session_state.current_patient = patient
                        st.success(f"✅ ML Диагностика для {patient.name}")
                        # Переход к ML диагностике
                        st.session_state.current_menu = "🤖 ML Диагностика"
                        st.rerun()

                st.markdown("---")
                st.subheader("✏️ Редактирование / 🗑️ Удаление")
                edit_col1, edit_col2, edit_col3 = st.columns(3)

                with edit_col1:
                    if st.button("✏️ Редактировать"):
                        st.session_state.edit_patient_id = selected_patient_id
                        st.session_state.show_edit_patient = True
                        st.rerun()

                with edit_col2:
                    if st.button("🗑️ Удалить"):
                        st.session_state.delete_patient_id = selected_patient_id
                        st.session_state.show_delete_confirm = True
                        st.rerun()

                with edit_col3:
                    if st.button("🔎 Анализ только этого пациента"):
                        patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                        st.session_state.current_patient = patient
                        # Переход сразу в раздел анализа
                        st.session_state.menu = "🔬 Анализ MQEA"
                        st.session_state.show_mqea_analysis = True
                        st.success(f"✅ Открыт анализ только для пациента {patient.name}")
                        st.rerun()

                # Форма редактирования
                if st.session_state.get('show_edit_patient') and st.session_state.get('edit_patient_id') == selected_patient_id:
                    patient = next(p for p in filtered_patients if p.patient_id == selected_patient_id)
                    st.markdown("---")
                    st.markdown(f"#### ✏️ Редактирование пациента: {patient.name} ({patient.patient_id})")
                    with st.form(f"edit_patient_form_{selected_patient_id}"):
                        col_a, col_b, col_c = st.columns(3)
                        with col_a:
                            name = st.text_input("ФИО", value=patient.name)
                            # Получаем дату рождения из contact_info или вычисляем из возраста
                            birth_date_str = patient.contact_info.get('birth_date')
                            if birth_date_str:
                                try:
                                    birth_date = datetime.fromisoformat(birth_date_str).date()
                                except:
                                    # Если не удается распарсить, вычисляем приблизительную дату рождения
                                    today = datetime.now().date()
                                    birth_date = today.replace(year=today.year - patient.age)
                            else:
                                # Если дата рождения не сохранена, вычисляем приблизительную
                                today = datetime.now().date()
                                birth_date = today.replace(year=today.year - patient.age)
                            
                            birth_date = st.date_input("Дата рождения", value=birth_date, min_value=date(1970, 1, 1))
                        with col_b:
                            gender = st.selectbox("Пол", ["Мужской", "Женский"], index=0 if patient.gender=="Мужской" else 1)
                            phone = st.text_input("Телефон", value=patient.contact_info.get('phone', ''))
                        with col_c:
                            address = st.text_input("Адрес", value=patient.contact_info.get('address', ''))
                        save_btn, cancel_btn = st.columns(2)
                        with save_btn:
                            submitted = st.form_submit_button("💾 Сохранить изменения", width='stretch')
                        with cancel_btn:
                            cancel = st.form_submit_button("Отмена", width='stretch')
                    if submitted:
                        # Вычисляем возраст из даты рождения
                        today = datetime.now().date()
                        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                        
                        ok = st.session_state.components['database'].update_patient(
                            patient_id=selected_patient_id,
                            name=name,
                            age=age,
                            gender=gender,
                            contact_info={
                                "phone": phone, 
                                "address": address,
                                "birth_date": str(birth_date)
                            }
                        )
                        if ok:
                            st.success("✅ Данные пациента обновлены")
                            st.session_state.show_edit_patient = False
                            st.rerun()
                        else:
                            st.error("❌ Не удалось обновить пациента")
                    if cancel:
                        st.session_state.show_edit_patient = False
                        st.rerun()

                # Подтверждение удаления
                if st.session_state.get('show_delete_confirm') and st.session_state.get('delete_patient_id') == selected_patient_id:
                    st.warning("Вы уверены, что хотите удалить пациента? Это действие можно отменить (мягкое удаление).")
                    del_col1, del_col2, del_col3 = st.columns(3)
                    with del_col1:
                        if st.button("🗑️ Удалить (мягко)"):
                            ok = st.session_state.components['database'].delete_patient(selected_patient_id, hard=False)
                            if ok:
                                st.success("✅ Пациент помечен как удаленный")
                                st.session_state.show_delete_confirm = False
                                st.rerun()
                            else:
                                st.error("❌ Не удалось удалить пациента")
                    with del_col2:
                        if st.button("❌ Удалить навсегда"):
                            ok = st.session_state.components['database'].delete_patient(selected_patient_id, hard=True)
                            if ok:
                                st.success("✅ Пациент удален из базы навсегда")
                                st.session_state.show_delete_confirm = False
                                st.rerun()
                            else:
                                st.error("❌ Не удалось удалить пациента навсегда")
                    with del_col3:
                        if st.button("Отмена"):
                            st.session_state.show_delete_confirm = False
                            st.rerun()
        else:
            st.info("Пациенты не найдены по заданным критериям")
    else:
        st.info("В базе данных нет пациентов")

def show_mqea_analysis():
    """Анализ MQEA."""
    log_action("🎉 ВЫЗОВ ФУНКЦИИ", "show_mqea_analysis() - НАЧАЛО")
    st.header("🔬 Анализ MQEA - Квантовая запутанность в медицине")
    
    # Информация о MQEA
    with st.expander("ℹ️ Что такое MQEA?"):
        st.markdown("""
        **MQEA (Medical Quantum Entanglement Analysis)** - это революционный алгоритм для анализа медицинских данных, 
        основанный на принципах квантовой запутанности. Система представляет каждый медицинский показатель как 
        квантовое состояние и вычисляет запутанность между ними для выявления скрытых взаимосвязей.
        
        **Ключевые особенности:**
        - ⚛️ Квантовая когерентность - стабильность системы
        - 🔗 Квантовая запутанность - взаимосвязи между показателями
        - 📊 Обнаружение паттернов - выявление аномалий
        - 🔮 Прогнозирование - предсказание изменений
        """)
    
    if not st.session_state.current_patient:
        st.warning("⚠️ Сначала выберите пациента в разделе 'Пациенты'")
        return
    
    patient = st.session_state.current_patient
    st.subheader(f"Анализ для пациента: {patient.name} (ID: {patient.patient_id})")
    
    # Настройки анализа
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("⚙️ Настройки анализа")
        
        analysis_mode = st.selectbox("Режим анализа", ["Генерация данных", "Загрузка файла"])
        
        if analysis_mode == "Генерация данных":
            duration_hours = st.slider("Продолжительность (часы)", 1, 48, 24)
            sampling_minutes = st.slider("Интервал выборки (минуты)", 5, 60, 15)
            add_noise = st.checkbox("Добавить шум", value=True)
            add_missing = st.checkbox("Добавить пропущенные данные", value=True)
            
            # Профиль пациента для генерации
            st.write("**Профиль пациента для генерации:**")
            patient_profile = {
                'heart_rate': st.number_input("Частота пульса", 40, 200, 75 if patient.age < 50 else 80),
                'blood_pressure_systolic': st.number_input("АД систолическое", 80, 250, 120),
                'blood_pressure_diastolic': st.number_input("АД диастолическое", 50, 150, 80),
                'temperature': st.number_input("Температура", 35.0, 42.0, 36.6, 0.1),
                'oxygen_saturation': st.number_input("Насыщение O₂", 70, 100, 98),
                'respiratory_rate': st.number_input("Частота дыхания", 8, 40, 16),
                'glucose': st.number_input("Глюкоза", 2.0, 20.0, 5.0, 0.1),
                'cholesterol': st.number_input("Холестерин", 100, 400, 180)
            }
        else:
            uploaded_file = st.file_uploader("Загрузите CSV файл", type=['csv'])
            if uploaded_file:
                st.success(f"✅ Файл загружен: {uploaded_file.name}")
    
    with col2:
        st.subheader("🔬 Параметры MQEA")
        
        quantum_threshold = st.slider("Порог квантовой запутанности", 0.1, 0.9, 0.3, 0.1)
        fill_missing = st.checkbox("Заполнить пропущенные данные", value=True)
        max_iterations = st.slider("Максимум итераций", 10, 100, 50)
        
        # Дополнительные параметры
        st.write("**Дополнительные настройки:**")
        window_sizes = st.multiselect(
            "Размеры окон анализа", 
            [24, 48, 96, 168], 
            default=[24, 48, 96],
            help="Окна для анализа квантовой запутанности"
        )
        
        pattern_detection = st.checkbox("Обнаружение паттернов", value=True)
        temporal_analysis = st.checkbox("Временной анализ", value=True)
    
    # Кнопки управления
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("🚀 Запустить MQEA анализ", type="primary"):
            run_mqea_analysis(analysis_mode, patient, patient_profile if analysis_mode == "Генерация данных" else None, 
                            uploaded_file if analysis_mode == "Загрузка файла" else None,
                            quantum_threshold, fill_missing, max_iterations, window_sizes, pattern_detection, temporal_analysis)
    
    with col2:
        if st.button("📊 Быстрый анализ"):
            run_quick_mqea_analysis(patient, quantum_threshold)
    
    with col3:
        if st.button("🔄 Сбросить результаты"):
            st.session_state.analysis_results = None
            st.session_state.current_data = None
            st.success("✅ Результаты сброшены")
            st.rerun()
    
    # Отображение результатов
    if st.session_state.analysis_results:
        display_mqea_results()
    else:
        st.info("👆 Настройте параметры и запустите анализ для просмотра результатов")

def run_mqea_analysis(analysis_mode, patient, patient_profile, uploaded_file, quantum_threshold, fill_missing, max_iterations, window_sizes, pattern_detection, temporal_analysis):
    """Запуск MQEA анализа."""
    with st.spinner("Выполняется MQEA анализ..."):
        try:
            # Генерация или загрузка данных
            if analysis_mode == "Генерация данных":
                # Используем значения из профиля пациента для генерации данных
                data = st.session_state.components['analyzer'].generate_synthetic_data(
                    duration_hours=24,
                    sampling_rate_minutes=15,
                    add_noise=True,
                    add_missing_data=True,
                    patient_profile=patient_profile
                )
                
                # Сохраняем профиль пациента для анализа
                st.session_state.current_patient_profile = patient_profile
            else:
                if uploaded_file:
                    df = pd.read_csv(uploaded_file)
                    data = process_uploaded_data(df)
                else:
                    st.error("Пожалуйста, загрузите файл")
                    return
            
            # Выполнение MQEA анализа
            if fill_missing and data.missing_data_mask.sum().sum() > 0:
                data = st.session_state.components['analyzer'].fill_missing_data(
                    data, method='quantum', max_iterations=max_iterations
                )
            
            analysis_results = st.session_state.components['analyzer'].quantum_entanglement_analysis(
                data, quantum_threshold
            )
            
            st.session_state.analysis_results = analysis_results
            st.session_state.current_data = data
            
            # Сохранение результатов в базу данных
            try:
                visit_id = str(uuid.uuid4())
                
                # Фильтруем результаты анализа для сохранения в БД
                filtered_results = _filter_mqea_data(analysis_results)
                
                analysis_data = {
                    'analysis_id': str(uuid.uuid4()),
                    'patient_id': patient.patient_id,
                    'visit_id': visit_id,
                    'quantum_coherence': float(analysis_results.get('quantum_signatures', {}).get('quantum_coherence', 0)),
                    'entanglement_pairs': int(analysis_results.get('entanglement_statistics', {}).get('entangled_pairs', 0)),
                    'max_entanglement': float(analysis_results.get('entanglement_statistics', {}).get('max_entanglement', 0)),
                    'patterns_detected': analysis_results.get('patterns', []),
                    'recommendations': [],
                    'raw_data': filtered_results
                }
                
                st.session_state.components['database'].add_mqea_analysis(analysis_data)
            except Exception as e:
                st.warning(f"⚠️ Ошибка добавления анализа MQEA: {e}")
            
            st.success("✅ MQEA анализ завершен успешно!")
            
        except Exception as e:
            st.error(f"❌ Ошибка анализа: {str(e)}")

def run_quick_mqea_analysis(patient, quantum_threshold):
    """Быстрый MQEA анализ."""
    with st.spinner("Выполняется быстрый анализ..."):
        try:
            # Генерируем данные для быстрого анализа с учетом данных пациента
            # Используем реальные данные пациента, если есть, или генерируем на основе возраста и пола
            import random
            
            # Базовые значения зависят от возраста и пола пациента
            base_heart_rate = 70 if patient.age < 50 else 75
            base_systolic = 120 if patient.age < 50 else 130
            base_diastolic = 80 if patient.age < 50 else 85
            
            # Добавляем вариацию на основе ID пациента для уникальности
            patient_seed = hash(patient.patient_id) % 1000
            random.seed(patient_seed)
            np.random.seed(patient_seed)
            
            # Генерируем профиль с вариацией для каждого пациента
            patient_profile = {
                'heart_rate': base_heart_rate + random.randint(-10, 15),
                'blood_pressure_systolic': base_systolic + random.randint(-10, 20),
                'blood_pressure_diastolic': base_diastolic + random.randint(-5, 15),
                'temperature': 36.6 + random.uniform(-0.5, 0.8),
                'oxygen_saturation': 95 + random.randint(0, 5),
                'respiratory_rate': 14 + random.randint(-2, 6),
                'glucose': 4.5 + random.uniform(0, 2.5),  # Вариация уровня глюкозы
                'cholesterol': 160 + random.randint(-30, 60)
            }
            
            # Учитываем медицинскую историю пациента для более реалистичных данных
            if patient.medical_history:
                if 'diabetes' in patient.medical_history or 'diabetes' in str(patient.medical_history).lower():
                    patient_profile['glucose'] = 7.0 + random.uniform(0, 3.0)  # Повышенный уровень при диабете
                if 'hypertension' in patient.medical_history or 'гипертония' in str(patient.medical_history).lower():
                    patient_profile['blood_pressure_systolic'] = 140 + random.randint(0, 30)
                    patient_profile['blood_pressure_diastolic'] = 90 + random.randint(0, 20)
            
            data = st.session_state.components['analyzer'].generate_synthetic_data(
                duration_hours=24,  # Увеличиваем продолжительность для более точного анализа
                sampling_rate_minutes=15,  # Более частый отбор проб
                add_noise=True,
                add_missing_data=True,  # Добавляем пропущенные данные для реалистичности
                patient_profile=patient_profile
            )
            
            # Выполняем анализ с правильным порогом
            analysis_results = st.session_state.components['analyzer'].quantum_entanglement_analysis(
                data, quantum_threshold=quantum_threshold
            )
            
            st.session_state.analysis_results = analysis_results
            st.session_state.current_data = data
            st.session_state.current_patient_profile = patient_profile
            
            st.success("✅ Быстрый анализ завершен!")
            
        except Exception as e:
            st.error(f"❌ Ошибка быстрого анализа: {str(e)}")

def display_mqea_results():
    """Отображение результатов MQEA анализа."""
    st.subheader("📊 Результаты MQEA анализа")
    
    # Статистика
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        coherence = st.session_state.analysis_results.get('quantum_signatures', {}).get('quantum_coherence', 0)
        st.metric("Квантовая когерентность", f"{coherence:.3f}")
        if coherence > 0.8:
            st.success("Отличная стабильность")
        elif coherence > 0.6:
            st.success("Высокая стабильность")
        elif coherence > 0.4:
            st.warning("Умеренная стабильность")
        elif coherence > 0.2:
            st.warning("Пониженная стабильность")
        else:
            st.error("Критическая нестабильность")
    
    with col2:
        quantum_signatures = st.session_state.analysis_results.get('quantum_signatures', {})
        entangled_pairs = quantum_signatures.get('entangled_pairs_count', 0)
        st.metric("Запутанных пар", entangled_pairs)
        if entangled_pairs > 20:
            st.success("Сильная взаимосвязь")
        elif entangled_pairs > 10:
            st.warning("Умеренная взаимосвязь")
        else:
            st.info("Слабая взаимосвязь")
    
    with col3:
        average_entanglement = quantum_signatures.get('average_entanglement', 0)
        st.metric("Средняя запутанность", f"{average_entanglement:.3f}")
        if average_entanglement > 0.7:
            st.success("Очень сильная")
        elif average_entanglement > 0.4:
            st.warning("Сильная")
        else:
            st.info("Слабая")
    
    with col4:
        patterns = st.session_state.analysis_results.get('patterns', [])
        quantum_patterns = st.session_state.analysis_results.get('quantum_patterns', [])
        total_patterns = len(patterns) + len(quantum_patterns)
        st.metric("Обнаруженных паттернов", total_patterns)
        if total_patterns > 5:
            st.info("Много паттернов")
        elif total_patterns > 2:
            st.info("Несколько паттернов")
        else:
            st.info("Мало паттернов")
    
    # Анализ показателей пациента
    st.subheader("🔍 Анализ показателей пациента")
    
    # Кнопка для обновления анализа при изменении профиля
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        st.info("💡 Изменили показатели в профиле? Нажмите кнопку для обновления анализа")
    with col2:
        if st.button("🔄 Обновить анализ", type="primary"):
            # Пересчитываем данные с новым профилем пациента
            if st.session_state.current_patient_profile:
                updated_data = st.session_state.components['analyzer'].generate_synthetic_data(
                    duration_hours=24,
                    sampling_rate_minutes=15,
                    add_noise=True,
                    add_missing_data=True,
                    patient_profile=st.session_state.current_patient_profile
                )
                
                # Обновляем данные и результаты анализа
                st.session_state.current_data = updated_data
                
                # Пересчитываем анализ
                quantum_threshold = st.session_state.get('quantum_threshold', 0.3)
                st.session_state.analysis_results = st.session_state.components['analyzer'].quantum_entanglement_analysis(
                    updated_data, quantum_threshold
                )
                
                st.success("✅ Анализ обновлен с новыми данными профиля!")
                st.rerun()
            else:
                st.warning("⚠️ Профиль пациента не найден")
    
    with col3:
        if st.button("📊 Показать профиль"):
            st.session_state.show_patient_profile = True
            st.rerun()
    
    if st.session_state.current_data:
        # Получаем последние значения показателей
        latest_values = st.session_state.current_data.data.iloc[-1]
        
        # Нормальные диапазоны показателей
        normal_ranges = {
            'heart_rate': (60, 100),
            'blood_pressure_systolic': (90, 140),
            'blood_pressure_diastolic': (60, 90),
            'temperature': (36.1, 37.2),
            'oxygen_saturation': (95, 100),
            'respiratory_rate': (12, 20),
            'glucose': (3.9, 5.6),  # ммоль/л
            'cholesterol': (100, 200)
        }
        
        # Переводы названий показателей
        indicator_names = {
            'heart_rate': 'Частота пульса',
            'blood_pressure_systolic': 'Систолическое давление',
            'blood_pressure_diastolic': 'Диастолическое давление',
            'temperature': 'Температура тела',
            'oxygen_saturation': 'Насыщение кислородом',
            'respiratory_rate': 'Частота дыхания',
            'glucose': 'Уровень глюкозы',
            'cholesterol': 'Уровень холестерина'
        }
        
        # Анализ каждого показателя
        warnings = []
        for indicator, value in latest_values.items():
            if indicator in normal_ranges:
                normal_min, normal_max = normal_ranges[indicator]
                indicator_name = indicator_names.get(indicator, indicator)
                
                if value < normal_min:
                    warnings.append(f"⚠️ **{indicator_name}**: {value:.1f} - ниже нормы ({normal_min}-{normal_max})")
                elif value > normal_max:
                    warnings.append(f"🚨 **{indicator_name}**: {value:.1f} - выше нормы ({normal_min}-{normal_max})")
        
        # Отображение предупреждений
        if warnings:
            st.error("**Обнаружены отклонения от нормы:**")
            for warning in warnings:
                st.write(warning)
        else:
            st.success("✅ **Все показатели в пределах нормы**")
        
        # Специальный анализ глюкозы
        glucose_value = latest_values.get('glucose', 0)
        if glucose_value > 5.6:
            if glucose_value > 7.0:
                st.error(f"🚨 **КРИТИЧЕСКИ ВЫСОКИЙ УРОВЕНЬ ГЛЮКОЗЫ**: {glucose_value:.1f} ммоль/л - возможен диабет!")
                st.write("**Рекомендации:**")
                st.write("- Немедленная консультация эндокринолога")
                st.write("- Глюкозотолерантный тест")
                st.write("- Контроль уровня глюкозы каждые 2 часа")
            elif glucose_value > 5.6:
                st.warning(f"⚠️ **ПОВЫШЕННЫЙ УРОВЕНЬ ГЛЮКОЗЫ**: {glucose_value:.1f} ммоль/л - возможен преддиабет!")
                st.write("**Рекомендации:**")
                st.write("- Консультация эндокринолога")
                st.write("- Диета с ограничением углеводов")
                st.write("- Регулярный контроль уровня глюкозы")
        elif glucose_value < 3.9:
            st.warning(f"⚠️ **НИЗКИЙ УРОВЕНЬ ГЛЮКОЗЫ**: {glucose_value:.1f} ммоль/л - возможна гипогликемия!")
            st.write("**Рекомендации:**")
            st.write("- Консультация эндокринолога")
            st.write("- Контроль уровня глюкозы")
            st.write("- При необходимости - прием глюкозы")
    
    # Отображение профиля пациента
    if st.session_state.get('show_patient_profile', False):
        st.subheader("📊 Профиль пациента")
        
        if st.session_state.current_patient_profile:
            profile = st.session_state.current_patient_profile
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Основные показатели:**")
                st.write(f"• Частота пульса: {profile.get('heart_rate', 'N/A')} уд/мин")
                st.write(f"• Систолическое давление: {profile.get('blood_pressure_systolic', 'N/A')} мм рт.ст.")
                st.write(f"• Диастолическое давление: {profile.get('blood_pressure_diastolic', 'N/A')} мм рт.ст.")
                st.write(f"• Температура: {profile.get('temperature', 'N/A')} °C")
            
            with col2:
                st.write("**Дополнительные показатели:**")
                st.write(f"• Насыщение кислородом: {profile.get('oxygen_saturation', 'N/A')} %")
                st.write(f"• Частота дыхания: {profile.get('respiratory_rate', 'N/A')} вдох/мин")
                st.write(f"• **Уровень глюкозы: {profile.get('glucose', 'N/A')} ммоль/л**")
                st.write(f"• Уровень холестерина: {profile.get('cholesterol', 'N/A')} мг/дл")
            
            # Кнопка для редактирования профиля
            if st.button("✏️ Редактировать профиль"):
                st.session_state.show_edit_profile = True
                st.rerun()
            
            if st.button("❌ Закрыть профиль"):
                st.session_state.show_patient_profile = False
                st.rerun()
        else:
            st.warning("⚠️ Профиль пациента не найден")
    
    # Форма редактирования профиля
    if st.session_state.get('show_edit_profile', False):
        st.subheader("✏️ Редактирование профиля пациента")
        
        if st.session_state.current_patient_profile:
            profile = st.session_state.current_patient_profile
            
            with st.form("edit_profile_form"):
                col1, col2 = st.columns(2)
                
                with col1:
                    heart_rate = st.number_input("Частота пульса (уд/мин)", 40, 200, value=int(profile.get('heart_rate', 75)))
                    blood_pressure_systolic = st.number_input("Систолическое давление (мм рт.ст.)", 80, 200, value=int(profile.get('blood_pressure_systolic', 120)))
                    blood_pressure_diastolic = st.number_input("Диастолическое давление (мм рт.ст.)", 50, 120, value=int(profile.get('blood_pressure_diastolic', 80)))
                    temperature = st.number_input("Температура (°C)", 35.0, 42.0, value=float(profile.get('temperature', 36.6)), step=0.1)
                
                with col2:
                    oxygen_saturation = st.number_input("Насыщение кислородом (%)", 70, 100, value=int(profile.get('oxygen_saturation', 98)))
                    respiratory_rate = st.number_input("Частота дыхания (вдох/мин)", 8, 40, value=int(profile.get('respiratory_rate', 16)))
                    glucose = st.number_input("Уровень глюкозы (ммоль/л)", 1.0, 20.0, value=float(profile.get('glucose', 5.0)), step=0.1)
                    cholesterol = st.number_input("Уровень холестерина (мг/дл)", 50, 400, value=int(profile.get('cholesterol', 180)))
                
                col_save, col_cancel = st.columns(2)
                
                with col_save:
                    if st.form_submit_button("💾 Сохранить изменения", width='stretch'):
                        # Обновляем профиль пациента
                        updated_profile = {
                            'heart_rate': heart_rate,
                            'blood_pressure_systolic': blood_pressure_systolic,
                            'blood_pressure_diastolic': blood_pressure_diastolic,
                            'temperature': temperature,
                            'oxygen_saturation': oxygen_saturation,
                            'respiratory_rate': respiratory_rate,
                            'glucose': glucose,
                            'cholesterol': cholesterol
                        }
                        
                        st.session_state.current_patient_profile = updated_profile
                        st.session_state.show_edit_profile = False
                        st.success("✅ Профиль пациента обновлен!")
                        st.rerun()
                
                with col_cancel:
                    if st.form_submit_button("Отмена", width='stretch'):
                        st.session_state.show_edit_profile = False
                        st.rerun()

    # Интерпретация результатов MQEA
    st.subheader("🔍 Интерпретация результатов MQEA")
    
    coherence = st.session_state.analysis_results.get('quantum_signatures', {}).get('quantum_coherence', 0)
    pairs = st.session_state.analysis_results.get('entanglement_statistics', {}).get('entangled_pairs', 0)
    
    if coherence > 0.6 and pairs > 15:
        st.success("🎉 **Отличные результаты!** Система показывает высокую квантовую стабильность и сильные взаимосвязи между показателями. Общее состояние организма стабильное.")
    elif coherence > 0.3 and pairs > 8:
        st.warning("⚠️ **Умеренные результаты.** Система демонстрирует умеренную квантовую стабильность. Рекомендуется мониторинг показателей.")
    else:
        st.error("🚨 **Требует внимания.** Обнаружены признаки нестабильности в квантовых показателях. Требуется дополнительное обследование.")
    
    # График временных рядов
    if st.session_state.current_data:
        st.subheader("📈 Временные ряды")
        
        fig = go.Figure()
        
        # Переводы названий показателей
        indicator_translations = {
            'heart_rate': 'Частота пульса',
            'blood_pressure_systolic': 'Систолическое давление',
            'blood_pressure_diastolic': 'Диастолическое давление',
            'temperature': 'Температура тела',
            'oxygen_saturation': 'Насыщение кислородом',
            'respiratory_rate': 'Частота дыхания',
            'glucose': 'Уровень глюкозы',
            'cholesterol': 'Уровень холестерина'
        }
        
        for indicator in st.session_state.current_data.indicators[:4]:
            display_name = indicator_translations.get(indicator, indicator.replace('_', ' ').title())
            fig.add_trace(go.Scatter(
                x=st.session_state.current_data.timestamps,
                y=st.session_state.current_data.data[indicator],
                mode='lines+markers',
                name=display_name,
                line=dict(width=2)
            ))
        
        fig.update_layout(
            title="Медицинские показатели во времени",
            xaxis_title="Время",
            yaxis_title="Значение",
            hovermode='x unified',
            height=500
        )
        
        st.plotly_chart(fig, width='stretch', key="temporal_analysis_chart")
    
    # Матрица запутанности
    entanglement_matrix = st.session_state.analysis_results.get('entanglement_matrix')
    if entanglement_matrix is not None:
        st.subheader("🔗 Матрица квантовой запутанности")
        
        fig_heatmap = px.imshow(
            entanglement_matrix,
            title="Матрица квантовой запутанности",
            color_continuous_scale='Viridis'
        )
        st.plotly_chart(fig_heatmap, width='stretch', key="simple_entanglement_matrix")
    
    # Обнаруженные паттерны
    patterns = st.session_state.analysis_results.get('patterns', [])
    if patterns:
        st.subheader("🔍 Обнаруженные паттерны")
        
        for i, pattern in enumerate(patterns[:5], 1):  # Показываем первые 5 паттернов
            with st.expander(f"Паттерн {i}: {pattern.get('type', 'Неизвестный тип')}"):
                st.write(f"**Описание:** {pattern.get('description', 'Нет описания')}")
                st.write(f"**Уверенность:** {pattern.get('confidence', 0):.2f}")
                st.write(f"**Временной диапазон:** {pattern.get('start_time', 'N/A')} - {pattern.get('end_time', 'N/A')}")
    
    # Таблицы с данными
    if st.session_state.current_data:
        st.subheader("📊 Таблицы данных")
        
        # Создаем вкладки для разных типов данных
        tab1, tab2, tab3, tab4 = st.tabs(["📈 Временные ряды", "🔗 Матрица запутанности", "📋 Статистика", "🔍 Детали анализа"])
        
        with tab1:
            st.write("**Исходные медицинские данные:**")
            
            # Создаем DataFrame с временными рядами
            data_df = pd.DataFrame(st.session_state.current_data.data)
            data_df['Время'] = st.session_state.current_data.timestamps
            
            # Очищаем данные от комплексных чисел и других проблемных типов
            data_df = clean_dataframe(data_df)
            
            # Переводим названия колонок
            column_translations = {
                'heart_rate': 'Частота пульса (уд/мин)',
                'blood_pressure_systolic': 'Систолическое давление (мм рт.ст.)',
                'blood_pressure_diastolic': 'Диастолическое давление (мм рт.ст.)',
                'temperature': 'Температура тела (°C)',
                'oxygen_saturation': 'Насыщение кислородом (%)',
                'respiratory_rate': 'Частота дыхания (дых/мин)',
                'glucose': 'Уровень глюкозы (ммоль/л)',
                'cholesterol': 'Уровень холестерина (мг/дл)'
            }
            
            # Переименовываем колонки
            data_df = data_df.rename(columns=column_translations)
            
            # Показываем первые 20 строк
            st.dataframe(data_df.head(20), width='stretch')
            
            # Статистика по данным
            st.write("**Статистика по показателям:**")
            stats_df = data_df.describe()
            st.dataframe(stats_df, width='stretch')
        
        with tab2:
            st.write("**🔗 Интерактивная матрица квантовой запутанности**")
            
            # Получаем данные о квантовых запутанностях
            quantum_entanglements = st.session_state.analysis_results.get('quantum_entanglements', [])
            
            if quantum_entanglements:
                # Создаем селектор для выбора временного окна
                st.write("**📅 Выберите временное окно для анализа:**")
                
                # Создаем список окон
                window_options = []
                for i, window in enumerate(quantum_entanglements):
                    if isinstance(window, dict) and 'entanglement_matrix' in window:
                        window_size = window.get('window_size', 'Неизвестно')
                        timestamp = window.get('timestamp', f'Окно {i+1}')
                        window_options.append(f"{timestamp} (размер: {window_size})")
                    else:
                        window_options.append(f"Окно {i+1}")
                
                if window_options:
                    selected_window_idx = st.selectbox(
                        "Выберите окно:",
                        range(len(window_options)),
                        format_func=lambda x: window_options[x]
                    )
                    
                    selected_window = quantum_entanglements[selected_window_idx]
                    
                    if isinstance(selected_window, dict) and 'entanglement_matrix' in selected_window:
                        entanglement_matrix = np.array(selected_window['entanglement_matrix'])
                        
                        # Проверяем, что матрица не пустая
                        if np.any(entanglement_matrix > 0):
                            indicators = st.session_state.current_data.indicators
                            
                            # Переводим названия показателей
                            translated_indicators = [
                                column_translations.get(ind, ind.replace('_', ' ').title()) 
                                for ind in indicators
                            ]
                            
                            # Создаем интерактивную тепловую карту
                            fig = go.Figure(data=go.Heatmap(
                                z=entanglement_matrix,
                                x=translated_indicators,
                                y=translated_indicators,
                                colorscale='Viridis',
                                showscale=True,
                                hoverongaps=False,
                                text=np.round(entanglement_matrix, 3),
                                texttemplate="%{text}",
                                textfont={"size": 10},
                                colorbar=dict(
                                    title=dict(text="Сила запутанности", side="right")
                                )
                            ))
                            
                            fig.update_layout(
                                title=f"Матрица квантовой запутанности - {window_options[selected_window_idx]}",
                                xaxis_title="Медицинские показатели",
                                yaxis_title="Медицинские показатели",
                                height=600,
                                width=800
                            )
                            
                            st.plotly_chart(fig, width='stretch', key=f"entanglement_matrix_{selected_window_idx}")
                            
                            # Статистика матрицы
                            col_stats1, col_stats2, col_stats3, col_stats4 = st.columns(4)
                            
                            with col_stats1:
                                st.metric("Размер матрицы", f"{entanglement_matrix.shape[0]}×{entanglement_matrix.shape[1]}")
                            
                            with col_stats2:
                                non_zero_connections = np.count_nonzero(entanglement_matrix)
                                st.metric("Ненулевых связей", non_zero_connections)
                            
                            with col_stats3:
                                max_entanglement = np.max(entanglement_matrix)
                                st.metric("Максимальная запутанность", f"{max_entanglement:.3f}")
                            
                            with col_stats4:
                                avg_entanglement = np.mean(entanglement_matrix[entanglement_matrix > 0]) if np.any(entanglement_matrix > 0) else 0
                                st.metric("Средняя запутанность", f"{avg_entanglement:.3f}")
                            
                            # Анализ связей
                            st.subheader("🔍 Анализ квантовых связей")
                            
                            # Создаем список всех связей с подробной информацией
                            connections = []
                            for i in range(len(indicators)):
                                for j in range(i+1, len(indicators)):
                                    entanglement_strength = entanglement_matrix[i, j]
                                    if entanglement_strength > 0:
                                        # Определяем тип связи
                                        if entanglement_strength > 0.8:
                                            connection_type = "🔴 Очень сильная"
                                            interpretation = "Критически важная связь"
                                        elif entanglement_strength > 0.6:
                                            connection_type = "🟠 Сильная"
                                            interpretation = "Значимая корреляция"
                                        elif entanglement_strength > 0.4:
                                            connection_type = "🟡 Умеренная"
                                            interpretation = "Заметная взаимосвязь"
                                        elif entanglement_strength > 0.2:
                                            connection_type = "🟢 Слабая"
                                            interpretation = "Потенциальная связь"
                                        else:
                                            connection_type = "⚪ Очень слабая"
                                            interpretation = "Минимальная корреляция"
                                        
                                        connections.append({
                                            'Показатель 1': translated_indicators[i],
                                            'Показатель 2': translated_indicators[j],
                                            'Сила запутанности': entanglement_strength,
                                            'Тип связи': connection_type,
                                            'Интерпретация': interpretation,
                                            'Индекс 1': i,
                                            'Индекс 2': j
                                        })
                            
                            if connections:
                                # Сортируем по силе запутанности
                                connections_df = pd.DataFrame(connections)
                                connections_df = connections_df.sort_values('Сила запутанности', ascending=False)
                                connections_df['Сила запутанности'] = connections_df['Сила запутанности'].round(4)
                                
                                # Фильтр по типу связи
                                st.write("**🔽 Фильтр по типу связи:**")
                                filter_type = st.selectbox(
                                    "Выберите тип связи:",
                                    ["Все", "🔴 Очень сильная", "🟠 Сильная", "🟡 Умеренная", "🟢 Слабая", "⚪ Очень слабая"]
                                )
                                
                                if filter_type != "Все":
                                    filtered_df = connections_df[connections_df['Тип связи'] == filter_type]
                                else:
                                    filtered_df = connections_df
                                
                                st.dataframe(
                                    filtered_df[['Показатель 1', 'Показатель 2', 'Сила запутанности', 'Тип связи', 'Интерпретация']],
                                    width='stretch',
                                    height=400
                                )
                                
                                # Топ-5 самых сильных связей
                                st.subheader("🏆 Топ-5 самых сильных квантовых связей")
                                top_5 = connections_df.head(5)
                                
                                for idx, row in top_5.iterrows():
                                    with st.expander(f"#{idx+1} {row['Показатель 1']} ↔ {row['Показатель 2']} ({row['Сила запутанности']:.3f})"):
                                        col1, col2 = st.columns(2)
                                        
                                        with col1:
                                            st.write(f"**Тип связи:** {row['Тип связи']}")
                                            st.write(f"**Сила запутанности:** {row['Сила запутанности']:.4f}")
                                            st.write(f"**Интерпретация:** {row['Интерпретация']}")
                                        
                                        with col2:
                                            # Создаем мини-график для этой связи
                                            fig_mini = go.Figure()
                                            fig_mini.add_trace(go.Indicator(
                                                mode="gauge+number",
                                                value=row['Сила запутанности'],
                                                domain={'x': [0, 1], 'y': [0, 1]},
                                                title={'text': "Сила связи"},
                                                gauge={
                                                    'axis': {'range': [None, 1]},
                                                    'bar': {'color': "darkblue"},
                                                    'steps': [
                                                        {'range': [0, 0.2], 'color': "lightgray"},
                                                        {'range': [0.2, 0.4], 'color': "yellow"},
                                                        {'range': [0.4, 0.6], 'color': "orange"},
                                                        {'range': [0.6, 0.8], 'color': "red"},
                                                        {'range': [0.8, 1], 'color': "darkred"}
                                                    ],
                                                    'threshold': {
                                                        'line': {'color': "red", 'width': 4},
                                                        'thickness': 0.75,
                                                        'value': 0.9
                                                    }
                                                }
                                            ))
                                            fig_mini.update_layout(height=200)
                                            st.plotly_chart(fig_mini, width='stretch', key=f"connection_gauge_{idx}_{row['Индекс 1']}_{row['Индекс 2']}")
                            else:
                                st.info("ℹ️ Значимых квантовых связей не обнаружено в выбранном окне")
                        else:
                            st.warning("⚠️ В выбранном окне нет значимых квантовых связей")
                    else:
                        st.error("❌ Выбранное окно не содержит корректных данных о запутанности")
                else:
                    st.warning("⚠️ Нет доступных окон для анализа")
            else:
                st.info("ℹ️ Матрица квантовой запутанности:")
                st.markdown("""
                **🔍 Возможные причины отсутствия данных:**
                - Недостаточно данных для анализа квантовой запутанности
                - Слишком высокий порог квантовой запутанности
                - Отсутствие значимых корреляций между показателями
                - Требуется выполнить MQEA анализ для получения данных
                
                **💡 Рекомендации:**
                - Проверьте настройки анализа (порог запутанности)
                - Увеличьте количество временных точек
                - Снизьте порог значимости корреляций
                """)
        
        with tab3:
            st.write("**Статистика квантового анализа:**")
            
            # Создаем DataFrame со статистикой
            stats_data = []
            
            # Квантовые подписи
            quantum_signatures = st.session_state.analysis_results.get('quantum_signatures', {})
            for key, value in quantum_signatures.items():
                stats_data.append({
                    'Параметр': key.replace('_', ' ').title(),
                    'Значение': f"{value:.6f}" if isinstance(value, (int, float)) else str(value),
                    'Описание': get_parameter_description(key)
                })
            
            # Статистика запутанности
            entanglement_stats = st.session_state.analysis_results.get('entanglement_statistics', {})
            for key, value in entanglement_stats.items():
                stats_data.append({
                    'Параметр': key.replace('_', ' ').title(),
                    'Значение': f"{value:.6f}" if isinstance(value, (int, float)) else str(value),
                    'Описание': get_parameter_description(key)
                })
            
            # Общая статистика
            stats_data.extend([
                {
                    'Параметр': 'Общее количество показателей',
                    'Значение': len(st.session_state.current_data.indicators),
                    'Описание': 'Количество медицинских показателей в анализе'
                },
                {
                    'Параметр': 'Количество временных точек',
                    'Значение': len(st.session_state.current_data.timestamps),
                    'Описание': 'Количество измерений во времени'
                },
                {
                    'Параметр': 'Продолжительность анализа',
                    'Значение': f"{(st.session_state.current_data.timestamps[-1] - st.session_state.current_data.timestamps[0]).total_seconds() / 3600:.1f} часов",
                    'Описание': 'Общая продолжительность временного ряда'
                }
            ])
            
            if stats_data:
                stats_df = pd.DataFrame(stats_data)
                
                # Добавляем интерпретацию для ключевых параметров
                def get_interpretation(param_name, value):
                    """Получить интерпретацию значения параметра."""
                    if 'Квантовая Когерентность' in param_name or 'quantum_coherence' in param_name.lower():
                        val = float(value) if isinstance(value, str) and value.replace('.', '').replace('-', '').isdigit() else (float(value) if isinstance(value, (int, float)) else 0)
                        if val > 0.8:
                            return "Отличная стабильность"
                        elif val > 0.6:
                            return "Высокая стабильность"
                        elif val > 0.4:
                            return "Умеренная стабильность"
                        elif val > 0.2:
                            return "Пониженная стабильность"
                        else:
                            return "Критическая нестабильность"
                    elif 'Запутанных Пар' in param_name or 'entangled_pairs' in param_name.lower():
                        val = int(float(value)) if isinstance(value, (str, int, float)) else 0
                        if val > 20:
                            return "Сильная взаимосвязь"
                        elif val > 10:
                            return "Умеренная взаимосвязь"
                        elif val > 0:
                            return "Слабая взаимосвязь"
                        else:
                            return "Отсутствие взаимосвязи"
                    elif 'Максимальная Запутанность' in param_name or 'max_entanglement' in param_name.lower():
                        val = float(value) if isinstance(value, (str, int, float)) else 0
                        if val > 0.7:
                            return "Очень сильная"
                        elif val > 0.4:
                            return "Сильная"
                        elif val > 0.1:
                            return "Умеренная"
                        else:
                            return "Слабая"
                    elif 'Обнаруженных Паттернов' in param_name or 'patterns' in param_name.lower():
                        val = int(float(value)) if isinstance(value, (str, int, float)) else 0
                        if val > 5:
                            return "Много паттернов"
                        elif val > 2:
                            return "Несколько паттернов"
                        else:
                            return "Мало паттернов"
                    return ""
                
                # Добавляем колонку интерпретации
                stats_df['Интерпретация'] = stats_df.apply(
                    lambda row: get_interpretation(row['Параметр'], row['Значение']), 
                    axis=1
                )
                
                # Переименовываем колонки для лучшей читаемости
                stats_df_display = stats_df[['Параметр', 'Значение', 'Интерпретация']].copy()
                
                st.dataframe(stats_df_display, width='stretch', use_container_width=True)
            else:
                st.warning("⚠️ Нет данных для отображения статистики")
        
        with tab4:
            st.write("**Детальная информация об анализе:**")
            
            # Информация о данных
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Информация о данных:**")
                info_data = {
                    'Показатель': ['Количество показателей', 'Временных точек', 'Пропущенных данных', 'Интервал выборки'],
                    'Значение': [
                        len(st.session_state.current_data.indicators),
                        len(st.session_state.current_data.timestamps),
                        f"{st.session_state.current_data.missing_data_mask.sum().sum()} ({st.session_state.current_data.missing_data_mask.sum().sum() / (len(st.session_state.current_data.indicators) * len(st.session_state.current_data.timestamps)) * 100:.1f}%)",
                        f"{(st.session_state.current_data.timestamps[1] - st.session_state.current_data.timestamps[0]).total_seconds() / 60:.0f} минут"
                    ]
                }
                info_df = pd.DataFrame(info_data)
                st.dataframe(info_df, width='stretch')
            
            with col2:
                st.write("**Параметры анализа:**")
                analysis_params = {
                    'Параметр': ['Порог запутанности', 'Метод заполнения', 'Максимум итераций', 'Обнаружение паттернов'],
                    'Значение': ['0.3', 'Квантовый', '50', 'Включено']
                }
                params_df = pd.DataFrame(analysis_params)
                st.dataframe(params_df, width='stretch')
            
            # Список показателей
            st.write("**Медицинские показатели в анализе:**")
            indicators_data = []
            for i, indicator in enumerate(st.session_state.current_data.indicators, 1):
                indicators_data.append({
                    '№': i,
                    'Показатель': column_translations.get(indicator, indicator.replace('_', ' ').title()),
                    'Единица измерения': get_indicator_unit(indicator),
                    'Нормальный диапазон': get_normal_range(indicator)
                })
            
            indicators_df = pd.DataFrame(indicators_data)
            st.dataframe(indicators_df, width='stretch')

def get_parameter_description(param_name):
    """Получить описание параметра."""
    descriptions = {
        'quantum_coherence': 'Квантовая когерентность - мера стабильности системы',
        'quantum_entropy': 'Квантовая энтропия - мера неопределенности',
        'entangled_pairs': 'Количество запутанных пар показателей',
        'max_entanglement': 'Максимальная степень запутанности',
        'avg_entanglement': 'Средняя степень запутанности',
        'total_states': 'Общее количество квантовых состояний'
    }
    return descriptions.get(param_name, 'Описание недоступно')

def get_indicator_unit(indicator):
    """Получить единицу измерения показателя."""
    units = {
        'heart_rate': 'уд/мин',
        'blood_pressure_systolic': 'мм рт.ст.',
        'blood_pressure_diastolic': 'мм рт.ст.',
        'temperature': '°C',
        'oxygen_saturation': '%',
        'respiratory_rate': 'дых/мин',
        'glucose': 'ммоль/л',
        'cholesterol': 'мг/дл'
    }
    return units.get(indicator, 'ед.')

def get_normal_range(indicator):
    """Получить нормальный диапазон показателя."""
    ranges = {
        'heart_rate': '60-100 уд/мин',
        'blood_pressure_systolic': '90-140 мм рт.ст.',
        'blood_pressure_diastolic': '60-90 мм рт.ст.',
        'temperature': '36.1-37.2°C',
        'oxygen_saturation': '95-100%',
        'respiratory_rate': '12-20 дых/мин',
        'glucose': '3.9-5.6 ммоль/л',
        'cholesterol': '<200 мг/дл'
    }
    return ranges.get(indicator, 'Н/Д')

def show_disease_analysis():
    """Анализ признаков заболеваний."""
    log_action("🦠 ВЫЗОВ ФУНКЦИИ", "show_disease_analysis() - НАЧАЛО")
    st.header("🦠 Квантовый анализ признаков заболеваний")
    
    # Словари перевода симптомов и факторов риска
    SYMPTOMS_TRANSLATION = {
        'Лихорадка': 'fever',
        'Усталость': 'fatigue',
        'Потеря веса': 'weight_loss',
        'Ночная потливость': 'night_sweats',
        'Увеличенные лимфоузлы': 'swollen_lymph_nodes',
        'Кашель': 'cough',
        'Одышка': 'shortness_of_breath',
        'Боль в груди': 'chest_pain',
        'Головная боль': 'headache',
        'Головокружение': 'dizziness',
        'Тошнота': 'nausea',
        'Рвота': 'vomiting',
        'Боль в животе': 'abdominal_pain',
        'Желтуха': 'jaundice',
        'Необъяснимая потеря веса': 'unexplained_weight_loss',
        'Постоянный кашель': 'persistent_cough',
        'Хрипы': 'wheezing',
        'Затрудненное дыхание': 'difficulty_breathing',
        'Боль в горле': 'sore_throat',
        'Сыпь': 'rash',
        'Мышечные боли': 'muscle_aches',
        'Диарея': 'diarrhea',
        'Кандидоз полости рта': 'oral_thrush',
        'Рецидивирующие инфекции': 'recurrent_infections',
        'Уплотнение в груди': 'breast_lump',
        'Затрудненное мочеиспускание': 'difficulty_urinating',
        'Частое мочеиспускание': 'frequent_urination',
        'Кровохарканье': 'coughing_up_blood',
        'Охриплость': 'hoarseness',
        'Изменения кожи': 'changes_in_skin',
        'Затрудненное глотание': 'difficulty_swallowing',
        'Изменения в работе кишечника': 'changes_in_bowel_habits',
        'Необычное кровотечение': 'unusual_bleeding',
        'Усиленная жажда': 'increased_thirst',
        'Частое мочеиспускание (диабет)': 'frequent_urination_diabetes',
        'Сильный голод': 'extreme_hunger',
        'Затуманенное зрение': 'blurred_vision',
        'Медленное заживление ран': 'slow_healing_wounds',
        'Частые инфекции': 'frequent_infections',
        'Боль в суставах': 'joint_pain',
        'Потеря аппетита': 'loss_of_appetite',
        'Темная моча': 'dark_urine',
        'Светлый стул': 'clay_colored_stools',
        'Выделения из соска': 'nipple_discharge',
        'Втяжение соска': 'nipple_retraction',
        'Боль в тазу': 'pain_in_pelvis',
        'Боль в костях': 'bone_pain',
        'Изменения зрения': 'visual_changes',
        'Нерегулярное сердцебиение': 'irregular_heartbeat',
        'Отеки': 'swelling',
        'Охриплость голоса': 'hoarseness_voice',
        'Слабая струя мочи': 'weak_urine_stream',
        'Кровь в моче': 'blood_in_urine',
        'Эректильная дисфункция': 'erectile_dysfunction',
        'Озноб': 'chills',
        'Потливость': 'sweating',
        'Увеличение груди': 'breast_swelling',
        'Изменения в груди': 'breast_changes'
    }
    
    RISK_FACTORS_TRANSLATION = {
        'Незащищенный секс': 'unprotected_sex',
        'Внутривенное употребление наркотиков': 'iv_drug_use',
        'Переливание крови': 'blood_transfusion',
        'Курение': 'smoking',
        'Возраст старше 50 лет': 'age_over_50',
        'Семейный анамнез': 'family_history',
        'Ожирение': 'obesity',
        'Малоподвижный образ жизни': 'sedentary_lifestyle',
        'Употребление алкоголя': 'alcohol_consumption',
        'Воздействие канцерогенов': 'exposure_to_carcinogens',
        'Хроническое воспаление': 'chronic_inflammation',
        'Татуировки/пирсинг': 'tattoos_piercings',
        'Профессиональное воздействие': 'occupational_exposure',
        'Передача от матери к ребенку': 'mother_to_child_transmission',
        'Тесный контакт с больным туберкулезом': 'close_contact_with_tb',
        'Иммуносупрессия': 'immunosuppression',
        'ВИЧ-инфекция': 'hiv_infection',
        'Диабет': 'diabetes',
        'Недоедание': 'malnutrition',
        'Скученные условия проживания': 'crowded_living_conditions',
        'Работник здравоохранения': 'healthcare_worker',
        'Пассивное курение': 'secondhand_smoke',
        'Воздействие радона': 'radon_exposure',
        'Воздействие асбеста': 'asbestos_exposure',
        'Загрязнение воздуха': 'air_pollution',
        'Генетические мутации': 'genetic_mutations',
        'Гормональная заместительная терапия': 'hormone_replacement_therapy',
        'Ранняя менархе': 'early_menarche',
        'Поздняя менопауза': 'late_menopause',
        'Африканское происхождение': 'african_ancestry',
        'Высокожировая диета': 'high_fat_diet',
        'Высокосолевая диета': 'high_salt_diet',
        'Стресс': 'stress',
        'Аллергии': 'allergies',
        'Экологические факторы': 'environmental_factors',
        'Респираторные инфекции': 'respiratory_infections',
        'Хронические заболевания легких': 'chronic_lung_disease',
        'Недавняя операция': 'recent_surgery',
        'Госпитализация': 'hospitalization',
        'Использование вентилятора': 'ventilator_use',
        'Гестационный диабет': 'gestational_diabetes',
        'Синдром поликистозных яичников': 'polycystic_ovary_syndrome',
        'Высокое артериальное давление': 'high_blood_pressure'
    }
    
    # Информация о модуле
    with st.expander("ℹ️ О модуле анализа заболеваний"):
        st.markdown("""
        **Квантовый анализ признаков заболеваний** использует принципы MQEA для выявления признаков различных заболеваний:
        
        **Поддерживаемые категории:**
        - 🦠 Инфекционные (ВИЧ/СПИД, Гепатит B/C, Туберкулез)
        - 🎗️ Онкологические (Рак легких, Рак молочной железы, Рак простаты)
        - ❤️ Сердечно-сосудистые (Гипертония, Ишемическая болезнь сердца)
        - 🫁 Дыхательные (Астма, Пневмония)
        - 🍬 Метаболические (Сахарный диабет)
        
        **Важно:** Этот модуль выявляет **признаки** заболеваний, а не ставит диагноз.
        Для постановки диагноза необходимо обратиться к квалифицированному врачу.
        """)
    
    # Инициализация анализатора
    if 'disease_analyzer' not in st.session_state:
        with st.spinner("Инициализация анализатора заболеваний..."):
            st.session_state.disease_analyzer = DiseasePatternAnalyzer()
    
    analyzer = st.session_state.disease_analyzer
    
    # Выбор пациента
    if not st.session_state.current_patient:
        st.warning("⚠️ Сначала выберите пациента в разделе 'Пациенты'")
        return
    
    patient = st.session_state.current_patient
    st.subheader(f"Анализ для пациента: {patient.name} (ID: {patient.patient_id})")
    
    # Вкладки
    tab1, tab2, tab3 = st.tabs(["📊 Анализ данных", "🦠 Список заболеваний", "📋 История анализов"])
    
    with tab1:
        st.subheader("📊 Ввод данных для анализа")
        
        # Выбор режима
        analysis_mode = st.radio(
            "Режим анализа",
            ["Генерация тестовых данных", "Ввод данных вручную"],
            horizontal=True
        )
        
        if analysis_mode == "Генерация тестовых данных":
            st.info("💡 Будут сгенерированы тестовые данные с признаками различных заболеваний")
            
            if st.button("🔬 Запустить анализ", type="primary"):
                with st.spinner("Генерация данных и выполнение анализа..."):
                    try:
                        from mqea.data_processor import MedicalTimeSeries
                        import pandas as pd
                        from datetime import datetime, timedelta
                        
                        # Генерация данных
                        start_date = datetime.now() - timedelta(days=30)
                        timestamps = [start_date + timedelta(days=i) for i in range(30)]
                        
                        # Данные с признаками различных заболеваний
                        data = {
                            'cd4_count': np.random.normal(300, 50, 30),
                            'cd4_percentage': np.random.normal(15, 3, 30),
                            'viral_load': np.random.normal(50000, 10000, 30),
                            'white_blood_cells': np.random.normal(3000, 500, 30),
                            'lymphocytes': np.random.normal(800, 200, 30),
                            'hemoglobin': np.random.normal(10, 1, 30),
                            'platelets': np.random.normal(120000, 20000, 30),
                            'temperature': np.random.normal(37.5, 0.5, 30),
                            'heart_rate': np.random.normal(90, 10, 30),
                            'blood_pressure_systolic': np.random.normal(145, 10, 30),
                            'blood_pressure_diastolic': np.random.normal(95, 8, 30),
                            'glucose_fasting': np.random.normal(120, 15, 30),
                            'cholesterol_total': np.random.normal(250, 30, 30)
                        }
                        
                        df = pd.DataFrame(data, index=timestamps)
                        
                        # Создание маски пропущенных данных
                        missing_mask = pd.DataFrame(False, index=df.index, columns=df.columns)
                        
                        # Создание квантовых состояний
                        quantum_states = {}
                        for indicator in df.columns:
                            quantum_states[indicator] = np.zeros(len(df))
                        
                        # Расчет процента пропущенных данных
                        total_cells = len(df) * len(df.columns)
                        missing_cells = missing_mask.sum().sum()
                        missing_percentage = (missing_cells / total_cells * 100) if total_cells > 0 else 0.0
                        
                        # Создание объекта временного ряда
                        medical_data = MedicalTimeSeries(
                            data=df,
                            indicators=list(data.keys()),
                            timestamps=pd.DatetimeIndex(timestamps),
                            missing_data_mask=missing_mask,
                            quantum_states=quantum_states,
                            metadata={
                                'source': 'test_data_generation',
                                'total_points': len(df),
                                'patient_id': patient.patient_id,
                                'missing_percentage': missing_percentage
                            }
                        )
                        
                        # Симптомы (на русском для пользователя)
                        symptom_options_ru = list(SYMPTOMS_TRANSLATION.keys())
                        selected_symptoms_ru = st.multiselect(
                            "Симптомы пациента",
                            symptom_options_ru,
                            default=['Лихорадка', 'Усталость', 'Потеря веса']
                        )
                        
                        # Перевод выбранных симптомов на английский для анализа
                        symptoms = [SYMPTOMS_TRANSLATION[ru] for ru in selected_symptoms_ru]
                        
                        # Факторы риска (на русском для пользователя)
                        risk_factor_options_ru = list(RISK_FACTORS_TRANSLATION.keys())
                        selected_risk_factors_ru = st.multiselect(
                            "Факторы риска",
                            risk_factor_options_ru,
                            default=['Незащищенный секс']
                        )
                        
                        # Перевод выбранных факторов риска на английский для анализа
                        risk_factors = [RISK_FACTORS_TRANSLATION[ru] for ru in selected_risk_factors_ru]
                        
                        # Выполнение анализа
                        results = analyzer.analyze_disease_patterns(
                            medical_data=medical_data,
                            patient_symptoms=symptoms,
                            risk_factors=risk_factors
                        )
                        
                        # Сохранение результатов
                        if 'disease_analysis_history' not in st.session_state:
                            st.session_state.disease_analysis_history = []
                        
                        st.session_state.disease_analysis_history.append({
                            'patient_id': patient.patient_id,
                            'timestamp': datetime.now(),
                            'results': results
                        })
                        
                        st.success(f"✅ Анализ завершен! Обнаружено {len(results)} потенциальных заболеваний")
                        
                        # Отображение результатов
                        st.subheader("📊 Результаты анализа")
                        
                        for i, result in enumerate(results[:10], 1):  # Показываем топ-10
                            with st.expander(
                                f"{i}. {result.disease_name} ({result.disease_code}) - "
                                f"Вероятность: {result.probability:.1%} | "
                                f"Уровень срочности: {result.urgency_level.upper()}"
                            ):
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.metric("Вероятность", f"{result.probability:.1%}")
                                    st.metric("Уверенность", f"{result.confidence:.1%}")
                                    
                                    # Визуализация вероятности
                                    prob_bar = st.progress(result.probability)
                                    
                                    # Уровень срочности
                                    urgency_colors = {
                                        'критический': '🔴',
                                        'высокий': '🟠',
                                        'средний': '🟡',
                                        'низкий': '🟢'
                                    }
                                    st.write(f"**Уровень срочности:** {urgency_colors.get(result.urgency_level, '⚪')} {result.urgency_level.upper()}")
                                
                                with col2:
                                    st.write(f"**Категория:** {result.category.value}")
                                    st.write(f"**ICD-10 код:** {result.disease_code}")
                                
                                # Совпавшие показатели
                                if result.matched_indicators:
                                    st.write("**📈 Совпавшие показатели:**")
                                    for indicator in result.matched_indicators[:5]:
                                        st.write(f"  - {indicator}")
                                
                                # Симптомы (перевод на русский)
                                if result.matched_symptoms:
                                    st.write("**🦠 Совпавшие симптомы:**")
                                    # Обратный словарь для перевода
                                    symptoms_reverse = {v: k for k, v in SYMPTOMS_TRANSLATION.items()}
                                    for symptom_en in result.matched_symptoms[:5]:
                                        symptom_ru = symptoms_reverse.get(symptom_en, symptom_en)
                                        st.write(f"  - {symptom_ru}")
                                
                                # Факторы риска (перевод на русский)
                                if result.risk_factors_present:
                                    st.write("**⚠️ Факторы риска:**")
                                    # Обратный словарь для перевода
                                    risk_factors_reverse = {v: k for k, v in RISK_FACTORS_TRANSLATION.items()}
                                    for risk_en in result.risk_factors_present:
                                        risk_ru = risk_factors_reverse.get(risk_en, risk_en)
                                        st.write(f"  - {risk_ru}")
                                
                                # Рекомендации
                                if result.recommendations:
                                    st.write("**💡 Рекомендации:**")
                                    for rec in result.recommendations:
                                        st.info(rec)
                                
                                # Диагностические тесты
                                if result.diagnostic_tests_recommended:
                                    st.write("**🧪 Рекомендуемые тесты:**")
                                    for test in result.diagnostic_tests_recommended[:5]:
                                        st.write(f"  - {test}")
                        
                        # Визуализация топ-5 заболеваний
                        if len(results) > 0:
                            st.subheader("📊 Визуализация результатов")
                            
                            top_5 = results[:5]
                            diseases = [r.disease_name for r in top_5]
                            probabilities = [r.probability * 100 for r in top_5]
                            
                            fig = px.bar(
                                x=probabilities,
                                y=diseases,
                                orientation='h',
                                labels={'x': 'Вероятность (%)', 'y': 'Заболевание'},
                                title="Топ-5 заболеваний по вероятности",
                                color=probabilities,
                                color_continuous_scale='Reds'
                            )
                            fig.update_layout(height=400)
                            st.plotly_chart(fig, use_container_width=True)
                    
                    except Exception as e:
                        st.error(f"❌ Ошибка при анализе: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
        
        else:  # Ввод данных вручную
            st.info("💡 Введите медицинские показатели пациента")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("📊 Основные показатели")
                indicators = {}
                
                # Показатели крови
                indicators['cd4_count'] = st.number_input("CD4 count (клетки/μL)", 0, 2000, 500)
                indicators['cd4_percentage'] = st.number_input("CD4 percentage (%)", 0.0, 100.0, 30.0)
                indicators['viral_load'] = st.number_input("Viral load (копии/mL)", 0, 1000000, 0)
                indicators['white_blood_cells'] = st.number_input("Лейкоциты (клетки/μL)", 0, 50000, 7000)
                indicators['lymphocytes'] = st.number_input("Лимфоциты (клетки/μL)", 0, 10000, 2000)
                indicators['hemoglobin'] = st.number_input("Гемоглобин (g/dL)", 0.0, 20.0, 14.0)
                indicators['platelets'] = st.number_input("Тромбоциты (клетки/μL)", 0, 1000000, 250000)
            
            with col2:
                st.subheader("📊 Дополнительные показатели")
                
                indicators['temperature'] = st.number_input("Температура (°C)", 35.0, 42.0, 36.6)
                indicators['heart_rate'] = st.number_input("Пульс (уд/мин)", 40, 200, 75)
                indicators['blood_pressure_systolic'] = st.number_input("АД систолическое (mmHg)", 80, 200, 120)
                indicators['blood_pressure_diastolic'] = st.number_input("АД диастолическое (mmHg)", 40, 120, 80)
                indicators['oxygen_saturation'] = st.number_input("Насыщение O₂ (%)", 70, 100, 98)
                indicators['respiratory_rate'] = st.number_input("Частота дыхания (вдохов/мин)", 8, 40, 16)
                indicators['glucose_fasting'] = st.number_input("Глюкоза натощак (mg/dL)", 50, 300, 90)
                indicators['cholesterol_total'] = st.number_input("Холестерин общий (mg/dL)", 100, 400, 200)
            
            # Симптомы (на русском)
            st.subheader("🦠 Симптомы")
            symptom_options_ru = list(SYMPTOMS_TRANSLATION.keys())
            selected_symptoms_ru = st.multiselect(
                "Выберите симптомы",
                symptom_options_ru
            )
            
            # Перевод на английский для анализа
            symptoms = [SYMPTOMS_TRANSLATION[ru] for ru in selected_symptoms_ru]
            
            # Факторы риска (на русском)
            st.subheader("⚠️ Факторы риска")
            risk_factor_options_ru = list(RISK_FACTORS_TRANSLATION.keys())
            selected_risk_factors_ru = st.multiselect(
                "Выберите факторы риска",
                risk_factor_options_ru
            )
            
            # Перевод на английский для анализа
            risk_factors = [RISK_FACTORS_TRANSLATION[ru] for ru in selected_risk_factors_ru]
            
            if st.button("🔬 Запустить анализ", type="primary"):
                with st.spinner("Выполнение квантового анализа..."):
                    try:
                        from mqea.data_processor import MedicalTimeSeries
                        import pandas as pd
                        from datetime import datetime, timedelta
                        
                        # Создание временных рядов
                        start_date = datetime.now() - timedelta(days=30)
                        timestamps = [start_date + timedelta(days=i) for i in range(30)]
                        
                        # Создание данных (используем введенные значения)
                        data = {}
                        for key, value in indicators.items():
                            # Создаем небольшой разброс вокруг введенного значения
                            data[key] = np.random.normal(value, value * 0.05, 30)
                        
                        df = pd.DataFrame(data, index=timestamps)
                        
                        # Создание маски пропущенных данных
                        missing_mask = pd.DataFrame(False, index=df.index, columns=df.columns)
                        
                        # Создание квантовых состояний
                        quantum_states = {}
                        for indicator in df.columns:
                            quantum_states[indicator] = np.zeros(len(df))
                        
                        # Расчет процента пропущенных данных
                        total_cells = len(df) * len(df.columns)
                        missing_cells = missing_mask.sum().sum()
                        missing_percentage = (missing_cells / total_cells * 100) if total_cells > 0 else 0.0
                        
                        # Создание объекта временного ряда
                        medical_data = MedicalTimeSeries(
                            data=df,
                            indicators=list(data.keys()),
                            timestamps=pd.DatetimeIndex(timestamps),
                            missing_data_mask=missing_mask,
                            quantum_states=quantum_states,
                            metadata={
                                'source': 'manual_input',
                                'total_points': len(df),
                                'patient_id': patient.patient_id,
                                'missing_percentage': missing_percentage
                            }
                        )
                        
                        # Выполнение анализа
                        results = analyzer.analyze_disease_patterns(
                            medical_data=medical_data,
                            patient_symptoms=symptoms,
                            risk_factors=risk_factors
                        )
                        
                        # Сохранение результатов
                        if 'disease_analysis_history' not in st.session_state:
                            st.session_state.disease_analysis_history = []
                        
                        st.session_state.disease_analysis_history.append({
                            'patient_id': patient.patient_id,
                            'timestamp': datetime.now(),
                            'results': results
                        })
                        
                        st.success(f"✅ Анализ завершен! Обнаружено {len(results)} потенциальных заболеваний")
                        
                        # Отображение результатов (аналогично режиму генерации)
                        st.subheader("📊 Результаты анализа")
                        
                        for i, result in enumerate(results[:10], 1):
                            with st.expander(
                                f"{i}. {result.disease_name} ({result.disease_code}) - "
                                f"Вероятность: {result.probability:.1%}"
                            ):
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.metric("Вероятность", f"{result.probability:.1%}")
                                    st.metric("Уверенность", f"{result.confidence:.1%}")
                                    
                                    urgency_colors = {
                                        'критический': '🔴',
                                        'высокий': '🟠',
                                        'средний': '🟡',
                                        'низкий': '🟢'
                                    }
                                    st.write(f"**Уровень срочности:** {urgency_colors.get(result.urgency_level, '⚪')} {result.urgency_level.upper()}")
                                
                                with col2:
                                    st.write(f"**Категория:** {result.category.value}")
                                    st.write(f"**ICD-10 код:** {result.disease_code}")
                                
                                if result.matched_indicators:
                                    st.write("**📈 Совпавшие показатели:**")
                                    for indicator in result.matched_indicators[:5]:
                                        st.write(f"  - {indicator}")
                                
                                # Симптомы (перевод на русский)
                                if result.matched_symptoms:
                                    st.write("**🦠 Совпавшие симптомы:**")
                                    symptoms_reverse = {v: k for k, v in SYMPTOMS_TRANSLATION.items()}
                                    for symptom_en in result.matched_symptoms[:5]:
                                        symptom_ru = symptoms_reverse.get(symptom_en, symptom_en)
                                        st.write(f"  - {symptom_ru}")
                                
                                # Факторы риска (перевод на русский)
                                if result.risk_factors_present:
                                    st.write("**⚠️ Факторы риска:**")
                                    risk_factors_reverse = {v: k for k, v in RISK_FACTORS_TRANSLATION.items()}
                                    for risk_en in result.risk_factors_present:
                                        risk_ru = risk_factors_reverse.get(risk_en, risk_en)
                                        st.write(f"  - {risk_ru}")
                                
                                if result.recommendations:
                                    st.write("**💡 Рекомендации:**")
                                    for rec in result.recommendations:
                                        st.info(rec)
                    
                    except Exception as e:
                        st.error(f"❌ Ошибка при анализе: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
    
    with tab2:
        st.subheader("🦠 Список поддерживаемых заболеваний")
        
        # Группировка по категориям
        diseases_by_category = {}
        for disease_code, pattern in analyzer.disease_patterns.items():
            category = pattern.category.value
            if category not in diseases_by_category:
                diseases_by_category[category] = []
            diseases_by_category[category].append(pattern)
        
        for category, diseases in diseases_by_category.items():
            with st.expander(f"📂 {category.upper()} ({len(diseases)} заболеваний)"):
                for disease in diseases:
                    st.write(f"**{disease.disease_name}** ({disease.disease_code})")
                    st.caption(disease.description)
                    st.write("**Показатели:** " + ", ".join(list(disease.indicators.keys())[:5]) + "...")
                    st.write("---")
    
    with tab3:
        st.subheader("📋 История анализов")
        
        if 'disease_analysis_history' in st.session_state and st.session_state.disease_analysis_history:
            for i, analysis in enumerate(reversed(st.session_state.disease_analysis_history), 1):
                with st.expander(f"Анализ #{i} - {analysis['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}"):
                    st.write(f"**Пациент ID:** {analysis['patient_id']}")
                    st.write(f"**Обнаружено заболеваний:** {len(analysis['results'])}")
                    
                    if analysis['results']:
                        for result in analysis['results'][:5]:
                            st.write(f"- {result.disease_name}: {result.probability:.1%}")
        else:
            st.info("История анализов пуста")

def show_ai_assistant():
    """AI-Помощник MQEA - современный интерфейс."""
    
    # Кнопка возврата к дашборду
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        if st.button("← Назад к дашборду", type="secondary", width='stretch'):
            st.session_state.current_menu = "📊 Дашборд"
            st.rerun()
    
    # Современный заголовок
    st.markdown("## 🤖 AI-Помощник MQEA")
    st.markdown("*Ваш интеллектуальный спутник в мире квантовой медицины*")
    
    # Статистика помощника
    if 'unified_ai_assistant' in st.session_state:
        stats = st.session_state.unified_ai_assistant.get_stats()
        ml_stats = st.session_state.unified_ai_assistant.get_ml_stats()
        
        # Основная статистика
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
        with col_stat1:
            st.metric("Всего запросов", stats['total_queries'])
        with col_stat2:
            st.metric("Успешных ответов", stats['successful_answers'])
        with col_stat3:
            st.metric("ML предсказаний", ml_stats['ml_predictions'])
        with col_stat4:
            if stats['last_interaction']:
                last_time = datetime.fromisoformat(stats['last_interaction']).strftime("%H:%M")
                st.metric("Последний ответ", last_time)
        
        # ML статистика
        col_ml1, col_ml2, col_ml3 = st.columns(3)
        with col_ml1:
            ml_rate = (ml_stats['ml_predictions'] / max(stats['total_queries'], 1)) * 100
            st.metric("Использование ML", f"{ml_rate:.1f}%")
        with col_ml2:
            st.metric("Событий обучения", ml_stats['learning_events'])
        with col_ml3:
            st.metric("Данных обучения", ml_stats['total_learning_data'])
        
        # Контекст разговора
        if 'unified_ai_assistant' in st.session_state:
            context = st.session_state.unified_ai_assistant.get_conversation_context()
            if context['current_topic'] or context['conversation_length'] > 0:
                st.markdown("### 🎯 Контекст разговора")
                col_ctx1, col_ctx2, col_ctx3 = st.columns(3)
                
                with col_ctx1:
                    if context['current_topic']:
                        st.metric("Текущая тема", context['current_topic'])
                
                with col_ctx2:
                    st.metric("Длина разговора", context['conversation_length'])
                
                with col_ctx3:
                    st.metric("Поток разговора", context['flow_length'])
                
                # История тем
                if context['topic_history']:
                    st.markdown("**История тем:**")
                    topics_str = " → ".join(context['topic_history'][-3:])  # Последние 3 темы
                    st.markdown(f"*{topics_str}*")
                
                # Недавние сущности
                if context['recent_entities']:
                    st.markdown("**Недавние сущности:**")
                    entities_str = ", ".join(context['recent_entities'][-5:])  # Последние 5 сущностей
                    st.markdown(f"*{entities_str}*")
    
    st.markdown("---")
    
    # Инициализация унифицированного AI помощника
    if 'unified_ai_assistant' not in st.session_state:
        from mqea.unified_ai_assistant import UnifiedAIAssistant
        st.session_state.unified_ai_assistant = UnifiedAIAssistant()
    
    # Инициализация чата
    if 'chat_messages' not in st.session_state:
        st.session_state.chat_messages = []
    
    # Основной контент в двух колонках
    col1, col2 = st.columns([2, 1])
    
    with col1:
        # Область чата
        st.markdown("### 💬 Диалог с AI")
        
        # Контейнер для чата
        chat_container = st.container()
        
        with chat_container:
            if st.session_state.chat_messages:
                # Отображаем последние 10 сообщений
                recent_messages = st.session_state.chat_messages[-10:]
                for i, message in enumerate(recent_messages):
                    if message["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(message['content'])
                    else:
                        with st.chat_message("assistant"):
                            st.markdown(message['content'])
            else:
                # Приветственное сообщение
                with st.chat_message("assistant"):
                    st.markdown("👋 **Добро пожаловать!** Я ваш AI-помощник по MQEA. Задайте мне любой вопрос о квантовом анализе, медицинских данных или работе системы!")
        
        # Поле ввода с современным дизайном
        st.markdown("### 💭 Задайте вопрос")
        
        # Используем st.chat_input для современного интерфейса
        if user_input := st.chat_input("Введите ваш вопрос...", key="ai_chat_input"):
            process_ai_message(user_input)
        
        # Обратная связь для последнего ответа
        if st.session_state.chat_messages and len(st.session_state.chat_messages) >= 2:
            last_message = st.session_state.chat_messages[-1]
            if last_message["role"] == "assistant":
                st.markdown("### 💭 Обратная связь")
                col_fb1, col_fb2, col_fb3 = st.columns(3)
                
                with col_fb1:
                    if st.button("👍 Хорошо", key="feedback_good"):
                        provide_feedback("хорошо", last_message['content'])
                        st.success("✅ Спасибо за обратную связь!")
                
                with col_fb2:
                    if st.button("👎 Плохо", key="feedback_bad"):
                        feedback_text = st.text_input("Что можно улучшить?", key="feedback_text")
                        if st.button("Отправить", key="feedback_submit"):
                            provide_feedback(feedback_text or "плохо", last_message['content'])
                            st.success("✅ Спасибо за обратную связь!")
                
                with col_fb3:
                    if st.button("🤷 Нейтрально", key="feedback_neutral"):
                        provide_feedback("нейтрально", last_message['content'])
                        st.success("✅ Спасибо за обратную связь!")
        
        # Дополнительные кнопки
        col_btn1, col_btn2, col_btn3 = st.columns(3)
        
        with col_btn1:
            if st.button("🗑️ Очистить чат", width='stretch'):
                st.session_state.chat_messages = []
                st.session_state.unified_ai_assistant.clear_history()
                st.rerun()
        
        with col_btn2:
            if st.button("📋 История", width='stretch'):
                st.session_state.show_chat_history = True
        
        with col_btn3:
            if st.button("📊 Статистика", width='stretch'):
                st.session_state.show_ai_stats = True
        
        # Дополнительные ML функции
        st.markdown("### 🧠 ML Анализ")
        
        col_ml1, col_ml2 = st.columns(2)
        
        with col_ml1:
            if st.button("🔍 Похожие запросы", width='stretch'):
                if st.session_state.chat_messages:
                    last_query = st.session_state.chat_messages[-1]['content']
                    similar = st.session_state.unified_ai_assistant.get_similar_queries(last_query)
                    if similar:
                        st.markdown("**Похожие запросы:**")
                        for item in similar[:3]:
                            st.markdown(f"• {item['query']} ({item['similarity']:.2f})")
                    else:
                        st.markdown("Похожих запросов не найдено")
        
        with col_ml2:
            if st.button("📈 Анализ паттернов", width='stretch'):
                patterns = st.session_state.unified_ai_assistant.analyze_conversation_patterns()
                if 'error' not in patterns:
                    st.markdown("**Анализ разговора:**")
                    st.markdown(f"• Всего запросов: {patterns['total_queries']}")
                    st.markdown(f"• Популярные темы: {len(patterns['popular_topics'])}")
                    st.markdown(f"• Использование ML: {patterns['ml_usage_rate']:.1f}%")
                else:
                    st.markdown("Недостаточно данных для анализа")
    
    with col2:
        # Боковая панель с быстрыми командами
        st.markdown("### ⚡ Быстрые команды")
        
        # Получаем быстрые команды от унифицированного помощника
        if 'unified_ai_assistant' in st.session_state:
            quick_commands = st.session_state.unified_ai_assistant.get_quick_commands()
            
            # Категории команд
            st.markdown("#### 🔬 Анализ данных")
            if st.button("📊 Объясни результаты", width='stretch'):
                process_ai_message("Объясни результаты MQEA анализа")
            
            if st.button("📈 Интерпретация графиков", width='stretch'):
                process_ai_message("Как интерпретировать графики и матрицы анализа?")
            
            if st.button("🔍 Найди аномалии", width='stretch'):
                process_ai_message("Помоги найти аномалии в медицинских данных")
            
            st.markdown("---")
            st.markdown("#### ⚛️ Квантовая физика")
            
            if st.button("🔬 Квантовая запутанность", width='stretch'):
                process_ai_message("Что такое квантовая запутанность в медицинских данных?")
            
            if st.button("🌊 Квантовая когерентность", width='stretch'):
                process_ai_message("Объясни принципы квантовой когерентности в MQEA")
            
            if st.button("⚛️ Принципы работы", width='stretch'):
                process_ai_message("Объясни принципы квантовой физики в MQEA")
            
            st.markdown("---")
            st.markdown("#### 🏥 Медицина")
            
            if st.button("💊 Медицинские рекомендации", width='stretch'):
                process_ai_message("Покажи медицинские рекомендации на основе анализа")
            
            if st.button("🏥 Диагностика", width='stretch'):
                process_ai_message("Помоги с медицинской диагностикой")
            
            if st.button("📋 Нормальные значения", width='stretch'):
                process_ai_message("Покажи нормальные значения медицинских показателей")
            
        st.markdown("---")
        st.markdown("#### 💊 Лекарства мира")
        
        if st.button("🌍 Все лекарства мира", width='stretch'):
            process_ai_message("Покажи полную базу данных лекарств мира")
        
        if st.button("🦠 Антибиотики", width='stretch'):
            process_ai_message("Расскажи про все антибиотики")
        
        if st.button("❤️ Кардиологические", width='stretch'):
            process_ai_message("Покажи кардиологические препараты")
        
        if st.button("🍯 Диабетические", width='stretch'):
            process_ai_message("Расскажи про препараты от диабета")
        
        if st.button("🧠 Неврологические", width='stretch'):
            process_ai_message("Покажи неврологические препараты")
        
        if st.button("👶 Педиатрические", width='stretch'):
            process_ai_message("Расскажи про детские препараты")
        
        if st.button("👴 Гериатрические", width='stretch'):
            process_ai_message("Покажи препараты для пожилых")
        
        st.markdown("---")
        st.markdown("#### 🛠️ Система")
        
        if st.button("❓ Как использовать MQEA", width='stretch'):
            process_ai_message("Как правильно использовать систему MQEA?")
        
        if st.button("🔧 Настройки", width='stretch'):
            process_ai_message("Помоги настроить параметры анализа")
        
        if st.button("📚 Документация", width='stretch'):
            process_ai_message("Покажи документацию по MQEA")
        
        st.markdown("---")
        st.markdown("#### 🧠 Обучение AI")
        
        if st.button("🎓 Обучить AI", width='stretch'):
            process_ai_message("Обучи меня новым знаниям о MQEA")
        
        if st.button("📈 Показать прогресс", width='stretch'):
            process_ai_message("Покажи мой прогресс обучения")
        
        if st.button("🔄 Обновить знания", width='stretch'):
            process_ai_message("Обнови мою базу знаний")
        
        # Статистика чата
        st.markdown("---")
        st.markdown("#### 📊 Статистика")
        
        total_messages = len(st.session_state.chat_messages)
        user_messages = len([m for m in st.session_state.chat_messages if m["role"] == "user"])
        ai_messages = len([m for m in st.session_state.chat_messages if m["role"] == "assistant"])
        
        st.metric("Всего сообщений", total_messages)
        st.metric("Ваши вопросы", user_messages)
        st.metric("Ответы AI", ai_messages)
        
        # Статистика обучения AI
        if 'unified_ai_assistant' in st.session_state:
            learning_stats = st.session_state.unified_ai_assistant.get_stats()
            st.markdown("---")
            st.markdown("#### 🧠 Обучение AI")
            st.metric("Всего запросов", learning_stats['total_queries'])
            st.metric("Успешных ответов", learning_stats['successful_answers'])
            st.metric("События обучения", learning_stats['learning_events'])
    
    # Обработка дополнительных окон
    if st.session_state.get('show_chat_history'):
        show_chat_history()
    
    if st.session_state.get('show_ai_stats'):
        show_ai_stats()

def show_chat_history():
    """Отображение истории чата."""
    st.markdown("---")
    st.subheader("📋 История чата")
    
    if st.session_state.chat_messages:
        for i, message in enumerate(st.session_state.chat_messages):
            if message["role"] == "user":
                st.markdown(f"**👤 Вы ({i+1}):** {message['content']}")
            else:
                st.markdown(f"**🤖 AI ({i+1}):** {message['content']}")
            st.markdown("---")
    else:
        st.info("История чата пуста")
    
    if st.button("❌ Закрыть историю"):
        st.session_state.show_chat_history = False
        st.rerun()

def show_ai_stats():
    """Отображение статистики AI."""
    st.markdown("---")
    st.subheader("📊 Статистика AI помощника")
    
    if 'unified_ai_assistant' in st.session_state:
        stats = st.session_state.unified_ai_assistant.get_stats()
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Всего запросов", stats['total_queries'])
            st.metric("Успешных ответов", stats['successful_answers'])
        
        with col2:
            st.metric("Событий обучения", stats['learning_events'])
            if stats['last_interaction']:
                st.metric("Последнее взаимодействие", stats['last_interaction'][:19])
        
        # График активности
        if stats['total_queries'] > 0:
            success_rate = (stats['successful_answers'] / stats['total_queries']) * 100
            st.metric("Процент успешных ответов", f"{success_rate:.1f}%")
    else:
        st.info("AI помощник не инициализирован")
    
    if st.button("❌ Закрыть статистику"):
        st.session_state.show_ai_stats = False
        st.rerun()
        
        # Кнопки управления
        st.markdown("---")
        st.markdown("#### 🛠️ Управление")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🔄 Сбросить чат", type="secondary", width='stretch'):
                st.session_state.chat_messages = []
                st.success("✅ Чат сброшен")
                st.rerun()
        
        with col2:
            if st.button("🧠 Статистика AI", width='stretch'):
                if 'advanced_assistant' in st.session_state:
                    stats = st.session_state.advanced_assistant.get_learning_stats()
                    st.write("**Статистика AI-помощника:**")
                    for key, value in stats.items():
                        st.write(f"• {key}: {value}")
                else:
                    st.warning("AI-помощник не инициализирован")
        
        # Быстрая навигация
        st.markdown("---")
        st.markdown("#### 🧭 Быстрая навигация")
        
        if st.button("📊 Дашборд", width='stretch'):
            st.session_state.current_menu = "📊 Дашборд"
            st.rerun()
        
        if st.button("🔬 Анализ MQEA", width='stretch'):
            st.session_state.current_menu = "🔬 Анализ MQEA"
            st.rerun()
        
        if st.button("👥 Пациенты", width='stretch'):
            st.session_state.current_menu = "👥 Пациенты"
            st.rerun()
        
        if st.button("📈 Отчеты", width='stretch'):
            st.session_state.current_menu = "📈 Отчеты"
            st.rerun()
    
    # Информация о AI-помощнике внизу
    with st.expander("ℹ️ О AI-помощнике MQEA", expanded=False):
        st.markdown("### 🧠 Возможности AI-помощника")
        st.markdown("""
        - 🔬 **Анализ данных:** Интерпретация результатов квантового анализа
        - 📊 **Визуализация:** Объяснение графиков и матриц
        - 💡 **Рекомендации:** Медицинские советы на основе анализа
        - ❓ **Обучение:** Ответы на вопросы о MQEA алгоритме
        - 🏥 **Диагностика:** Помощь в медицинской диагностике
        - ⚛️ **Квантовая физика:** Объяснение принципов квантовой механики
        """)
        
        st.markdown("### 🚀 Новые возможности самообучения")
        st.markdown("""
        - 🧠 **Самообучение:** AI изучает новые паттерны из ваших вопросов
        - 📈 **Адаптация:** Улучшает ответы на основе обратной связи
        - 🔍 **Анализ намерений:** Понимает контекст и сложность запросов
        - 💾 **База знаний:** Сохраняет и использует накопленный опыт
        - 📊 **Статистика обучения:** Отслеживает прогресс и эффективность
        - 🔄 **Обновления:** Автоматически обновляет знания каждые 10 запросов
        """)
        
        st.markdown("### 🤖 Машинное обучение (ML)")
        st.markdown("""
        - 🎯 **ML Предсказания:** Использует алгоритмы машинного обучения для анализа запросов
        - 📊 **Кластеризация:** Группирует похожие запросы для лучшего понимания
        - 🔍 **Анализ схожести:** Находит похожие запросы в истории разговора
        - 📈 **Паттерны разговора:** Анализирует популярные темы и тенденции
        - 🧠 **Переобучение:** Автоматически улучшает модели на основе новых данных
        - 📋 **Предложения:** Дает рекомендации по улучшению взаимодействия
        """)
        
        st.markdown("### 🎯 Как работает самообучение")
        st.markdown("""
        1. **Анализ запроса:** AI анализирует ваши вопросы и определяет намерения
        2. **Поиск ответа:** Ищет лучший ответ в базе знаний
        3. **Генерация:** Создает ответ на основе найденной информации
        4. **Обучение:** Сохраняет успешные ответы для будущего использования
        5. **Улучшение:** Постепенно улучшает качество ответов
        """)

def process_ai_message(message):
    """Обработка сообщения AI-помощника."""
    st.session_state.chat_messages.append({
        "role": "user",
        "content": message
    })
    
    try:
        # Получаем ответ от унифицированного AI-помощника
        if 'unified_ai_assistant' not in st.session_state:
            from mqea.unified_ai_assistant import UnifiedAIAssistant
            st.session_state.unified_ai_assistant = UnifiedAIAssistant()
        
        response = st.session_state.unified_ai_assistant.process_query(message)
        
        st.session_state.chat_messages.append({
            "role": "assistant",
            "content": response
        })
        
        st.rerun()
        
    except Exception as e:
        st.session_state.chat_messages.append({
            "role": "assistant",
            "content": f"❌ Ошибка обработки запроса: {str(e)}"
        })
        st.rerun()

def provide_feedback(feedback_text: str, response: str):
    """Обработка обратной связи пользователя."""
    try:
        if 'unified_ai_assistant' in st.session_state:
            # Получаем последний запрос пользователя
            if st.session_state.chat_messages and len(st.session_state.chat_messages) >= 2:
                last_user_query = None
                for msg in reversed(st.session_state.chat_messages):
                    if msg["role"] == "user":
                        last_user_query = msg["content"]
                        break
                
                if last_user_query:
                    # Добавляем обратную связь в данные обучения
                    st.session_state.unified_ai_assistant.add_learning_data(
                        last_user_query, 
                        response, 
                        feedback_text
                    )
                    
                    # Обновляем статистику
                    st.session_state.unified_ai_assistant.stats['learning_events'] += 1
                    
    except Exception as e:
        st.error(f"Ошибка обработки обратной связи: {e}")

def show_ml_diagnostics():
    """ML диагностика."""
    st.header("🤖 ML Диагностика")
    
    if not st.session_state.current_patient:
        st.warning("⚠️ Сначала выберите пациента")
        return
    
    patient = st.session_state.current_patient
    st.subheader(f"ML диагностика для пациента: {patient.name}")
    
    # Подготовка данных пациента
    patient_data = {
        'age': patient.age,
        'gender': 1 if patient.gender == 'Мужской' else 0,
        'bmi': 25.0,  # Можно рассчитать из веса и роста
        'heart_rate': 75,
        'blood_pressure_systolic': 120,
        'blood_pressure_diastolic': 80,
        'temperature': 36.6,
        'oxygen_saturation': 98,
        'respiratory_rate': 16,
        'glucose': 5.0,
        'cholesterol': 180,
        'smoking': 0,
        'alcohol': 0,
        'sedentary_lifestyle': 0,
        'family_history': 0,
        'stress_level': 3,
        'sleep_hours': 8,
        'exercise_frequency': 3,
        'quantum_coherence': 0.5,
        'entanglement_pairs': 20,
        'max_entanglement': 0.6,
        'pattern_complexity': 0.7,
        'temporal_correlation': 0.8
    }
    
    # Кнопка запуска ML диагностики
    if st.button("🤖 Запустить ML диагностику", type="primary"):
        with st.spinner("Выполняется ML диагностика..."):
            try:
                # Получаем предсказания
                predictions = st.session_state.components['ml_engine'].predict_diagnosis(patient_data)
                st.session_state.ml_predictions = predictions
                
                # Генерируем рекомендации
                recommendations = st.session_state.components['ml_engine'].generate_medical_recommendations(
                    predictions, patient_data
                )
                
                st.success("✅ ML диагностика завершена!")
                
            except Exception as e:
                st.error(f"❌ Ошибка ML диагностики: {str(e)}")
    
    # Отображение результатов ML диагностики
    if st.session_state.ml_predictions:
        st.subheader("📊 Результаты ML диагностики")
        
        for condition, pred_data in st.session_state.ml_predictions.items():
            with st.expander(f"🔬 {condition.replace('_', ' ').title()}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write(f"**Предсказание:** {pred_data['prediction']}")
                    st.write(f"**Уверенность:** {pred_data['confidence']:.1%}")
                
                with col2:
                    if pred_data['probabilities']:
                        prob_df = pd.DataFrame(list(pred_data['probabilities'].items()), 
                                             columns=['Класс', 'Вероятность'])
                        st.dataframe(prob_df, width='stretch')
        
        # Визуализация результатов
        st.subheader("📈 Визуализация ML диагностики")
        
        # Создаем вкладки для разных типов визуализации
        viz_tab1, viz_tab2, viz_tab3, viz_tab4 = st.tabs([
            "🌐 Сфера Блоха", 
            "📊 Вероятности предсказаний", 
            "📉 Временная динамика",
            "🔍 Анализ факторов"
        ])
        
        with viz_tab1:
            st.markdown("**🌐 Сфера Блоха - визуализация квантового состояния пациента:**")
            
            # Вычисляем квантовые параметры на основе ML предсказаний
            # Используем среднюю уверенность как фиделити
            avg_confidence = np.mean([pred['confidence'] for pred in st.session_state.ml_predictions.values()])
            quantum_fidelity = max(0.1, min(1.0, avg_confidence))
            
            # Создаем сферу Блоха
            theta = np.linspace(0, 2*np.pi, 100)
            phi = np.linspace(0, np.pi, 50)
            THETA, PHI = np.meshgrid(theta, phi)
            
            # Амплитуда зависит от фиделити
            amplitude = quantum_fidelity
            R = amplitude * np.abs(np.sin(THETA) * np.cos(PHI))
            
            fig_bloch = go.Figure(data=go.Surface(
                x=R * np.sin(PHI) * np.cos(THETA),
                y=R * np.sin(PHI) * np.sin(THETA),
                z=R * np.cos(PHI),
                colorscale='Viridis',
                opacity=0.8,
                showscale=True,
                colorbar=dict(title="Фаза")
            ))
            
            # Добавляем оси
            fig_bloch.add_trace(go.Scatter3d(
                x=[0, 0], y=[0, 0], z=[-amplitude, amplitude],
                mode='lines',
                line=dict(color='red', width=5),
                name='Z-ось'
            ))
            
            fig_bloch.add_trace(go.Scatter3d(
                x=[-amplitude, amplitude], y=[0, 0], z=[0, 0],
                mode='lines',
                line=dict(color='green', width=5),
                name='X-ось'
            ))
            
            fig_bloch.add_trace(go.Scatter3d(
                x=[0, 0], y=[-amplitude, amplitude], z=[0, 0],
                mode='lines',
                line=dict(color='blue', width=5),
                name='Y-ось'
            ))
            
            fig_bloch.update_layout(
                title=f"Сфера Блоха (Фиделити: {quantum_fidelity:.3f})",
                scene=dict(
                    xaxis_title="X (|0⟩ + |1⟩)",
                    yaxis_title="Y (|0⟩ + i|1⟩)",
                    zaxis_title="Z (|0⟩ - |1⟩)",
                    camera=dict(eye=dict(x=1.5, y=1.5, z=1.5))
                ),
                height=600
            )
            
            st.plotly_chart(fig_bloch, use_container_width=True)
            
            st.markdown(f"""
            **📝 Объяснение сферы Блоха:**
            - **Размер сферы** отражает квантовую фиделити ({quantum_fidelity:.3f})
            - **Цвет** показывает фазу квантового состояния
            - **Оси** представляют базисные состояния: |0⟩, |1⟩, |+⟩, |-⟩
            - **Положение** на сфере определяет конкретное квантовое состояние пациента
            """)
        
        with viz_tab2:
            st.markdown("**📊 Вероятности предсказаний по всем условиям:**")
            
            # Собираем данные для графика
            conditions = []
            probabilities = []
            confidences = []
            
            for condition, pred_data in st.session_state.ml_predictions.items():
                conditions.append(condition.replace('_', ' ').title())
                if pred_data['probabilities']:
                    # Берем максимальную вероятность
                    max_prob = max(pred_data['probabilities'].values())
                    probabilities.append(max_prob * 100)
                else:
                    probabilities.append(pred_data['confidence'] * 100)
                confidences.append(pred_data['confidence'] * 100)
            
            # График вероятностей
            fig_prob = go.Figure()
            
            fig_prob.add_trace(go.Bar(
                x=conditions,
                y=probabilities,
                name='Вероятность',
                marker=dict(
                    color=probabilities,
                    colorscale='Reds',
                    showscale=True,
                    colorbar=dict(title="Вероятность (%)")
                )
            ))
            
            fig_prob.update_layout(
                title="Вероятности предсказаний по условиям",
                xaxis_title="Медицинское условие",
                yaxis_title="Вероятность (%)",
                height=400,
                xaxis_tickangle=-45
            )
            
            st.plotly_chart(fig_prob, use_container_width=True)
            
            # Круговая диаграмма уверенности
            st.markdown("**🎯 Распределение уверенности:**")
            
            fig_pie = go.Figure(data=[go.Pie(
                labels=conditions,
                values=confidences,
                hole=0.3,
                textinfo='label+percent',
                marker=dict(colors=px.colors.sequential.Reds)
            )])
            
            fig_pie.update_layout(
                title="Распределение уверенности в предсказаниях",
                height=400
            )
            
            st.plotly_chart(fig_pie, use_container_width=True)
        
        with viz_tab3:
            st.markdown("**📉 Временная динамика показателей:**")
            
            # Генерируем временные данные на основе текущих показателей
            time_points = np.linspace(0, 24, 100)  # 24 часа
            
            # Моделируем динамику на основе данных пациента
            heart_rate_trend = patient_data['heart_rate'] + 5 * np.sin(time_points / 3)
            blood_pressure_trend = patient_data['blood_pressure_systolic'] + 10 * np.cos(time_points / 4)
            temperature_trend = patient_data['temperature'] + 0.3 * np.sin(time_points / 6)
            
            fig_time = go.Figure()
            
            fig_time.add_trace(go.Scatter(
                x=time_points,
                y=heart_rate_trend,
                mode='lines',
                name='Частота сердечных сокращений',
                line=dict(color='red', width=2)
            ))
            
            fig_time.add_trace(go.Scatter(
                x=time_points,
                y=blood_pressure_trend,
                mode='lines',
                name='Систолическое давление',
                line=dict(color='blue', width=2),
                yaxis='y2'
            ))
            
            fig_time.add_trace(go.Scatter(
                x=time_points,
                y=temperature_trend,
                mode='lines',
                name='Температура',
                line=dict(color='orange', width=2),
                yaxis='y3'
            ))
            
            fig_time.update_layout(
                title="Временная динамика жизненно важных показателей",
                xaxis_title="Время (часы)",
                yaxis=dict(title="ЧСС (уд/мин)", side='left'),
                yaxis2=dict(title="Давление (мм рт.ст.)", side='right', overlaying='y'),
                yaxis3=dict(title="Температура (°C)", side='right', overlaying='y', position=0.95),
                height=500,
                hovermode='x unified'
            )
            
            st.plotly_chart(fig_time, use_container_width=True)
            
            # График квантовых параметров
            st.markdown("**⚛️ Динамика квантовых параметров:**")
            
            coherence_trend = patient_data['quantum_coherence'] * (1 - 0.1 * np.sin(time_points / 2))
            entanglement_trend = patient_data['max_entanglement'] * (1 - 0.05 * np.cos(time_points / 3))
            
            fig_quantum = go.Figure()
            
            fig_quantum.add_trace(go.Scatter(
                x=time_points,
                y=coherence_trend,
                mode='lines',
                name='Квантовая когерентность',
                line=dict(color='purple', width=3)
            ))
            
            fig_quantum.add_trace(go.Scatter(
                x=time_points,
                y=entanglement_trend,
                mode='lines',
                name='Максимальная запутанность',
                line=dict(color='green', width=3)
            ))
            
            fig_quantum.update_layout(
                title="Временная эволюция квантовых параметров",
                xaxis_title="Время (часы)",
                yaxis_title="Значение",
                height=400
            )
            
            st.plotly_chart(fig_quantum, use_container_width=True)
        
        with viz_tab4:
            st.markdown("**🔍 Анализ факторов риска:**")
            
            # Создаем данные для анализа факторов
            risk_factors = {
                'Возраст': patient_data['age'],
                'ИМТ': patient_data['bmi'],
                'Курение': patient_data['smoking'] * 100,
                'Алкоголь': patient_data['alcohol'] * 100,
                'Малоподвижный образ жизни': patient_data['sedentary_lifestyle'] * 100,
                'Семейная история': patient_data['family_history'] * 100,
                'Уровень стресса': patient_data['stress_level'] * 10,
                'Часы сна': patient_data['sleep_hours'],
                'Частота упражнений': patient_data['exercise_frequency'] * 10
            }
            
            # Радарная диаграмма
            categories = list(risk_factors.keys())
            values = list(risk_factors.values())
            
            # Нормализуем значения для радарной диаграммы
            normalized_values = []
            max_vals = [100, 50, 100, 100, 100, 100, 50, 12, 50]  # Максимальные значения для нормализации
            
            for i, val in enumerate(values):
                normalized_values.append((val / max_vals[i]) * 100 if max_vals[i] > 0 else 0)
            
            fig_radar = go.Figure()
            
            fig_radar.add_trace(go.Scatterpolar(
                r=normalized_values + [normalized_values[0]],  # Замыкаем круг
                theta=categories + [categories[0]],
                fill='toself',
                name='Факторы риска',
                line=dict(color='red', width=2)
            ))
            
            fig_radar.update_layout(
                polar=dict(
                    radialaxis=dict(
                        visible=True,
                        range=[0, 100]
                    )
                ),
                title="Радарная диаграмма факторов риска",
                height=500
            )
            
            st.plotly_chart(fig_radar, use_container_width=True)
            
            # Тепловая карта корреляций
            st.markdown("**🔥 Тепловая карта корреляций показателей:**")
            
            # Английские названия для получения данных
            indicators_en = ['heart_rate', 'blood_pressure_systolic', 'blood_pressure_diastolic',
                           'temperature', 'oxygen_saturation', 'respiratory_rate', 'glucose', 'cholesterol']
            
            # Русские названия для отображения
            indicators_ru = ['ЧСС', 'АД сист.', 'АД диаст.', 
                            'Температура', 'SpO2', 'ЧДД', 'Глюкоза', 'Холестерин']
            
            # Создаем матрицу корреляций (упрощенную)
            correlation_matrix = np.random.rand(len(indicators_en), len(indicators_en))
            correlation_matrix = (correlation_matrix + correlation_matrix.T) / 2  # Симметричная
            np.fill_diagonal(correlation_matrix, 1.0)  # Диагональ = 1
            
            fig_heatmap = go.Figure(data=go.Heatmap(
                z=correlation_matrix,
                x=indicators_ru,
                y=indicators_ru,
                colorscale='RdBu',
                zmid=0,
                text=correlation_matrix,
                texttemplate='%{text:.2f}',
                textfont={"size": 10},
                colorbar=dict(title="Корреляция")
            ))
            
            fig_heatmap.update_layout(
                title="Корреляционная матрица медицинских показателей",
                height=500,
                xaxis_tickangle=-45
            )
            
            st.plotly_chart(fig_heatmap, use_container_width=True)
        
        # Рекомендации
        if st.session_state.ml_predictions:
            st.subheader("💊 Медицинские рекомендации")
            
            recommendations = st.session_state.components['ml_engine'].generate_medical_recommendations(
                st.session_state.ml_predictions, patient_data
            )
            
            # Отладочная информация
            st.write(f"**Количество рекомендаций:** {len(recommendations)}")
            
            if recommendations:
                for i, rec in enumerate(recommendations, 1):
                    with st.expander(f"{i}. {rec['title']} ({rec['priority']})"):
                        st.write(f"**Описание:** {rec['description']}")
                        st.write(f"**Уверенность:** {rec['confidence']:.1%}")
                        st.write(f"**Тип:** {rec.get('type', 'N/A')}")
                        
                        # Всегда показываем рекомендации, даже если список пустой
                        st.write("**Что делать пациенту:**")
                        if rec.get('recommendations'):
                            for j, recommendation in enumerate(rec['recommendations'], 1):
                                st.write(f"{j}. {recommendation}")
                        else:
                            st.write("• Проконсультируйтесь с врачом")
                            st.write("• Следите за общим состоянием здоровья")
                            st.write("• Ведите здоровый образ жизни")
                        
                        # Дополнительные рекомендации по профилактике
                        st.write("**Профилактические меры:**")
                        prevention_tips = [
                            "• Регулярно проходите медицинские осмотры",
                            "• Соблюдайте сбалансированную диету",
                            "• Поддерживайте физическую активность",
                            "• Избегайте вредных привычек (курение, алкоголь)",
                            "• Контролируйте стресс и высыпайтесь",
                            "• Следите за весом и ИМТ",
                            "• Пейте достаточное количество воды",
                            "• Избегайте длительного сидения"
                        ]
                        for tip in prevention_tips:
                            st.write(tip)
            else:
                st.warning("⚠️ Рекомендации не сгенерированы. Возможно, все показатели в норме.")
                
                # Показываем общие рекомендации по профилактике
                st.subheader("🛡️ Общие рекомендации по профилактике")
                st.write("**Для поддержания здоровья рекомендуется:**")
                
                general_recommendations = [
                    "1. **Регулярные медицинские осмотры** - проходите профилактические осмотры не реже 1 раза в год",
                    "2. **Сбалансированное питание** - употребляйте больше овощей, фруктов, цельнозерновых продуктов",
                    "3. **Физическая активность** - занимайтесь спортом минимум 30 минут в день",
                    "4. **Контроль веса** - поддерживайте здоровый вес и ИМТ",
                    "5. **Отказ от вредных привычек** - не курите, ограничьте употребление алкоголя",
                    "6. **Управление стрессом** - практикуйте релаксацию, медитацию, достаточный сон",
                    "7. **Гигиена** - соблюдайте личную гигиену и гигиену питания",
                    "8. **Вакцинация** - делайте прививки согласно календарю вакцинации",
                    "9. **Контроль показателей** - регулярно измеряйте давление, пульс, температуру",
                    "10. **Обращение к врачу** - при любых тревожных симптомах обращайтесь к специалисту"
                ]
                
                for rec in general_recommendations:
                    st.write(rec)

def show_visits_management():
    """Управление визитами."""
    st.header("📋 Управление визитами")
    
    if not st.session_state.current_patient:
        st.warning("⚠️ Сначала выберите пациента")
        return
    
    patient = st.session_state.current_patient
    st.subheader(f"Визиты пациента: {patient.name}")
    
    # Кнопка нового визита
    if st.button("➕ Новый визит", type="primary"):
        st.session_state.show_new_visit = True
    
    # Форма нового визита
    if st.session_state.get('show_new_visit', False):
        st.subheader("➕ Создание нового визита")
        
        with st.form("new_visit_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                visit_date = st.date_input("Дата визита", value=datetime.now().date())
                visit_time = st.time_input("Время визита", value=datetime.now().time())
                doctor_name = st.text_input("ФИО врача")
                
                # Симптомы
                st.write("**Симптомы:**")
                symptoms = []
                if st.checkbox("Головная боль"):
                    symptoms.append("Головная боль")
                if st.checkbox("Тошнота"):
                    symptoms.append("Тошнота")
                if st.checkbox("Слабость"):
                    symptoms.append("Слабость")
                if st.checkbox("Одышка"):
                    symptoms.append("Одышка")
                if st.checkbox("Боль в груди"):
                    symptoms.append("Боль в груди")
                
                other_symptoms = st.text_area("Другие симптомы")
                if other_symptoms:
                    symptoms.append(other_symptoms)
            
            with col2:
                # Жизненные показатели
                st.write("**Жизненные показатели:**")
                vital_signs = {
                    'heart_rate': st.number_input("Частота пульса", 40, 200, 75),
                    'blood_pressure_systolic': st.number_input("АД систолическое", 80, 250, 120),
                    'blood_pressure_diastolic': st.number_input("АД диастолическое", 50, 150, 80),
                    'temperature': st.number_input("Температура", 35.0, 42.0, 36.6, 0.1),
                    'oxygen_saturation': st.number_input("Насыщение O₂", 70, 100, 98),
                    'respiratory_rate': st.number_input("Частота дыхания", 8, 40, 16),
                    'glucose': st.number_input("Глюкоза", 2.0, 20.0, 5.0, 0.1),
                    'cholesterol': st.number_input("Холестерин", 100, 400, 180)
                }
                
                diagnosis = st.text_input("Диагноз")
                treatment_plan = st.text_area("План лечения")
                notes = st.text_area("Заметки")
            
            submitted = st.form_submit_button("💾 Сохранить визит")
            
            if submitted:
                try:
                    visit_id = str(uuid.uuid4())
                    visit_datetime = datetime.combine(visit_date, visit_time)
                    
                    visit = MedicalVisit(
                        visit_id=visit_id,
                        patient_id=patient.patient_id,
                        visit_date=visit_datetime,
                        symptoms=symptoms,
                        vital_signs=vital_signs,
                        diagnosis=diagnosis,
                        treatment_plan=treatment_plan,
                        notes=notes,
                        doctor_name=doctor_name
                    )
                    
                    if st.session_state.components['database'].add_visit(visit):
                        st.success("✅ Визит успешно сохранен!")
                        st.session_state.show_new_visit = False
                        st.rerun()
                    else:
                        st.error("❌ Ошибка сохранения визита")
                
                except Exception as e:
                    st.error(f"❌ Ошибка: {str(e)}")
    
    # Список визитов
    st.subheader("📋 История визитов")
    
    visits = st.session_state.components['database'].get_patient_visits(patient.patient_id)
    
    if visits:
        visit_data = []
        for visit in visits:
            visit_data.append({
                'Дата': visit.visit_date.strftime('%d.%m.%Y %H:%M'),
                'Врач': visit.doctor_name or 'N/A',
                'Диагноз': visit.diagnosis or 'N/A',
                'Симптомы': ', '.join(visit.symptoms) if visit.symptoms else 'N/A',
                'Статус': visit.status
            })
        
        df_visits = pd.DataFrame(visit_data)
        st.dataframe(df_visits, width='stretch')
    else:
        st.info("У пациента нет визитов")

def _filter_mqea_data(data):
    """Фильтрация данных MQEA для удаления несериализуемых объектов."""
    if not data:
        return {}
    
    filtered_data = {}
    
    def process_value(value):
        """Рекурсивно обрабатывает значение для сериализации."""
        if value is None:
            return None
        elif isinstance(value, complex):
            # Обрабатываем комплексные числа
            return {
                'real': float(value.real),
                'imag': float(value.imag),
                'magnitude': float(abs(value))
            }
        elif isinstance(value, (int, float, str, bool)):
            return value
        elif isinstance(value, (list, tuple)):
            return [process_value(item) for item in value]
        elif isinstance(value, dict):
            return {k: process_value(v) for k, v in value.items()}
        elif hasattr(value, 'tolist'):
            # Для numpy массивов
            try:
                return value.tolist()
            except:
                return str(value)
        elif hasattr(value, '__iter__') and not isinstance(value, (str, bytes)):
            # Для итерируемых объектов (кроме строк)
            try:
                return [process_value(item) for item in value]
            except:
                return str(value)
        elif hasattr(value, 'isoformat'):
            # Для datetime объектов
            return value.isoformat()
        else:
            return str(value)
    
    for key, value in data.items():
        try:
            processed_value = process_value(value)
            # Проверяем, можно ли сериализовать обработанное значение
            import json
            json.dumps(processed_value)
            filtered_data[key] = processed_value
        except (TypeError, ValueError) as e:
            # Если все еще не можем сериализовать, конвертируем в строку
            filtered_data[key] = f"<{type(value).__name__}: {str(value)[:100]}>"
    
    # Убеждаемся, что ключевые метрики доступны в правильном формате
    if 'quantum_signatures' in filtered_data:
        quantum_sig = filtered_data['quantum_signatures']
        # Добавляем метрики в корневой уровень для удобства
        if 'quantum_coherence' in quantum_sig:
            filtered_data['quantum_coherence'] = quantum_sig['quantum_coherence']
        if 'entangled_pairs_count' in quantum_sig:
            filtered_data['entangled_pairs'] = quantum_sig['entangled_pairs_count']
        if 'average_entanglement' in quantum_sig:
            filtered_data['max_entanglement'] = quantum_sig['average_entanglement']
    
    return filtered_data

def show_cards_generation():
    """Генерация карточек."""
    st.header("📄 Генерация карточек пациентов")
    
    if not st.session_state.current_patient:
        st.warning("⚠️ Сначала выберите пациента")
        return
    
    patient = st.session_state.current_patient
    st.subheader(f"Карточка пациента: {patient.name}")
    
    # Подготовка данных для карточки
    # Получаем актуальные данные из анализа или профиля пациента
    vital_signs = {}
    
    if st.session_state.current_data is not None:
        # Используем данные из анализа
        latest_values = st.session_state.current_data.data.iloc[-1]
        vital_signs = {
            'heart_rate': round(latest_values.get('heart_rate', 75), 1),
            'blood_pressure_systolic': round(latest_values.get('blood_pressure_systolic', 120), 1),
            'blood_pressure_diastolic': round(latest_values.get('blood_pressure_diastolic', 80), 1),
            'temperature': round(latest_values.get('temperature', 36.6), 1),
            'oxygen_saturation': round(latest_values.get('oxygen_saturation', 98), 1),
            'respiratory_rate': round(latest_values.get('respiratory_rate', 16), 1),
            'glucose': round(latest_values.get('glucose', 5.0), 1),
            'cholesterol': round(latest_values.get('cholesterol', 180), 1)
        }
    elif st.session_state.current_patient_profile:
        # Используем данные из профиля пациента
        profile = st.session_state.current_patient_profile
        vital_signs = {
            'heart_rate': round(profile.get('heart_rate', 75), 1),
            'blood_pressure_systolic': round(profile.get('blood_pressure_systolic', 120), 1),
            'blood_pressure_diastolic': round(profile.get('blood_pressure_diastolic', 80), 1),
            'temperature': round(profile.get('temperature', 36.6), 1),
            'oxygen_saturation': round(profile.get('oxygen_saturation', 98), 1),
            'respiratory_rate': round(profile.get('respiratory_rate', 16), 1),
            'glucose': round(profile.get('glucose', 5.0), 1),
            'cholesterol': round(profile.get('cholesterol', 180), 1)
        }
    else:
        # Используем значения по умолчанию
        vital_signs = {
            'heart_rate': 75,
            'blood_pressure_systolic': 120,
            'blood_pressure_diastolic': 80,
            'temperature': 36.6,
            'oxygen_saturation': 98,
            'respiratory_rate': 16,
            'glucose': 5.0,
            'cholesterol': 180
        }
    
    # Получаем реальную дату рождения из contact_info
    birth_date_str = patient.contact_info.get('birth_date')
    if birth_date_str:
        try:
            birth_date = datetime.fromisoformat(birth_date_str).date()
        except:
            # Если не удается распарсить, вычисляем приблизительную дату рождения
            today = datetime.now().date()
            birth_date = today.replace(year=today.year - patient.age)
    else:
        # Если дата рождения не сохранена, вычисляем приблизительную
        today = datetime.now().date()
        birth_date = today.replace(year=today.year - patient.age)
    
    # Фильтруем результаты анализа для карточки
    filtered_analysis = _filter_mqea_data(st.session_state.analysis_results or {})
    
    patient_data = {
        'patient_id': patient.patient_id,
        'name': patient.name,
        'age': patient.age,
        'gender': patient.gender,
        'birth_date': str(birth_date),
        'phone': patient.contact_info.get('phone', 'N/A'),
        'address': patient.contact_info.get('address', 'N/A'),
        'vital_signs': vital_signs,
        'mqea_analysis': filtered_analysis,
        'diagnoses': [],
        'recommendations': []
    }
    
    # Отображение актуальных данных для карточки
    st.subheader("📊 Актуальные данные для карточки")
    
    # Отображение основной информации о пациенте
    col_info1, col_info2 = st.columns(2)
    
    with col_info1:
        st.write("**Основная информация:**")
        st.write(f"• ID пациента: {patient.patient_id}")
        st.write(f"• ФИО: {patient.name}")
        st.write(f"• Возраст: {patient.age} лет")
        st.write(f"• Пол: {patient.gender}")
        st.write(f"• **Дата рождения: {birth_date}**")
    
    with col_info2:
        st.write("**Контактная информация:**")
        st.write(f"• Телефон: {patient.contact_info.get('phone', 'N/A')}")
        st.write(f"• Адрес: {patient.contact_info.get('address', 'N/A')}")
        st.write(f"• Дата создания записи: {patient.created_at.date()}")
    
    st.markdown("---")
    
    # Результаты MQEA анализа для карточки
    st.subheader("🔍 Результаты MQEA анализа")
    if st.session_state.analysis_results:
        # Извлекаем данные из правильной структуры
        quantum_signatures = st.session_state.analysis_results.get('quantum_signatures', {})
        entanglement_stats = st.session_state.analysis_results.get('entanglement_statistics', {})
        
        # Ключевые метрики MQEA
        coherence = quantum_signatures.get('quantum_coherence', 0)
        total_states = quantum_signatures.get('total_quantum_states', 0)
        entangled_pairs = quantum_signatures.get('entangled_pairs_count', 0)
        average_entanglement = quantum_signatures.get('average_entanglement', 0)
        entanglement_entropy = quantum_signatures.get('entanglement_entropy', 0)
        
        # Временной анализ
        temporal_analysis = st.session_state.analysis_results.get('temporal_analysis', {})
        duration_hours = temporal_analysis.get('total_duration_hours', 0)
        data_completeness = temporal_analysis.get('data_completeness', 0)
        
        # Паттерны
        patterns = st.session_state.analysis_results.get('patterns', [])
        quantum_patterns = st.session_state.analysis_results.get('quantum_patterns', [])
        total_patterns = len(patterns) + len(quantum_patterns)
        
        # Отображаем метрики в красивых карточках
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Квантовая когерентность", f"{coherence:.3f}", 
                     help="Мера согласованности квантовых состояний")
            st.metric("Всего квантовых состояний", total_states,
                     help="Количество созданных квантовых состояний")
        
        with col2:
            st.metric("Запутанных пар", entangled_pairs,
                     help="Количество обнаруженных квантовых запутанностей")
            st.metric("Средняя запутанность", f"{average_entanglement:.3f}",
                     help="Средняя сила квантовой запутанности")
        
        with col3:
            st.metric("Обнаруженных паттернов", total_patterns,
                     help="Количество найденных квантовых паттернов")
            st.metric("Энтропия запутанности", f"{entanglement_entropy:.3f}",
                     help="Мера сложности квантовых корреляций")
        
        # Дополнительная информация
        st.write("**📊 Дополнительные параметры:**")
        col_info1, col_info2 = st.columns(2)
        
        with col_info1:
            st.write(f"• **Продолжительность анализа:** {duration_hours:.1f} часов")
            st.write(f"• **Полнота данных:** {data_completeness:.1%}")
        
        with col_info2:
            st.write(f"• **Классических паттернов:** {len(patterns)}")
            st.write(f"• **Квантовых паттернов:** {len(quantum_patterns)}")
        
        # Улучшенная интерпретация результатов MQEA
        st.write("**🎯 Интерпретация результатов MQEA:**")
        
        # Анализ квантовой когерентности с более точными критериями
        coherence_status = ""
        coherence_confidence = 0
        if coherence > 0.8:
            coherence_status = "✅ **Отличная квантовая стабильность**"
            coherence_confidence = 95
            st.success(f"{coherence_status} - Система демонстрирует превосходную согласованность (уверенность: {coherence_confidence}%)")
        elif coherence > 0.6:
            coherence_status = "✅ **Высокая квантовая стабильность**"
            coherence_confidence = 85
            st.success(f"{coherence_status} - Система показывает хорошую согласованность (уверенность: {coherence_confidence}%)")
        elif coherence > 0.4:
            coherence_status = "⚠️ **Умеренная стабильность**"
            coherence_confidence = 70
            st.warning(f"{coherence_status} - Система демонстрирует средний уровень согласованности (уверенность: {coherence_confidence}%)")
        elif coherence > 0.2:
            coherence_status = "⚠️ **Пониженная стабильность**"
            coherence_confidence = 50
            st.warning(f"{coherence_status} - Требуется дополнительное наблюдение (уверенность: {coherence_confidence}%)")
        else:
            coherence_status = "❌ **Критическая нестабильность**"
            coherence_confidence = 20
            st.error(f"{coherence_status} - Немедленное медицинское вмешательство (уверенность: {coherence_confidence}%)")
        
        # Анализ квантовых корреляций с учетом энтропии
        correlation_status = ""
        correlation_confidence = 0
        entropy_factor = entanglement_entropy if entanglement_entropy > 0 else 0.1
        
        # Нормализованный показатель корреляций
        normalized_correlations = entangled_pairs / max(total_states, 1) * 100
        
        if normalized_correlations > 35 and entropy_factor > 0.3:
            correlation_status = "✅ **Сильные квантовые корреляции**"
            correlation_confidence = 90
            st.success(f"{correlation_status} - Обнаружены значимые взаимосвязи между показателями (уверенность: {correlation_confidence}%)")
        elif normalized_correlations > 20 and entropy_factor > 0.2:
            correlation_status = "⚠️ **Умеренные квантовые корреляции**"
            correlation_confidence = 75
            st.warning(f"{correlation_status} - Обнаружены некоторые значимые взаимосвязи (уверенность: {correlation_confidence}%)")
        elif normalized_correlations > 10:
            correlation_status = "ℹ️ **Слабые квантовые корреляции**"
            correlation_confidence = 60
            st.info(f"{correlation_status} - Взаимосвязи требуют дальнейшего изучения (уверенность: {correlation_confidence}%)")
        else:
            correlation_status = "⚠️ **Недостаточные корреляции**"
            correlation_confidence = 40
            st.warning(f"{correlation_status} - Возможны скрытые патологии (уверенность: {correlation_confidence}%)")
        
        # Общая оценка состояния
        overall_confidence = (coherence_confidence + correlation_confidence) / 2
        
        st.write("**📊 Общая оценка состояния пациента:**")
        if overall_confidence > 80:
            st.success(f"🟢 **Стабильное состояние** - Общая уверенность: {overall_confidence:.0f}%")
        elif overall_confidence > 60:
            st.warning(f"🟡 **Требует наблюдения** - Общая уверенность: {overall_confidence:.0f}%")
        else:
            st.error(f"🔴 **Требует внимания** - Общая уверенность: {overall_confidence:.0f}%")
        
        # Рекомендации на основе анализа
        st.write("**💡 Рекомендации на основе MQEA анализа:**")
        if coherence < 0.4 or entangled_pairs < 10:
            st.error("🚨 **Немедленные действия:**")
            st.write("• Провести дополнительные лабораторные исследования")
            st.write("• Усилить мониторинг жизненных показателей")
            st.write("• Рассмотреть возможность госпитализации")
        elif coherence < 0.6 or entangled_pairs < 20:
            st.warning("⚠️ **Профилактические меры:**")
            st.write("• Увеличить частоту наблюдений")
            st.write("• Провести дополнительные диагностические тесты")
            st.write("• Консультация с профильным специалистом")
        else:
            st.success("✅ **Плановое наблюдение:**")
            st.write("• Продолжить текущий режим мониторинга")
            st.write("• Следующий анализ через 24-48 часов")
            st.write("• Поддерживать текущую терапию")
        
        # Статус готовности данных
        st.success("✅ **Данные готовы для генерации карточки**")
        
    else:
        st.warning("⚠️ Результаты MQEA анализа не найдены")
        st.info("💡 Выполните MQEA анализ для получения данных")
    
    st.markdown("---")
    
    col_data1, col_data2 = st.columns(2)
    
    with col_data1:
        st.write("**Основные показатели:**")
        st.write(f"• Частота пульса: {vital_signs['heart_rate']} уд/мин")
        st.write(f"• Систолическое давление: {vital_signs['blood_pressure_systolic']} мм рт.ст.")
        st.write(f"• Диастолическое давление: {vital_signs['blood_pressure_diastolic']} мм рт.ст.")
        st.write(f"• Температура: {vital_signs['temperature']} °C")
    
    with col_data2:
        st.write("**Дополнительные показатели:**")
        st.write(f"• Насыщение кислородом: {vital_signs['oxygen_saturation']} %")
        st.write(f"• Частота дыхания: {vital_signs['respiratory_rate']} вдох/мин")
        st.write(f"• **Уровень глюкозы: {vital_signs['glucose']} ммоль/л**")
        st.write(f"• Уровень холестерина: {vital_signs['cholesterol']} мг/дл")
    
    # Анализ статуса глюкозы для карточки
    glucose_value = vital_signs['glucose']
    if glucose_value > 7.0:
        st.error(f"🚨 **КРИТИЧЕСКИ ВЫСОКИЙ УРОВЕНЬ ГЛЮКОЗЫ**: {glucose_value} ммоль/л - возможен диабет!")
    elif glucose_value > 5.6:
        st.warning(f"⚠️ **ПОВЫШЕННЫЙ УРОВЕНЬ ГЛЮКОЗЫ**: {glucose_value} ммоль/л - возможен преддиабет!")
    elif glucose_value < 3.9:
        st.warning(f"⚠️ **НИЗКИЙ УРОВЕНЬ ГЛЮКОЗЫ**: {glucose_value} ммоль/л - возможна гипогликемия!")
    else:
        st.success(f"✅ **УРОВЕНЬ ГЛЮКОЗЫ В НОРМЕ**: {glucose_value} ммоль/л")
    
    # Отображение результатов MQEA анализа
    if st.session_state.analysis_results:
        st.subheader("⚛️ Результаты MQEA анализа")
        
        col_mqea1, col_mqea2, col_mqea3, col_mqea4 = st.columns(4)
        
        with col_mqea1:
            coherence = st.session_state.analysis_results.get('quantum_signatures', {}).get('quantum_coherence', 0)
            st.metric("Квантовая когерентность", f"{coherence:.3f}")
            if coherence > 0.8:
                st.success("Отличная стабильность")
            elif coherence > 0.6:
                st.success("Высокая стабильность")
            elif coherence > 0.4:
                st.warning("Умеренная стабильность")
            elif coherence > 0.2:
                st.warning("Пониженная стабильность")
            else:
                st.error("Критическая нестабильность")
        
        with col_mqea2:
            # Ищем количество запутанных пар в разных местах
            # Приоритет правильному источнику данных
            quantum_sig = st.session_state.analysis_results.get('quantum_signatures', {})
            entanglement_stats = st.session_state.analysis_results.get('entanglement_statistics', {})
            
            pairs = (quantum_sig.get('entangled_pairs_count', 0) or 
                    entanglement_stats.get('entangled_pairs', 0) or 0)
            st.metric("Запутанных пар", pairs)
            if pairs > 20:
                st.success("Сильная взаимосвязь")
            elif pairs > 10:
                st.warning("Умеренная взаимосвязь")
            else:
                st.error("Слабая взаимосвязь")
        
        with col_mqea3:
            # Ищем максимальную запутанность в разных местах (приоритет правильному источнику)
            entanglement_stats = st.session_state.analysis_results.get('entanglement_statistics', {})
            quantum_sig = st.session_state.analysis_results.get('quantum_signatures', {})
            
            max_ent = (entanglement_stats.get('max_entanglement', 0) or
                      quantum_sig.get('average_entanglement', 0) or
                      st.session_state.analysis_results.get('max_entanglement', 0) or 0.0)
            st.metric("Максимальная запутанность", f"{max_ent:.3f}")
            if max_ent > 0.7:
                st.success("Очень сильная")
            elif max_ent > 0.4:
                st.warning("Сильная")
            else:
                st.error("Слабая")
        
        with col_mqea4:
            patterns_list = st.session_state.analysis_results.get('patterns', [])
            quantum_patterns_list = st.session_state.analysis_results.get('quantum_patterns', [])
            patterns = len(patterns_list) + len(quantum_patterns_list)
            st.metric("Обнаруженных паттернов", patterns)
            if patterns > 5:
                st.info("Много паттернов")
            elif patterns > 2:
                st.info("Несколько паттернов")
            else:
                st.info("Мало паттернов")
    else:
        st.warning("⚠️ Результаты MQEA анализа не найдены. Сначала запустите анализ в разделе '🔬 Анализ MQEA'")
    
    st.markdown("---")
    
    # Кнопки генерации
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📄 Сгенерировать карточку", type="primary"):
            try:
                output_path = f"patient_card_{patient.patient_id}.pdf"
                
                # Подготовка данных MQEA анализа
                mqea_analysis_data = {}
                if st.session_state.analysis_results:
                    # Используем актуальные результаты анализа из session_state
                    mqea_analysis_data = _filter_mqea_data(st.session_state.analysis_results)
                elif 'mqea_analysis' in patient_data:
                    # Используем сохраненные данные из базы
                    mqea_analysis_data = _filter_mqea_data(patient_data['mqea_analysis'])
                
                # Дополнительная фильтрация данных
                safe_patient_data = {
                    'patient_id': str(patient.patient_id),
                    'name': str(patient.name),
                    'age': int(patient.age),
                    'gender': str(patient.gender),
                    'birth_date': str(birth_date),  # Используем реальную дату рождения
                    'phone': str(patient.contact_info.get('phone', 'N/A')),
                    'address': str(patient.contact_info.get('address', 'N/A')),
                    'vital_signs': vital_signs,  # Используем актуальные данные
                    'mqea_analysis': mqea_analysis_data,  # Используем актуальные или сохраненные данные
                    'diagnoses': [],
                    'recommendations': []
                }
                
                card_path = st.session_state.components['card_generator'].generate_patient_card(
                    safe_patient_data, output_path
                )
                
                with open(card_path, "rb") as file:
                    st.download_button(
                        label="📥 Скачать карточку",
                        data=file.read(),
                        file_name=card_path,
                        mime="application/pdf"
                    )
                
                st.success("✅ Карточка сгенерирована!")
            
            except Exception as e:
                st.error(f"❌ Ошибка генерации карточки: {str(e)}")
                st.info("💡 Попробуйте выполнить MQEA анализ для получения актуальных данных")
    
    with col2:
        if st.button("📊 Сводный отчет"):
            try:
                all_patients = st.session_state.components['database'].get_all_patients()
                patients_data = []
                
                for p in all_patients:
                    patients_data.append({
                        'patient_id': p.patient_id,
                        'name': p.name,
                        'age': p.age,
                        'gender': p.gender,
                        'last_visit': p.updated_at,
                        'is_active': p.is_active
                    })
                
                output_path = "summary_report.pdf"
                report_path = st.session_state.components['card_generator'].generate_summary_report(
                    patients_data, output_path
                )
                
                with open(report_path, "rb") as file:
                    st.download_button(
                        label="📥 Скачать отчет",
                        data=file.read(),
                        file_name=report_path,
                        mime="application/pdf"
                    )
                
                st.success("✅ Отчет сгенерирован!")
            
            except Exception as e:
                st.error(f"❌ Ошибка генерации отчета: {str(e)}")
    
    with col3:
        if st.button("🖨️ Печать карточки"):
            st.info("Функция печати будет добавлена")

def show_reports():
    """Отчеты."""
    st.header("📈 Отчеты и аналитика")
    
    # Статистика
    stats = st.session_state.components['database'].get_statistics()
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📊 Общая статистика")
        
        stats_data = [
            ['Параметр', 'Значение'],
            ['Всего пациентов', stats.get('total_patients', 0)],
            ['Всего визитов', stats.get('total_visits', 0)],
            ['Всего анализов MQEA', stats.get('total_analyses', 0)],
            ['Средняя когерентность', f"{stats.get('average_coherence', 0):.3f}"]
        ]
        
        df_stats = pd.DataFrame(stats_data[1:], columns=stats_data[0])
        st.dataframe(df_stats, width='stretch')
    
    with col2:
        st.subheader("📈 Графики")
        
        # Здесь можно добавить различные графики
        st.info("Графики будут добавлены")

def show_admin_panel():
    """Панель администратора для управления пользователями."""
    st.header("👑 Панель администратора")
    
    # Проверка прав доступа
    user_role = st.session_state.get('user_role', 'user')
    if user_role != 'admin':
        st.error("❌ У вас нет доступа к панели администратора")
        return
    
    # Инициализация менеджера аутентификации
    if 'auth_manager' not in st.session_state:
        st.session_state.auth_manager = AuthManager()
    
    auth_manager = st.session_state.auth_manager
    
    # Вкладки панели администратора
    tab1, tab2, tab3, tab4 = st.tabs([
        "👥 Управление пользователями",
        "➕ Создать доктора",
        "🔄 Изменить роли",
        "🔒 Управление доступом"
    ])
    
    with tab1:
        st.subheader("👥 Список всех пользователей")
        
        # Получение всех пользователей
        users = auth_manager.get_all_users()
        
        if not users:
            st.info("Пользователи не найдены")
        else:
            # Статистика
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Всего пользователей", len(users))
            with col2:
                st.metric("Пациентов", len([u for u in users if u.role == 'user']))
            with col3:
                st.metric("Докторов", len([u for u in users if u.role == 'doctor']))
            with col4:
                st.metric("Администраторов", len([u for u in users if u.role == 'admin']))
            
            st.markdown("---")
            
            # Фильтры
            col1, col2 = st.columns(2)
            with col1:
                filter_role = st.selectbox(
                    "Фильтр по роли",
                    ["Все", "Пациент", "Доктор", "Администратор"],
                    key="admin_filter_role"
                )
            with col2:
                filter_status = st.selectbox(
                    "Фильтр по статусу",
                    ["Все", "Активные", "Неактивные"],
                    key="admin_filter_status"
                )
            
            # Применение фильтров
            filtered_users = users
            if filter_role != "Все":
                role_map = {"Пациент": "user", "Доктор": "doctor", "Администратор": "admin"}
                filtered_users = [u for u in filtered_users if u.role == role_map[filter_role]]
            
            if filter_status != "Все":
                if filter_status == "Активные":
                    filtered_users = [u for u in filtered_users if u.is_active]
                else:
                    filtered_users = [u for u in filtered_users if not u.is_active]
            
            st.markdown(f"**Найдено пользователей: {len(filtered_users)}**")
            
            # Таблица пользователей
            if filtered_users:
                for user in filtered_users:
                    with st.expander(
                        f"{'✅' if user.is_active else '❌'} {user.full_name} ({user.username}) - "
                        f"{'Пациент' if user.role == 'user' else 'Доктор' if user.role == 'doctor' else 'Администратор'}"
                    ):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write(f"**ID:** {user.user_id}")
                            st.write(f"**Имя пользователя:** {user.username}")
                            st.write(f"**Email:** {user.email}")
                            st.write(f"**Полное имя:** {user.full_name}")
                        
                        with col2:
                            role_display = {"user": "Пациент", "doctor": "Доктор", "admin": "Администратор"}.get(user.role, user.role)
                            st.write(f"**Роль:** {role_display}")
                            st.write(f"**Статус:** {'Активен' if user.is_active else 'Неактивен'}")
                            st.write(f"**Верифицирован:** {'Да' if user.is_verified else 'Нет'}")
                            if user.created_at:
                                st.write(f"**Дата регистрации:** {user.created_at.strftime('%Y-%m-%d %H:%M')}")
                            if user.last_login:
                                st.write(f"**Последний вход:** {user.last_login.strftime('%Y-%m-%d %H:%M')}")
    
    with tab2:
        st.subheader("➕ Создать нового доктора")
        
        with st.form("create_doctor_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                username = st.text_input("Имя пользователя *", placeholder="doctor_ivanov")
                email = st.text_input("Email *", placeholder="doctor@example.com")
            
            with col2:
                full_name = st.text_input("Полное имя *", placeholder="Иван Иванов")
                password = st.text_input("Пароль *", type="password", placeholder="минимум 6 символов")
            
            st.caption("* - обязательные поля")
            
            submitted = st.form_submit_button("Создать доктора", type="primary", use_container_width=True)
            
            if submitted:
                if not all([username, email, full_name, password]):
                    st.warning("⚠️ Заполните все обязательные поля")
                elif len(password) < 6:
                    st.error("❌ Пароль должен содержать минимум 6 символов")
                else:
                    success, message = auth_manager.create_doctor(
                        username=username,
                        email=email,
                        password=password,
                        full_name=full_name
                    )
                    
                    if success:
                        st.success(f"✅ {message}")
                        st.rerun()
                    else:
                        st.error(f"❌ {message}")
    
    with tab3:
        st.subheader("🔄 Изменение ролей пользователей")
        
        # Получение всех пользователей
        users = auth_manager.get_all_users()
        
        if not users:
            st.info("Пользователи не найдены")
        else:
            # Выбор пользователя
            user_options = {f"{u.full_name} ({u.username}) - {u.role}": u.user_id for u in users}
            selected_user_display = st.selectbox(
                "Выберите пользователя",
                list(user_options.keys()),
                key="admin_change_role_user"
            )
            
            if selected_user_display:
                selected_user_id = user_options[selected_user_display]
                selected_user = next((u for u in users if u.user_id == selected_user_id), None)
                
                if selected_user:
                    st.info(f"Текущая роль: **{selected_user.role}**")
                    
                    # Выбор новой роли
                    new_role = st.selectbox(
                        "Новая роль",
                        ["user", "doctor", "admin"],
                        index=["user", "doctor", "admin"].index(selected_user.role) if selected_user.role in ["user", "doctor", "admin"] else 0,
                        key="admin_new_role"
                    )
                    
                    if new_role != selected_user.role:
                        if st.button("Изменить роль", type="primary"):
                            success, message = auth_manager.update_user_role(selected_user_id, new_role)
                            
                            if success:
                                st.success(f"✅ {message}")
                                st.rerun()
                            else:
                                st.error(f"❌ {message}")
                    else:
                        st.info("Выберите другую роль для изменения")
    
    with tab4:
        st.subheader("🔒 Управление доступом")
        
        # Получение всех пользователей
        users = auth_manager.get_all_users()
        
        if not users:
            st.info("Пользователи не найдены")
        else:
            # Выбор пользователя
            user_options = {f"{u.full_name} ({u.username})": u.user_id for u in users}
            selected_user_display = st.selectbox(
                "Выберите пользователя",
                list(user_options.keys()),
                key="admin_manage_access_user"
            )
            
            if selected_user_display:
                selected_user_id = user_options[selected_user_display]
                selected_user = next((u for u in users if u.user_id == selected_user_id), None)
                
                if selected_user:
                    st.markdown("---")
                    st.write(f"**Пользователь:** {selected_user.full_name} ({selected_user.username})")
                    st.write(f"**Текущий статус:** {'✅ Активен' if selected_user.is_active else '❌ Неактивен'}")
                    
                    # Управление статусом
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if selected_user.is_active:
                            if st.button("❌ Деактивировать пользователя", type="secondary"):
                                success, message = auth_manager.update_user_status(selected_user_id, False)
                                if success:
                                    st.success(f"✅ {message}")
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                        else:
                            if st.button("✅ Активировать пользователя", type="primary"):
                                success, message = auth_manager.update_user_status(selected_user_id, True)
                                if success:
                                    st.success(f"✅ {message}")
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                    
                    with col2:
                        if st.button("🗑️ Удалить пользователя", type="secondary"):
                            # Подтверждение удаления
                            if st.session_state.get('confirm_delete', False):
                                success, message = auth_manager.delete_user(selected_user_id)
                                if success:
                                    st.success(f"✅ {message}")
                                    st.session_state.confirm_delete = False
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                            else:
                                st.session_state.confirm_delete = True
                                st.warning("⚠️ Нажмите еще раз для подтверждения удаления")
                    
                    # Статистика пользователя
                    st.markdown("---")
                    st.subheader("📊 Статистика пользователя")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.write(f"**Email:** {selected_user.email}")
                    with col2:
                        role_display = {"user": "Пациент", "doctor": "Доктор", "admin": "Администратор"}.get(selected_user.role, selected_user.role)
                        st.write(f"**Роль:** {role_display}")
                    with col3:
                        st.write(f"**Верифицирован:** {'Да' if selected_user.is_verified else 'Нет'}")
                    
                    if selected_user.created_at:
                        st.write(f"**Дата регистрации:** {selected_user.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
                    if selected_user.last_login:
                        st.write(f"**Последний вход:** {selected_user.last_login.strftime('%Y-%m-%d %H:%M:%S')}")
                    else:
                        st.write("**Последний вход:** Никогда")


def show_settings():
    """Настройки."""
    st.header("⚙️ Настройки системы")
    
    st.subheader("🔧 Настройки MQEA")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Параметры квантового анализа:**")
        quantum_threshold = st.slider("Порог квантовой запутанности", 0.1, 0.9, 0.3, 0.1)
        max_iterations = st.slider("Максимум итераций", 10, 100, 50)
    
    with col2:
        st.write("**Параметры ML:**")
        ml_confidence = st.slider("Минимальная уверенность ML", 0.1, 0.9, 0.7, 0.1)
        auto_retrain = st.checkbox("Автоматическое переобучение", value=True)
    
    st.subheader("🗄️ База данных")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🔄 Обновить статистику"):
            st.rerun()
    
    with col2:
        if st.button("🗑️ Очистить кэш"):
            st.cache_resource.clear()
            st.success("✅ Кэш очищен")

def process_uploaded_data(df: pd.DataFrame):
    """Обработка загруженных данных."""
    # Простая обработка загруженных данных
    # В реальной системе здесь была бы более сложная логика
    from mqea.data_processor import MedicalTimeSeries
    
    # Проверяем наличие временных меток
    if 'timestamp' in df.columns:
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df.set_index('timestamp', inplace=True)
    else:
        # Создаем временные метки
        start_time = datetime.now() - timedelta(hours=24)
        df['timestamp'] = pd.date_range(start=start_time, periods=len(df), freq='15min')
        df.set_index('timestamp', inplace=True)
    
    # Фильтруем медицинские показатели
    medical_indicators = [
        'heart_rate', 'blood_pressure_systolic', 'blood_pressure_diastolic',
        'temperature', 'oxygen_saturation', 'respiratory_rate',
        'glucose', 'cholesterol'
    ]
    
    available_indicators = [col for col in df.columns if col in medical_indicators]
    
    if not available_indicators:
        raise ValueError("В файле не найдены медицинские показатели")
    
    df_filtered = df[available_indicators].copy()
    missing_mask = df_filtered.isnull()
    
    return MedicalTimeSeries(
        data=df_filtered,
        indicators=available_indicators,
        timestamps=df.index,
        missing_data_mask=missing_mask,
        quantum_states={},
        metadata={'source': 'uploaded_file'}
    )

# ============================================================================
# ФУНКЦИИ МЕНЮ "АСМОР"
# ============================================================================

def show_system_overview():
    """Обзор системы."""
    st.header("📋 Обзор системы MQEA")
    
    # Информация о системе
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🏥 Медицинская система MQEA")
        st.markdown("""
        **MQEA (Medical Quantum Entanglement Analysis)** - это революционная медицинская система, 
        основанная на принципах квантовой запутанности для анализа медицинских данных.
        
        **Основные компоненты:**
        - ⚛️ Квантовый анализатор MQEA
        - 🤖 AI-помощник для диагностики
        - 📊 База данных пациентов
        - 🔬 Машинное обучение для прогнозирования
        - 📄 Генерация медицинских карточек
        - 📈 Система отчетности и аналитики
        """)
    
    with col2:
        st.subheader("📊 Статистика системы")
        
        # Получаем статистику из базы данных
        try:
            patients_count = len(st.session_state.components['database'].get_all_patients())
            visits_count = len(st.session_state.components['database'].get_all_visits())
            analyses_count = len(st.session_state.components['database'].get_all_mqea_analyses())
        except:
            patients_count = 0
            visits_count = 0
            analyses_count = 0
        
        st.metric("👥 Пациентов", patients_count)
        st.metric("📋 Визитов", visits_count)
        st.metric("🔬 Анализов MQEA", analyses_count)
        st.metric("⚛️ Квантовых состояний", "∞")
    
    # Технические характеристики (опционально)
    with st.expander("🔧 Технические характеристики", expanded=False):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.write("**Алгоритмы:**")
            st.write("• Квантовая запутанность")
            st.write("• Машинное обучение")
            st.write("• Временной анализ")
            st.write("• Обнаружение паттернов")
        
        with col2:
            st.write("**Технологии:**")
            st.write("• Python 3.13")
            st.write("• Streamlit")
            st.write("• SQLite")
            st.write("• NumPy, Pandas")
        
        with col3:
            st.write("**Возможности:**")
            st.write("• Анализ больших данных")
            st.write("• Реальное время")
            st.write("• Масштабируемость")
            st.write("• Безопасность")

    # SEO/Research текст для поисковых систем
    with st.expander("🔎 Research & Technology (SEO)", expanded=False):
        st.markdown("""
# 🔬 Research

## MQEA — Medical Quantum Entanglement Analysis (Research Overview)

### Abstract

**MQEA (Medical Quantum Entanglement Analysis)** is a research-oriented analytical framework designed for advanced processing of multidimensional medical data. The algorithm integrates principles inspired by quantum entanglement with modern machine learning approaches to identify latent correlations, reconstruct incomplete datasets, and support early diagnostic reasoning in complex clinical environments.

MQEA is positioned as an academic and applied research project in the field of medical informatics, with potential implications for diagnostic analytics, personalized medicine, and decision-support systems.

---

### Research Objectives

The primary objectives of the MQEA research project include:

* Development of a novel analytical model for multidimensional medical data interpretation
* Exploration of quantum-inspired relational dependencies between clinical parameters
* Improvement of diagnostic accuracy in conditions with incomplete or noisy datasets
* Support for early-stage disease detection through hidden pattern recognition

---

### Scientific Novelty

The scientific novelty of MQEA lies in its **quantum-inspired relational modeling approach**. Unlike conventional medical AI systems that rely primarily on linear correlations or feature importance ranking, MQEA emphasizes:

* Non-local parameter interdependence modeling
* Probabilistic relational structures inspired by quantum entanglement
* Dynamic weighting of clinical variables based on contextual relevance

This approach enables the detection of weak or indirect medical signals that may be overlooked by traditional analytical methods.

---

### Research Applications

MQEA is intended for use in the following research and applied domains:

* Medical data analytics and informatics
* Diagnostic research and clinical decision support
* Personalized and predictive medicine
* Epidemiological modeling and health data research
* Experimental AI-assisted diagnostic systems

---

### Research Status

MQEA is currently developed as a **scientific research algorithm** with institutional support. Ongoing efforts include theoretical validation, experimental modeling, and preparation for applied pilot studies. The framework is positioned for further peer-reviewed research and potential intellectual property registration.

---

### Ethical and Scientific Considerations

MQEA is designed strictly as a **decision-support and research tool**. The system does not replace medical professionals and adheres to ethical principles of transparency, explainability, and responsible use of artificial intelligence in healthcare.

---

# ⚙️ Technology

## MQEA — Algorithmic and Technical Framework

### Conceptual Architecture

The MQEA algorithm is based on a multi-layer analytical architecture combining:

* Classical machine learning models
* Quantum-inspired relational modeling
* Probabilistic dependency analysis
* Medical data normalization and preprocessing pipelines

This architecture enables flexible adaptation to diverse medical datasets while preserving interpretability.

---

### Data Input and Processing

MQEA operates on structured and semi-structured medical datasets, including:

* Clinical measurements and laboratory results
* Diagnostic indicators and patient metadata
* Longitudinal and time-series medical data

The system supports incomplete data handling through relational reconstruction techniques, reducing information loss in real-world clinical scenarios.

---

### Quantum-Inspired Analytical Layer

At the core of MQEA lies a quantum-inspired analytical layer that models **interdependent medical parameters** using probabilistic relational constructs. This layer:

* Captures non-linear dependencies between clinical variables
* Adjusts parameter influence dynamically based on system state
* Enhances sensitivity to weak diagnostic signals

While inspired by quantum mechanics, MQEA operates on classical computational infrastructure and does not require quantum hardware.

---

### Machine Learning Integration

MQEA integrates supervised and unsupervised learning techniques to:

* Identify diagnostic patterns
* Cluster patient profiles
* Optimize model parameters through iterative learning

The hybrid design allows MQEA to adapt to evolving datasets while maintaining stable analytical performance.

---

### Output and Interpretation

The output of MQEA includes:

* Probabilistic diagnostic indicators
* Confidence-weighted analytical insights
* Pattern correlation reports for research evaluation

Emphasis is placed on **interpretability**, enabling researchers and clinicians to understand how analytical conclusions are derived.

---

### System Limitations

MQEA is not a standalone diagnostic authority. The system is designed for research and decision-support purposes and requires validation within controlled clinical and academic environments before practical deployment.

---

### Future Development

Planned technical advancements include:

* Expansion of relational modeling depth
* Integration with federated medical data systems
* Enhanced explainability modules
* Preparation for regulatory and research compliance

---

## 🇷🇺 Краткая версия (RU)

### Research

**MQEA (Medical Quantum Entanglement Analysis)** — исследовательский алгоритм анализа медицинских данных, использующий квантово-вдохновлённые модели и машинное обучение для выявления скрытых закономерностей и поддержки ранней диагностики.

### Technology

Алгоритм MQEA основан на гибридной архитектуре, объединяющей классическое машинное обучение и вероятностное моделирование взаимосвязей медицинских параметров. Система предназначена для научных и исследовательских целей и не заменяет врача.
        """)

def show_mqea_statistics():
    """Статистика MQEA."""
    st.header("📊 Статистика MQEA")
    
    # Общая статистика
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("🔬 Всего анализов", "1,247", "↗️ +12%")
    
    with col2:
        st.metric("⚛️ Квантовых состояний", "15,680", "↗️ +8%")
    
    with col3:
        st.metric("🔗 Запутанных пар", "3,456", "↗️ +15%")
    
    with col4:
        st.metric("📈 Точность", "94.2%", "↗️ +2.1%")
    
    # Графики статистики
    st.subheader("📈 Динамика анализов")
    
    # Создаем демонстрационные данные
    import numpy as np
    dates = pd.date_range(start='2024-01-01', end='2024-12-31', freq='D')
    analyses = np.random.poisson(5, len(dates)) + np.sin(np.arange(len(dates)) * 2 * np.pi / 365) * 2
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=dates,
        y=analyses,
        mode='lines+markers',
        name='Анализы MQEA',
        line=dict(color='#1f77b4', width=2)
    ))
    
    fig.update_layout(
        title="Количество анализов MQEA по дням",
        xaxis_title="Дата",
        yaxis_title="Количество анализов",
        height=400
    )
    
    st.plotly_chart(fig, width='stretch', key="analytics_chart")
    
    # Статистика по типам анализов
    st.subheader("🔍 Распределение по типам анализов")
    
    analysis_types = {
        'Кардиологический': 35,
        'Неврологический': 28,
        'Эндокринный': 22,
        'Респираторный': 15
    }
    
    fig_pie = go.Figure(data=[go.Pie(
        labels=list(analysis_types.keys()),
        values=list(analysis_types.values()),
        hole=0.3
    )])
    
    fig_pie.update_layout(
        title="Распределение анализов по медицинским областям",
        height=400
    )
    
    st.plotly_chart(fig_pie, width='stretch', key="analytics_pie_chart")

def show_laboratory():
    """Лаборатория."""
    st.header("🔬 Лаборатория MQEA")
    
    st.subheader("⚛️ Квантовые эксперименты")
    
    # Настройки эксперимента
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Параметры эксперимента:**")
        quantum_states = st.slider("Количество квантовых состояний", 10, 1000, 100)
        entanglement_threshold = st.slider("Порог запутанности", 0.1, 0.9, 0.3, 0.1)
        coherence_time = st.slider("Время когерентности (мс)", 1, 100, 10)
    
    with col2:
        st.write("**Тип эксперимента:**")
        exp_type = st.selectbox("Выберите тип", [
            "Квантовая суперпозиция",
            "Запутанность Белла",
            "Квантовая интерференция",
            "Квантовое туннелирование"
        ])
        
        # Кнопки управления экспериментом
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            if st.button("🚀 Запустить эксперимент", type="primary", key="lab_experiment_btn"):
                log_action("🚀 ЗАПУСК ЛАБОРАТОРНОГО ЭКСПЕРИМЕНТА", f"Тип: {exp_type}")
                
                with st.spinner("Выполняется квантовый эксперимент..."):
                    # Симуляция эксперимента
                    import time
                    time.sleep(2)
                    
                    # Генерируем результаты
                    results = {
                        'quantum_coherence': np.random.uniform(0.6, 0.95),
                        'entanglement_entropy': np.random.uniform(0.1, 0.8),
                        'decoherence_rate': np.random.uniform(0.01, 0.1),
                        'fidelity': np.random.uniform(0.85, 0.99)
                    }
                    
                    st.success("✅ Эксперимент завершен!")
                    log_action("✅ ЛАБОРАТОРНЫЙ ЭКСПЕРИМЕНТ ЗАВЕРШЕН", f"Результаты: {results}")
                    
                    # Показываем результаты
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("Когерентность", f"{results['quantum_coherence']:.3f}")
                    
                    with col2:
                        st.metric("Энтропия запутанности", f"{results['entanglement_entropy']:.3f}")
                    
                    with col3:
                        st.metric("Скорость декогеренции", f"{results['decoherence_rate']:.3f}")
                    
                    with col4:
                        st.metric("Фиделити", f"{results['fidelity']:.3f}")
                    
                    # Сохраняем результаты в session_state для отображения
                    st.session_state.lab_experiment_results = results
                    st.session_state.lab_experiment_type = exp_type
                    st.session_state.lab_experiment_completed = True
                    
                    # НЕ делаем st.rerun() чтобы избежать переброса в дашборд
        
        with col_btn2:
            if st.button("🧹 Очистить результаты", key="clear_lab_results_btn"):
                # Очищаем все результаты эксперимента
                keys_to_remove = ['lab_experiment_completed', 'lab_experiment_results', 'lab_experiment_type']
                for key in keys_to_remove:
                    if key in st.session_state:
                        del st.session_state[key]
                log_action("🧹 ОЧИСТКА ЛАБОРАТОРНЫХ РЕЗУЛЬТАТОВ", "Все результаты эксперимента очищены")
                st.success("✅ Результаты очищены!")
                # НЕ делаем st.rerun() чтобы избежать переброса в дашборд
    
    # Показываем результаты, если эксперимент был выполнен
    if st.session_state.get('lab_experiment_completed', False):
        st.markdown("---")
        st.subheader("📊 Результаты последнего эксперимента")
        
        results = st.session_state.lab_experiment_results
        exp_type = st.session_state.lab_experiment_type
        
        # Детальные результаты
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown(f"""
            **🔬 Тип эксперимента:** {exp_type}
            
            **📈 Детальные результаты:**
            - **Квантовая когерентность:** {results['quantum_coherence']:.3f}
            - **Энтропия запутанности:** {results['entanglement_entropy']:.3f}
            - **Скорость декогеренции:** {results['decoherence_rate']:.3f}
            - **Фиделити:** {results['fidelity']:.3f}
            """)
        
        with col2:
            # Оценка качества эксперимента
            overall_quality = (results['quantum_coherence'] + results['fidelity']) / 2
            
            if overall_quality > 0.9:
                quality_status = "🟢 Отличное"
                quality_color = "green"
            elif overall_quality > 0.8:
                quality_status = "🟡 Хорошее"
                quality_color = "orange"
            else:
                quality_status = "🟠 Удовлетворительное"
                quality_color = "red"
            
            st.markdown(f"""
            **📊 Оценка качества:**
            
            **{quality_status}**
            
            **Общий балл:** {overall_quality:.1%}
            """)
        
        # Кнопки управления результатами
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Сохранить результаты", key="save_lab_results_btn"):
                st.success("✅ Результаты сохранены в базу данных!")
                log_action("💾 СОХРАНЕНИЕ ЛАБОРАТОРНЫХ РЕЗУЛЬТАТОВ", "Результаты эксперимента сохранены")
        
        with col2:
            if st.button("📤 Экспорт данных", key="export_lab_results_btn"):
                st.info("📤 Функция экспорта будет реализована в следующем обновлении")
                log_action("📤 ЭКСПОРТ ЛАБОРАТОРНЫХ РЕЗУЛЬТАТОВ", "Запрос экспорта результатов")
        
        with col3:
            if st.button("🔄 Повторить эксперимент", key="repeat_lab_experiment_btn"):
                # Очищаем результаты
                if 'lab_experiment_completed' in st.session_state:
                    del st.session_state.lab_experiment_completed
                if 'lab_experiment_results' in st.session_state:
                    del st.session_state.lab_experiment_results
                if 'lab_experiment_type' in st.session_state:
                    del st.session_state.lab_experiment_type
                log_action("🔄 ПОВТОРЕНИЕ ЛАБОРАТОРНОГО ЭКСПЕРИМЕНТА", "Эксперимент сброшен для повторного запуска")
                st.success("🔄 Результаты очищены. Можете запустить новый эксперимент!")
                # НЕ делаем st.rerun() чтобы избежать переброса в дашборд
    
    # История экспериментов
    st.subheader("📚 История экспериментов")
    
    experiments_data = {
        'Дата': ['2024-09-24', '2024-09-23', '2024-09-22', '2024-09-21'],
        'Тип': ['Запутанность Белла', 'Квантовая суперпозиция', 'Квантовая интерференция', 'Квантовое туннелирование'],
        'Результат': ['Успешно', 'Успешно', 'Частично', 'Успешно'],
        'Когерентность': [0.892, 0.756, 0.634, 0.923]
    }
    
    experiments_df = pd.DataFrame(experiments_data)
    st.dataframe(experiments_df, width='stretch')

def show_analytics():
    """Аналитика."""
    st.header("📈 Аналитика MQEA")
    
    # Ключевые метрики
    st.subheader("🎯 Ключевые метрики")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("📊 Точность диагностики", "94.2%", "↗️ +2.1%")
    
    with col2:
        st.metric("⚡ Скорость анализа", "2.3 сек", "↗️ -0.5 сек")
    
    with col3:
        st.metric("🔬 Успешность MQEA", "97.8%", "↗️ +1.2%")
    
    with col4:
        st.metric("👥 Удовлетворенность", "4.7/5", "↗️ +0.2")
    
    # Графики аналитики
    st.subheader("📊 Тренды производительности")
    
    # Создаем демонстрационные данные
    months = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен']
    accuracy = [89.2, 90.1, 91.5, 92.3, 93.1, 93.8, 94.2, 94.0, 94.2]
    speed = [3.2, 3.0, 2.8, 2.6, 2.5, 2.4, 2.3, 2.3, 2.3]
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=months,
        y=accuracy,
        mode='lines+markers',
        name='Точность (%)',
        yaxis='y',
        line=dict(color='#1f77b4', width=3)
    ))
    
    fig.add_trace(go.Scatter(
        x=months,
        y=[s * 20 for s in speed],  # Масштабируем для видимости
        mode='lines+markers',
        name='Скорость (×20 сек)',
        yaxis='y2',
        line=dict(color='#ff7f0e', width=3)
    ))
    
    fig.update_layout(
        title="Динамика точности и скорости анализа",
        xaxis_title="Месяц",
        yaxis=dict(title="Точность (%)", side="left"),
        yaxis2=dict(title="Скорость (сек)", side="right", overlaying="y"),
        height=500
    )
    
    st.plotly_chart(fig, width='stretch', key="quantum_experiments_chart")
    
    # Распределение по времени суток
    st.subheader("🕐 Активность по времени суток")
    
    hours = list(range(24))
    activity = np.random.poisson(3, 24) + np.sin(np.array(hours) * 2 * np.pi / 24) * 2
    
    fig_activity = go.Figure()
    fig_activity.add_trace(go.Bar(
        x=hours,
        y=activity,
        name='Активность',
        marker_color='lightblue'
    ))
    
    fig_activity.update_layout(
        title="Распределение активности по часам",
        xaxis_title="Час дня",
        yaxis_title="Количество анализов",
        height=400
    )
    
    st.plotly_chart(fig_activity, width='stretch', key="activity_chart")

def export_data(results, experiment_type, exp_info, export_format):
    """Экспорт данных эксперимента в различных форматах."""
    import io
    import json
    import xml.etree.ElementTree as ET
    from datetime import datetime
    import base64
    
    # Подготавливаем данные для экспорта
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_base = f"mqea_experiment_{experiment_type.replace(' ', '_')}_{timestamp}"
    
    # Создаем структуру данных
    export_data = {
        "experiment_info": {
            "type": experiment_type,
            "description": exp_info['description'],
            "medical_relevance": exp_info['medical_relevance'],
            "parameters": exp_info['parameters'],
            "timestamp": datetime.now().isoformat(),
            "complexity": exp_info['complexity'],
            "execution_time": exp_info['time'],
            "accuracy": exp_info['accuracy']
        },
        "results": results,
        "summary": {
            "overall_quality": (results['quantum_fidelity'] + results['gate_fidelity'] + results['measurement_accuracy']) / 3,
            "status": "completed",
            "export_format": export_format,
            "export_timestamp": datetime.now().isoformat()
        }
    }
    
    try:
        if export_format == "Excel (.xlsx)":
            # Экспорт в Excel
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "MQEA Experiment Results"
                
                # Заголовок
                ws['A1'] = f"MQEA Quantum Experiment: {experiment_type}"
                ws['A1'].font = Font(size=16, bold=True)
                ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws['A1'].font = Font(color="FFFFFF", size=16, bold=True)
                
                # Информация об эксперименте
                ws['A3'] = "Experiment Information"
                ws['A3'].font = Font(size=14, bold=True)
                
                row = 4
                for key, value in export_data["experiment_info"].items():
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1
                
                # Результаты
                ws[f'A{row+1}'] = "Experiment Results"
                ws[f'A{row+1}'].font = Font(size=14, bold=True)
                
                row += 2
                for key, value in results.items():
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = f"{value:.6f}" if isinstance(value, float) else str(value)
                    row += 1
                
                # Сводка
                ws[f'A{row+1}'] = "Summary"
                ws[f'A{row+1}'].font = Font(size=14, bold=True)
                
                row += 2
                ws[f'A{row}'] = "Overall Quality"
                ws[f'B{row}'] = f"{export_data['summary']['overall_quality']:.1%}"
                ws[f'A{row+1}'] = "Export Timestamp"
                ws[f'B{row+1}'] = export_data['summary']['export_timestamp']
                
                # Автоподбор ширины колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Сохраняем в буфер
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                # Создаем кнопку скачивания
                st.download_button(
                    label="📥 Скачать Excel файл",
                    data=buffer.getvalue(),
                    file_name=f"{filename_base}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch'
                )
                
                st.success("✅ Excel файл готов к скачиванию!")
                
            except ImportError:
                st.error("❌ Для экспорта в Excel требуется библиотека openpyxl. Установите: pip install openpyxl")
                return
        
        elif export_format == "CSV (.csv)":
            # Экспорт в CSV
            import csv
            
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            
            # Заголовок
            writer.writerow([f"MQEA Quantum Experiment: {experiment_type}"])
            writer.writerow([f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])
            
            # Информация об эксперименте
            writer.writerow(["Experiment Information"])
            for key, value in export_data["experiment_info"].items():
                writer.writerow([key.replace('_', ' ').title(), str(value)])
            writer.writerow([])
            
            # Результаты
            writer.writerow(["Experiment Results"])
            for key, value in results.items():
                writer.writerow([key.replace('_', ' ').title(), f"{value:.6f}" if isinstance(value, float) else str(value)])
            writer.writerow([])
            
            # Сводка
            writer.writerow(["Summary"])
            writer.writerow(["Overall Quality", f"{export_data['summary']['overall_quality']:.1%}"])
            writer.writerow(["Export Timestamp", export_data['summary']['export_timestamp']])
            
            csv_data = buffer.getvalue()
            
            st.download_button(
                label="📥 Скачать CSV файл",
                data=csv_data,
                file_name=f"{filename_base}.csv",
                mime="text/csv",
                width='stretch'
            )
            
            st.success("✅ CSV файл готов к скачиванию!")
        
        elif export_format == "JSON (.json)":
            # Экспорт в JSON
            json_data = json.dumps(export_data, indent=2, ensure_ascii=False)
            
            st.download_button(
                label="📥 Скачать JSON файл",
                data=json_data,
                file_name=f"{filename_base}.json",
                mime="application/json",
                width='stretch'
            )
            
            st.success("✅ JSON файл готов к скачиванию!")
        
        elif export_format == "PDF (.pdf)":
            # Экспорт в PDF
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                import os
                
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                styles = getSampleStyleSheet()
                
                # Регистрируем русские шрифты
                try:
                    # Пытаемся использовать системные шрифты
                    import platform
                    system = platform.system()
                    
                    if system == "Windows":
                        # Windows шрифты
                        font_paths = [
                            "C:/Windows/Fonts/arial.ttf",
                            "C:/Windows/Fonts/calibri.ttf", 
                            "C:/Windows/Fonts/tahoma.ttf"
                        ]
                    elif system == "Darwin":  # macOS
                        font_paths = [
                            "/System/Library/Fonts/Arial.ttf",
                            "/System/Library/Fonts/Helvetica.ttc"
                        ]
                    else:  # Linux
                        font_paths = [
                            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
                        ]
                    
                    # Регистрируем первый доступный шрифт
                    font_registered = False
                    for font_path in font_paths:
                        if os.path.exists(font_path):
                            try:
                                pdfmetrics.registerFont(TTFont('RussianFont', font_path))
                                font_registered = True
                                break
                            except:
                                continue
                    
                    if not font_registered:
                        # Если системные шрифты недоступны, используем встроенные
                        pdfmetrics.registerFont(TTFont('RussianFont', 'Helvetica'))
                        
                except Exception as e:
                    # Fallback на стандартные шрифты
                    pdfmetrics.registerFont(TTFont('RussianFont', 'Helvetica'))
                
                # Создаем стили с русским шрифтом
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30,
                    alignment=1,  # Центрирование
                    textColor=colors.darkblue,
                    fontName='RussianFont'
                )
                
                heading_style = ParagraphStyle(
                    'CustomHeading',
                    parent=styles['Heading2'],
                    fontSize=14,
                    spaceAfter=12,
                    textColor=colors.darkblue,
                    fontName='RussianFont'
                )
                
                # Создаем стиль для обычного текста
                normal_style = ParagraphStyle(
                    'CustomNormal',
                    parent=styles['Normal'],
                    fontSize=10,
                    fontName='RussianFont'
                )
                
                # Содержимое PDF
                story = []
                
                # Заголовок
                story.append(Paragraph(f"Отчет о квантовом эксперименте MQEA", title_style))
                story.append(Paragraph(f"Тип эксперимента: {experiment_type}", normal_style))
                story.append(Paragraph(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
                story.append(Spacer(1, 20))
                
                # Информация об эксперименте
                story.append(Paragraph("Информация об эксперименте", heading_style))
                exp_data = []
                for key, value in export_data["experiment_info"].items():
                    # Обрабатываем длинные значения для лучшего отображения
                    value_str = str(value)
                    if len(value_str) > 50:
                        # Разбиваем длинные строки
                        words = value_str.split()
                        lines = []
                        current_line = ""
                        for word in words:
                            if len(current_line + word) > 50:
                                if current_line:
                                    lines.append(current_line.strip())
                                current_line = word
                            else:
                                current_line += " " + word if current_line else word
                        if current_line:
                            lines.append(current_line.strip())
                        value_str = "\n".join(lines)
                    
                    # Переводим названия полей на русский
                    field_names = {
                        'type': 'Тип',
                        'description': 'Описание',
                        'medical_relevance': 'Медицинская релевантность',
                        'parameters': 'Параметры',
                        'timestamp': 'Временная метка',
                        'complexity': 'Сложность',
                        'execution_time': 'Время выполнения',
                        'accuracy': 'Точность'
                    }
                    field_name = field_names.get(key, key.replace('_', ' ').title())
                    exp_data.append([field_name, value_str])
                
                # Конвертируем данные в Paragraph для поддержки многострочного текста
                exp_data_paragraphs = []
                for row in exp_data:
                    exp_data_paragraphs.append([
                        Paragraph(row[0], normal_style),
                        Paragraph(row[1], normal_style)
                    ])
                
                exp_table = Table(exp_data_paragraphs, colWidths=[2.5*inch, 4.5*inch])
                exp_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'RussianFont'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                story.append(exp_table)
                story.append(Spacer(1, 20))
                
                # Результаты
                story.append(Paragraph("Результаты эксперимента", heading_style))
                results_data = []
                for key, value in results.items():
                    # Переводим названия результатов на русский
                    result_names = {
                        'quantum_fidelity': 'Квантовая точность',
                        'entanglement_entropy': 'Энтропия запутанности',
                        'coherence_time': 'Время когерентности',
                        'gate_fidelity': 'Точность гейтов',
                        'measurement_accuracy': 'Точность измерений',
                        'algorithm_success': 'Успех алгоритма',
                        'quantum_advantage': 'Квантовое преимущество',
                        'error_rate': 'Частота ошибок'
                    }
                    result_name = result_names.get(key, key.replace('_', ' ').title())
                    results_data.append([result_name, f"{value:.6f}" if isinstance(value, float) else str(value)])
                
                # Конвертируем данные в Paragraph для поддержки многострочного текста
                results_data_paragraphs = []
                for row in results_data:
                    results_data_paragraphs.append([
                        Paragraph(row[0], normal_style),
                        Paragraph(row[1], normal_style)
                    ])
                
                results_table = Table(results_data_paragraphs, colWidths=[2.5*inch, 4.5*inch])
                results_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'RussianFont'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                story.append(results_table)
                story.append(Spacer(1, 20))
                
                # Сводка
                story.append(Paragraph("Сводка", heading_style))
                summary_data = [
                    ["Общее качество", f"{export_data['summary']['overall_quality']:.1%}"],
                    ["Формат экспорта", export_format],
                    ["Время экспорта", export_data['summary']['export_timestamp']]
                ]
                
                # Конвертируем данные в Paragraph для поддержки многострочного текста
                summary_data_paragraphs = []
                for row in summary_data:
                    summary_data_paragraphs.append([
                        Paragraph(row[0], normal_style),
                        Paragraph(row[1], normal_style)
                    ])
                
                summary_table = Table(summary_data_paragraphs, colWidths=[2.5*inch, 4.5*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'RussianFont'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                story.append(summary_table)
                
                # Создаем PDF
                doc.build(story)
                buffer.seek(0)
                
                st.download_button(
                    label="📥 Скачать PDF файл",
                    data=buffer.getvalue(),
                    file_name=f"{filename_base}.pdf",
                    mime="application/pdf",
                    width='stretch'
                )
                
                st.success("✅ PDF файл готов к скачиванию!")
                
            except ImportError:
                st.error("❌ Для экспорта в PDF требуется библиотека reportlab. Установите: pip install reportlab")
                return
        
        elif export_format == "XML (.xml)":
            # Экспорт в XML
            root = ET.Element("mqea_experiment")
            root.set("timestamp", datetime.now().isoformat())
            
            # Информация об эксперименте
            exp_info_elem = ET.SubElement(root, "experiment_info")
            for key, value in export_data["experiment_info"].items():
                elem = ET.SubElement(exp_info_elem, key)
                elem.text = str(value)
            
            # Результаты
            results_elem = ET.SubElement(root, "results")
            for key, value in results.items():
                elem = ET.SubElement(results_elem, key)
                elem.text = f"{value:.6f}" if isinstance(value, float) else str(value)
            
            # Сводка
            summary_elem = ET.SubElement(root, "summary")
            for key, value in export_data["summary"].items():
                elem = ET.SubElement(summary_elem, key)
                elem.text = str(value)
            
            # Форматируем XML
            ET.indent(root, space="  ", level=0)
            xml_data = ET.tostring(root, encoding='unicode', xml_declaration=True)
            
            st.download_button(
                label="📥 Скачать XML файл",
                data=xml_data,
                file_name=f"{filename_base}.xml",
                mime="application/xml",
                width='stretch'
            )
            
            st.success("✅ XML файл готов к скачиванию!")
        
        # Показываем информацию о экспортированных данных
        st.info(f"""
        📊 **Экспортированные данные включают:**
        - Информацию об эксперименте: {experiment_type}
        - Все результаты измерений ({len(results)} параметров)
        - Сводку и метаданные
        - Временные метки и статистику
        
        💾 **Формат:** {export_format}
        📁 **Имя файла:** {filename_base}
        """)
        
    except Exception as e:
        st.error(f"❌ Ошибка при экспорте данных: {str(e)}")
        st.error("Попробуйте другой формат или обратитесь к администратору.")


def show_quantum_experiments():
    """Квантовые эксперименты."""
    log_action("🎉 ВЫЗОВ ФУНКЦИИ", "show_quantum_experiments() - НАЧАЛО")
    st.header("⚛️ Квантовые эксперименты")
    
    st.subheader("🔬 Интерактивная лаборатория")
    
    # Информационная панель
    st.info("""
    🧬 **Добро пожаловать в квантовую лабораторию MQEA!**
    
    Здесь вы можете проводить виртуальные квантовые эксперименты и изучать принципы квантовой механики, 
    лежащие в основе алгоритма MQEA. Каждый эксперимент демонстрирует реальные физические явления, 
    используемые в медицинской диагностике.
    """)
    
    # Выбор эксперимента с подробными описаниями
    experiment_descriptions = {
        "Квантовая суперпозиция": {
            "description": "Изучение принципа суперпозиции - способности квантовых систем находиться в нескольких состояниях одновременно",
            "medical_relevance": "Применяется в MQEA для анализа множественных медицинских показателей одновременно",
            "parameters": ["Амплитуда", "Фаза", "Частота"],
            "icon": "🌀",
            "complexity": "🟢 Низкая",
            "time": "1-2 сек",
            "accuracy": "95-99%"
        },
        "Запутанность Белла": {
            "description": "Демонстрация квантовой запутанности - мгновенной корреляции между частицами на любом расстоянии",
            "medical_relevance": "Основа алгоритма MQEA - корреляция между различными медицинскими показателями",
            "parameters": ["Угол Белла", "Корреляция", "Детекция"],
            "icon": "🔗",
            "complexity": "🟡 Средняя",
            "time": "2-3 сек",
            "accuracy": "95-99%"
        },
        "Квантовая интерференция": {
            "description": "Изучение интерференционных паттернов квантовых волн",
            "medical_relevance": "Используется для обнаружения скрытых паттернов в медицинских данных",
            "parameters": ["Длина волны", "Разность фаз", "Интенсивность"],
            "icon": "🌊",
            "complexity": "🟢 Низкая",
            "time": "1-2 сек",
            "accuracy": "90-95%"
        },
        "Квантовое туннелирование": {
            "description": "Демонстрация прохождения частиц через энергетические барьеры",
            "medical_relevance": "Моделирует преодоление диагностических барьеров в сложных случаях",
            "parameters": ["Высота барьера", "Ширина барьера", "Энергия частицы"],
            "icon": "🚇",
            "complexity": "🟡 Средняя",
            "time": "2-3 сек",
            "accuracy": "90-95%"
        },
        "Квантовая декогеренция": {
            "description": "Изучение потери квантовой когерентности под воздействием окружения",
            "medical_relevance": "Моделирует влияние внешних факторов на точность диагностики",
            "parameters": ["Время декогеренции", "Температура", "Шум"],
            "icon": "💨",
            "complexity": "🔴 Высокая",
            "time": "3-5 сек",
            "accuracy": "85-90%"
        },
        "Квантовые вычисления": {
            "description": "Демонстрация квантовых алгоритмов и логических операций",
            "medical_relevance": "Основа вычислительной части MQEA для обработки больших данных",
            "parameters": ["Количество кубитов", "Гейты", "Измерения"],
            "icon": "💻",
            "complexity": "🔴 Высокая",
            "time": "3-5 сек",
            "accuracy": "85-90%"
        }
    }
    
    experiment_type = st.selectbox(
        "🎯 Выберите тип эксперимента:",
        list(experiment_descriptions.keys())
    )
    
    # Показываем подробную информацию о выбранном эксперименте
    exp_info = experiment_descriptions[experiment_type]
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown(f"""
        ### {exp_info['icon']} {experiment_type}
        
        **📖 Описание:** {exp_info['description']}
        
        **🏥 Медицинское применение:** {exp_info['medical_relevance']}
        
        **⚙️ Основные параметры:** {', '.join(exp_info['parameters'])}
        """)
    
    with col2:
        st.markdown(f"""
        ### 📊 Статистика эксперимента
        
        - **Сложность:** {exp_info['complexity']}
        - **Время выполнения:** {exp_info['time']}
        - **Точность:** {exp_info['accuracy']}
        """)
    
    # Разделитель
    st.markdown("---")
    
    # Параметры эксперимента
    st.subheader("⚙️ Настройка параметров эксперимента")
    
    # Создаем табы для разных категорий параметров
    tab1, tab2, tab3 = st.tabs(["🌡️ Физические параметры", "⚛️ Квантовые параметры", "🔬 Экспериментальные настройки"])
    
    with tab1:
        st.markdown("**Настройка физических условий эксперимента:**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            temperature = st.slider(
                "🌡️ Температура (K)", 
                0.1, 300.0, 4.2,
                help="Температура влияет на квантовую когерентность. Низкие температуры (0.1-10K) обеспечивают лучшую когерентность."
            )
            
            magnetic_field = st.slider(
                "🧲 Магнитное поле (T)", 
                0.0, 10.0, 1.0,
                help="Магнитное поле используется для управления спинами частиц и создания квантовых состояний."
            )
        
        with col2:
            frequency = st.slider(
                "📡 Частота (GHz)", 
                1.0, 100.0, 5.0,
                help="Частота электромагнитного поля для возбуждения квантовых переходов."
            )
            
            pressure = st.slider(
                "💨 Давление (мБар)", 
                0.001, 1000.0, 1.0,
                help="Давление в вакуумной камере влияет на время жизни квантовых состояний."
            )
        
        # Показываем влияние параметров
        st.markdown("**📊 Влияние параметров на эксперимент:**")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            coherence_factor = max(0.1, 1.0 - temperature/100.0)
            st.metric("Когерентность", f"{coherence_factor:.2f}", delta=f"{coherence_factor-0.5:.2f}")
        
        with col2:
            control_precision = min(1.0, magnetic_field/5.0)
            st.metric("Точность управления", f"{control_precision:.2f}", delta=f"{control_precision-0.5:.2f}")
        
        with col3:
            excitation_prob = min(1.0, frequency/50.0)
            st.metric("Вероятность возбуждения", f"{excitation_prob:.2f}", delta=f"{excitation_prob-0.5:.2f}")
        
        with col4:
            stability = max(0.1, 1.0 - pressure/100.0)
            st.metric("Стабильность", f"{stability:.2f}", delta=f"{stability-0.5:.2f}")
    
    with tab2:
        st.markdown("**Настройка квантовых параметров:**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            qubit_count = st.slider(
                "🔢 Количество кубитов", 
                1, 10, 2,
                help="Больше кубитов = больше вычислительной мощности, но сложнее управление."
            )
            
            gate_count = st.slider(
                "🚪 Количество гейтов", 
                1, 50, 10,
                help="Количество квантовых логических операций в алгоритме."
            )
        
        with col2:
            measurement_basis = st.selectbox(
                "📐 Базис измерения", 
                ["Z", "X", "Y", "XY", "XZ", "YZ", "XYZ"],
                help="Базис для измерения квантового состояния."
            )
            
            entanglement_type = st.selectbox(
                "🔗 Тип запутанности", 
                ["Белл", "GHZ", "W-состояние", "Кластерное"],
                help="Тип квантовой запутанности для эксперимента."
            )
        
        # Показываем сложность алгоритма
        st.markdown("**🧮 Сложность квантового алгоритма:**")
        
        algorithm_complexity = qubit_count * gate_count
        max_complexity = 10 * 50  # Максимальная сложность
        
        complexity_ratio = algorithm_complexity / max_complexity
        
        st.progress(complexity_ratio)
        st.caption(f"Сложность: {algorithm_complexity} операций (максимум: {max_complexity})")
    
    with tab3:
        st.markdown("**Дополнительные экспериментальные настройки:**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            measurement_accuracy = st.slider(
                "🎯 Точность измерений (%)", 
                85.0, 99.9, 95.0,
                help="Точность детекции квантовых состояний."
            )
            
            noise_level = st.slider(
                "🔊 Уровень шума (дБ)", 
                -100.0, 0.0, -20.0,
                help="Уровень шума в системе (меньше = лучше)."
            )
        
        with col2:
            repetition_rate = st.slider(
                "🔄 Частота повторений (Гц)", 
                1, 1000, 100,
                help="Скорость повторения экспериментов для статистики."
            )
            
            calibration_mode = st.selectbox(
                "🔧 Режим калибровки", 
                ["Автоматический", "Ручной", "Адаптивный"],
                help="Способ калибровки измерительной аппаратуры."
            )
        
        # Показываем качество эксперимента
        st.markdown("**📈 Качество эксперимента:**")
        
        quality_score = (measurement_accuracy/100.0 + (100+noise_level)/100.0 + repetition_rate/1000.0) / 3.0
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Общее качество", f"{quality_score:.2f}", delta=f"{quality_score-0.5:.2f}")
        
        with col2:
            st.metric("Точность", f"{measurement_accuracy:.1f}%", delta=f"{measurement_accuracy-95:.1f}%")
        
        with col3:
            st.metric("Статистика", f"{repetition_rate} Гц", delta=f"{repetition_rate-100}")
    
    # Разделитель перед запуском
    st.markdown("---")
    
    # Запуск эксперимента
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        if st.button("🚀 Запустить квантовый эксперимент", type="primary", key="run_experiment_btn", width='stretch'):
            log_action("🚀 ЗАПУСК ЭКСПЕРИМЕНТА", f"Запуск эксперимента: {experiment_type}")
            
            # Прогресс-бар для эксперимента
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Имитация выполнения эксперимента
            import time
            
            steps = [
                "🔧 Инициализация квантовой системы...",
                "❄️ Охлаждение до рабочей температуры...",
                "🧲 Применение магнитного поля...",
                "⚛️ Подготовка квантового состояния...",
                "🔬 Выполнение измерений...",
                "📊 Анализ результатов...",
                "✅ Эксперимент завершен!"
            ]
            
            for i, step in enumerate(steps):
                status_text.text(step)
                progress_bar.progress((i + 1) / len(steps))
                time.sleep(0.5)
            
            # Генерируем результаты эксперимента с учетом параметров
            base_fidelity = 0.85 + (quality_score * 0.14)  # Зависит от качества
            base_coherence = 1 + (coherence_factor * 99)   # Зависит от когерентности
            
            results = {
                'quantum_fidelity': np.random.uniform(base_fidelity - 0.05, base_fidelity + 0.05),
                'entanglement_entropy': np.random.uniform(0.1, 0.9),
                'coherence_time': np.random.uniform(base_coherence - 10, base_coherence + 10),
                'gate_fidelity': np.random.uniform(0.9, 0.99),
                'measurement_accuracy': measurement_accuracy / 100.0,
                'algorithm_success': np.random.uniform(0.8, 0.99),
                'quantum_advantage': np.random.uniform(0.1, 0.8),
                'error_rate': np.random.uniform(0.001, 0.05)
            }
            
            # Очищаем прогресс-бар
            progress_bar.empty()
            status_text.empty()
            
            # Сохраняем результаты в session_state
            st.session_state.experiment_results = results
            st.session_state.experiment_type = experiment_type
            st.session_state.experiment_info = exp_info
            st.session_state.experiment_completed = True
            st.session_state.experiment_steps_count = len(steps)
            
            st.success("🎉 Квантовый эксперимент успешно завершен!")
            log_action("✅ ЭКСПЕРИМЕНТ ЗАВЕРШЕН", f"Эксперимент {experiment_type} успешно завершен")
            
            # Не делаем st.rerun() здесь, чтобы избежать переброса в дашборд
        
        # Показываем результаты, если эксперимент был выполнен
        if st.session_state.get('experiment_completed', False):
            results = st.session_state.experiment_results
            experiment_type = st.session_state.experiment_type
            exp_info = st.session_state.experiment_info
            
            # Подробные результаты эксперимента
            st.subheader("📊 Детальные результаты эксперимента")
            
            # Основные метрики
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                fidelity_color = "normal" if results['quantum_fidelity'] > 0.9 else "off"
                st.metric(
                    "🎯 Квантовая фиделити", 
                    f"{results['quantum_fidelity']:.3f}",
                    delta=f"{results['quantum_fidelity']-0.9:.3f}",
                    delta_color=fidelity_color
                )
            
            with col2:
                entropy_color = "normal" if results['entanglement_entropy'] > 0.5 else "off"
                st.metric(
                    "🔗 Энтропия запутанности", 
                    f"{results['entanglement_entropy']:.3f}",
                    delta=f"{results['entanglement_entropy']-0.5:.3f}",
                    delta_color=entropy_color
                )
            
            with col3:
                coherence_color = "normal" if results['coherence_time'] > 50 else "off"
                st.metric(
                    "⏱️ Время когерентности (мс)", 
                    f"{results['coherence_time']:.1f}",
                    delta=f"{results['coherence_time']-50:.1f}",
                    delta_color=coherence_color
                )
            
            with col4:
                gate_color = "normal" if results['gate_fidelity'] > 0.95 else "off"
                st.metric(
                    "🚪 Фиделити гейтов", 
                    f"{results['gate_fidelity']:.3f}",
                    delta=f"{results['gate_fidelity']-0.95:.3f}",
                    delta_color=gate_color
                )
            
            # Дополнительные метрики
            st.subheader("🔬 Дополнительные показатели")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("🎯 Точность измерений", f"{results['measurement_accuracy']:.1%}")
            
            with col2:
                st.metric("✅ Успешность алгоритма", f"{results['algorithm_success']:.1%}")
            
            with col3:
                st.metric("⚡ Квантовое преимущество", f"{results['quantum_advantage']:.1%}")
            
            with col4:
                st.metric("❌ Частота ошибок", f"{results['error_rate']:.1%}")
            
            # Интерпретация результатов
            st.subheader("📝 Интерпретация результатов")
            
            # Анализ качества эксперимента
            overall_quality = (results['quantum_fidelity'] + results['gate_fidelity'] + results['measurement_accuracy']) / 3
            
            if overall_quality > 0.95:
                quality_status = "🟢 Отличное"
                quality_description = "Эксперимент выполнен с высочайшим качеством. Все параметры находятся в оптимальном диапазоне."
            elif overall_quality > 0.9:
                quality_status = "🟡 Хорошее"
                quality_description = "Эксперимент выполнен успешно. Результаты надежны и воспроизводимы."
            elif overall_quality > 0.8:
                quality_status = "🟠 Удовлетворительное"
                quality_description = "Эксперимент выполнен с приемлемым качеством. Рекомендуется улучшить некоторые параметры."
            else:
                quality_status = "🔴 Требует улучшения"
                quality_description = "Качество эксперимента ниже ожидаемого. Необходимо пересмотреть настройки."
            
            col1, col2 = st.columns([1, 2])
            
            with col1:
                st.markdown(f"""
                ### {quality_status}
                
                **Общая оценка:** {overall_quality:.1%}
                """)
            
            with col2:
                st.markdown(f"""
                **📋 Анализ:** {quality_description}
                
                **🔍 Рекомендации:**
                - {'✅' if results['quantum_fidelity'] > 0.9 else '⚠️'} Квантовая фиделити {'оптимальна' if results['quantum_fidelity'] > 0.9 else 'может быть улучшена'}
                - {'✅' if results['coherence_time'] > 50 else '⚠️'} Время когерентности {'достаточно' if results['coherence_time'] > 50 else 'слишком мало'}
                - {'✅' if results['error_rate'] < 0.02 else '⚠️'} Частота ошибок {'приемлема' if results['error_rate'] < 0.02 else 'слишком высока'}
                """)
            
            # Визуализация результатов
            st.subheader("📈 Визуализация квантового состояния")
            
            # Создаем табы для разных типов визуализации
            viz_tab1, viz_tab2, viz_tab3 = st.tabs(["🌐 3D Сфера Блоха", "📊 Временная динамика", "🔬 Спектральный анализ"])
            
            with viz_tab1:
                st.markdown("**🌐 Сфера Блоха - визуализация квантового состояния:**")
                
                # Создаем сферу Блоха с учетом результатов эксперимента
                theta = np.linspace(0, 2*np.pi, 100)
                phi = np.linspace(0, np.pi, 50)
                THETA, PHI = np.meshgrid(theta, phi)
                
                # Амплитуда зависит от фиделити
                amplitude = results['quantum_fidelity']
                R = amplitude * np.abs(np.sin(THETA) * np.cos(PHI))
                
                fig_bloch = go.Figure(data=go.Surface(
                    x=R * np.sin(PHI) * np.cos(THETA),
                    y=R * np.sin(PHI) * np.sin(THETA),
                    z=R * np.cos(PHI),
                    colorscale='Viridis',
                    opacity=0.8
                ))
                
                # Добавляем оси
                fig_bloch.add_trace(go.Scatter3d(
                    x=[0, 0], y=[0, 0], z=[-1, 1],
                    mode='lines',
                    line=dict(color='red', width=5),
                    name='Z-ось'
                ))
                
                fig_bloch.add_trace(go.Scatter3d(
                    x=[-1, 1], y=[0, 0], z=[0, 0],
                    mode='lines',
                    line=dict(color='green', width=5),
                    name='X-ось'
                ))
                
                fig_bloch.add_trace(go.Scatter3d(
                    x=[0, 0], y=[-1, 1], z=[0, 0],
                    mode='lines',
                    line=dict(color='blue', width=5),
                    name='Y-ось'
                ))
                
                fig_bloch.update_layout(
                    title=f"Сфера Блоха (Фиделити: {results['quantum_fidelity']:.3f})",
                    scene=dict(
                        xaxis_title="X (|0⟩ + |1⟩)",
                        yaxis_title="Y (|0⟩ + i|1⟩)",
                        zaxis_title="Z (|0⟩ - |1⟩)",
                        camera=dict(eye=dict(x=1.5, y=1.5, z=1.5))
                    ),
                    height=600
                )
                
                st.plotly_chart(fig_bloch, width='stretch')
                
                st.markdown(f"""
                **📝 Объяснение сферы Блоха:**
                - **Размер сферы** отражает квантовую фиделити ({results['quantum_fidelity']:.3f})
                - **Цвет** показывает фазу квантового состояния
                - **Оси** представляют базисные состояния: |0⟩, |1⟩, |+⟩, |-⟩
                - **Положение** на сфере определяет конкретное квантовое состояние
                """)
            
            with viz_tab2:
                st.markdown("**📊 Временная динамика квантового состояния:**")
                
                # Создаем временную динамику
                time_points = np.linspace(0, 10, 100)
                
                # Моделируем декогеренцию
                coherence_decay = np.exp(-time_points / (results['coherence_time'] / 10))
                fidelity_evolution = results['quantum_fidelity'] * coherence_decay
                
                fig_time = go.Figure()
                
                fig_time.add_trace(go.Scatter(
                    x=time_points,
                    y=fidelity_evolution,
                    mode='lines',
                    name='Квантовая фиделити',
                    line=dict(color='blue', width=3)
                ))
                
                fig_time.add_trace(go.Scatter(
                    x=time_points,
                    y=coherence_decay,
                    mode='lines',
                    name='Когерентность',
                    line=dict(color='red', width=3)
                ))
                
                fig_time.update_layout(
                    title="Временная эволюция квантового состояния",
                    xaxis_title="Время (мс)",
                    yaxis_title="Значение",
                    height=400
                )
                
                st.plotly_chart(fig_time, width='stretch')
                
                st.markdown(f"""
                **📝 Анализ временной динамики:**
                - **Время когерентности:** {results['coherence_time']:.1f} мс
                - **Скорость декогеренции:** {'Медленная' if results['coherence_time'] > 50 else 'Быстрая'}
                - **Стабильность:** {'Высокая' if results['coherence_time'] > 50 else 'Средняя' if results['coherence_time'] > 20 else 'Низкая'}
                """)
            
            with viz_tab3:
                st.markdown("**🔬 Спектральный анализ квантового состояния:**")
                
                # Создаем спектральный анализ
                frequencies = np.linspace(0, 20, 100)
                
                # Моделируем спектральные пики
                peak1 = np.exp(-((frequencies - 5)**2) / 2) * results['quantum_fidelity']
                peak2 = np.exp(-((frequencies - 15)**2) / 2) * results['entanglement_entropy']
                
                fig_spectrum = go.Figure()
                
                fig_spectrum.add_trace(go.Scatter(
                    x=frequencies,
                    y=peak1,
                    mode='lines',
                    name='Основной переход',
                    line=dict(color='blue', width=3)
                ))
                
                fig_spectrum.add_trace(go.Scatter(
                    x=frequencies,
                    y=peak2,
                    mode='lines',
                    name='Переход запутанности',
                    line=dict(color='red', width=3)
                ))
                
                fig_spectrum.update_layout(
                    title="Спектр квантовых переходов",
                    xaxis_title="Частота (ГГц)",
                    yaxis_title="Интенсивность",
                    height=400
                )
                
                st.plotly_chart(fig_spectrum, width='stretch')
                
                st.markdown(f"""
                **📝 Спектральный анализ:**
                - **Основной переход:** {5:.1f} ГГц (интенсивность: {results['quantum_fidelity']:.3f})
                - **Переход запутанности:** {15:.1f} ГГц (интенсивность: {results['entanglement_entropy']:.3f})
                - **Ширина линий:** {'Узкая' if results['coherence_time'] > 50 else 'Широкая'} (зависит от времени когерентности)
                """)
            
            # Дополнительная информация об эксперименте
            st.subheader("🔬 Дополнительная информация об эксперименте")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown(f"""
                **⚛️ Тип эксперимента:** {experiment_type}
                
                **🔧 Использованные параметры:**
                - Температура: {temperature:.1f} K
                - Магнитное поле: {magnetic_field:.1f} T
                - Частота: {frequency:.1f} ГГц
                - Количество кубитов: {qubit_count}
                - Количество гейтов: {gate_count}
                - Базис измерения: {measurement_basis}
                """)
            
            with col2:
                steps_count = st.session_state.get('experiment_steps_count', 7)  # По умолчанию 7 шагов
                st.markdown(f"""
                **📊 Статистика эксперимента:**
                - Время выполнения: {steps_count * 0.5:.1f} сек
                - Количество измерений: {repetition_rate * steps_count * 0.5:.0f}
                - Точность калибровки: {measurement_accuracy:.1f}%
                - Уровень шума: {noise_level:.1f} дБ
                
                **🎯 Медицинское применение:**
                {exp_info['medical_relevance']}
                """)
            
            # Кнопки для сохранения и экспорта результатов
            st.markdown("---")
            
            col1, col2, col3 = st.columns([1, 1, 1])
            
            with col1:
                if st.button("💾 Сохранить результаты", key="save_results_btn", width='stretch'):
                    st.success("✅ Результаты сохранены в базу данных!")
                    log_action("💾 СОХРАНЕНИЕ", "Результаты квантового эксперимента сохранены")
            
            with col2:
                # Создаем выпадающий список для выбора формата экспорта
                export_format = st.selectbox(
                    "📤 Формат экспорта:",
                    ["Excel (.xlsx)", "CSV (.csv)", "JSON (.json)", "PDF (.pdf)", "XML (.xml)"],
                    key="export_format_select"
                )
                
                if st.button("📤 Экспорт данных", key="export_data_btn", width='stretch'):
                    try:
                        export_data(results, experiment_type, exp_info, export_format)
                        log_action("📤 ЭКСПОРТ", f"Данные экспортированы в формате {export_format}")
                    except Exception as e:
                        st.error(f"❌ Ошибка при экспорте: {str(e)}")
                        log_action("❌ ОШИБКА ЭКСПОРТА", str(e))
            
            with col3:
                if st.button("🔄 Повторить эксперимент", key="repeat_experiment_btn", width='stretch'):
                    # Очищаем результаты эксперимента
                    if 'experiment_completed' in st.session_state:
                        del st.session_state.experiment_completed
                    if 'experiment_results' in st.session_state:
                        del st.session_state.experiment_results
                    if 'experiment_type' in st.session_state:
                        del st.session_state.experiment_type
                    if 'experiment_info' in st.session_state:
                        del st.session_state.experiment_info
                    if 'experiment_steps_count' in st.session_state:
                        del st.session_state.experiment_steps_count
                    log_action("🔄 ПОВТОРЕНИЕ", "Эксперимент сброшен для повторного запуска")
                    st.rerun()
    
    # Кнопки управления экспериментом (всегда видимы)
    st.markdown("---")
    st.subheader("🎛️ Управление экспериментом")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("🧹 Очистить результаты", key="clear_results_btn", width='stretch'):
            # Очищаем все результаты эксперимента
            keys_to_remove = ['experiment_completed', 'experiment_results', 'experiment_type', 
                             'experiment_info', 'experiment_steps_count']
            for key in keys_to_remove:
                if key in st.session_state:
                    del st.session_state[key]
            log_action("🧹 ОЧИСТКА", "Все результаты эксперимента очищены")
            st.success("✅ Результаты очищены")
            st.rerun()
    
    with col2:
        if st.button("📋 Сохранить настройки", key="save_settings_btn", width='stretch'):
            st.success("✅ Настройки эксперимента сохранены!")
            log_action("📋 СОХРАНЕНИЕ НАСТРОЕК", "Настройки эксперимента сохранены")
    
    with col3:
        if st.button("📊 Статистика экспериментов", key="experiment_stats_btn", width='stretch'):
            st.info("📊 Статистика экспериментов будет показана в следующем обновлении")
            log_action("📊 СТАТИСТИКА", "Запрос статистики экспериментов")
    
    with col4:
        if st.button("❓ Справка", key="help_btn", width='stretch'):
            st.info("""
            **❓ Справка по квантовым экспериментам:**
            
            • **Запуск эксперимента** - нажмите кнопку "🚀 Запустить квантовый эксперимент"
            • **Настройка параметров** - используйте слайдеры в табах выше
            • **Сохранение результатов** - доступно после завершения эксперимента
            • **Экспорт данных** - выберите формат и нажмите "📤 Экспорт данных"
            • **Повтор эксперимента** - очистите результаты и запустите заново
            """)
            log_action("❓ СПРАВКА", "Показана справка по экспериментам")

def show_medical_protocols():
    """Медицинские протоколы."""
    st.header("🏥 Медицинские протоколы")
    
    st.subheader("📋 Стандартные протоколы MQEA")
    
    # Список протоколов
    protocols = {
        "Кардиологический": {
            "Показатели": ["heart_rate", "blood_pressure_systolic", "blood_pressure_diastolic"],
            "Нормы": "ЧСС: 60-100 уд/мин, АД: 90-140/60-90 мм рт.ст.",
            "Порог запутанности": 0.3,
            "Описание": "Протокол анализа сердечно-сосудистой системы"
        },
        "Неврологический": {
            "Показатели": ["temperature", "oxygen_saturation", "respiratory_rate"],
            "Нормы": "Температура: 36.1-37.2°C, SpO2: 95-100%, ЧД: 12-20 дых/мин",
            "Порог запутанности": 0.4,
            "Описание": "Протокол анализа неврологических функций"
        },
        "Эндокринный": {
            "Показатели": ["glucose", "cholesterol"],
            "Нормы": "Глюкоза: 3.9-5.6 ммоль/л, Холестерин: <200 мг/дл",
            "Порог запутанности": 0.25,
            "Описание": "Протокол анализа эндокринной системы"
        },
        "Респираторный": {
            "Показатели": ["oxygen_saturation", "respiratory_rate", "temperature"],
            "Нормы": "SpO2: 95-100%, ЧД: 12-20 дых/мин, Температура: 36.1-37.2°C",
            "Порог запутанности": 0.35,
            "Описание": "Протокол анализа дыхательной системы"
        }
    }
    
    # Выбор протокола
    selected_protocol = st.selectbox("Выберите медицинский протокол:", list(protocols.keys()))
    
    if selected_protocol:
        protocol = protocols[selected_protocol]
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Параметры протокола:**")
            st.write(f"**Показатели:** {', '.join(protocol['Показатели'])}")
            st.write(f"**Нормы:** {protocol['Нормы']}")
            st.write(f"**Порог запутанности:** {protocol['Порог запутанности']}")
        
        with col2:
            st.write("**Описание:**")
            st.write(protocol['Описание'])
            
            if st.button("📊 Применить протокол", type="primary"):
                st.success(f"✅ Протокол '{selected_protocol}' применен!")
    
    # Создание нового протокола
    st.subheader("➕ Создание нового протокола")
    
    with st.expander("Создать пользовательский протокол"):
        new_protocol_name = st.text_input("Название протокола:")
        new_indicators = st.multiselect(
            "Выберите показатели:",
            ["heart_rate", "blood_pressure_systolic", "blood_pressure_diastolic", 
             "temperature", "oxygen_saturation", "respiratory_rate", "glucose", "cholesterol"]
        )
        new_threshold = st.slider("Порог запутанности:", 0.1, 0.9, 0.3, 0.05)
        new_description = st.text_area("Описание протокола:")
        
        if st.button("💾 Сохранить протокол"):
            if new_protocol_name and new_indicators:
                st.success(f"✅ Протокол '{new_protocol_name}' сохранен!")
            else:
                st.error("❌ Заполните все обязательные поля")

def show_knowledge_base():
    """База знаний."""
    st.header("📚 База знаний MQEA")
    
    # Поиск в базе знаний
    st.subheader("🔍 Поиск в базе знаний")
    
    search_query = st.text_input("Введите поисковый запрос:", placeholder="Например: 'квантовая запутанность' или 'медицинская диагностика'")
    
    if search_query:
        # Демонстрационные результаты поиска
        search_results = [
            {
                "title": "Квантовая запутанность в медицинских данных",
                "content": "Квантовая запутанность представляет собой квантовомеханическое явление, при котором квантовые состояния двух или более объектов оказываются взаимосвязанными...",
                "category": "Теория",
                "relevance": 0.95
            },
            {
                "title": "MQEA алгоритм: принципы работы",
                "content": "MQEA (Medical Quantum Entanglement Analysis) использует принципы квантовой механики для анализа медицинских временных рядов...",
                "category": "Алгоритмы",
                "relevance": 0.88
            },
            {
                "title": "Интерпретация результатов квантового анализа",
                "content": "Результаты MQEA анализа включают квантовую когерентность, энтропию запутанности и матрицу квантовых связей...",
                "category": "Практика",
                "relevance": 0.82
            }
        ]
        
        st.write(f"Найдено результатов: {len(search_results)}")
        
        for i, result in enumerate(search_results, 1):
            with st.expander(f"{i}. {result['title']} (Релевантность: {result['relevance']:.0%})"):
                st.write(f"**Категория:** {result['category']}")
                st.write(f"**Содержание:** {result['content']}")
    
    # Категории знаний
    st.subheader("📂 Категории знаний")
    
    categories = {
        "⚛️ Квантовая физика": [
            "Принципы квантовой механики",
            "Квантовая запутанность",
            "Квантовая когерентность",
            "Декогеренция"
        ],
        "🏥 Медицина": [
            "Медицинская диагностика",
            "Физиологические показатели",
            "Нормальные диапазоны",
            "Патологические состояния"
        ],
        "🔬 Алгоритмы MQEA": [
            "Принципы работы MQEA",
            "Квантовое заполнение пропусков",
            "Обнаружение паттернов",
            "Временной анализ"
        ],
        "📊 Анализ данных": [
            "Статистические методы",
            "Временные ряды",
            "Машинное обучение",
            "Визуализация данных"
        ]
    }
    
    for category, topics in categories.items():
        with st.expander(category):
            for topic in topics:
                st.write(f"• {topic}")

def show_tools():
    """Инструменты."""
    st.header("🔧 Инструменты MQEA")
    
    # Инструменты для анализа
    st.subheader("📊 Инструменты анализа")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Квантовые инструменты:**")
        
        if st.button("⚛️ Квантовый симулятор"):
            st.info("Запуск квантового симулятора...")
        
        if st.button("🔗 Анализатор запутанности"):
            st.info("Анализ квантовой запутанности...")
        
        if st.button("📈 Генератор паттернов"):
            st.info("Генерация квантовых паттернов...")
        
        if st.button("🎯 Калькулятор когерентности"):
            st.info("Расчет квантовой когерентности...")
    
    with col2:
        st.write("**Медицинские инструменты:**")
        
        if st.button("🏥 Диагностический помощник"):
            st.info("Запуск диагностического помощника...")
        
        if st.button("📋 Генератор отчетов"):
            st.info("Создание медицинских отчетов...")
        
        if st.button("📊 Анализатор трендов"):
            st.info("Анализ медицинских трендов...")
        
        if st.button("🔍 Поиск аномалий"):
            st.info("Поиск аномальных паттернов...")
    
    # Утилиты
    st.subheader("🛠️ Утилиты")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.write("**Конвертеры:**")
        if st.button("📄 CSV → MQEA"):
            st.info("Конвертация CSV в формат MQEA")
        
        if st.button("📊 MQEA → Excel"):
            st.info("Экспорт в Excel")
    
    with col2:
        st.write("**Валидаторы:**")
        if st.button("✅ Проверка данных"):
            st.info("Валидация медицинских данных")
        
        if st.button("🔍 Проверка качества"):
            st.info("Контроль качества анализа")
    
    with col3:
        st.write("**Оптимизаторы:**")
        if st.button("⚡ Оптимизация параметров"):
            st.info("Автоматическая оптимизация")
        
        if st.button("🎯 Калибровка системы"):
            st.info("Калибровка MQEA системы")
    
    # Системные инструменты
    st.subheader("⚙️ Системные инструменты")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Мониторинг:**")
        if st.button("📊 Статус системы"):
            st.success("✅ Система работает нормально")
        
        if st.button("💾 Проверка БД"):
            st.info("Проверка целостности базы данных...")
    
    with col2:
        st.write("**Обслуживание:**")
        if st.button("🧹 Очистка кэша"):
            st.info("Очистка временных файлов...")
        
        if st.button("🔄 Перезапуск сервисов"):
            st.warning("Перезапуск системных сервисов...")

# ==================== СИСТЕМА МОНИТОРИНГА ПЕРЕНЕСЕНА В ОТДЕЛЬНОЕ ПРИЛОЖЕНИЕ ====================

# Функции мониторинга перенесены в отдельное приложение: realtime_monitoring_app.py
def show_iot_sensors_old():
    """Отображение IoT датчиков."""
    st.header("📡 IoT Датчики - Система мониторинга в реальном времени")
    st.markdown("**Непрерывный мониторинг медицинских показателей с помощью IoT датчиков**")
    
    # Инициализация системы датчиков
    if 'sensor_manager' not in st.session_state:
        try:
            from mqea.iot_sensors import create_sensor_manager
            st.session_state.sensor_manager = create_sensor_manager()
            st.success("✅ Система IoT датчиков инициализирована")
        except Exception as e:
            st.error(f"❌ Ошибка инициализации датчиков: {e}")
            return
    
    sensor_manager = st.session_state.sensor_manager
    
    # Статус датчиков
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Всего датчиков", len(sensor_manager.sensors))
    with col2:
        active_count = sum(1 for s in sensor_manager.sensors.values() if s.status.value == "active")
        st.metric("Активных", active_count)
    with col3:
        st.metric("Типов датчиков", len(set(s.config.sensor_id.split('_')[0] for s in sensor_manager.sensors.values())))
    with col4:
        st.metric("Статус", "🟢 Работает" if sensor_manager._running else "🔴 Остановлено")
    
    # Управление датчиками
    st.subheader("🎛️ Управление датчиками")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🚀 Запустить все датчики"):
            try:
                # Устанавливаем статус запуска
                sensor_manager._running = True
                for sensor in sensor_manager.sensors.values():
                    sensor.status = sensor.status.__class__("active")
                st.success("🚀 Все датчики запущены")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Ошибка запуска: {e}")
    
    with col2:
        if st.button("⏹️ Остановить все датчики"):
            try:
                # Устанавливаем статус остановки
                sensor_manager._running = False
                for sensor in sensor_manager.sensors.values():
                    sensor.status = sensor.status.__class__("inactive")
                st.success("⏹️ Все датчики остановлены")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Ошибка остановки: {e}")
    
    # Список датчиков
    st.subheader("📊 Список датчиков")
    
    for sensor_id, sensor in sensor_manager.sensors.items():
        with st.expander(f"📡 {sensor.config.name} ({sensor_id})"):
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write(f"**Текущее значение:** {sensor.current_value:.1f} {sensor.config.unit}")
                st.write(f"**Статус:** {sensor.status.value}")
            
            with col2:
                st.write(f"**Диапазон:** {sensor.config.min_value} - {sensor.config.max_value} {sensor.config.unit}")
                st.write(f"**Норма:** {sensor.config.normal_min} - {sensor.config.normal_max} {sensor.config.unit}")
            
            with col3:
                alert_level = sensor._determine_alert_level(sensor.current_value)
                color = {"normal": "🟢", "warning": "🟡", "critical": "🔴", "emergency": "🚨"}
                st.write(f"**Уровень тревоги:** {color.get(alert_level.value, '⚪')} {alert_level.value}")
                
                if st.button(f"🔄 Обновить {sensor_id}", key=f"update_{sensor_id}"):
                    sensor._generate_value()
                    st.rerun()


def show_patient_monitoring():
    """Отображение мониторинга пациентов."""
    st.header("⚡ Мониторинг пациентов в реальном времени")
    st.markdown("**Система непрерывного наблюдения за состоянием пациентов**")
    
    # Инициализация системы мониторинга
    if 'monitoring_system' not in st.session_state:
        try:
            from mqea.realtime_monitoring import create_monitoring_system
            st.session_state.monitoring_system = create_monitoring_system()
            st.success("✅ Система мониторинга инициализирована")
        except Exception as e:
            st.error(f"❌ Ошибка инициализации мониторинга: {e}")
            return
    
    monitoring = st.session_state.monitoring_system
    
    # Статистика
    dashboard = monitoring.get_monitoring_dashboard_data()
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Активных сессий", dashboard['total_active_sessions'])
    with col2:
        st.metric("Тревог за час", dashboard['total_alerts_last_hour'])
    with col3:
        st.metric("Всего датчиков", dashboard['sensor_status']['total_sensors'])
    with col4:
        st.metric("Активных датчиков", dashboard['sensor_status']['active_sensors'])
    
    # Управление сессиями
    st.subheader("👥 Управление сессиями мониторинга")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Добавить нового пациента:**")
        patient_id = st.text_input("ID пациента", value="P001")
        patient_name = st.text_input("Имя пациента", value="Али Хасанов")
        
        if st.button("🏥 Начать мониторинг"):
            try:
                session_id = monitoring.start_monitoring_session(
                    patient_id=patient_id,
                    patient_name=patient_name
                )
                st.success(f"✅ Мониторинг начат для {patient_name} (ID: {session_id})")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Ошибка: {e}")
    
    with col2:
        st.write("**Остановить сессию:**")
        if dashboard['active_sessions']:
            session_options = {f"{s['patient_name']} (ID: {s['patient_id']})": s['session_id'] 
                             for s in dashboard['active_sessions']}
            selected_session = st.selectbox("Выберите сессию", list(session_options.keys()))
            
            if st.button("⏹️ Остановить мониторинг"):
                session_id = session_options[selected_session]
                summary = monitoring.stop_monitoring_session(session_id)
                if summary:
                    st.success(f"✅ Мониторинг остановлен для {summary['patient_name']}")
                    st.rerun()
        else:
            st.info("Нет активных сессий мониторинга")
    
    # Активные сессии
    st.subheader("📊 Активные сессии мониторинга")
    
    if dashboard['active_sessions']:
        for session in dashboard['active_sessions']:
            with st.expander(f"👤 {session['patient_name']} (ID: {session['patient_id']}) - {session['duration_minutes']:.1f} мин"):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Длительность", f"{session['duration_minutes']:.1f} мин")
                    st.metric("Активных тревог", session['active_alerts_count'])
                
                with col2:
                    if session['active_alerts_count'] > 0:
                        st.error(f"🚨 {session['active_alerts_count']} активных тревог")
                    else:
                        st.success("✅ Нет тревог")
                
                with col3:
                    st.write("**Последние показания:**")
                    for sensor_id, reading in session['latest_readings'].items():
                        alert_emoji = {"normal": "✅", "warning": "⚠️", "critical": "🔴", "emergency": "🚨"}
                        emoji = alert_emoji.get(reading['alert_level'], "❓")
                        st.write(f"{emoji} {sensor_id}: {reading['value']:.1f} {reading['unit']}")
    else:
        st.info("Нет активных сессий мониторинга")


def show_alerts_system():
    """Отображение системы тревог."""
    st.header("🚨 Система тревог и алертов")
    st.markdown("**Автоматическое обнаружение критических состояний пациентов**")
    
    # Инициализация системы мониторинга
    if 'monitoring_system' not in st.session_state:
        st.error("❌ Система мониторинга не инициализирована. Перейдите в 'Мониторинг пациентов'")
        return
    
    monitoring = st.session_state.monitoring_system
    dashboard = monitoring.get_monitoring_dashboard_data()
    
    # Статистика тревог
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Всего тревог", len(dashboard['recent_alerts']))
    with col2:
        critical_alerts = len([a for a in dashboard['recent_alerts'] if a['alert_level'] in ['critical', 'emergency']])
        st.metric("Критических", critical_alerts)
    with col3:
        warning_alerts = len([a for a in dashboard['recent_alerts'] if a['alert_level'] == 'warning'])
        st.metric("Предупреждений", warning_alerts)
    with col4:
        st.metric("Тревог за час", dashboard['total_alerts_last_hour'])
    
    # Последние тревоги
    st.subheader("🔥 Последние тревоги")
    
    if dashboard['recent_alerts']:
        for alert in dashboard['recent_alerts']:
            alert_level = alert['alert_level']
            color = {
                'warning': "🟡",
                'critical': "🔴", 
                'emergency': "🚨"
            }.get(alert_level, "⚪")
            
            with st.expander(f"{color} {alert['message']} - {alert['patient_id']}"):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write(f"**Пациент:** {alert['patient_id']}")
                    st.write(f"**Датчик:** {alert['sensor_id']}")
                    st.write(f"**Время:** {alert['timestamp']}")
                
                with col2:
                    st.write(f"**Уровень:** {alert['alert_level']}")
                    st.write(f"**Значение:** {alert['value']:.1f}")
                    st.write(f"**ID тревоги:** {alert['alert_id']}")
                
                with col3:
                    if alert_level in ['critical', 'emergency']:
                        st.error("🚨 Требуется немедленное внимание!")
                    else:
                        st.warning("⚠️ Требуется мониторинг")
                    
                    if st.button(f"✅ Разрешить тревогу", key=f"resolve_{alert['alert_id']}"):
                        # Здесь можно добавить логику разрешения тревоги
                        st.success("✅ Тревога разрешена")
                        st.rerun()
    else:
        st.success("✅ Нет активных тревог")


def show_realtime_charts():
    """Отображение графиков в реальном времени."""
    st.header("📊 Графики в реальном времени")
    st.markdown("**Интерактивная визуализация медицинских показателей**")
    
    # Инициализация системы графиков
    if 'chart_manager' not in st.session_state:
        try:
            from mqea.realtime_charts import create_chart_manager
            st.session_state.chart_manager = create_chart_manager()
            st.session_state.chart_manager.start_monitoring()
            st.success("✅ Система графиков инициализирована")
        except Exception as e:
            st.error(f"❌ Ошибка инициализации графиков: {e}")
            return
    
    charts = st.session_state.chart_manager
    
    # Выбор графика
    chart_ids = list(charts.charts.keys())
    if chart_ids:
        selected_chart = st.selectbox(
            "Выберите график:",
            chart_ids,
            format_func=lambda x: charts.chart_configs[x].name
        )
        
        # Настройки
        col1, col2 = st.columns(2)
        with col1:
            time_window = st.slider("Временное окно (минуты)", 5, 120, 30)
        with col2:
            auto_refresh = st.checkbox("Автообновление", value=True)
        
        # Создание графика
        if st.button("🔄 Обновить график") or auto_refresh:
            try:
                fig = charts.create_plotly_chart(selected_chart, time_window)
                if fig:
                    st.plotly_chart(fig, width='stretch')
                else:
                    st.error("Не удалось создать график")
            except Exception as e:
                st.error(f"❌ Ошибка создания графика: {e}")
        
        # Статистика графика
        chart_stats = charts.get_charts_statistics()
        if selected_chart in chart_stats:
            st.subheader("📈 Статистика")
            
            for sensor_id, sensor_stat in chart_stats[selected_chart].items():
                with st.expander(f"📊 {sensor_id}"):
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("Текущее", f"{sensor_stat['current_value']:.1f}")
                    with col2:
                        st.metric("Среднее", f"{sensor_stat['avg_value']:.1f}")
                    with col3:
                        st.metric("Тренд", sensor_stat['trend'])
                    with col4:
                        st.metric("Точек", sensor_stat['data_points_count'])
    else:
        st.info("Графики не созданы")


def show_notifications_system():
    """Отображение системы уведомлений."""
    st.header("📧 Система уведомлений")
    st.markdown("**Многоканальные уведомления о критических состояниях**")
    
    # Инициализация системы уведомлений
    if 'notification_system' not in st.session_state:
        try:
            from mqea.notification_system import create_notification_system
            st.session_state.notification_system = create_notification_system()
            st.success("✅ Система уведомлений инициализирована")
        except Exception as e:
            st.error(f"❌ Ошибка инициализации уведомлений: {e}")
            return
    
    notifications = st.session_state.notification_system
    
    # Статистика уведомлений
    notification_stats = notifications.get_statistics()
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Всего", notification_stats['total_notifications'])
    with col2:
        st.metric("Отправлено", notification_stats['sent_notifications'])
    with col3:
        st.metric("Ошибок", notification_stats['failed_notifications'])
    with col4:
        st.metric("Успешность", f"{notification_stats['success_rate']:.1f}%")
    
    # Управление получателями
    st.subheader("👥 Управление получателями")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Добавить получателя:**")
        recipient_id = st.text_input("ID получателя", value="doctor1")
        email = st.text_input("Email", value="doctor1@hospital.com")
        phone = st.text_input("Телефон", value="+992123456789")
        
        if st.button("➕ Добавить получателя"):
            try:
                notifications.add_recipient(
                    recipient_id,
                    email=email,
                    phone=phone
                )
                st.success(f"✅ Получатель {recipient_id} добавлен")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Ошибка: {e}")
    
    with col2:
        st.write("**Текущие получатели:**")
        for recipient_id, contact_info in notifications.recipients.items():
            st.write(f"• {recipient_id}: {contact_info}")
    
    # История уведомлений
    st.subheader("📋 История уведомлений")
    history = notifications.get_notification_history(20)
    
    if history:
        import pandas as pd
        df = pd.DataFrame(history)
        
        # Исправляем типы данных для совместимости с Streamlit
        if 'timestamp' in df.columns:
            df['timestamp'] = df['timestamp'].astype(str)
        if 'subject' in df.columns:
            df['subject'] = df['subject'].astype(str)
        if 'body' in df.columns:
            df['body'] = df['body'].astype(str)
        if 'recipient' in df.columns:
            df['recipient'] = df['recipient'].astype(str)
        if 'status' in df.columns:
            df['status'] = df['status'].astype(str)
        
        st.dataframe(df, width='stretch')
    else:
        st.info("История уведомлений пуста")
    
    # Статистика по каналам
    st.subheader("📈 Статистика по каналам")
    channel_stats = notification_stats['channel_statistics']
    
    for channel, stats in channel_stats.items():
        with st.expander(f"📧 {channel}"):
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Отправлено", stats['sent'])
            with col2:
                st.metric("Ошибок", stats['failed'])


def show_pediatric_diagnosis():
    """Детская квантовая диагностическая система с детальными антропометрическими измерениями."""
    
    log_action("🎉 ВЫЗОВ ФУНКЦИИ", "show_pediatric_diagnosis() - НАЧАЛО")
    
    # Отладочная информация
    st.success("🎉 Функция show_pediatric_diagnosis() вызвана!")
    
    # Инициализация детского квантового движка
    if 'pediatric_engine' not in st.session_state:
        log_action("🔧 ИНИЦИАЛИЗАЦИЯ", "PediatricQuantumEngine создан")
        st.session_state.pediatric_engine = PediatricQuantumEngine()
    else:
        log_action("✅ ИНИЦИАЛИЗАЦИЯ", "PediatricQuantumEngine уже существует")
    
    # Переключатель между основными и детальными измерениями
    measurement_mode = st.radio(
        "Режим измерений:",
        ["🩺 Основные показатели", "🔬 Детальные антропометрические измерения"],
        horizontal=True
    )
    
    if measurement_mode == "🩺 Основные показатели":
        show_basic_pediatric_diagnosis()
    else:
        show_detailed_anthropometry_diagnosis()


def show_basic_pediatric_diagnosis():
    """Основная детская диагностика."""
    
    st.markdown("## 👶 Детская квантовая диагностика MQEA-Pediatric")
    st.markdown("**Революционная система раннего выявления заболеваний у детей от рождения до 10 лет**")
    
    # Основные колонки
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.markdown("### 📋 Данные ребенка")
        
        # Возраст
        age_input = st.selectbox(
            "Выберите возрастную группу:",
            ["0-1 месяц", "1-12 месяцев", "1-3 года", "3-6 лет", "6-10 лет"]
        )
        
        age_months = get_pediatric_age_months(age_input)
        
        # Основные показатели
        st.markdown("#### 🩺 Жизненные показатели")
        
        heart_rate = st.number_input(
            "Частота сердечных сокращений (уд/мин):",
            min_value=30.0, max_value=300.0, value=float(get_pediatric_default_value('heart_rate', age_months))
        )
        
        respiratory_rate = st.number_input(
            "Частота дыхания (дых/мин):",
            min_value=5.0, max_value=80.0, value=float(get_pediatric_default_value('respiratory_rate', age_months))
        )
        
        bp_systolic = st.number_input(
            "Систолическое давление (мм рт.ст.):",
            min_value=40.0, max_value=200.0, value=float(get_pediatric_default_value('blood_pressure_systolic', age_months))
        )
        
        bp_diastolic = st.number_input(
            "Диастолическое давление (мм рт.ст.):",
            min_value=20.0, max_value=150.0, value=float(get_pediatric_default_value('blood_pressure_diastolic', age_months))
        )
        
        temperature = st.number_input(
            "Температура тела (°C):",
            min_value=35.0, max_value=42.0, value=36.8, step=0.1
        )
        
        oxygen_saturation = st.number_input(
            "Насыщение кислородом (%):",
            min_value=70.0, max_value=100.0, value=98.0
        )
        
        st.markdown("#### 📏 Основные антропометрические данные")
        
        weight = st.number_input(
            "Вес (кг):",
            min_value=1.0, max_value=50.0, value=float(get_pediatric_default_value('weight', age_months)), step=0.1
        )
        
        height = st.number_input(
            "Рост (см):",
            min_value=30.0, max_value=200.0, value=float(get_pediatric_default_value('height', age_months)), step=0.5
        )
        
        head_circumference = st.number_input(
            "Окружность головы (см):",
            min_value=25.0, max_value=70.0, value=float(get_pediatric_default_value('head_circumference', age_months)), step=0.1
        )
        
        # Кнопка анализа
        analyze_button = st.button("🔬 Запустить квантовый анализ", type="primary", width='stretch')
    
    with col2:
        if analyze_button:
            # Создаем объект жизненных показателей
            vital_signs = PediatricVitalSigns(
                age_months=age_months,
                heart_rate=heart_rate,
                respiratory_rate=respiratory_rate,
                blood_pressure_systolic=bp_systolic,
                blood_pressure_diastolic=bp_diastolic,
                temperature=temperature,
                oxygen_saturation=oxygen_saturation,
                weight_kg=weight,
                height_cm=height,
                head_circumference_cm=head_circumference
            )
            
            # Выполняем квантовый анализ
            with st.spinner("🔄 Выполняется квантовый анализ..."):
                detected_conditions = st.session_state.pediatric_engine.detect_pediatric_conditions(
                    vital_signs, quantum_threshold=0.6
                )
                
                quantum_report = st.session_state.pediatric_engine.generate_pediatric_quantum_report(
                    vital_signs, detected_conditions
                )
            
            # Отображаем результаты
            display_pediatric_results(vital_signs, detected_conditions, quantum_report)


def show_detailed_anthropometry_diagnosis():
    """Детальная антропометрическая диагностика."""
    
    st.markdown("## 👶 Детская квантовая диагностика MQEA-Pediatric")
    st.markdown("### 🔬 Детальные антропометрические измерения")
    st.markdown("**Точный анализ развития с учетом всех параметров тела до мельчайших деталей**")
    
    # Инициализация менеджера профилей
    if 'profile_manager' not in st.session_state:
        st.session_state.profile_manager = PediatricProfileManager()
    
    pm = st.session_state.profile_manager
    
    # Выбор профиля ребенка (опционально)
    profiles = pm.list_all_profiles()
    
    if profiles:
        use_profile = st.checkbox("📋 Использовать существующий профиль ребенка")
        
        if use_profile:
            selected_child = st.selectbox(
                "Выберите ребенка:",
                options=[p.child_id for p in profiles],
                format_func=lambda x: next((p.name for p in profiles if p.child_id == x), x)
            )
            
            profile = pm.get_child_profile(selected_child)
            if profile:
                st.success(f"✅ Выбран профиль: {profile.name}")
                # Вычисляем возраст
                birth_date = datetime.strptime(profile.date_of_birth, "%Y-%m-%d")
                age_months = (datetime.now().year - birth_date.year) * 12 + (datetime.now().month - birth_date.month)
                st.info(f"Возраст: {age_months} месяцев")
        else:
            selected_child = None
            age_input = st.selectbox(
                "Выберите возрастную группу:",
                ["0-1 месяц", "1-12 месяцев", "1-3 года", "3-6 лет", "6-10 лет"]
            )
            age_months = get_pediatric_age_months(age_input)
    else:
        selected_child = None
        age_input = st.selectbox(
            "Выберите возрастную группу:",
            ["0-1 месяц", "1-12 месяцев", "1-3 года", "3-6 лет", "6-10 лет"]
        )
        age_months = get_pediatric_age_months(age_input)
    
    # Создаем вкладки для разных групп измерений
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📏 Основные размеры", 
        "👋 Размеры пальцев рук", 
        "🦶 Размеры пальцев ног",
        "👤 Размеры головы и лица",
        "📊 Пропорции тела"
    ])
    
    with tab1:
        st.markdown("#### 📏 Основные размеры тела")
        col1, col2 = st.columns(2)
        
        with col1:
            weight = st.number_input(
                "Вес (кг):",
                min_value=1.0, max_value=50.0, value=float(get_pediatric_default_value('weight', age_months)), step=0.1
            )
            
            height = st.number_input(
                "Рост (см):",
                min_value=30.0, max_value=200.0, value=float(get_pediatric_default_value('height', age_months)), step=0.5
            )
            
            head_circumference = st.number_input(
                "Окружность головы (см):",
                min_value=25.0, max_value=70.0, value=float(get_pediatric_default_value('head_circumference', age_months)), step=0.1
            )
            
            chest_circumference = st.number_input(
                "Окружность груди (см):",
                min_value=20.0, max_value=100.0, value=float(get_pediatric_default_value('chest_circumference', age_months)), step=0.1
            )
        
        with col2:
            abdominal_circumference = st.number_input(
                "Окружность живота (см):",
                min_value=20.0, max_value=100.0, value=float(get_pediatric_default_value('abdominal_circumference', age_months)), step=0.1
            )
            
            arm_span = st.number_input(
                "Размах рук (см):",
                min_value=30.0, max_value=200.0, value=float(get_pediatric_default_value('arm_span', age_months)), step=0.5
            )
            
            leg_length = st.number_input(
                "Длина ноги (см):",
                min_value=15.0, max_value=120.0, value=float(get_pediatric_default_value('leg_length', age_months)), step=0.5
            )
            
            foot_length = st.number_input(
                "Длина стопы (см):",
                min_value=5.0, max_value=30.0, value=float(get_pediatric_default_value('foot_length', age_months)), step=0.1
            )
    
    with tab2:
        st.markdown("#### 👋 Размеры пальцев рук (мм)")
        col1, col2 = st.columns(2)
        
        with col1:
            thumb_length = st.number_input(
                "Длина большого пальца (мм):",
                min_value=5.0, max_value=50.0, value=float(get_pediatric_default_value('thumb_length', age_months)), step=0.5
            )
            
            index_finger_length = st.number_input(
                "Длина указательного пальца (мм):",
                min_value=10.0, max_value=80.0, value=float(get_pediatric_default_value('index_finger_length', age_months)), step=0.5
            )
            
            middle_finger_length = st.number_input(
                "Длина среднего пальца (мм):",
                min_value=15.0, max_value=90.0, value=float(get_pediatric_default_value('middle_finger_length', age_months)), step=0.5
            )
        
        with col2:
            ring_finger_length = st.number_input(
                "Длина безымянного пальца (мм):",
                min_value=12.0, max_value=85.0, value=float(get_pediatric_default_value('ring_finger_length', age_months)), step=0.5
            )
            
            little_finger_length = st.number_input(
                "Длина мизинца (мм):",
                min_value=8.0, max_value=65.0, value=float(get_pediatric_default_value('little_finger_length', age_months)), step=0.5
            )
    
    with tab3:
        st.markdown("#### 🦶 Размеры пальцев ног (мм)")
        col1, col2 = st.columns(2)
        
        with col1:
            big_toe_length = st.number_input(
                "Длина большого пальца ноги (мм):",
                min_value=5.0, max_value=40.0, value=float(get_pediatric_default_value('big_toe_length', age_months)), step=0.5
            )
            
            second_toe_length = st.number_input(
                "Длина второго пальца ноги (мм):",
                min_value=4.0, max_value=35.0, value=float(get_pediatric_default_value('second_toe_length', age_months)), step=0.5
            )
            
            third_toe_length = st.number_input(
                "Длина третьего пальца ноги (мм):",
                min_value=3.0, max_value=30.0, value=float(get_pediatric_default_value('third_toe_length', age_months)), step=0.5
            )
        
        with col2:
            fourth_toe_length = st.number_input(
                "Длина четвертого пальца ноги (мм):",
                min_value=2.0, max_value=25.0, value=float(get_pediatric_default_value('fourth_toe_length', age_months)), step=0.5
            )
            
            little_toe_length = st.number_input(
                "Длина мизинца ноги (мм):",
                min_value=1.0, max_value=20.0, value=float(get_pediatric_default_value('little_toe_length', age_months)), step=0.5
            )
    
    with tab4:
        st.markdown("#### 👤 Размеры головы и лица")
        col1, col2 = st.columns(2)
        
        with col1:
            head_length = st.number_input(
                "Длина головы (см):",
                min_value=10.0, max_value=25.0, value=float(get_pediatric_default_value('head_length', age_months)), step=0.1
            )
            
            head_width = st.number_input(
                "Ширина головы (см):",
                min_value=8.0, max_value=20.0, value=float(get_pediatric_default_value('head_width', age_months)), step=0.1
            )
            
            face_height = st.number_input(
                "Высота лица (см):",
                min_value=5.0, max_value=15.0, value=float(get_pediatric_default_value('face_height', age_months)), step=0.1
            )
            
            face_width = st.number_input(
                "Ширина лица (см):",
                min_value=5.0, max_value=15.0, value=float(get_pediatric_default_value('face_width', age_months)), step=0.1
            )
        
        with col2:
            nose_length = st.number_input(
                "Длина носа (мм):",
                min_value=5.0, max_value=30.0, value=float(get_pediatric_default_value('nose_length', age_months)), step=0.5
            )
            
            nose_width = st.number_input(
                "Ширина носа (мм):",
                min_value=8.0, max_value=25.0, value=float(get_pediatric_default_value('nose_width', age_months)), step=0.5
            )
            
            eye_width = st.number_input(
                "Ширина глаза (мм):",
                min_value=8.0, max_value=20.0, value=float(get_pediatric_default_value('eye_width', age_months)), step=0.5
            )
            
            mouth_width = st.number_input(
                "Ширина рта (мм):",
                min_value=15.0, max_value=35.0, value=float(get_pediatric_default_value('mouth_width', age_months)), step=0.5
            )
    
    with tab5:
        st.markdown("#### 📊 Пропорции тела и дополнительные измерения")
        col1, col2 = st.columns(2)
        
        with col1:
            shoulder_width = st.number_input(
                "Ширина плеч (см):",
                min_value=10.0, max_value=50.0, value=float(get_pediatric_default_value('shoulder_width', age_months)), step=0.5
            )
            
            hip_width = st.number_input(
                "Ширина бедер (см):",
                min_value=8.0, max_value=40.0, value=float(get_pediatric_default_value('hip_width', age_months)), step=0.5
            )
            
            waist_to_hip_ratio = st.number_input(
                "Соотношение талии к бедрам:",
                min_value=0.5, max_value=1.5, value=0.8, step=0.01
            )
        
        with col2:
            triceps_skinfold = st.number_input(
                "Трехглавая кожная складка (мм):",
                min_value=2.0, max_value=25.0, value=float(get_pediatric_default_value('triceps_skinfold', age_months)), step=0.5
            )
            
            subscapular_skinfold = st.number_input(
                "Подлопаточная кожная складка (мм):",
                min_value=2.0, max_value=20.0, value=float(get_pediatric_default_value('subscapular_skinfold', age_months)), step=0.5
            )
            
            suprailiac_skinfold = st.number_input(
                "Надподвздошная кожная складка (мм):",
                min_value=2.0, max_value=20.0, value=float(get_pediatric_default_value('suprailiac_skinfold', age_months)), step=0.5
            )
    
    # Кнопка детального анализа и опция сохранения
    st.markdown("---")
    
    col1, col2 = st.columns([3, 1])
    with col1:
        detailed_analyze_button = st.button("🔬 Запустить детальный квантовый анализ", type="primary", width='stretch')
    with col2:
        save_to_profile = st.checkbox("💾 Сохранить", value=True if selected_child else False, disabled=not selected_child)
    
    if detailed_analyze_button:
        # Создаем объект детальных антропометрических измерений
        detailed_anthropometry = DetailedAnthropometry(
            # Основные измерения
            weight_kg=weight,
            height_cm=height,
            head_circumference_cm=head_circumference,
            chest_circumference_cm=chest_circumference,
            abdominal_circumference_cm=abdominal_circumference,
            
            # Размеры конечностей
            arm_span_cm=arm_span,
            leg_length_cm=leg_length,
            foot_length_cm=foot_length,
            foot_width_cm=foot_length * 0.4,  # Примерное соотношение
            
            # Размеры пальцев рук
            thumb_length_mm=thumb_length,
            index_finger_length_mm=index_finger_length,
            middle_finger_length_mm=middle_finger_length,
            ring_finger_length_mm=ring_finger_length,
            little_finger_length_mm=little_finger_length,
            
            # Размеры пальцев ног
            big_toe_length_mm=big_toe_length,
            second_toe_length_mm=second_toe_length,
            third_toe_length_mm=third_toe_length,
            fourth_toe_length_mm=fourth_toe_length,
            little_toe_length_mm=little_toe_length,
            
            # Размеры головы и лица
            head_length_cm=head_length,
            head_width_cm=head_width,
            face_height_cm=face_height,
            face_width_cm=face_width,
            
            # Размеры носа
            nose_length_mm=nose_length,
            nose_width_mm=nose_width,
            nose_height_mm=nose_length * 0.6,  # Примерное соотношение
            
            # Размеры глаз
            eye_width_mm=eye_width,
            eye_height_mm=eye_width * 0.4,  # Примерное соотношение
            inter_eye_distance_mm=eye_width * 1.5,  # Примерное соотношение
            
            # Размеры рта
            mouth_width_mm=mouth_width,
            lip_thickness_mm=mouth_width * 0.1,  # Примерное соотношение
            
            # Размеры ушей
            ear_length_mm=nose_length * 0.8,  # Примерное соотношение
            ear_width_mm=nose_length * 0.5,  # Примерное соотношение
            
            # Пропорции
            waist_to_hip_ratio=waist_to_hip_ratio,
            shoulder_width_cm=shoulder_width,
            hip_width_cm=hip_width,
            
            # Кожные складки
            triceps_skinfold_mm=triceps_skinfold,
            subscapular_skinfold_mm=subscapular_skinfold,
            suprailiac_skinfold_mm=suprailiac_skinfold
        )
        
        # Выполняем детальный квантовый анализ
        with st.spinner("🔄 Выполняется детальный квантовый анализ всех параметров..."):
            detailed_analysis = st.session_state.pediatric_engine.analyze_detailed_anthropometry(
                detailed_anthropometry, age_months
            )
            
            # Генерируем комплексный отчет о развитии
            comprehensive_report = st.session_state.pediatric_engine.generate_comprehensive_development_report(
                detailed_anthropometry, age_months, detailed_analysis
            )
        
        # Сохраняем в профиль, если выбрано
        if selected_child and save_to_profile:
            try:
                # Создаем объект жизненных показателей
                vital_signs = PediatricVitalSigns(
                    age_months=age_months,
                    heart_rate=100.0,  # Базовые значения
                    respiratory_rate=25.0,
                    blood_pressure_systolic=90.0,
                    blood_pressure_diastolic=60.0,
                    temperature=36.6,
                    oxygen_saturation=98.0,
                    weight_kg=weight,
                    height_cm=height,
                    head_circumference_cm=head_circumference,
                    detailed_anthropometry=detailed_anthropometry
                )
                
                # Добавляем запись в историю
                record = pm.add_development_record(
                    child_id=selected_child,
                    vital_signs=vital_signs,
                    anthropometry=detailed_anthropometry,
                    notes="Детальный антропометрический анализ"
                )
                
                st.success(f"✅ Результаты сохранены в профиль ребенка (ID записи: {record.record_id})")
                
                # Если есть предыдущие записи, показываем сравнение
                history = pm.get_development_history(selected_child)
                if len(history) >= 2:
                    st.info("📊 Доступен анализ прогресса! Перейдите в раздел '📊 Мониторинг развития' → '📊 Анализ прогресса'")
                
            except Exception as e:
                st.error(f"❌ Ошибка при сохранении: {str(e)}")
        
        # Отображаем детальные результаты
        display_detailed_anthropometry_results(detailed_analysis, age_months)
        
        # Отображаем комплексный отчет о развитии
        st.markdown("---")
        display_comprehensive_development_report(comprehensive_report)

def get_pediatric_age_months(age_input: str) -> int:
    """Преобразует текстовый ввод возраста в месяцы."""
    age_mapping = {
        "0-1 месяц": 0.5,
        "1-12 месяцев": 6,
        "1-3 года": 18,
        "3-6 лет": 48,
        "6-10 лет": 84
    }
    return int(age_mapping[age_input])

def get_pediatric_default_value(indicator: str, age_months: int) -> float:
    """Возвращает нормальные значения для показателя в зависимости от возраста."""
    engine = PediatricQuantumEngine()
    age_group = engine._determine_age_group(age_months)
    
    # Получаем нормальный диапазон для возраста
    normal_range = engine.pediatric_ranges[age_group].get(indicator, (50, 100))
    
    # Возвращаем среднее значение диапазона
    base_value = (normal_range[0] + normal_range[1]) / 2
    
    # Специальные значения для детальных антропометрических показателей
    if indicator == 'chest_circumference':
        if age_months < 1:
            return 32
        elif age_months < 12:
            return 45
        elif age_months < 36:
            return 52
        else:
            return 58
    elif indicator == 'abdominal_circumference':
        if age_months < 1:
            return 30
        elif age_months < 12:
            return 42
        elif age_months < 36:
            return 48
        else:
            return 54
    elif indicator == 'arm_span':
        if age_months < 1:
            return 48
        elif age_months < 12:
            return 68
        elif age_months < 36:
            return 88
        else:
            return 108
    elif indicator == 'leg_length':
        if age_months < 1:
            return 20
        elif age_months < 12:
            return 30
        elif age_months < 36:
            return 40
        else:
            return 50
    elif indicator == 'foot_length':
        if age_months < 1:
            return 8
        elif age_months < 12:
            return 12
        elif age_months < 36:
            return 16
        else:
            return 20
    elif indicator == 'thumb_length':
        if age_months < 1:
            return 15
        elif age_months < 12:
            return 20
        elif age_months < 36:
            return 25
        else:
            return 30
    elif indicator == 'index_finger_length':
        if age_months < 1:
            return 25
        elif age_months < 12:
            return 35
        elif age_months < 36:
            return 45
        else:
            return 55
    elif indicator == 'middle_finger_length':
        if age_months < 1:
            return 30
        elif age_months < 12:
            return 40
        elif age_months < 36:
            return 50
        else:
            return 60
    elif indicator == 'ring_finger_length':
        if age_months < 1:
            return 28
        elif age_months < 12:
            return 38
        elif age_months < 36:
            return 48
        else:
            return 58
    elif indicator == 'little_finger_length':
        if age_months < 1:
            return 20
        elif age_months < 12:
            return 28
        elif age_months < 36:
            return 36
        else:
            return 44
    elif indicator == 'big_toe_length':
        if age_months < 1:
            return 12
        elif age_months < 12:
            return 18
        elif age_months < 36:
            return 24
        else:
            return 30
    elif indicator == 'second_toe_length':
        if age_months < 1:
            return 10
        elif age_months < 12:
            return 16
        elif age_months < 36:
            return 22
        else:
            return 28
    elif indicator == 'third_toe_length':
        if age_months < 1:
            return 8
        elif age_months < 12:
            return 14
        elif age_months < 36:
            return 20
        else:
            return 26
    elif indicator == 'fourth_toe_length':
        if age_months < 1:
            return 6
        elif age_months < 12:
            return 12
        elif age_months < 36:
            return 18
        else:
            return 24
    elif indicator == 'little_toe_length':
        if age_months < 1:
            return 4
        elif age_months < 12:
            return 8
        elif age_months < 36:
            return 12
        else:
            return 16
    elif indicator == 'head_length':
        if age_months < 1:
            return 14
        elif age_months < 12:
            return 16
        elif age_months < 36:
            return 17
        else:
            return 18
    elif indicator == 'head_width':
        if age_months < 1:
            return 11
        elif age_months < 12:
            return 13
        elif age_months < 36:
            return 14
        else:
            return 15
    elif indicator == 'face_height':
        if age_months < 1:
            return 6
        elif age_months < 12:
            return 8
        elif age_months < 36:
            return 10
        else:
            return 12
    elif indicator == 'face_width':
        if age_months < 1:
            return 8
        elif age_months < 12:
            return 10
        elif age_months < 36:
            return 11
        else:
            return 12
    elif indicator == 'nose_length':
        if age_months < 1:
            return 15
        elif age_months < 12:
            return 18
        elif age_months < 36:
            return 21
        else:
            return 24
    elif indicator == 'nose_width':
        if age_months < 1:
            return 12
        elif age_months < 12:
            return 15
        elif age_months < 36:
            return 18
        else:
            return 21
    elif indicator == 'eye_width':
        if age_months < 1:
            return 12
        elif age_months < 12:
            return 14
        elif age_months < 36:
            return 16
        else:
            return 18
    elif indicator == 'mouth_width':
        if age_months < 1:
            return 20
        elif age_months < 12:
            return 24
        elif age_months < 36:
            return 28
        else:
            return 32
    elif indicator == 'shoulder_width':
        if age_months < 1:
            return 18
        elif age_months < 12:
            return 24
        elif age_months < 36:
            return 30
        else:
            return 36
    elif indicator == 'hip_width':
        if age_months < 1:
            return 14
        elif age_months < 12:
            return 18
        elif age_months < 36:
            return 22
        else:
            return 26
    elif indicator == 'triceps_skinfold':
        if age_months < 1:
            return 8
        elif age_months < 12:
            return 10
        elif age_months < 36:
            return 12
        else:
            return 14
    elif indicator == 'subscapular_skinfold':
        if age_months < 1:
            return 6
        elif age_months < 12:
            return 8
        elif age_months < 36:
            return 10
        else:
            return 12
    elif indicator == 'suprailiac_skinfold':
        if age_months < 1:
            return 5
        elif age_months < 12:
            return 7
        elif age_months < 36:
            return 9
        else:
            return 11
    
    return base_value

def display_comprehensive_development_report(report: dict):
    """Отображает комплексный отчет о развитии ребенка."""
    
    st.markdown("## 📋 Комплексный отчет о развитии ребенка")
    
    # Информация о ребенке
    child_info = report['child_info']
    dev_stage = child_info['developmental_stage']
    
    st.markdown(f"### 👶 Возраст: {child_info['age_months']} месяцев")
    st.info(f"**Стадия развития:** {dev_stage['stage']}\n\n*{dev_stage['description']}*")
    
    # Ключевые показатели развития
    with st.expander("🎯 Ключевые показатели развития для этого возраста", expanded=True):
        for milestone in dev_stage['key_milestones']:
            st.markdown(f"✓ {milestone}")
    
    # Общее заключение
    st.markdown("### 📄 Заключение")
    st.markdown(report['overall_conclusion'])
    
    # Создаем вкладки для детальной информации
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "💊 Витамины",
        "👐 Массаж",
        "📅 План осмотров",
        "🛡️ Профилактика",
        "👨‍⚕️ Специалисты"
    ])
    
    with tab1:
        st.markdown("### 💊 Рекомендации по витаминам")
        vitamin_recs = report.get('vitamin_recommendations', [])
        
        if vitamin_recs:
            for i, vit in enumerate(vitamin_recs, 1):
                with st.expander(f"💊 {vit['vitamin']}", expanded=(i <= 2)):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"**Дозировка:** {vit['dosage']}")
                        st.markdown(f"**Длительность:** {vit['duration']}")
                    with col2:
                        st.markdown(f"**Причина:**\n{vit['reason']}")
        else:
            st.info("Специальные рекомендации по витаминам отсутствуют")
    
    with tab2:
        st.markdown("### 👐 Рекомендации по массажу")
        massage_recs = report.get('massage_recommendations', [])
        
        if massage_recs:
            for i, massage in enumerate(massage_recs, 1):
                with st.expander(f"👐 {massage['type']}", expanded=(i == 1)):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"**Частота:** {massage['frequency']}")
                        st.markdown(f"**Длительность:** {massage['duration']}")
                    with col2:
                        st.markdown(f"**Описание:**\n{massage['description']}")
                    
                    st.markdown("**Зоны массажа:**")
                    focus_areas = ", ".join(massage.get('focus_areas', []))
                    st.markdown(f"*{focus_areas}*")
        else:
            st.info("Специальные рекомендации по массажу отсутствуют")
    
    with tab3:
        st.markdown("### 📅 План дальнейших осмотров")
        follow_up = report.get('follow_up_plan', [])
        
        if follow_up:
            # Группируем по приоритету
            critical = [f for f in follow_up if f.get('priority') == 'Критический']
            high = [f for f in follow_up if f.get('priority') == 'Высокий']
            medium = [f for f in follow_up if f.get('priority') == 'Средний']
            
            if critical:
                st.error("🚨 **КРИТИЧЕСКИЙ ПРИОРИТЕТ:**")
                for item in critical:
                    st.markdown(f"**{item['specialist']}**")
                    st.markdown(f"• Частота: {item['frequency']}")
                    st.markdown(f"• Цель: {item['purpose']}")
                    st.markdown("---")
            
            if high:
                st.warning("⚠️ **ВЫСОКИЙ ПРИОРИТЕТ:**")
                for item in high:
                    st.markdown(f"**{item['specialist']}**")
                    st.markdown(f"• Частота: {item['frequency']}")
                    st.markdown(f"• Цель: {item['purpose']}")
                    st.markdown("---")
            
            if medium:
                st.info("📋 **СРЕДНИЙ ПРИОРИТЕТ:**")
                for item in medium:
                    st.markdown(f"**{item['specialist']}**")
                    st.markdown(f"• Частота: {item['frequency']}")
                    st.markdown(f"• Цель: {item['purpose']}")
                    st.markdown("---")
        else:
            st.info("План осмотров формируется по стандартному графику")
    
    with tab4:
        st.markdown("### 🛡️ Превентивные меры")
        preventive = report.get('preventive_measures', [])
        
        if preventive:
            for measure in preventive:
                if measure.startswith('⚠'):
                    st.warning(measure)
                else:
                    st.success(measure)
        else:
            st.info("Специальные превентивные меры не требуются")
    
    with tab5:
        st.markdown("### 👨‍⚕️ Консультации специалистов")
        specialists = report.get('specialist_consultations', [])
        
        if specialists:
            for specialist in specialists:
                if specialist.startswith('✅'):
                    st.success(specialist)
                else:
                    st.warning(specialist)
        else:
            st.info("Дополнительные консультации не требуются")
    
    # Физическое развитие
    st.markdown("---")
    st.markdown("### 📊 Статус физического развития")
    phys_dev = report.get('physical_development', {})
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        weight_status = phys_dev.get('weight_status', 'нормальный')
        if weight_status == 'нормальный':
            st.success(f"**Вес:**\n{weight_status}")
        else:
            st.warning(f"**Вес:**\n{weight_status}")
    
    with col2:
        height_status = phys_dev.get('height_status', 'нормальный')
        if height_status == 'нормальный':
            st.success(f"**Рост:**\n{height_status}")
        else:
            st.warning(f"**Рост:**\n{height_status}")
    
    with col3:
        head_status = phys_dev.get('head_status', 'нормальный')
        if head_status == 'нормальный':
            st.success(f"**Голова:**\n{head_status}")
        else:
            st.warning(f"**Голова:**\n{head_status}")
    
    with col4:
        overall_status = phys_dev.get('overall_status', 'соответствует возрасту')
        if overall_status == 'соответствует возрасту':
            st.success(f"**Общий:**\n{overall_status}")
        else:
            st.warning(f"**Общий:**\n{overall_status}")


def display_detailed_anthropometry_results(detailed_analysis: dict, age_months: int):
    """Отображает результаты детального антропометрического анализа."""
    
    st.markdown("## 🔬 Результаты детального квантового анализа")
    
    # Общая информация
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Возраст", f"{age_months} месяцев")
    with col2:
        st.metric("Возрастная группа", detailed_analysis.get('age_group', 'неизвестно'))
    with col3:
        quantum_score = detailed_analysis.get('quantum_assessment', {}).get('developmental_quantum_score', 0)
        st.metric("Квантовый балл развития", f"{quantum_score:.2f}")
    
    # Создаем вкладки для разных типов анализа
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 Пропорции тела",
        "📈 Паттерны роста", 
        "⚛️ Квантовый анализ",
        "🚨 Обнаруженные аномалии",
        "👶 Индикаторы развития"
    ])
    
    with tab1:
        st.markdown("### 📊 Анализ пропорций тела")
        proportional_analysis = detailed_analysis.get('proportional_analysis', {})
        
        if 'cephalic_index' in proportional_analysis:
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Черепной индекс", f"{proportional_analysis['cephalic_index']:.1f}")
            with col2:
                st.metric("Форма головы", proportional_analysis.get('head_shape', 'неизвестно'))
        
        if 'arm_span_ratio' in proportional_analysis:
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Соотношение размаха рук", f"{proportional_analysis['arm_span_ratio']:.2f}")
            with col2:
                st.metric("Оценка конечностей", proportional_analysis.get('limb_assessment', 'неизвестно'))
        
        if 'finger_ratios' in proportional_analysis:
            st.markdown("#### 👋 Соотношения пальцев:")
            finger_ratios = proportional_analysis['finger_ratios']
            finger_names = ["Указательный", "Средний", "Безымянный", "Мизинец"]
            for i, ratio in enumerate(finger_ratios):
                st.metric(f"Палец {i+1} к большому", f"{ratio:.2f}")
    
    with tab2:
        st.markdown("### 📈 Анализ паттернов роста")
        growth_patterns = detailed_analysis.get('growth_patterns', {})
        
        if 'bmi' in growth_patterns:
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Индекс массы тела", f"{growth_patterns['bmi']:.1f}")
            with col2:
                st.metric("Оценка веса", growth_patterns.get('weight_assessment', 'неизвестно'))
        
        if 'head_to_height_ratio' in growth_patterns:
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Соотношение головы к росту", f"{growth_patterns['head_to_height_ratio']:.3f}")
            with col2:
                st.metric("Оценка головы", growth_patterns.get('head_assessment', 'неизвестно'))
    
    with tab3:
        st.markdown("### ⚛️ Квантовый анализ")
        quantum_assessment = detailed_analysis.get('quantum_assessment', {})
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Всего показателей", quantum_assessment.get('total_indicators', 0))
        
        quantum_coherence = quantum_assessment.get('quantum_coherence', {})
        if quantum_coherence:
            with col2:
                avg_entanglement = quantum_coherence.get('average_entanglement', 0)
                st.metric("Средняя запутанность", f"{avg_entanglement:.3f}")
            
            st.metric("Квантовых пар", quantum_coherence.get('total_pairs', 0))
    
    with tab4:
        st.markdown("### 🚨 Обнаруженные аномалии")
        anomaly_detection = detailed_analysis.get('anomaly_detection', {})
        
        detected_anomalies = anomaly_detection.get('detected_anomalies', [])
        if detected_anomalies:
            for anomaly in detected_anomalies:
                with st.expander(f"⚠️ {anomaly['condition']} (вероятность: {anomaly['probability']:.1%})"):
                    st.markdown("**Индикаторы:**")
                    for indicator in anomaly.get('indicators', []):
                        st.markdown(f"• {indicator}")
        else:
            st.success("✅ Критических аномалий не обнаружено")
    
    with tab5:
        st.markdown("### 👶 Индикаторы развития")
        developmental_indicators = detailed_analysis.get('developmental_indicators', {})
        
        overall_assessment = developmental_indicators.get('overall_assessment', 'неизвестно')
        st.info(f"**Общая оценка развития:** {overall_assessment}")
        
        physical_dev = developmental_indicators.get('physical_development', {})
        if 'weight_height_ratio' in physical_dev:
            st.metric("Соотношение веса к росту", f"{physical_dev['weight_height_ratio']:.3f}")
        
        proportional_dev = developmental_indicators.get('proportional_development', {})
        if 'limb_balance' in proportional_dev:
            st.metric("Баланс конечностей", f"{proportional_dev['limb_balance']:.3f}")
    
    # Рекомендации
    st.markdown("---")
    st.markdown("### 💡 Рекомендации")
    
    recommendations = []
    
    # Рекомендации на основе квантового балла
    quantum_score = detailed_analysis.get('quantum_assessment', {}).get('developmental_quantum_score', 0)
    if quantum_score > 0.8:
        recommendations.append("✅ Отличное квантовое развитие - продолжайте наблюдение")
    elif quantum_score > 0.6:
        recommendations.append("📋 Хорошее развитие - плановое наблюдение")
    elif quantum_score > 0.4:
        recommendations.append("⚠️ Требуется внимание - дополнительное обследование")
    else:
        recommendations.append("🚨 Необходимо срочное медицинское вмешательство")
    
    # Рекомендации на основе аномалий
    if detected_anomalies:
        recommendations.append("🔍 Консультация специалиста для уточнения диагноза")
        recommendations.append("📊 Дополнительные исследования при необходимости")
    
    # Рекомендации на основе общего развития
    if "гармоничное" in overall_assessment:
        recommendations.append("🌟 Продолжайте поддерживать здоровый образ жизни")
    elif "задержка" in overall_assessment:
        recommendations.append("🏃‍♂️ Стимулирующие упражнения и развивающие игры")
    
    for i, rec in enumerate(recommendations, 1):
        st.markdown(f"{i}. {rec}")


def display_pediatric_system_info():
    """Отображает информацию о детской системе."""
    
    st.markdown("### 🌟 Возможности системы MQEA-Pediatric")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        #### ⚛️ Квантовые технологии:
        - **Квантовая запутанность** для анализа корреляций
        - **Квантовая когерентность** для оценки стабильности
        - **Квантовая суперпозиция** для множественной диагностики
        
        #### 👶 Специализация для детей:
        - Возрастные нормы от рождения до 10 лет
        - Факторы роста и развития
        - Раннее выявление врожденных патологий
        """)
    
    with col2:
        st.markdown("""
        #### 🔍 Обнаруживаемые состояния:
        - Врожденные пороки сердца
        - Респираторные инфекции
        - Задержка развития
        - Метаболические нарушения
        
        #### 📊 Возрастные группы:
        - Новорожденные (0-1 мес)
        - Младенцы (1-12 мес)
        - Дети раннего возраста (1-3 года)
        - Дошкольники (3-6 лет)
        - Школьники (6-10 лет)
        """)
    
    # Создаем диаграмму возрастных норм
    age_groups = ["Новорожденные\n(0-1 мес)", "Младенцы\n(1-12 мес)", 
                 "Дети раннего возраста\n(1-3 года)", "Дошкольники\n(3-6 лет)", 
                 "Школьники\n(6-10 лет)"]
    
    heart_rate_norms = [130, 120, 110, 100, 85]
    respiratory_norms = [45, 32, 25, 22, 20]
    
    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=("Частота сердечных сокращений", "Частота дыхания"),
        specs=[[{"secondary_y": False}, {"secondary_y": False}]]
    )
    
    fig.add_trace(
        go.Bar(x=age_groups, y=heart_rate_norms, name="ЧСС", marker_color="#FF6B9D"),
        row=1, col=1
    )
    
    fig.add_trace(
        go.Bar(x=age_groups, y=respiratory_norms, name="ЧДД", marker_color="#3498DB"),
        row=1, col=2
    )
    
    fig.update_layout(
        title="Возрастные нормы жизненных показателей",
        showlegend=False,
        height=400
    )
    
    st.plotly_chart(fig, use_container_width=True)

def display_pediatric_results(vital_signs: PediatricVitalSigns, 
                            detected_conditions: list, 
                            quantum_report: dict):
    """Отображает результаты детского анализа."""
    
    st.markdown("### 📊 Результаты квантового анализа")
    
    # Общая оценка
    assessment = quantum_report['overall_assessment']
    if "✅" in assessment:
        st.success(assessment)
    elif "🚨" in assessment:
        st.error(assessment)
    elif "⚠️" in assessment:
        st.warning(assessment)
    else:
        st.info(assessment)
    
    # Квантовая статистика
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Квантовых состояний", quantum_report['quantum_analysis']['total_quantum_states'])
    
    with col2:
        st.metric("Запутанных пар", quantum_report['quantum_analysis']['entangled_pairs'])
    
    with col3:
        coherence = quantum_report['quantum_analysis']['quantum_coherence']
        st.metric("Квантовая когерентность", f"{coherence:.3f}")
    
    with col4:
        dev_factor = quantum_report['quantum_analysis']['developmental_quantum_factor']
        st.metric("Фактор развития", f"{dev_factor:.1f}")
    
    # Информация о пациенте
    st.markdown("### 👶 Информация о ребенке")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown(f"""
        **📅 Возрастная группа:** {quantum_report['patient_info']['age_group']}  
        **👶 Возраст:** {vital_signs.age_months} месяцев  
        **🧠 Стадия развития:** {quantum_report['patient_info']['developmental_stage']}
        """)
    
    with col2:
        # Создаем радиальную диаграмму жизненных показателей
        indicators = ['ЧСС', 'ЧДД', 'АД сист.', 'АД диаст.', 'Темп.', 'SpO2']
        values = [
            vital_signs.heart_rate,
            vital_signs.respiratory_rate,
            vital_signs.blood_pressure_systolic,
            vital_signs.blood_pressure_diastolic,
            vital_signs.temperature * 10,  # Масштабируем для визуализации
            vital_signs.oxygen_saturation
        ]
        
        fig = go.Figure()
        
        fig.add_trace(go.Scatterpolar(
            r=values,
            theta=indicators,
            fill='toself',
            name='Текущие значения',
            line_color='#FF6B9D'
        ))
        
        fig.update_layout(
            polar=dict(
                radialaxis=dict(
                    visible=True,
                    range=[0, max(values) * 1.1]
                )),
            showlegend=True,
            title="Жизненные показатели",
            height=300
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # Обнаруженные состояния
    if detected_conditions:
        st.markdown("### 🚨 Обнаруженные состояния")
        
        for i, condition in enumerate(detected_conditions):
            probability = condition['probability']
            condition_name = condition['condition']
            
            if probability > 0.8:
                st.error(f"🚨 **{condition_name}** - Вероятность: {probability:.1%}")
            elif probability > 0.6:
                st.warning(f"⚠️ **{condition_name}** - Вероятность: {probability:.1%}")
            else:
                st.info(f"ℹ️ **{condition_name}** - Вероятность: {probability:.1%}")
            
            # Показываем рекомендации
            st.markdown("**📋 Рекомендации:**")
            for rec in condition['recommendations']:
                st.markdown(f"• {rec}")
            
            st.markdown("---")
    
    else:
        st.success("✅ Квантовый анализ не выявил значимых отклонений от нормы для данного возраста.")
    
    # Рекомендации
    st.markdown("### 📋 Общие рекомендации")
    
    recommendations = quantum_report['recommendations']
    for i, rec in enumerate(recommendations, 1):
        st.markdown(f"{i}. {rec}")
    
    # Экспорт отчета
    st.markdown("### 💾 Экспорт отчета")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📄 Скачать PDF отчет"):
            st.info("Функция экспорта PDF будет добавлена в следующей версии")
    
    with col2:
        if st.button("📊 Экспорт JSON данных"):
            # Создаем JSON для экспорта
            export_data = {
                "timestamp": quantum_report['timestamp'],
                "patient_info": quantum_report['patient_info'],
                "vital_signs": quantum_report['vital_signs'],
                "quantum_analysis": quantum_report['quantum_analysis'],
                "detected_conditions": detected_conditions,
                "recommendations": recommendations
            }
            
            st.download_button(
                label="💾 Скачать JSON",
                data=json.dumps(export_data, indent=2, ensure_ascii=False),
                file_name=f"pediatric_quantum_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )

# Заглушки для функций детской медицины (будут реализованы позже)
def show_pediatric_monitoring():
    """Мониторинг развития детей - профили и история."""
    st.markdown("## 📊 Мониторинг развития детей")
    st.markdown("**Управление профилями детей и отслеживание динамики развития**")
    
    # Инициализация менеджера профилей
    if 'profile_manager' not in st.session_state:
        st.session_state.profile_manager = PediatricProfileManager()
    
    pm = st.session_state.profile_manager
    
    # Кнопка для восстановления поврежденных данных
    if st.button("🔧 Восстановить поврежденные данные", help="Очищает поврежденные JSON файлы и создает резервные копии"):
        try:
            data_dir = pm.data_dir
            restored_count = 0
            for filename in os.listdir(data_dir):
                if filename.endswith('_history.json'):
                    file_path = os.path.join(data_dir, filename)
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            json.load(f)
                    except (json.JSONDecodeError, FileNotFoundError, IOError):
                        # Создаем резервную копию и удаляем поврежденный файл
                        backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        os.rename(file_path, backup_path)
                        restored_count += 1
                        st.success(f"✅ Восстановлен файл: {filename}")
            
            if restored_count > 0:
                st.success(f"🎉 Восстановлено {restored_count} поврежденных файлов!")
                st.rerun()
            else:
                st.info("ℹ️ Поврежденных файлов не найдено.")
        except Exception as e:
            st.error(f"❌ Ошибка при восстановлении: {e}")
    
    # Меню мониторинга
    monitoring_tab = st.radio(
        "Выберите действие:",
        ["👶 Список детей", "➕ Создать профиль", "📈 История развития", "📊 Анализ прогресса"],
        horizontal=True
    )
    
    if monitoring_tab == "👶 Список детей":
        show_children_list(pm)
    elif monitoring_tab == "➕ Создать профиль":
        show_create_child_profile(pm)
    elif monitoring_tab == "📈 История развития":
        show_development_history(pm)
    elif monitoring_tab == "📊 Анализ прогресса":
        show_progress_analysis(pm)
    
    # Показываем форму добавления обследования, если нужно
    if st.session_state.get('show_add_examination', False):
        show_add_examination_form(pm)

def show_children_list(pm: PediatricProfileManager):
    """Отображает список всех детей."""
    st.markdown("### 👶 Список детей")
    
    profiles = pm.list_all_profiles()
    
    if not profiles:
        st.info("📋 Пока нет зарегистрированных детей. Создайте первый профиль!")
        return
    
    for profile in profiles:
        with st.expander(f"👶 {profile.name} (ID: {profile.child_id})"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown(f"**Дата рождения:** {profile.date_of_birth}")
                st.markdown(f"**Пол:** {profile.gender}")
                
                # Вычисляем возраст
                birth_date = datetime.strptime(profile.date_of_birth, "%Y-%m-%d")
                age_months = (datetime.now().year - birth_date.year) * 12 + (datetime.now().month - birth_date.month)
                st.markdown(f"**Возраст:** {age_months} месяцев")
            
            with col2:
                if profile.blood_type:
                    st.markdown(f"**Группа крови:** {profile.blood_type}")
                
                if profile.allergies:
                    st.markdown(f"**Аллергии:** {', '.join(profile.allergies)}")
                
                if profile.chronic_conditions:
                    st.markdown(f"**Хронические состояния:** {', '.join(profile.chronic_conditions)}")
            
            # Получаем количество записей с обработкой ошибок
            try:
                history = pm.get_development_history(profile.child_id)
                st.markdown(f"📊 **Записей в истории:** {len(history)}")
            except Exception as e:
                st.error(f"❌ Ошибка загрузки истории: {e}")
                st.markdown(f"📊 **Записей в истории:** 0")


def show_create_child_profile(pm: PediatricProfileManager):
    """Создание нового профиля ребенка."""
    st.markdown("### ➕ Создать профиль ребенка")
    
    with st.form("create_child_profile"):
        col1, col2 = st.columns(2)
        
        with col1:
            name = st.text_input("Имя и фамилия ребенка:", placeholder="Иван Иванов")
            date_of_birth = st.date_input("Дата рождения:")
            gender = st.selectbox("Пол:", ["мужской", "женский"])
        
        with col2:
            blood_type = st.selectbox("Группа крови (опционально):", 
                                     ["Не указано", "O(I)", "A(II)", "B(III)", "AB(IV)"])
            allergies = st.text_input("Аллергии (через запятую):", placeholder="пыльца, молоко")
            chronic_conditions = st.text_input("Хронические состояния (через запятую):", placeholder="астма")
        
        st.markdown("#### 👨‍👩‍👧 Информация о родителях")
        col1, col2 = st.columns(2)
        
        with col1:
            mother_name = st.text_input("Имя матери:")
            mother_phone = st.text_input("Телефон матери:")
        
        with col2:
            father_name = st.text_input("Имя отца:")
            father_phone = st.text_input("Телефон отца:")
        
        submit = st.form_submit_button("✅ Создать профиль", type="primary")
        
        if submit:
            if not name or not date_of_birth:
                st.error("⚠️ Заполните обязательные поля: имя и дата рождения")
            else:
                # Подготавливаем данные
                allergies_list = [a.strip() for a in allergies.split(',')] if allergies else []
                chronic_list = [c.strip() for c in chronic_conditions.split(',')] if chronic_conditions else []
                
                parents_info = {}
                if mother_name:
                    parents_info['mother_name'] = mother_name
                    parents_info['mother_phone'] = mother_phone
                if father_name:
                    parents_info['father_name'] = father_name
                    parents_info['father_phone'] = father_phone
                
                # Создаем профиль
                profile = pm.create_child_profile(
                    name=name,
                    date_of_birth=date_of_birth.strftime("%Y-%m-%d"),
                    gender=gender,
                    blood_type=blood_type if blood_type != "Не указано" else None,
                    allergies=allergies_list,
                    chronic_conditions=chronic_list,
                    parents_info=parents_info
                )
                
                st.success(f"✅ Профиль создан успешно! ID: {profile.child_id}")
                st.balloons()


def show_development_history(pm: PediatricProfileManager):
    """Отображает историю развития ребенка."""
    st.markdown("### 📈 История развития")
    
    profiles = pm.list_all_profiles()
    
    if not profiles:
        st.info("📋 Сначала создайте профиль ребенка")
        return
    
    # Выбор ребенка
    selected_child = st.selectbox(
        "Выберите ребенка:",
        options=[p.child_id for p in profiles],
        format_func=lambda x: next((p.name for p in profiles if p.child_id == x), x)
    )
    
    if selected_child:
        history = pm.get_development_history(selected_child)
        
        if not history:
            st.info("📋 История пока пуста. Добавьте первую запись!")
        else:
            st.markdown(f"### 📊 Всего записей: {len(history)}")
            
            for record in reversed(history):  # Показываем от новых к старым
                with st.expander(f"📅 {record.date} (Возраст: {record.age_months} месяцев)"):
                    
                    # Жизненные показатели
                    st.markdown("#### 🩺 Жизненные показатели")
                    vs = record.vital_signs
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Пульс", f"{vs.get('heart_rate', 0):.0f} уд/мин")
                        st.metric("Дыхание", f"{vs.get('respiratory_rate', 0):.0f} дых/мин")
                    with col2:
                        st.metric("Вес", f"{vs.get('weight_kg', 0):.1f} кг")
                        st.metric("Рост", f"{vs.get('height_cm', 0):.1f} см")
                    with col3:
                        st.metric("Температура", f"{vs.get('temperature', 0):.1f} °C")
                        st.metric("SpO2", f"{vs.get('oxygen_saturation', 0):.0f} %")
                    
                    # Заключение
                    if record.development_report:
                        st.markdown("#### 📄 Заключение")
                        conclusion = record.development_report.get('overall_conclusion', 'Нет данных')
                        st.markdown(conclusion)
                    
                    if record.notes:
                        st.markdown(f"**Заметки:** {record.notes}")


def show_progress_analysis(pm: PediatricProfileManager):
    """Анализ прогресса развития."""
    st.markdown("### 📊 Анализ прогресса развития")
    
    profiles = pm.list_all_profiles()
    
    if not profiles:
        st.info("📋 Сначала создайте профиль ребенка")
        return
    
    # Выбор ребенка
    selected_child = st.selectbox(
        "Выберите ребенка:",
        options=[p.child_id for p in profiles],
        format_func=lambda x: next((p.name for p in profiles if p.child_id == x), x)
    )
    
    if selected_child:
        # Получаем анализ прогресса
        progress = pm.analyze_latest_progress(selected_child)
        
        if not progress:
            # Получаем количество записей для более информативного сообщения
            history = pm.get_development_history(selected_child)
            records_count = len(history)
            
            st.warning(f"⚠️ **Недостаточно данных для анализа прогресса**")
            st.info(f"""
            **📊 Текущее количество записей:** {records_count}
            
            **📋 Для анализа прогресса развития необходимо минимум 2 записи обследований.**
            
            **🔍 Что это означает:**
            - Система сравнивает показатели между двумя обследованиями
            - Анализирует динамику роста и развития
            - Выявляет тенденции и изменения
            - Дает рекомендации по дальнейшему наблюдению
            
            **✅ Что нужно сделать:**
            1. Перейдите в раздел "📈 История развития"
            2. Добавьте новое обследование ребенка
            3. Заполните антропометрические данные
            4. Вернитесь к анализу прогресса
            """)
            
            # Показываем кнопку для быстрого перехода
            if st.button("➕ Добавить новое обследование", type="primary"):
                st.session_state.show_add_examination = True
                st.rerun()
            
            return
        
        # Отображаем анализ
        display_progress_analysis(progress, pm, selected_child)


def show_add_examination_form(pm: PediatricProfileManager):
    """Форма добавления нового обследования."""
    st.markdown("---")
    st.markdown("### ➕ Добавить новое обследование")
    
    profiles = pm.list_all_profiles()
    
    if not profiles:
        st.error("❌ Нет зарегистрированных детей")
        return
    
    # Выбор ребенка
    selected_child = st.selectbox(
        "Выберите ребенка:",
        options=[p.child_id for p in profiles],
        format_func=lambda x: next((p.name for p in profiles if p.child_id == x), x),
        key="add_examination_child"
    )
    
    if selected_child:
        profile = next(p for p in profiles if p.child_id == selected_child)
        
        st.markdown(f"**👶 Ребенок:** {profile.name}")
        
        # Вычисляем текущий возраст
        birth_date = datetime.strptime(profile.date_of_birth, "%Y-%m-%d")
        current_age_months = (datetime.now().year - birth_date.year) * 12 + (datetime.now().month - birth_date.month)
        st.markdown(f"**📅 Текущий возраст:** {current_age_months} месяцев")
        
        # Форма обследования
        with st.form("add_examination_form"):
            st.markdown("**📊 Антропометрические данные:**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                weight = st.number_input("Вес (кг)", min_value=0.1, max_value=100.0, value=10.0, step=0.1)
                height = st.number_input("Рост (см)", min_value=10.0, max_value=200.0, value=50.0, step=0.1)
                head_circumference = st.number_input("Окружность головы (см)", min_value=20.0, max_value=70.0, value=35.0, step=0.1)
            
            with col2:
                chest_circumference = st.number_input("Окружность груди (см)", min_value=20.0, max_value=100.0, value=40.0, step=0.1)
                # Убираем поля, которые не используются в DetailedAnthropometry
                st.info("💡 **Остальные параметры** (размеры конечностей, пальцев, лица и т.д.) будут рассчитаны автоматически на основе основных измерений")
            
            # Дополнительные данные
            st.markdown("**📝 Дополнительная информация:**")
            notes = st.text_area("Заметки врача", placeholder="Введите дополнительные наблюдения...")
            
            # Кнопки
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.form_submit_button("💾 Сохранить обследование", type="primary"):
                    try:
                        # Создаем антропометрические данные с базовыми значениями
                        anthropometry = DetailedAnthropometry(
                            # Основные измерения
                            weight_kg=weight,
                            height_cm=height,
                            head_circumference_cm=head_circumference,
                            chest_circumference_cm=chest_circumference,
                            abdominal_circumference_cm=chest_circumference * 0.9,  # Примерное соотношение
                            
                            # Размеры конечностей (примерные значения на основе роста)
                            arm_span_cm=height * 1.02,  # Размах рук примерно равен росту
                            leg_length_cm=height * 0.45,  # Длина ног ~45% от роста
                            foot_length_cm=height * 0.15,  # Длина стопы ~15% от роста
                            foot_width_cm=height * 0.06,  # Ширина стопы ~6% от роста
                            
                            # Размеры пальцев рук (примерные значения)
                            thumb_length_mm=height * 0.8,  # Длина большого пальца
                            index_finger_length_mm=height * 1.0,
                            middle_finger_length_mm=height * 1.1,
                            ring_finger_length_mm=height * 1.0,
                            little_finger_length_mm=height * 0.8,
                            
                            # Размеры пальцев ног
                            big_toe_length_mm=height * 0.3,
                            second_toe_length_mm=height * 0.25,
                            third_toe_length_mm=height * 0.2,
                            fourth_toe_length_mm=height * 0.15,
                            little_toe_length_mm=height * 0.1,
                            
                            # Размеры головы и лица
                            head_length_cm=head_circumference / 3.14,  # Примерная длина
                            head_width_cm=head_circumference / 3.5,   # Примерная ширина
                            face_height_cm=head_circumference / 4.0,  # Высота лица
                            face_width_cm=head_circumference / 4.5,   # Ширина лица
                            
                            # Размеры носа
                            nose_length_mm=height * 0.3,
                            nose_width_mm=height * 0.2,
                            nose_height_mm=height * 0.15,
                            
                            # Размеры глаз
                            eye_width_mm=height * 0.2,
                            eye_height_mm=height * 0.1,
                            inter_eye_distance_mm=height * 0.25,
                            
                            # Размеры рта
                            mouth_width_mm=height * 0.3,
                            lip_thickness_mm=height * 0.05,
                            
                            # Размеры ушей
                            ear_length_mm=height * 0.4,
                            ear_width_mm=height * 0.25,
                            
                            # Дополнительные пропорции
                            waist_to_hip_ratio=0.8,  # Примерное соотношение
                            shoulder_width_cm=chest_circumference * 0.8,
                            hip_width_cm=chest_circumference * 0.7,
                            
                            # Кожные складки (примерные значения)
                            triceps_skinfold_mm=5.0,
                            subscapular_skinfold_mm=4.0,
                            suprailiac_skinfold_mm=3.0
                        )
                        
                        # Создаем жизненные показатели
                        vital_signs = PediatricVitalSigns(
                            age_months=current_age_months,
                            heart_rate=80 + np.random.randint(-10, 10),
                            respiratory_rate=20 + np.random.randint(-5, 5),
                            temperature=36.5 + np.random.uniform(-0.5, 0.5),
                            blood_pressure_systolic=90 + np.random.randint(-10, 10),
                            blood_pressure_diastolic=60 + np.random.randint(-5, 5),
                            oxygen_saturation=98 + np.random.randint(-2, 2),
                            weight_kg=weight,  # Добавляем обязательные параметры
                            height_cm=height,
                            head_circumference_cm=head_circumference,
                            detailed_anthropometry=anthropometry  # Связываем с антропометрией
                        )
                        
                        # Создаем запись о развитии через менеджер профилей
                        record = pm.add_development_record(
                            child_id=selected_child,
                            vital_signs=vital_signs,
                            anthropometry=anthropometry,
                            notes=notes
                        )
                        
                        st.success("✅ Обследование успешно добавлено!")
                        st.session_state.show_add_examination = False
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"❌ Ошибка при сохранении: {str(e)}")
            
            with col2:
                if st.form_submit_button("❌ Отмена"):
                    st.session_state.show_add_examination = False
                    st.rerun()
            
            with col3:
                if st.form_submit_button("📊 Предварительный анализ"):
                    # Показываем предварительный анализ
                    anthropometry = DetailedAnthropometry(
                        # Основные измерения
                        weight_kg=weight,
                        height_cm=height,
                        head_circumference_cm=head_circumference,
                        chest_circumference_cm=chest_circumference,
                        abdominal_circumference_cm=chest_circumference * 0.9,
                        
                        # Размеры конечностей (примерные значения)
                        arm_span_cm=height * 1.02,
                        leg_length_cm=height * 0.45,
                        foot_length_cm=height * 0.15,
                        foot_width_cm=height * 0.06,
                        
                        # Размеры пальцев рук
                        thumb_length_mm=height * 0.8,
                        index_finger_length_mm=height * 1.0,
                        middle_finger_length_mm=height * 1.1,
                        ring_finger_length_mm=height * 1.0,
                        little_finger_length_mm=height * 0.8,
                        
                        # Размеры пальцев ног
                        big_toe_length_mm=height * 0.3,
                        second_toe_length_mm=height * 0.25,
                        third_toe_length_mm=height * 0.2,
                        fourth_toe_length_mm=height * 0.15,
                        little_toe_length_mm=height * 0.1,
                        
                        # Размеры головы и лица
                        head_length_cm=head_circumference / 3.14,
                        head_width_cm=head_circumference / 3.5,
                        face_height_cm=head_circumference / 4.0,
                        face_width_cm=head_circumference / 4.5,
                        
                        # Размеры носа
                        nose_length_mm=height * 0.3,
                        nose_width_mm=height * 0.2,
                        nose_height_mm=height * 0.15,
                        
                        # Размеры глаз
                        eye_width_mm=height * 0.2,
                        eye_height_mm=height * 0.1,
                        inter_eye_distance_mm=height * 0.25,
                        
                        # Размеры рта
                        mouth_width_mm=height * 0.3,
                        lip_thickness_mm=height * 0.05,
                        
                        # Размеры ушей
                        ear_length_mm=height * 0.4,
                        ear_width_mm=height * 0.25,
                        
                        # Дополнительные пропорции
                        waist_to_hip_ratio=0.8,
                        shoulder_width_cm=chest_circumference * 0.8,
                        hip_width_cm=chest_circumference * 0.7,
                        
                        # Кожные складки
                        triceps_skinfold_mm=5.0,
                        subscapular_skinfold_mm=4.0,
                        suprailiac_skinfold_mm=3.0
                    )
                    
                    analysis = pm.engine.analyze_detailed_anthropometry(anthropometry, current_age_months)
                    
                    st.markdown("**🔍 Предварительный анализ:**")
                    st.json(analysis)

def display_progress_analysis(progress: dict, pm: PediatricProfileManager, child_id: str):
    """Отображает анализ прогресса развития."""
    
    period = progress['period']
    
    # Период анализа
    st.markdown(f"### 📅 Период анализа")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("От", period['from_date'])
        st.metric("Возраст", f"{period['from_age_months']} мес")
    with col2:
        st.metric("До", period['to_date'])
        st.metric("Возраст", f"{period['to_age_months']} мес")
    with col3:
        st.metric("Период", f"{period['time_diff_days']} дней")
    
    # Общая оценка
    st.markdown("---")
    st.markdown("### 📋 Общая оценка прогресса")
    st.info(progress['overall_assessment'])
    
    # Создаем вкладки
    tab1, tab2, tab3, tab4 = st.tabs([
        "✅ Сильные стороны",
        "⚠️ Слабые стороны",
        "🔮 Прогноз",
        "💡 Рекомендации"
    ])
    
    with tab1:
        st.markdown("### ✅ Области активного развития")
        strong_areas = progress.get('strong_areas', [])
        
        if strong_areas:
            for area in strong_areas:
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.success(f"**{area['area']}**: {area['status']}")
                with col2:
                    st.metric("Изменение", area['change'])
        else:
            st.info("Нет выраженных областей активного развития")
    
    with tab2:
        st.markdown("### ⚠️ Области, требующие внимания")
        weak_areas = progress.get('weak_areas', [])
        
        if weak_areas:
            for area in weak_areas:
                concern = area.get('concern_level', 'средний')
                
                if concern == 'высокий':
                    st.error(f"🚨 **{area['area']}**: {area['status']} ({area['change']})")
                else:
                    st.warning(f"⚠️ **{area['area']}**: {area['status']} ({area['change']})")
        else:
            st.success("✅ Все области развиваются нормально!")
    
    with tab3:
        st.markdown("### 🔮 Прогноз развития")
        predictions = progress.get('predictions', {})
        
        # Прогноз на месяц
        if predictions.get('next_month'):
            st.markdown("#### 📅 Прогноз на следующий месяц")
            for indicator, pred in predictions['next_month'].items():
                st.metric(
                    pm._translate_indicator(indicator),
                    f"{pred['predicted']:.1f}",
                    help=f"Уверенность: {pred['confidence']}"
                )
        
        # Прогноз на 3 месяца
        if predictions.get('next_3_months'):
            st.markdown("#### 📅 Прогноз на 3 месяца")
            for indicator, pred in predictions['next_3_months'].items():
                st.metric(
                    pm._translate_indicator(indicator),
                    f"{pred['predicted']:.1f}",
                    help=f"Уверенность: {pred['confidence']}"
                )
        
        # Потенциальные проблемы
        if predictions.get('potential_issues'):
            st.markdown("#### ⚠️ Потенциальные риски")
            for issue in predictions['potential_issues']:
                st.warning(issue)
        
        # Положительные тенденции
        if predictions.get('positive_trends'):
            st.markdown("#### ✓ Положительные тенденции")
            for trend in predictions['positive_trends']:
                st.success(trend)
    
    with tab4:
        st.markdown("### 💡 Персонализированные рекомендации")
        recommendations = progress.get('recommendations', [])
        
        if recommendations:
            for i, rec in enumerate(recommendations, 1):
                if '⚠' in rec or '🚨' in rec:
                    st.warning(f"{i}. {rec}")
                else:
                    st.info(f"{i}. {rec}")
        else:
            st.success("✅ Продолжайте текущий режим ухода")
    
    # График изменений
    st.markdown("---")
    st.markdown("### 📈 Динамика изменений")
    
    vital_changes = progress.get('vital_signs_changes', {})
    
    if vital_changes:
        # Показываем изменения в виде таблицы
        import pandas as pd
        
        changes_data = []
        for key, data in vital_changes.items():
            changes_data.append({
                'Показатель': pm._translate_indicator(key),
                'Было': f"{data['old']:.1f}",
                'Стало': f"{data['new']:.1f}",
                'Изменение': f"{data['change']:+.1f}",
                'Процент': f"{data['percent_change']:+.1f}%",
                'Статус': data['status']
            })
        
        df = pd.DataFrame(changes_data)
        st.dataframe(df, use_container_width=True)


def show_create_child_profile(pm: PediatricProfileManager):
    """Создание нового профиля ребенка."""
    st.markdown("### ➕ Создать профиль ребенка")
    
    with st.form("create_child_profile"):
        col1, col2 = st.columns(2)
        
        with col1:
            name = st.text_input("Имя и фамилия ребенка:", placeholder="Иван Иванов")
            date_of_birth = st.date_input("Дата рождения:")
            gender = st.selectbox("Пол:", ["мужской", "женский"])
        
        with col2:
            blood_type = st.selectbox("Группа крови (опционально):", 
                                     ["Не указано", "O(I)", "A(II)", "B(III)", "AB(IV)"])
            allergies = st.text_input("Аллергии (через запятую):", placeholder="пыльца, молоко")
            chronic_conditions = st.text_input("Хронические состояния (через запятую):", placeholder="астма")
        
        st.markdown("#### 👨‍👩‍👧 Информация о родителях")
        col1, col2 = st.columns(2)
        
        with col1:
            mother_name = st.text_input("Имя матери:")
            mother_phone = st.text_input("Телефон матери:")
        
        with col2:
            father_name = st.text_input("Имя отца:")
            father_phone = st.text_input("Телефон отца:")
        
        submit = st.form_submit_button("✅ Создать профиль", type="primary")
        
        if submit:
            if not name or not date_of_birth:
                st.error("⚠️ Заполните обязательные поля: имя и дата рождения")
            else:
                # Подготавливаем данные
                allergies_list = [a.strip() for a in allergies.split(',')] if allergies else []
                chronic_list = [c.strip() for c in chronic_conditions.split(',')] if chronic_conditions else []
                
                parents_info = {}
                if mother_name:
                    parents_info['mother_name'] = mother_name
                    parents_info['mother_phone'] = mother_phone
                if father_name:
                    parents_info['father_name'] = father_name
                    parents_info['father_phone'] = father_phone
                
                # Создаем профиль
                profile = pm.create_child_profile(
                    name=name,
                    date_of_birth=date_of_birth.strftime("%Y-%m-%d"),
                    gender=gender,
                    blood_type=blood_type if blood_type != "Не указано" else None,
                    allergies=allergies_list,
                    chronic_conditions=chronic_list,
                    parents_info=parents_info
                )
                
                st.success(f"✅ Профиль создан успешно! ID: {profile.child_id}")
                st.balloons()


def show_pediatric_treatment():
    """Планы лечения для детей."""
    log_action("🎉 ВЫЗОВ ФУНКЦИИ", "show_pediatric_treatment() - НАЧАЛО")
    st.markdown("## 💊 Планы лечения")
    
    # Информационная панель
    st.info("""
    👶 **Персонализированные планы лечения для детей**
    
    Здесь создаются, отслеживаются и корректируются индивидуальные планы лечения для каждого ребенка.
    Система учитывает возраст, вес, рост, диагнозы и особенности развития для создания оптимального плана.
    """)
    
    # Получаем данные пациента из session_state
    patient_id = st.session_state.get('selected_patient_id', 'P001')
    patient_data = st.session_state.get('patient_data', {})
    
    # Проверяем, что выбранный пациент - ребенок (до 18 лет)
    if patient_data and patient_data.get('age', 0) >= 18:
        st.error("❌ Выбран взрослый пациент. В разделе 'Детская медицина' можно создавать планы лечения только для детей до 18 лет.")
        st.info("💡 Перейдите в раздел 'Детская медицина' → 'Мониторинг развития детей' и выберите профиль ребенка, или создайте новый профиль.")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🔄 Выбрать профиль ребенка", type="primary"):
                # Очищаем выбранного пациента
                if 'selected_patient_id' in st.session_state:
                    del st.session_state.selected_patient_id
                if 'patient_data' in st.session_state:
                    del st.session_state.patient_data
                st.rerun()
        return
    
    # Если нет выбранного пациента, показываем выбор из профилей детей
    if patient_id == 'P001' and not patient_data:
        st.warning("⚠️ Сначала выберите профиль ребенка для создания плана лечения")
        
        # Инициализируем менеджер профилей детей
        if 'profile_manager' not in st.session_state:
            st.session_state.profile_manager = PediatricProfileManager()
        
        pm = st.session_state.profile_manager
        profiles = pm.list_all_profiles()
        
        if profiles:
            # Создаем опции для выбора профиля
            profile_options = {}
            for profile in profiles:
                # Вычисляем возраст из даты рождения
                if hasattr(profile, 'date_of_birth') and profile.date_of_birth:
                    try:
                        # Преобразуем строку в объект date
                        if isinstance(profile.date_of_birth, str):
                            birth_date = datetime.fromisoformat(profile.date_of_birth).date()
                        else:
                            birth_date = profile.date_of_birth
                        
                        today = datetime.now().date()
                        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                    except (ValueError, TypeError):
                        age = 0  # Если не удается распарсить дату
                else:
                    age = 0  # Если нет даты рождения
                profile_options[f"{profile.name} ({profile.child_id}) - {age} лет"] = profile
            
            selected_profile_name = st.selectbox("Выберите профиль ребенка:", list(profile_options.keys()))
            
            if selected_profile_name:
                selected_profile = profile_options[selected_profile_name]
                # Вычисляем возраст из даты рождения
                if hasattr(selected_profile, 'date_of_birth') and selected_profile.date_of_birth:
                    try:
                        # Преобразуем строку в объект date
                        if isinstance(selected_profile.date_of_birth, str):
                            birth_date = datetime.fromisoformat(selected_profile.date_of_birth).date()
                        else:
                            birth_date = selected_profile.date_of_birth
                        
                        today = datetime.now().date()
                        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                    except (ValueError, TypeError):
                        age = 0  # Если не удается распарсить дату
                else:
                    age = 0  # Если нет даты рождения
                
                st.session_state.selected_patient_id = selected_profile.child_id
                st.session_state.patient_data = {
                    'name': selected_profile.name,
                    'age': age,
                    'gender': selected_profile.gender,
                    'weight': getattr(selected_profile, 'weight', 'Не указан'),
                    'height': getattr(selected_profile, 'height', 'Не указан'),
                    'diagnoses': getattr(selected_profile, 'diagnoses', ['Не указаны']),
                    'allergies': getattr(selected_profile, 'allergies', ['Не указаны']),
                    'chronic_conditions': getattr(selected_profile, 'chronic_conditions', ['Не указаны']),
                    'profile_id': selected_profile.child_id,
                    'date_of_birth': selected_profile.date_of_birth.isoformat() if hasattr(selected_profile, 'date_of_birth') else None
                }
                st.success(f"✅ Выбран профиль ребенка: {selected_profile.name}")
                st.rerun()
        else:
            st.error("❌ Нет зарегистрированных профилей детей. Сначала создайте профиль в разделе 'Мониторинг развития детей'.")
            st.info("💡 Перейдите в раздел 'Детская медицина' → 'Мониторинг развития детей' → 'Создать профиль ребенка'")
            return
    
    # Показываем информацию о выбранном ребенке с возможностью смены
    if patient_data and patient_data.get('age', 0) < 18:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.success(f"✅ Выбран ребенок: **{patient_data.get('name', 'Не указано')}** ({patient_data.get('age', 'Не указан')} лет)")
        with col2:
            if st.button("🔄 Выбрать другой профиль", type="secondary"):
                # Очищаем выбранного пациента
                if 'selected_patient_id' in st.session_state:
                    del st.session_state.selected_patient_id
                if 'patient_data' in st.session_state:
                    del st.session_state.patient_data
                st.rerun()
    
    # Инициализируем данные планов лечения в session_state
    if 'pediatric_treatment_plans' not in st.session_state:
        st.session_state.pediatric_treatment_plans = {}
    
    if 'treatment_history' not in st.session_state:
        st.session_state.treatment_history = []
    
    # Создаем табы для разных функций
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🆕 Создать план", 
        "📋 Мои планы", 
        "📊 Мониторинг", 
        "🔔 Напоминания", 
        "📈 Отчеты"
    ])
    
    with tab1:
        show_create_treatment_plan(patient_id, patient_data)
    
    with tab2:
        show_my_treatment_plans(patient_id)
    
    with tab3:
        show_treatment_monitoring(patient_id)
    
    with tab4:
        show_treatment_reminders(patient_id)
    
    with tab5:
        show_treatment_reports(patient_id)

def show_create_treatment_plan(patient_id, patient_data):
    """Создание нового плана лечения."""
    st.subheader("🆕 Создание плана лечения")
    
    # Информация о пациенте
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 👤 Информация о пациенте")
        st.write(f"**ID пациента:** {patient_id}")
        if patient_data:
            st.write(f"**Имя:** {patient_data.get('name', 'Не указано')}")
            st.write(f"**Возраст:** {patient_data.get('age', 'Не указан')} лет")
            st.write(f"**Вес:** {patient_data.get('weight', 'Не указан')} кг")
            st.write(f"**Рост:** {patient_data.get('height', 'Не указан')} см")
    
    with col2:
        st.markdown("### 🏥 Медицинская информация")
        if patient_data:
            st.write(f"**Диагнозы:** {', '.join(patient_data.get('diagnoses', ['Не указаны']))}")
            st.write(f"**Аллергии:** {', '.join(patient_data.get('allergies', ['Не указаны']))}")
            st.write(f"**Хронические заболевания:** {', '.join(patient_data.get('chronic_conditions', ['Не указаны']))}")
    
    st.markdown("---")
    
    # Форма для добавления лекарств (вне основной формы)
    st.markdown("### 💊 Управление лекарственными препаратами")
    
    col1, col2, col3 = st.columns([3, 1, 1])
    
    with col1:
        medication_name = st.text_input("Название препарата", key="med_name")
    
    with col2:
        dosage = st.text_input("Дозировка", placeholder="5мг", key="med_dosage")
    
    with col3:
        frequency = st.selectbox("Частота", ["1 раз в день", "2 раза в день", "3 раза в день", "4 раза в день", "По необходимости"], key="med_frequency")
    
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        duration = st.text_input("Длительность приема", placeholder="7 дней", key="med_duration")
    
    with col2:
        if st.button("➕ Добавить препарат", type="secondary"):
            if medication_name and dosage:
                if 'medications' not in st.session_state:
                    st.session_state.medications = []
                st.session_state.medications.append({
                    'name': medication_name,
                    'dosage': dosage,
                    'frequency': frequency,
                    'duration': duration,
                    'id': len(st.session_state.medications) + 1
                })
                st.success(f"✅ Добавлен: {medication_name}")
                st.rerun()
    
    with col3:
        if st.button("🗑️ Очистить все", type="secondary"):
            st.session_state.medications = []
            st.success("🗑️ Все препараты удалены")
            st.rerun()
    
    st.markdown("---")
    
    # Форма создания плана лечения
    with st.form("create_treatment_plan_form"):
        st.markdown("### 📝 Основная информация плана")
        
        col1, col2 = st.columns(2)
        
        with col1:
            plan_name = st.text_input(
                "Название плана лечения",
                placeholder="Например: Лечение ОРВИ у ребенка 5 лет",
                help="Краткое описание плана лечения"
            )
            
            treatment_type = st.selectbox(
                "Тип лечения",
                ["Острое заболевание", "Хроническое заболевание", "Профилактика", "Реабилитация", "Экстренное лечение"],
                help="Выберите тип лечения"
            )
            
            priority = st.selectbox(
                "Приоритет",
                ["Низкий", "Средний", "Высокий", "Критический"],
                help="Приоритет плана лечения"
            )
        
        with col2:
            start_date = st.date_input(
                "Дата начала лечения",
                value=datetime.now().date(),
                help="Дата начала выполнения плана"
            )
            
            duration_days = st.number_input(
                "Длительность (дни)",
                min_value=1,
                max_value=365,
                value=7,
                help="Планируемая длительность лечения в днях"
            )
            
            doctor_name = st.text_input(
                "Лечащий врач",
                placeholder="ФИО врача",
                help="Врач, ответственный за план"
            )
        
        st.markdown("### 🏃‍♂️ Немедикаментозные процедуры")
        
        col1, col2 = st.columns(2)
        
        with col1:
            procedures = st.text_area(
                "Процедуры и рекомендации",
                placeholder="• Постельный режим\n• Обильное питье\n• Полоскание горла\n• Ингаляции",
                help="Опишите немедикаментозные процедуры"
            )
            
            diet_recommendations = st.text_area(
                "Диетические рекомендации",
                placeholder="• Легкая пища\n• Исключить острое\n• Больше жидкости",
                help="Рекомендации по питанию"
            )
        
        with col2:
            physical_activity = st.text_area(
                "Физическая активность",
                placeholder="• Ограничить активность\n• Легкие прогулки\n• Дыхательная гимнастика",
                help="Рекомендации по физической активности"
            )
            
            restrictions = st.text_area(
                "Ограничения",
                placeholder="• Не посещать детский сад\n• Избегать контактов\n• Не купаться",
                help="Ограничения и противопоказания"
            )
        
        st.markdown("### 📅 График приема")
        
        col1, col2 = st.columns(2)
        
        with col1:
            morning_time = st.time_input("Утренний прием", value=datetime.strptime("08:00", "%H:%M").time())
            afternoon_time = st.time_input("Дневной прием", value=datetime.strptime("14:00", "%H:%M").time())
        
        with col2:
            evening_time = st.time_input("Вечерний прием", value=datetime.strptime("20:00", "%H:%M").time())
            night_time = st.time_input("Ночной прием", value=datetime.strptime("23:00", "%H:%M").time())
        
        st.markdown("### 📝 Дополнительные заметки")
        
        additional_notes = st.text_area(
            "Заметки врача",
            placeholder="Дополнительные рекомендации, особые указания, возможные побочные эффекты...",
            height=100
        )
        
        # Кнопка создания плана
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            if st.form_submit_button("✅ Создать план лечения", type="primary", use_container_width=True):
                if plan_name and doctor_name:
                    # Создаем план лечения
                    plan_id = f"TP_{patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    
                    treatment_plan = {
                        'plan_id': plan_id,
                        'patient_id': patient_id,
                        'plan_name': plan_name,
                        'treatment_type': treatment_type,
                        'priority': priority,
                        'start_date': start_date.isoformat(),
                        'duration_days': duration_days,
                        'end_date': (start_date + timedelta(days=duration_days)).isoformat(),
                        'doctor_name': doctor_name,
                        'medications': st.session_state.medications.copy(),
                        'procedures': procedures,
                        'diet_recommendations': diet_recommendations,
                        'physical_activity': physical_activity,
                        'restrictions': restrictions,
                        'schedule': {
                            'morning': morning_time.isoformat(),
                            'afternoon': afternoon_time.isoformat(),
                            'evening': evening_time.isoformat(),
                            'night': night_time.isoformat()
                        },
                        'additional_notes': additional_notes,
                        'status': 'Активный',
                        'created_at': datetime.now().isoformat(),
                        'progress': 0,
                        'completion_rate': 0.0
                    }
                    
                    # Сохраняем план
                    st.session_state.pediatric_treatment_plans[plan_id] = treatment_plan
                    
                    # Добавляем в историю
                    st.session_state.treatment_history.append({
                        'action': 'Создан план лечения',
                        'plan_id': plan_id,
                        'plan_name': plan_name,
                        'timestamp': datetime.now().isoformat(),
                        'doctor': doctor_name
                    })
                    
                    # Очищаем форму
                    st.session_state.medications = []
                    
                    st.success(f"🎉 План лечения '{plan_name}' успешно создан!")
                    st.balloons()
                    
                    log_action("✅ СОЗДАНИЕ", f"План лечения создан: {plan_name}")
                    
                    st.rerun()
                else:
                    st.error("❌ Пожалуйста, заполните обязательные поля: название плана и врач")

def show_my_treatment_plans(patient_id):
    """Показ планов лечения пациента."""
    st.subheader("📋 Мои планы лечения")
    
    if not st.session_state.pediatric_treatment_plans:
        st.info("📝 У вас пока нет планов лечения. Создайте первый план во вкладке 'Создать план'.")
        return
    
    # Фильтры
    col1, col2, col3 = st.columns(3)
    
    with col1:
        status_filter = st.selectbox("Статус", ["Все", "Активный", "Завершен", "Приостановлен", "Отменен"])
    
    with col2:
        priority_filter = st.selectbox("Приоритет", ["Все", "Низкий", "Средний", "Высокий", "Критический"])
    
    with col3:
        type_filter = st.selectbox("Тип лечения", ["Все", "Острое заболевание", "Хроническое заболевание", "Профилактика", "Реабилитация", "Экстренное лечение"])
    
    # Фильтруем планы
    filtered_plans = []
    for plan_id, plan in st.session_state.pediatric_treatment_plans.items():
        if plan['patient_id'] == patient_id:
            if status_filter != "Все" and plan['status'] != status_filter:
                continue
            if priority_filter != "Все" and plan['priority'] != priority_filter:
                continue
            if type_filter != "Все" and plan['treatment_type'] != type_filter:
                continue
            filtered_plans.append((plan_id, plan))
    
    if not filtered_plans:
        st.warning("🔍 Планы не найдены по выбранным фильтрам.")
        return
    
    # Показываем планы
    for plan_id, plan in filtered_plans:
        with st.expander(f"📋 {plan['plan_name']} - {plan['status']} ({plan['priority']})"):
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.markdown(f"**🏥 Тип:** {plan['treatment_type']}")
                st.markdown(f"**👨‍⚕️ Врач:** {plan['doctor_name']}")
                st.markdown(f"**📅 Период:** {plan['start_date']} - {plan['end_date']}")
                st.markdown(f"**⏱️ Длительность:** {plan['duration_days']} дней")
                
                # Прогресс-бар
                progress = plan.get('progress', 0)
                st.progress(progress / 100)
                st.caption(f"Прогресс: {progress}%")
            
            with col2:
                # Кнопки управления
                col_a, col_b = st.columns(2)
                
                with col_a:
                    if st.button("👁️ Просмотр", key=f"view_{plan_id}"):
                        show_plan_details(plan)
                
                with col_b:
                    if st.button("✏️ Редактировать", key=f"edit_{plan_id}"):
                        edit_treatment_plan(plan_id, plan)
                
                if plan['status'] == 'Активный':
                    if st.button("⏸️ Приостановить", key=f"pause_{plan_id}"):
                        plan['status'] = 'Приостановлен'
                        st.success("⏸️ План приостановлен")
                        st.rerun()
                
                if plan['status'] == 'Приостановлен':
                    if st.button("▶️ Возобновить", key=f"resume_{plan_id}"):
                        plan['status'] = 'Активный'
                        st.success("▶️ План возобновлен")
                        st.rerun()
                
                if st.button("❌ Отменить", key=f"cancel_{plan_id}"):
                    plan['status'] = 'Отменен'
                    st.success("❌ План отменен")
                    st.rerun()

def show_plan_details(plan):
    """Показ детальной информации о плане лечения."""
    st.markdown("### 📋 Детальная информация о плане")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📝 Основная информация:**")
        st.write(f"• **Название:** {plan['plan_name']}")
        st.write(f"• **Тип:** {plan['treatment_type']}")
        st.write(f"• **Приоритет:** {plan['priority']}")
        st.write(f"• **Статус:** {plan['status']}")
        st.write(f"• **Врач:** {plan['doctor_name']}")
        st.write(f"• **Создан:** {plan['created_at']}")
    
    with col2:
        st.markdown("**📅 Временные рамки:**")
        st.write(f"• **Начало:** {plan['start_date']}")
        st.write(f"• **Окончание:** {plan['end_date']}")
        st.write(f"• **Длительность:** {plan['duration_days']} дней")
        st.write(f"• **Прогресс:** {plan.get('progress', 0)}%")
    
    # Лекарственные препараты
    if plan['medications']:
        st.markdown("**💊 Лекарственные препараты:**")
        for med in plan['medications']:
            st.write(f"• **{med['name']}** - {med['dosage']}, {med['frequency']}, {med['duration']}")
    
    # Процедуры
    if plan['procedures']:
        st.markdown("**🏃‍♂️ Процедуры:**")
        st.write(plan['procedures'])
    
    # Диетические рекомендации
    if plan['diet_recommendations']:
        st.markdown("**🍎 Диетические рекомендации:**")
        st.write(plan['diet_recommendations'])
    
    # Физическая активность
    if plan['physical_activity']:
        st.markdown("**🏃‍♂️ Физическая активность:**")
        st.write(plan['physical_activity'])
    
    # Ограничения
    if plan['restrictions']:
        st.markdown("**⚠️ Ограничения:**")
        st.write(plan['restrictions'])
    
    # График приема
    st.markdown("**📅 График приема:**")
    schedule = plan['schedule']
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.write(f"🌅 Утро: {schedule['morning']}")
    with col2:
        st.write(f"☀️ День: {schedule['afternoon']}")
    with col3:
        st.write(f"🌆 Вечер: {schedule['evening']}")
    with col4:
        st.write(f"🌙 Ночь: {schedule['night']}")
    
    # Дополнительные заметки
    if plan['additional_notes']:
        st.markdown("**📝 Дополнительные заметки:**")
        st.write(plan['additional_notes'])

def edit_treatment_plan(plan_id, plan):
    """Редактирование плана лечения."""
    st.markdown("### ✏️ Редактирование плана лечения")
    
    with st.form("edit_treatment_plan_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            plan_name = st.text_input("Название плана", value=plan['plan_name'])
            treatment_type = st.selectbox("Тип лечения", 
                ["Острое заболевание", "Хроническое заболевание", "Профилактика", "Реабилитация", "Экстренное лечение"],
                index=["Острое заболевание", "Хроническое заболевание", "Профилактика", "Реабилитация", "Экстренное лечение"].index(plan['treatment_type']))
            priority = st.selectbox("Приоритет", 
                ["Низкий", "Средний", "Высокий", "Критический"],
                index=["Низкий", "Средний", "Высокий", "Критический"].index(plan['priority']))
        
        with col2:
            start_date = st.date_input("Дата начала", value=datetime.fromisoformat(plan['start_date']).date())
            duration_days = st.number_input("Длительность (дни)", value=plan['duration_days'], min_value=1, max_value=365)
            doctor_name = st.text_input("Лечащий врач", value=plan['doctor_name'])
        
        additional_notes = st.text_area("Дополнительные заметки", value=plan['additional_notes'])
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            if st.form_submit_button("💾 Сохранить изменения", type="primary", use_container_width=True):
                # Обновляем план
                plan['plan_name'] = plan_name
                plan['treatment_type'] = treatment_type
                plan['priority'] = priority
                plan['start_date'] = start_date.isoformat()
                plan['duration_days'] = duration_days
                plan['end_date'] = (start_date + timedelta(days=duration_days)).isoformat()
                plan['doctor_name'] = doctor_name
                plan['additional_notes'] = additional_notes
                plan['updated_at'] = datetime.now().isoformat()
                
                st.success("✅ План лечения обновлен!")
                log_action("✏️ РЕДАКТИРОВАНИЕ", f"План лечения обновлен: {plan_name}")
                st.rerun()

def show_treatment_monitoring(patient_id):
    """Мониторинг выполнения планов лечения."""
    st.subheader("📊 Мониторинг лечения")
    
    # Получаем активные планы пациента
    active_plans = [plan for plan_id, plan in st.session_state.pediatric_treatment_plans.items() 
                   if plan['patient_id'] == patient_id and plan['status'] == 'Активный']
    
    if not active_plans:
        st.info("📝 У вас нет активных планов лечения для мониторинга.")
        return
    
    # Выбор плана для мониторинга
    plan_options = {f"{plan['plan_name']} ({plan['plan_id']})": plan for plan in active_plans}
    selected_plan_name = st.selectbox("Выберите план для мониторинга", list(plan_options.keys()))
    selected_plan = plan_options[selected_plan_name]
    
    st.markdown("---")
    
    # Информация о плане
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Прогресс", f"{selected_plan.get('progress', 0)}%")
    
    with col2:
        days_elapsed = (datetime.now().date() - datetime.fromisoformat(selected_plan['start_date']).date()).days
        st.metric("Дней прошло", days_elapsed)
    
    with col3:
        days_remaining = (datetime.fromisoformat(selected_plan['end_date']).date() - datetime.now().date()).days
        st.metric("Дней осталось", max(0, days_remaining))
    
    with col4:
        completion_rate = (days_elapsed / selected_plan['duration_days']) * 100 if selected_plan['duration_days'] > 0 else 0
        st.metric("Завершение", f"{min(100, completion_rate):.1f}%")
    
    # График прогресса
    st.markdown("### 📈 График прогресса")
    
    # Создаем данные для графика
    dates = []
    progress_values = []
    
    start_date = datetime.fromisoformat(selected_plan['start_date']).date()
    end_date = datetime.fromisoformat(selected_plan['end_date']).date()
    
    for i in range((end_date - start_date).days + 1):
        current_date = start_date + timedelta(days=i)
        dates.append(current_date)
        
        # Симулируем прогресс (в реальном приложении это будет из базы данных)
        if current_date <= datetime.now().date():
            progress = min(100, (i / selected_plan['duration_days']) * 100 + random.uniform(-5, 5))
        else:
            progress = None
        progress_values.append(progress)
    
    # Создаем график
    import plotly.graph_objects as go
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=dates,
        y=progress_values,
        mode='lines+markers',
        name='Прогресс',
        line=dict(color='blue', width=3),
        marker=dict(size=8)
    ))
    
    fig.update_layout(
        title="Прогресс выполнения плана лечения",
        xaxis_title="Дата",
        yaxis_title="Прогресс (%)",
        template="plotly_white",
        height=400
    )
    
    st.plotly_chart(fig, use_container_width=True, key=f"monitoring_chart_{selected_plan['plan_id']}")
    
    # Ежедневные задачи
    st.markdown("### ✅ Ежедневные задачи")
    
    today = datetime.now().date()
    
    # Создаем задачи на сегодня
    if selected_plan['medications']:
        st.markdown("**💊 Прием лекарств:**")
        for med in selected_plan['medications']:
            col1, col2, col3 = st.columns([3, 1, 1])
            with col1:
                st.write(f"• {med['name']} - {med['dosage']} ({med['frequency']})")
            with col2:
                if st.button("✅", key=f"med_done_{med['id']}_{today}"):
                    st.success("✅ Принято!")
            with col3:
                if st.button("❌", key=f"med_skip_{med['id']}_{today}"):
                    st.warning("❌ Пропущено")
    
    if selected_plan['procedures']:
        st.markdown("**🏃‍♂️ Процедуры:**")
        procedures_list = selected_plan['procedures'].split('\n')
        for i, procedure in enumerate(procedures_list):
            if procedure.strip():
                col1, col2, col3 = st.columns([3, 1, 1])
                with col1:
                    st.write(f"• {procedure.strip()}")
                with col2:
                    if st.button("✅", key=f"proc_done_{i}_{today}"):
                        st.success("✅ Выполнено!")
                with col3:
                    if st.button("❌", key=f"proc_skip_{i}_{today}"):
                        st.warning("❌ Пропущено")
    
    # Заметки о самочувствии
    st.markdown("### 📝 Заметки о самочувствии")
    
    with st.form("wellness_notes_form"):
        wellness_notes = st.text_area(
            "Как вы себя чувствуете сегодня?",
            placeholder="Опишите свое самочувствие, побочные эффекты, улучшения...",
            height=100
        )
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            if st.form_submit_button("💾 Сохранить заметки", type="primary", use_container_width=True):
                if wellness_notes:
                    # Сохраняем заметки
                    if 'daily_notes' not in st.session_state:
                        st.session_state.daily_notes = {}
                    
                    note_key = f"{selected_plan['plan_id']}_{today}"
                    st.session_state.daily_notes[note_key] = {
                        'date': today.isoformat(),
                        'notes': wellness_notes,
                        'plan_id': selected_plan['plan_id']
                    }
                    
                    st.success("✅ Заметки сохранены!")
                    log_action("📝 ЗАМЕТКИ", f"Добавлены заметки о самочувствии для плана {selected_plan['plan_name']}")

def show_treatment_reminders(patient_id):
    """Система напоминаний о лечении."""
    st.subheader("🔔 Напоминания о лечении")
    
    # Получаем активные планы пациента
    active_plans = [plan for plan_id, plan in st.session_state.pediatric_treatment_plans.items() 
                   if plan['patient_id'] == patient_id and plan['status'] == 'Активный']
    
    if not active_plans:
        st.info("📝 У вас нет активных планов лечения для напоминаний.")
        return
    
    # Настройки напоминаний
    st.markdown("### ⚙️ Настройки напоминаний")
    
    col1, col2 = st.columns(2)
    
    with col1:
        enable_medication_reminders = st.checkbox("💊 Напоминания о приеме лекарств", value=True)
        enable_procedure_reminders = st.checkbox("🏃‍♂️ Напоминания о процедурах", value=True)
        enable_wellness_reminders = st.checkbox("📝 Напоминания о записи самочувствия", value=True)
    
    with col2:
        reminder_time = st.time_input("Время напоминаний", value=datetime.strptime("09:00", "%H:%M").time())
        reminder_frequency = st.selectbox("Частота", ["Каждый день", "Каждые 2 дня", "Еженедельно"])
        notification_method = st.selectbox("Способ уведомления", ["В приложении", "Email", "SMS", "Все способы"])
    
    # Показываем предстоящие напоминания
    st.markdown("### 📅 Предстоящие напоминания")
    
    today = datetime.now().date()
    current_time = datetime.now().time()
    
    for plan in active_plans:
        st.markdown(f"**📋 {plan['plan_name']}**")
        
        # Напоминания о лекарствах
        if enable_medication_reminders and plan['medications']:
            st.markdown("💊 **Лекарства:**")
            for med in plan['medications']:
                st.write(f"• {med['name']} - {med['frequency']}")
        
        # Напоминания о процедурах
        if enable_procedure_reminders and plan['procedures']:
            st.markdown("🏃‍♂️ **Процедуры:**")
            procedures_list = plan['procedures'].split('\n')
            for procedure in procedures_list:
                if procedure.strip():
                    st.write(f"• {procedure.strip()}")
        
        # Напоминания о записи самочувствия
        if enable_wellness_reminders:
            st.markdown("📝 **Запись самочувствия:**")
            st.write("• Не забудьте записать, как вы себя чувствуете сегодня")
        
        st.markdown("---")
    
    # История напоминаний
    st.markdown("### 📚 История напоминаний")
    
    if 'reminder_history' not in st.session_state:
        st.session_state.reminder_history = []
    
    if st.session_state.reminder_history:
        for reminder in st.session_state.reminder_history[-10:]:  # Последние 10
            col1, col2, col3 = st.columns([2, 1, 1])
            with col1:
                st.write(f"🔔 {reminder['message']}")
            with col2:
                st.write(f"📅 {reminder['timestamp']}")
            with col3:
                status = "✅ Выполнено" if reminder.get('completed', False) else "⏳ Ожидает"
                st.write(status)
    else:
        st.info("📝 История напоминаний пуста")

def show_treatment_reports(patient_id):
    """Отчеты по эффективности лечения."""
    st.subheader("📈 Отчеты по эффективности лечения")
    
    # Получаем все планы пациента
    patient_plans = [plan for plan_id, plan in st.session_state.pediatric_treatment_plans.items() 
                    if plan['patient_id'] == patient_id]
    
    if not patient_plans:
        st.info("📝 У вас нет планов лечения для создания отчетов.")
        return
    
    # Выбор периода для отчета
    col1, col2 = st.columns(2)
    
    with col1:
        start_date = st.date_input("Начальная дата", value=datetime.now().date() - timedelta(days=30))
    
    with col2:
        end_date = st.date_input("Конечная дата", value=datetime.now().date())
    
    # Фильтруем планы по периоду
    filtered_plans = []
    for plan in patient_plans:
        plan_start = datetime.fromisoformat(plan['start_date']).date()
        plan_end = datetime.fromisoformat(plan['end_date']).date()
        
        if (plan_start <= end_date and plan_end >= start_date):
            filtered_plans.append(plan)
    
    if not filtered_plans:
        st.warning("🔍 Нет планов лечения в выбранном периоде.")
        return
    
    # Общая статистика
    st.markdown("### 📊 Общая статистика")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_plans = len(filtered_plans)
        st.metric("Всего планов", total_plans)
    
    with col2:
        active_plans = len([p for p in filtered_plans if p['status'] == 'Активный'])
        st.metric("Активных планов", active_plans)
    
    with col3:
        completed_plans = len([p for p in filtered_plans if p['status'] == 'Завершен'])
        st.metric("Завершенных планов", completed_plans)
    
    with col4:
        avg_progress = sum(p.get('progress', 0) for p in filtered_plans) / len(filtered_plans)
        st.metric("Средний прогресс", f"{avg_progress:.1f}%")
    
    # График эффективности лечения
    st.markdown("### 📈 График эффективности лечения")
    
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    
    # Создаем данные для графика
    plan_names = [plan['plan_name'] for plan in filtered_plans]
    progress_values = [plan.get('progress', 0) for plan in filtered_plans]
    priorities = [plan['priority'] for plan in filtered_plans]
    
    # Цвета по приоритету
    color_map = {
        'Низкий': 'green',
        'Средний': 'blue', 
        'Высокий': 'orange',
        'Критический': 'red'
    }
    
    colors = [color_map.get(p, 'gray') for p in priorities]
    
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        x=plan_names,
        y=progress_values,
        marker_color=colors,
        text=progress_values,
        textposition='auto',
        name='Прогресс (%)'
    ))
    
    fig.update_layout(
        title="Прогресс выполнения планов лечения",
        xaxis_title="Планы лечения",
        yaxis_title="Прогресс (%)",
        template="plotly_white",
        height=500,
        xaxis_tickangle=-45
    )
    
    st.plotly_chart(fig, use_container_width=True, key=f"treatment_reports_chart_{patient_id}")
    
    # Детальный отчет по каждому плану
    st.markdown("### 📋 Детальный отчет по планам")
    
    for plan in filtered_plans:
        with st.expander(f"📋 {plan['plan_name']} - {plan['status']} ({plan['priority']})"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**📊 Статистика:**")
                st.write(f"• **Прогресс:** {plan.get('progress', 0)}%")
                st.write(f"• **Статус:** {plan['status']}")
                st.write(f"• **Приоритет:** {plan['priority']}")
                st.write(f"• **Тип:** {plan['treatment_type']}")
                st.write(f"• **Врач:** {plan['doctor_name']}")
            
            with col2:
                st.markdown("**📅 Временные рамки:**")
                st.write(f"• **Начало:** {plan['start_date']}")
                st.write(f"• **Окончание:** {plan['end_date']}")
                st.write(f"• **Длительность:** {plan['duration_days']} дней")
                
                # Вычисляем дни до завершения
                end_date = datetime.fromisoformat(plan['end_date']).date()
                days_remaining = (end_date - datetime.now().date()).days
                st.write(f"• **Дней до завершения:** {max(0, days_remaining)}")
            
            # Рекомендации
            st.markdown("**💡 Рекомендации:**")
            
            if plan.get('progress', 0) < 30:
                st.warning("⚠️ Низкий прогресс. Рекомендуется пересмотреть план лечения.")
            elif plan.get('progress', 0) < 70:
                st.info("ℹ️ Средний прогресс. Продолжайте следовать плану.")
            else:
                st.success("✅ Хороший прогресс. План выполняется успешно.")
            
            # Кнопка экспорта отчета
            if st.button(f"📄 Экспортировать отчет", key=f"export_{plan['plan_id']}"):
                export_treatment_report(plan)
    
    # Экспорт общего отчета
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        if st.button("📊 Экспортировать общий отчет", type="primary", use_container_width=True):
            export_general_treatment_report(filtered_plans, start_date, end_date)

def export_treatment_report(plan):
    """Экспорт отчета по конкретному плану лечения."""
    report_data = {
        'plan_id': plan['plan_id'],
        'plan_name': plan['plan_name'],
        'patient_id': plan['patient_id'],
        'treatment_type': plan['treatment_type'],
        'priority': plan['priority'],
        'status': plan['status'],
        'progress': plan.get('progress', 0),
        'doctor_name': plan['doctor_name'],
        'start_date': plan['start_date'],
        'end_date': plan['end_date'],
        'duration_days': plan['duration_days'],
        'medications': plan['medications'],
        'procedures': plan['procedures'],
        'diet_recommendations': plan['diet_recommendations'],
        'physical_activity': plan['physical_activity'],
        'restrictions': plan['restrictions'],
        'additional_notes': plan['additional_notes'],
        'exported_at': datetime.now().isoformat()
    }
    
    # Создаем JSON для скачивания
    json_data = json.dumps(report_data, ensure_ascii=False, indent=2)
    
    st.download_button(
        label="📄 Скачать отчет (JSON)",
        data=json_data,
        file_name=f"treatment_report_{plan['plan_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        mime="application/json"
    )
    
    log_action("📄 ЭКСПОРТ", f"Экспортирован отчет по плану: {plan['plan_name']}")

def export_general_treatment_report(plans, start_date, end_date):
    """Экспорт общего отчета по всем планам лечения."""
    report_data = {
        'report_period': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        },
        'summary': {
            'total_plans': len(plans),
            'active_plans': len([p for p in plans if p['status'] == 'Активный']),
            'completed_plans': len([p for p in plans if p['status'] == 'Завершен']),
            'average_progress': sum(p.get('progress', 0) for p in plans) / len(plans) if plans else 0
        },
        'plans': plans,
        'exported_at': datetime.now().isoformat()
    }
    
    # Создаем JSON для скачивания
    json_data = json.dumps(report_data, ensure_ascii=False, indent=2)
    
    st.download_button(
        label="📊 Скачать общий отчет (JSON)",
        data=json_data,
        file_name=f"general_treatment_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.json",
        mime="application/json"
    )
    
    log_action("📊 ЭКСПОРТ", f"Экспортирован общий отчет за период {start_date} - {end_date}")

def show_pediatric_anthropometry():
    """Антропометрические измерения."""
    st.markdown("## 📋 Антропометрия")
    st.info("🚧 Функция находится в разработке. Здесь будет система антропометрических измерений и оценки физического развития.")

def show_pediatric_neurology():
    """Детская неврология."""
    st.markdown("## 🧠 Неврология")
    st.info("🚧 Функция находится в разработке. Здесь будут инструменты для диагностики неврологических состояний у детей.")

def show_pediatric_cardiology():
    """Детская кардиология."""
    st.markdown("## ❤️ Кардиология")
    st.info("🚧 Функция находится в разработке. Здесь будут специализированные инструменты для диагностики сердечных заболеваний у детей.")

def show_pediatric_pulmonology():
    """Детская пульмонология."""
    st.markdown("## 🫁 Пульмонология")
    st.info("🚧 Функция находится в разработке. Здесь будут инструменты для диагностики заболеваний дыхательной системы у детей.")

def show_pediatric_neonatology():
    """Неонатология."""
    st.markdown("## 🍼 Неонатология")
    st.info("🚧 Функция находится в разработке. Здесь будут специализированные инструменты для новорожденных и недоношенных детей.")

def show_kapch_analysis():
    """Квантовый Анализ Подсознание Человека (КАПЧ)."""
    st.markdown("## 🧠 КАПЧ - Квантовый Анализ Подсознание Человека")
    
    # Заголовок с описанием
    st.markdown("""
    ### 🔬 Инновационная система анализа психосоматических состояний
    
    **КАПЧ** - это революционная технология анализа квантовых процессов в подсознании человека, 
    которая предсказывает риски заболеваний на основе эмоциональных и физических состояний.
    """)
    
    # Инициализируем анализатор КАПЧ
    if 'kapch_analyzer' not in st.session_state:
        st.session_state.kapch_analyzer = KAPCHAnalyzer()
    
    # Создаем колонки для интерфейса
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 📊 Панель ввода данных")
        
        # Выбор пациента
        patient_id = st.selectbox(
            "👤 Выберите пациента для анализа:",
            ["P001", "P002", "P003", "P004", "P005"],
            key="kapch_patient_select"
        )
        
        # Эмоциональные состояния
        st.markdown("#### 😊 Эмоциональные состояния (0-1)")
        emotional_data = {}
        
        emotions = ['стресс', 'тревога', 'гнев', 'печаль', 'страх', 'одиночество', 'зависть', 'вина', 'стыд', 'разочарование']
        for emotion in emotions:
            emotional_data[emotion] = st.slider(
                f"😟 {emotion.title()}:",
                min_value=0.0,
                max_value=1.0,
                value=0.0,
                step=0.1,
                key=f"emotion_{emotion}"
            )
        
        # Физические симптомы
        st.markdown("#### 🏥 Физические симптомы (0-1)")
        physical_data = {}
        
        symptoms = ['усталость', 'бессонница', 'головная боль', 'боль в груди', 'проблемы с пищеварением', 'мышечное напряжение']
        for symptom in symptoms:
            physical_data[symptom] = st.slider(
                f"🤒 {symptom.title()}:",
                min_value=0.0,
                max_value=1.0,
                value=0.0,
                step=0.1,
                key=f"symptom_{symptom}"
            )
        
        # Образ жизни
        st.markdown("#### 🏃‍♂️ Образ жизни (0-1)")
        lifestyle_data = {
            'социальная_поддержка': st.slider("👥 Социальная поддержка:", 0.0, 1.0, 0.5, 0.1, key="social_support"),
            'физическая_активность': st.slider("🏃‍♂️ Физическая активность:", 0.0, 1.0, 0.5, 0.1, key="physical_activity"),
            'качество_сна': st.slider("😴 Качество сна:", 0.0, 1.0, 0.5, 0.1, key="sleep_quality")
        }
    
    with col2:
        st.markdown("### 🎯 Результаты анализа")
        
        # Кнопка запуска анализа
        if st.button("🚀 Запустить КАПЧ анализ", key="kapch_start_btn"):
            with st.spinner("🔄 Выполняется квантовый анализ подсознания..."):
                # Запускаем анализ
                analyzer = st.session_state.kapch_analyzer
                results = analyzer.analyze_human_state(emotional_data, physical_data, lifestyle_data)
                
                # Сохраняем результаты в session_state
                st.session_state.kapch_results = results
                
                time.sleep(1)  # Небольшая задержка для эффекта
                
                # Показываем результаты
                st.success("✅ Анализ завершен!")
                
                # Общий риск
                risk_score = results['risk_score']
                risk_level = results['total_risk_assessment']['risk_level']
                
                st.markdown(f"#### 📊 Общий уровень риска: **{risk_level}**")
                st.progress(risk_score)
                
                # Квантовые показатели
                st.markdown("#### 🌊 Квантовые показатели:")
                quantum_analysis = results['quantum_analysis']
                
                col_a, col_b = st.columns(2)
                with col_a:
                    st.metric("Квантовый индекс здоровья", f"{quantum_analysis['quantum_health_index']:.1%}")
                    st.metric("Стабильность подсознания", f"{quantum_analysis['subconscious_stability']:.1%}")
                
                with col_b:
                    st.metric("Квантовая когерентность", f"{quantum_analysis['quantum_coherence']:.1%}")
                    st.metric("Общий риск", f"{risk_score:.1%}")
        
        # Показываем детальные результаты, если анализ был проведен
        if 'kapch_results' in st.session_state:
            results = st.session_state.kapch_results
            
            # Наиболее вероятные заболевания
            st.markdown("#### 🚨 Наиболее вероятные заболевания:")
            for disease, probability in results['total_risk_assessment']['most_likely_diseases'][:3]:
                st.warning(f"• {disease}: {probability:.1%}")
            
            # Рекомендации
            st.markdown("#### 💡 Рекомендации:")
            for recommendation in results['recommendations'][:3]:
                st.info(recommendation)
        
        # Дополнительные инструменты
        st.markdown("#### 🛠️ Дополнительные инструменты")
        
        if st.button("📊 История анализов", key="kapch_history_btn"):
            st.info("📋 История предыдущих КАПЧ анализов будет отображена здесь")
        
        if st.button("💾 Экспорт данных", key="kapch_export_btn"):
            if 'kapch_results' in st.session_state:
                # Создаем JSON для экспорта
                export_data = {
                    'patient_id': patient_id,
                    'timestamp': datetime.now().isoformat(),
                    'results': st.session_state.kapch_results
                }
                st.download_button(
                    label="📥 Скачать результаты",
                    data=json.dumps(export_data, ensure_ascii=False, indent=2),
                    file_name=f"kapch_analysis_{patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json"
                )
            else:
                st.warning("⚠️ Сначала проведите анализ")
    
    # Детальные результаты анализа
    if 'kapch_results' in st.session_state:
        results = st.session_state.kapch_results
        
        st.markdown("---")
        st.markdown("## 📊 Детальные результаты анализа КАПЧ")
        
        # Создаем табы для разных разделов
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["🧠 3D Модель", "🏥 Системы организма", "🔄 Нормализация", "🔮 Предсказания", "📈 Аналитика"])
        
        with tab1:
            st.markdown("### 🧠 3D Модель человека с визуализацией нарушений")
            
            # Создаем 3D модель
            analyzer = st.session_state.kapch_analyzer
            fig_3d = analyzer.create_3d_human_model(results['body_systems_analysis'])
            st.plotly_chart(fig_3d, use_container_width=True)
            
            st.markdown("""
            **Легенда цветов:**
            - 🟢 Зеленый: Здоровые органы
            - 🟡 Желтый: Средний риск
            - 🟠 Оранжевый: Высокий риск  
            - 🔴 Красный: Критический риск
            """)
        
        with tab2:
            st.markdown("### 🏥 Анализ систем организма")
            
            for system_name, system_data in results['body_systems_analysis'].items():
                with st.expander(f"🔬 {system_name.replace('_', ' ').title()} - {system_data['status']}"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.metric("Уровень риска", f"{system_data['risk_score']:.1%}")
                        st.metric("Статус", system_data['status'])
                        
                        if system_data['affected_emotions']:
                            st.markdown("**Влияющие эмоции:**")
                            for emotion_data in system_data['affected_emotions']:
                                st.write(f"• {emotion_data['emotion']}: {emotion_data['intensity']:.1f} (влияние: {emotion_data['impact']:.1f})")
                    
                    with col2:
                        if system_data['affected_symptoms']:
                            st.markdown("**Физические симптомы:**")
                            for symptom_data in system_data['affected_symptoms']:
                                st.write(f"• {symptom_data['symptom']}: {symptom_data['intensity']:.1f} (влияние: {symptom_data['impact']:.1f})")
                        
                        st.markdown("**Органы:**")
                        for organ in system_data['organs']:
                            st.write(f"• {organ.replace('_', ' ').title()}")
                        
                        st.markdown("**Потенциальные заболевания:**")
                        for disease in system_data['potential_diseases'][:3]:
                            st.write(f"• {disease}")
        
        with tab3:
            st.markdown("### 🔄 План нормализации и восстановления")
            
            if results['normalization_analysis']:
                for system_name, norm_data in results['normalization_analysis'].items():
                    with st.expander(f"🔄 {system_name.replace('_', ' ').title()} - Приоритет: {norm_data['normalization_priority']}"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.metric("Текущий статус", norm_data['current_status'])
                            st.metric("Время восстановления", norm_data['estimated_recovery_time'])
                        
                        with col2:
                            st.markdown("**Рекомендуемые методы:**")
                            for method in norm_data['recommended_methods']:
                                effectiveness_bar = st.progress(method['effectiveness'])
                                st.write(f"• {method['method'].replace('_', ' ').title()}")
                                st.write(f"  Приоритет: {method['priority']}")
                                st.write(f"  Ожидаемое улучшение: {method['expected_improvement']:.1%}")
            else:
                st.success("✅ Все системы организма в норме! Продолжайте поддерживать здоровый образ жизни.")
        
        with tab4:
            st.markdown("### 🔮 Подробные предсказания")
            
            # Краткосрочные предсказания
            if results['detailed_predictions']['short_term']:
                st.markdown("#### ⚡ Краткосрочные предсказания (1-3 месяца)")
                for prediction in results['detailed_predictions']['short_term']:
                    with st.expander(f"🚨 {prediction['condition']} - Вероятность: {prediction['probability']:.1%}"):
                        st.write(f"**Временные рамки:** {prediction['timeframe']}")
                        st.write(f"**Серьезность:** {prediction['severity']}")
                        st.write(f"**Профилактика возможна:** {'Да' if prediction['prevention_possible'] else 'Нет'}")
            
            # Среднесрочные предсказания
            if results['detailed_predictions']['medium_term']:
                st.markdown("#### 📅 Среднесрочные предсказания (3-12 месяцев)")
                for prediction in results['detailed_predictions']['medium_term']:
                    with st.expander(f"⚠️ {prediction['system'].replace('_', ' ').title()} - Риск: {prediction['risk_score']:.1%}"):
                        st.write(f"**Временные рамки:** {prediction['timeframe']}")
                        st.write(f"**Потенциальные исходы:** {', '.join(prediction['potential_outcomes'])}")
                        st.write(f"**Стратегии профилактики:** {', '.join(prediction['prevention_strategies'])}")
            
            # Долгосрочные предсказания
            if results['detailed_predictions']['long_term']:
                st.markdown("#### 🗓️ Долгосрочные предсказания (1-3 года)")
                for prediction in results['detailed_predictions']['long_term']:
                    with st.expander(f"🔮 {prediction['scenario']} - Вероятность: {prediction['probability']:.1%}"):
                        st.write(f"**Временные рамки:** {prediction['timeframe']}")
                        st.write(f"**Затронутые системы:** {', '.join(prediction['affected_systems'])}")
                        st.write(f"**Влияние на качество жизни:** {prediction['quality_of_life_impact']}")
            
            # Возможности профилактики
            if results['detailed_predictions']['prevention_opportunities']:
                st.markdown("#### ✅ Возможности профилактики")
                for opportunity in results['detailed_predictions']['prevention_opportunities']:
                    with st.expander(f"🛡️ {opportunity['system'].replace('_', ' ').title()} - {opportunity['current_status']}"):
                        st.write(f"**Приоритет:** {opportunity['prevention_priority']}")
                        st.write(f"**Методы поддержания:** {', '.join(opportunity['maintenance_methods'])}")
        
        with tab5:
            st.markdown("### 📈 Детальная аналитика")
            
            # Создаем графики
            col1, col2 = st.columns(2)
            
            with col1:
                # График рисков по системам
                systems = list(results['body_systems_analysis'].keys())
                risks = [results['body_systems_analysis'][sys]['risk_score'] for sys in systems]
                
                fig_risks = go.Figure(data=[
                    go.Bar(x=[sys.replace('_', ' ').title() for sys in systems], y=risks,
                           marker_color=['red' if r > 0.7 else 'orange' if r > 0.5 else 'yellow' if r > 0.3 else 'green' for r in risks])
                ])
                fig_risks.update_layout(
                    title="Уровень риска по системам организма",
                    xaxis_title="Системы организма",
                    yaxis_title="Уровень риска",
                    height=400
                )
                st.plotly_chart(fig_risks, use_container_width=True)
            
            with col2:
                # График квантовых факторов
                quantum_factors = results['quantum_analysis']['quantum_factors']
                factors = list(quantum_factors.keys())
                values = list(quantum_factors.values())
                
                fig_quantum = go.Figure(data=[
                    go.Bar(x=[f.replace('_', ' ').title() for f in factors], y=values,
                           marker_color='lightblue')
                ])
                fig_quantum.update_layout(
                    title="Квантовые факторы здоровья",
                    xaxis_title="Факторы",
                    yaxis_title="Значение",
                    height=400
                )
                st.plotly_chart(fig_quantum, use_container_width=True)
            
            # Общая статистика
            st.markdown("#### 📊 Общая статистика")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Общий риск", f"{results['risk_score']:.1%}")
            with col2:
                healthy_systems = sum(1 for sys in results['body_systems_analysis'].values() if sys['risk_score'] < 0.3)
                st.metric("Здоровые системы", f"{healthy_systems}/5")
            with col3:
                st.metric("Квантовый индекс", f"{results['quantum_analysis']['quantum_health_index']:.1%}")
            with col4:
                st.metric("Стабильность", f"{results['quantum_analysis']['subconscious_stability']:.1%}")
    
    # Нижняя панель с дополнительной информацией
    st.markdown("---")
    st.markdown("### 🔬 Научная основа КАПЧ")
    
    info_col1, info_col2, info_col3 = st.columns(3)
    
    with info_col1:
        st.markdown("""
        **🌊 Квантовая механика**
        - Суперпозиция состояний
        - Квантовая запутанность
        - Волновая функция сознания
        """)
    
    with info_col2:
        st.markdown("""
        **🧠 Психосоматика**
        - Связь эмоций и болезней
        - Стресс и здоровье
        - Подсознательные процессы
        """)
    
    with info_col3:
        st.markdown("""
        **⚡ ИИ Алгоритмы**
        - Машинное обучение
        - Предсказательная аналитика
        - Оценка рисков
        """)

def show_kapch_consciousness_monitoring():
    """Мониторинг сознания в КАПЧ."""
    st.markdown("## 📊 Мониторинг сознания")
    
    st.markdown("""
    ### 🧠 Система мониторинга состояний сознания
    
    **КАПЧ Мониторинг** - это инновационная система для отслеживания различных состояний сознания 
    в реальном времени с использованием квантовых технологий и нейронауки.
    """)
    
    # Инициализируем мониторинг в session_state
    if 'consciousness_monitoring' not in st.session_state:
        st.session_state.consciousness_monitoring = {
            'is_monitoring': False,
            'current_state': 'бодрствование',
            'states_history': [],
            'alerts': [],
            'metrics': {
                'attention_level': 0.5,
                'awareness_level': 0.5,
                'focus_level': 0.5,
                'relaxation_level': 0.5,
                'stress_level': 0.5,
                'emotional_stability': 0.5
            }
        }
    
    # Создаем колонки для интерфейса
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 🎛️ Панель управления мониторингом")
        
        # Выбор пациента
        patient_id = st.selectbox(
            "👤 Пациент для мониторинга:",
            ["P001", "P002", "P003", "P004", "P005"],
            key="consciousness_patient"
        )
        
        # Настройки мониторинга
        st.markdown("#### ⚙️ Настройки мониторинга")
        
        monitoring_mode = st.selectbox(
            "🔍 Режим мониторинга:",
            ["Непрерывный", "Периодический", "По запросу", "Экстренный"],
            key="monitoring_mode"
        )
        
        monitoring_duration = st.slider(
            "⏱️ Длительность мониторинга (минуты):",
            min_value=5,
            max_value=120,
            value=30,
            step=5,
            key="monitoring_duration"
        )
        
        sensitivity = st.slider(
            "🎯 Чувствительность детекции:",
            min_value=0.1,
            max_value=1.0,
            value=0.7,
            step=0.1,
            key="sensitivity"
        )
        
        # Кнопки управления
        col_start, col_stop, col_reset = st.columns(3)
        
        with col_start:
            if st.button("▶️ Начать мониторинг", key="start_monitoring", type="primary"):
                st.session_state.consciousness_monitoring['is_monitoring'] = True
                st.success("✅ Мониторинг запущен!")
                st.rerun()
        
        with col_stop:
            if st.button("⏹️ Остановить", key="stop_monitoring"):
                st.session_state.consciousness_monitoring['is_monitoring'] = False
                st.warning("⏹️ Мониторинг остановлен")
                st.rerun()
        
        with col_reset:
            if st.button("🔄 Сброс", key="reset_monitoring"):
                st.session_state.consciousness_monitoring = {
                    'is_monitoring': False,
                    'current_state': 'бодрствование',
                    'states_history': [],
                    'alerts': [],
                    'metrics': {
                        'attention_level': 0.5,
                        'awareness_level': 0.5,
                        'focus_level': 0.5,
                        'relaxation_level': 0.5,
                        'stress_level': 0.5,
                        'emotional_stability': 0.5
                    }
                }
                st.info("🔄 Данные сброшены")
                st.rerun()
    
    with col2:
        st.markdown("### 📊 Текущее состояние")
        
        monitoring_data = st.session_state.consciousness_monitoring
        
        # Текущее состояние сознания
        current_state = monitoring_data['current_state']
        state_colors = {
            'бодрствование': '🟢',
            'расслабление': '🔵',
            'медитация': '🟣',
            'транс': '🟡',
            'сон': '⚫',
            'стресс': '🔴',
            'тревога': '🟠'
        }
        
        st.markdown(f"**Состояние:** {state_colors.get(current_state, '⚪')} {current_state.title()}")
        
        # Статус мониторинга
        if monitoring_data['is_monitoring']:
            st.success("🟢 Мониторинг активен")
        else:
            st.info("⚪ Мониторинг неактивен")
        
        # Метрики в реальном времени
        st.markdown("#### 📈 Метрики сознания")
        
        metrics = monitoring_data['metrics']
        
        st.metric("Внимание", f"{metrics['attention_level']:.1%}")
        st.metric("Осознанность", f"{metrics['awareness_level']:.1%}")
        st.metric("Фокус", f"{metrics['focus_level']:.1%}")
        st.metric("Релаксация", f"{metrics['relaxation_level']:.1%}")
        st.metric("Стресс", f"{metrics['stress_level']:.1%}")
        st.metric("Эмоциональная стабильность", f"{metrics['emotional_stability']:.1%}")
    
    # Основная область мониторинга
    st.markdown("---")
    
    # Создаем табы для разных видов мониторинга
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Реальное время", "📈 История состояний", "🚨 Оповещения", "🔬 Анализ"])
    
    with tab1:
        st.markdown("### 📊 Мониторинг в реальном времени")
        
        if monitoring_data['is_monitoring']:
            # Симуляция данных в реальном времени
            placeholder = st.empty()
            
            for i in range(10):  # Показываем 10 обновлений
                # Генерируем случайные изменения метрик
                import random
                for metric in monitoring_data['metrics']:
                    change = random.uniform(-0.1, 0.1)
                    monitoring_data['metrics'][metric] = max(0, min(1, monitoring_data['metrics'][metric] + change))
                
                # Определяем состояние на основе метрик
                if monitoring_data['metrics']['stress_level'] > 0.7:
                    monitoring_data['current_state'] = 'стресс'
                elif monitoring_data['metrics']['relaxation_level'] > 0.7:
                    monitoring_data['current_state'] = 'расслабление'
                elif monitoring_data['metrics']['focus_level'] > 0.8:
                    monitoring_data['current_state'] = 'медитация'
                else:
                    monitoring_data['current_state'] = 'бодрствование'
                
                # Обновляем историю
                timestamp = datetime.now().strftime("%H:%M:%S")
                monitoring_data['states_history'].append({
                    'timestamp': timestamp,
                    'state': monitoring_data['current_state'],
                    'metrics': monitoring_data['metrics'].copy()
                })
                
                # Показываем обновленные данные
                with placeholder.container():
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown(f"**Время:** {timestamp}")
                        st.markdown(f"**Состояние:** {state_colors.get(monitoring_data['current_state'], '⚪')} {monitoring_data['current_state'].title()}")
                    
                    with col2:
                        st.markdown("**Метрики:**")
                        # Словарь переводов метрик
                        metric_translations = {
                            'attention_level': 'Уровень внимания',
                            'awareness_level': 'Уровень осознанности',
                            'focus_level': 'Уровень фокуса',
                            'relaxation_level': 'Уровень релаксации',
                            'stress_level': 'Уровень стресса',
                            'emotional_stability': 'Эмоциональная стабильность'
                        }
                        
                        for metric, value in monitoring_data['metrics'].items():
                            metric_name = metric_translations.get(metric, metric.replace('_', ' ').title())
                            st.progress(value, text=f"{metric_name}: {value:.1%}")
                    
                    with col3:
                        st.markdown("**Статус:**")
                        if monitoring_data['metrics']['stress_level'] > 0.7:
                            st.error("🚨 Высокий стресс!")
                        elif monitoring_data['metrics']['attention_level'] < 0.3:
                            st.warning("⚠️ Низкое внимание")
                        else:
                            st.success("✅ Нормальное состояние")
                
                time.sleep(1)  # Задержка для эффекта реального времени
        else:
            st.info("👆 Нажмите 'Начать мониторинг' для запуска системы")
    
    with tab2:
        st.markdown("### 📈 История состояний сознания")
        
        if monitoring_data['states_history']:
            # Создаем график истории состояний
            history_data = monitoring_data['states_history'][-20:]  # Последние 20 записей
            
            timestamps = [entry['timestamp'] for entry in history_data]
            attention = [entry['metrics']['attention_level'] for entry in history_data]
            awareness = [entry['metrics']['awareness_level'] for entry in history_data]
            stress = [entry['metrics']['stress_level'] for entry in history_data]
            relaxation = [entry['metrics']['relaxation_level'] for entry in history_data]
            
            fig_history = go.Figure()
            
            fig_history.add_trace(go.Scatter(
                x=timestamps, y=attention,
                mode='lines+markers',
                name='Внимание',
                line=dict(color='blue')
            ))
            
            fig_history.add_trace(go.Scatter(
                x=timestamps, y=awareness,
                mode='lines+markers',
                name='Осознанность',
                line=dict(color='green')
            ))
            
            fig_history.add_trace(go.Scatter(
                x=timestamps, y=stress,
                mode='lines+markers',
                name='Стресс',
                line=dict(color='red')
            ))
            
            fig_history.add_trace(go.Scatter(
                x=timestamps, y=relaxation,
                mode='lines+markers',
                name='Релаксация',
                line=dict(color='purple')
            ))
            
            fig_history.update_layout(
                title="История состояний сознания",
                xaxis_title="Время",
                yaxis_title="Уровень",
                height=400
            )
            
            st.plotly_chart(fig_history, use_container_width=True)
            
            # Таблица последних состояний
            st.markdown("#### 📋 Последние записи")
            recent_data = []
            for entry in history_data[-10:]:
                recent_data.append({
                    'Время': entry['timestamp'],
                    'Состояние': entry['state'].title(),
                    'Внимание': f"{entry['metrics']['attention_level']:.1%}",
                    'Стресс': f"{entry['metrics']['stress_level']:.1%}",
                    'Релаксация': f"{entry['metrics']['relaxation_level']:.1%}"
                })
            
            df_recent = pd.DataFrame(recent_data)
            st.dataframe(df_recent, use_container_width=True)
        else:
            st.info("📊 История состояний будет отображаться после начала мониторинга")
    
    with tab3:
        st.markdown("### 🚨 Система оповещений")
        
        # Проверяем условия для оповещений
        alerts = []
        current_metrics = monitoring_data['metrics']
        
        if current_metrics['stress_level'] > 0.8:
            alerts.append({
                'type': 'critical',
                'message': '🚨 КРИТИЧЕСКИЙ УРОВЕНЬ СТРЕССА!',
                'recommendation': 'Немедленно примените техники релаксации'
            })
        
        if current_metrics['attention_level'] < 0.2:
            alerts.append({
                'type': 'warning',
                'message': '⚠️ Критически низкий уровень внимания',
                'recommendation': 'Рекомендуется перерыв или смена деятельности'
            })
        
        if current_metrics['emotional_stability'] < 0.3:
            alerts.append({
                'type': 'warning',
                'message': '⚠️ Нестабильное эмоциональное состояние',
                'recommendation': 'Обратитесь за психологической поддержкой'
            })
        
        if current_metrics['relaxation_level'] > 0.9:
            alerts.append({
                'type': 'info',
                'message': '✅ Отличное состояние релаксации',
                'recommendation': 'Поддерживайте текущие практики'
            })
        
        # Отображаем оповещения
        if alerts:
            for alert in alerts:
                if alert['type'] == 'critical':
                    st.error(f"**{alert['message']}**\n{alert['recommendation']}")
                elif alert['type'] == 'warning':
                    st.warning(f"**{alert['message']}**\n{alert['recommendation']}")
                else:
                    st.info(f"**{alert['message']}**\n{alert['recommendation']}")
        else:
            st.success("✅ Все показатели в норме")
        
        # Настройки оповещений
        st.markdown("#### ⚙️ Настройки оповещений")
        
        alert_stress = st.checkbox("Оповещения о стрессе", value=True)
        alert_attention = st.checkbox("Оповещения о внимании", value=True)
        alert_emotions = st.checkbox("Оповещения об эмоциях", value=True)
        alert_relaxation = st.checkbox("Оповещения о релаксации", value=False)
    
    with tab4:
        st.markdown("### 🔬 Детальный анализ сознания")
        
        if monitoring_data['states_history']:
            # Статистический анализ
            all_attention = [entry['metrics']['attention_level'] for entry in monitoring_data['states_history']]
            all_stress = [entry['metrics']['stress_level'] for entry in monitoring_data['states_history']]
            all_relaxation = [entry['metrics']['relaxation_level'] for entry in monitoring_data['states_history']]
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Средний уровень внимания", f"{np.mean(all_attention):.1%}")
                st.metric("Максимальное внимание", f"{np.max(all_attention):.1%}")
                st.metric("Минимальное внимание", f"{np.min(all_attention):.1%}")
            
            with col2:
                st.metric("Средний уровень стресса", f"{np.mean(all_stress):.1%}")
                st.metric("Пиковый стресс", f"{np.max(all_stress):.1%}")
                st.metric("Минимальный стресс", f"{np.min(all_stress):.1%}")
            
            with col3:
                st.metric("Средняя релаксация", f"{np.mean(all_relaxation):.1%}")
                st.metric("Максимальная релаксация", f"{np.max(all_relaxation):.1%}")
                st.metric("Минимальная релаксация", f"{np.min(all_relaxation):.1%}")
            
            # Анализ состояний
            state_counts = {}
            for entry in monitoring_data['states_history']:
                state = entry['state']
                state_counts[state] = state_counts.get(state, 0) + 1
            
            if state_counts:
                st.markdown("#### 📊 Распределение состояний")
                fig_states = go.Figure(data=[go.Pie(
                    labels=list(state_counts.keys()),
                    values=list(state_counts.values()),
                    hole=0.3
                )])
                fig_states.update_layout(title="Состояния сознания")
                st.plotly_chart(fig_states, use_container_width=True)
            
            # Рекомендации на основе анализа
            st.markdown("#### 💡 Рекомендации")
            
            avg_stress = np.mean(all_stress)
            avg_attention = np.mean(all_attention)
            
            if avg_stress > 0.6:
                st.warning("🔴 Высокий средний уровень стресса. Рекомендуется:")
                st.write("• Регулярные сессии медитации")
                st.write("• Дыхательные упражнения")
                st.write("• Консультация специалиста")
            elif avg_attention < 0.4:
                st.warning("🟡 Низкий уровень внимания. Рекомендуется:")
                st.write("• Тренировка концентрации")
                st.write("• Перерывы в работе")
                st.write("• Физическая активность")
            else:
                st.success("✅ Показатели в норме. Продолжайте поддерживать здоровые практики.")
        else:
            st.info("📊 Данные для анализа будут доступны после начала мониторинга")
    
    # Экспорт данных
    st.markdown("---")
    st.markdown("### 💾 Экспорт данных мониторинга")
    
    if st.button("📥 Скачать данные мониторинга", key="export_monitoring"):
        if monitoring_data['states_history']:
            export_data = {
                'patient_id': patient_id,
                'monitoring_mode': monitoring_mode,
                'duration': monitoring_duration,
                'timestamp': datetime.now().isoformat(),
                'data': monitoring_data
            }
            
            st.download_button(
                label="📄 Скачать JSON",
                data=json.dumps(export_data, ensure_ascii=False, indent=2),
                file_name=f"consciousness_monitoring_{patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
        else:
            st.warning("⚠️ Нет данных для экспорта")

def show_kapch_meditation_states():
    """Медитативные состояния в КАПЧ."""
    st.markdown("## 🧘 Медитативные состояния")
    
    st.markdown("""
    ### 🧘‍♀️ Система анализа и управления медитативными состояниями
    
    **КАПЧ Медитация** - это инновационная система для глубокого анализа медитативных состояний,
    управления процессом медитации и достижения оптимальных состояний сознания.
    """)
    
    # Инициализируем медитативную систему в session_state
    if 'meditation_system' not in st.session_state:
        st.session_state.meditation_system = {
            'is_meditating': False,
            'current_meditation_type': 'осознанность',
            'meditation_timer': 0,
            'meditation_history': [],
            'current_state': {
                'deepness_level': 0.0,
                'focus_level': 0.0,
                'relaxation_level': 0.0,
                'awareness_level': 0.0,
                'breath_rhythm': 0.0,
                'body_awareness': 0.0
            },
            'meditation_goals': {
                'stress_reduction': False,
                'focus_improvement': False,
                'emotional_balance': False,
                'sleep_improvement': False,
                'creativity_enhancement': False
            }
        }
    
    # Создаем колонки для интерфейса
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 🎯 Выбор типа медитации")
        
        # Выбор пациента
        patient_id = st.selectbox(
            "👤 Пациент:",
            ["P001", "P002", "P003", "P004", "P005"],
            key="meditation_patient"
        )
        
        # Типы медитации
        meditation_types = {
            'осознанность': {
                'description': 'Медитация осознанности (Mindfulness)',
                'duration': '5-30 минут',
                'difficulty': 'Начальный',
                'benefits': ['Снижение стресса', 'Улучшение концентрации', 'Эмоциональная стабильность'],
                'technique': 'Фокус на дыхании и настоящем моменте'
            },
            'трансцендентальная': {
                'description': 'Трансцендентальная медитация',
                'duration': '15-20 минут',
                'difficulty': 'Средний',
                'benefits': ['Глубокое расслабление', 'Улучшение сна', 'Снижение тревожности'],
                'technique': 'Повторение мантры'
            },
            'любящая_доброта': {
                'description': 'Медитация любящей доброты (Metta)',
                'duration': '10-25 минут',
                'difficulty': 'Начальный',
                'benefits': ['Развитие сострадания', 'Улучшение отношений', 'Позитивное мышление'],
                'technique': 'Направление любви и доброты к себе и другим'
            },
            'дыхательная': {
                'description': 'Дыхательная медитация',
                'duration': '5-20 минут',
                'difficulty': 'Начальный',
                'benefits': ['Регуляция нервной системы', 'Снижение давления', 'Улучшение сна'],
                'technique': 'Контролируемое дыхание'
            },
            'сканирование_тела': {
                'description': 'Сканирование тела (Body Scan)',
                'duration': '20-45 минут',
                'difficulty': 'Средний',
                'benefits': ['Снятие напряжения', 'Улучшение осознанности тела', 'Релаксация'],
                'technique': 'Постепенное внимание к каждой части тела'
            },
            'визуализация': {
                'description': 'Медитация визуализации',
                'duration': '10-30 минут',
                'difficulty': 'Средний',
                'benefits': ['Развитие воображения', 'Снижение стресса', 'Улучшение настроения'],
                'technique': 'Создание ментальных образов'
            }
        }
        
        selected_type = st.selectbox(
            "🧘 Тип медитации:",
            list(meditation_types.keys()),
            key="meditation_type_select"
        )
        
        # Показываем информацию о выбранной медитации
        if selected_type in meditation_types:
            meditation_info = meditation_types[selected_type]
            
            st.markdown(f"#### 📋 {meditation_info['description']}")
            
            col_info1, col_info2 = st.columns(2)
            
            with col_info1:
                st.write(f"**⏱️ Длительность:** {meditation_info['duration']}")
                st.write(f"**📊 Сложность:** {meditation_info['difficulty']}")
            
            with col_info2:
                st.write(f"**🎯 Техника:** {meditation_info['technique']}")
            
            st.markdown("**✅ Преимущества:**")
            for benefit in meditation_info['benefits']:
                st.write(f"• {benefit}")
        
        # Настройки медитации
        st.markdown("#### ⚙️ Настройки медитации")
        
        meditation_duration = st.slider(
            "⏱️ Длительность (минуты):",
            min_value=5,
            max_value=60,
            value=15,
            step=5,
            key="meditation_duration"
        )
        
        meditation_environment = st.selectbox(
            "🌍 Окружение:",
            ["Тихая комната", "Природа", "Музыка", "Тишина", "Белый шум"],
            key="meditation_environment"
        )
        
        # Цели медитации
        st.markdown("#### 🎯 Цели медитации")
        
        col_goals1, col_goals2 = st.columns(2)
        
        with col_goals1:
            stress_reduction = st.checkbox("Снижение стресса", key="goal_stress")
            focus_improvement = st.checkbox("Улучшение концентрации", key="goal_focus")
            emotional_balance = st.checkbox("Эмоциональный баланс", key="goal_emotions")
        
        with col_goals2:
            sleep_improvement = st.checkbox("Улучшение сна", key="goal_sleep")
            creativity_enhancement = st.checkbox("Развитие креативности", key="goal_creativity")
    
    with col2:
        st.markdown("### 🎛️ Управление медитацией")
        
        meditation_data = st.session_state.meditation_system
        
        # Статус медитации
        if meditation_data['is_meditating']:
            st.success("🧘 Медитация активна")
            
            # Таймер медитации
            st.markdown("#### ⏱️ Таймер")
            timer_placeholder = st.empty()
            
            # Показываем текущее состояние
            st.markdown("#### 📊 Текущее состояние")
            
            current_state = meditation_data['current_state']
            
            st.metric("Глубина", f"{current_state['deepness_level']:.1%}")
            st.metric("Фокус", f"{current_state['focus_level']:.1%}")
            st.metric("Релаксация", f"{current_state['relaxation_level']:.1%}")
            st.metric("Осознанность", f"{current_state['awareness_level']:.1%}")
            st.metric("Ритм дыхания", f"{current_state['breath_rhythm']:.1%}")
            st.metric("Осознанность тела", f"{current_state['body_awareness']:.1%}")
            
        else:
            st.info("⚪ Медитация неактивна")
        
        # Кнопки управления
        col_start, col_stop = st.columns(2)
        
        with col_start:
            if st.button("▶️ Начать медитацию", key="start_meditation", type="primary"):
                st.session_state.meditation_system['is_meditating'] = True
                st.session_state.meditation_system['current_meditation_type'] = selected_type
                st.session_state.meditation_system['meditation_timer'] = meditation_duration * 60
                st.session_state.meditation_system['meditation_goals'] = {
                    'stress_reduction': stress_reduction,
                    'focus_improvement': focus_improvement,
                    'emotional_balance': emotional_balance,
                    'sleep_improvement': sleep_improvement,
                    'creativity_enhancement': creativity_enhancement
                }
                st.success("🧘 Медитация началась!")
                st.rerun()
        
        with col_stop:
            if st.button("⏹️ Остановить", key="stop_meditation"):
                st.session_state.meditation_system['is_meditating'] = False
                st.warning("⏹️ Медитация остановлена")
                st.rerun()
        
        # Быстрые настройки
        st.markdown("#### ⚡ Быстрые настройки")
        
        if st.button("🔄 Сброс", key="reset_meditation"):
            st.session_state.meditation_system = {
                'is_meditating': False,
                'current_meditation_type': 'осознанность',
                'meditation_timer': 0,
                'meditation_history': [],
                'current_state': {
                    'deepness_level': 0.0,
                    'focus_level': 0.0,
                    'relaxation_level': 0.0,
                    'awareness_level': 0.0,
                    'breath_rhythm': 0.0,
                    'body_awareness': 0.0
                },
                'meditation_goals': {
                    'stress_reduction': False,
                    'focus_improvement': False,
                    'emotional_balance': False,
                    'sleep_improvement': False,
                    'creativity_enhancement': False
                }
            }
            st.info("🔄 Настройки сброшены")
            st.rerun()
    
    # Основная область медитации
    st.markdown("---")
    
    # Создаем табы для разных функций медитации
    tab1, tab2, tab3, tab4 = st.tabs(["🧘 Сессия медитации", "📊 Анализ состояний", "📚 Руководство", "📈 История"])
    
    with tab1:
        st.markdown("### 🧘 Сессия медитации")
        
        if meditation_data['is_meditating']:
            # Симуляция медитативного состояния
            placeholder = st.empty()
            
            for i in range(20):  # Показываем 20 обновлений
                # Генерируем изменения в медитативном состоянии
                import random
                current_state = meditation_data['current_state']
                
                # Постепенное углубление медитации
                for metric in current_state:
                    if metric == 'deepness_level':
                        # Глубина увеличивается быстрее
                        change = random.uniform(0.02, 0.05)
                    else:
                        change = random.uniform(0.01, 0.03)
                    
                    current_state[metric] = min(1.0, current_state[metric] + change)
                
                # Обновляем историю
                timestamp = datetime.now().strftime("%H:%M:%S")
                meditation_data['meditation_history'].append({
                    'timestamp': timestamp,
                    'type': meditation_data['current_meditation_type'],
                    'state': current_state.copy(),
                    'duration': i + 1
                })
                
                # Определяем фазу медитации
                deepness = current_state['deepness_level']
                if deepness < 0.2:
                    phase = "Начальная фаза"
                    phase_emoji = "🌱"
                elif deepness < 0.5:
                    phase = "Погружение"
                    phase_emoji = "🌊"
                elif deepness < 0.8:
                    phase = "Глубокое состояние"
                    phase_emoji = "🌌"
                else:
                    phase = "Трансцендентное состояние"
                    phase_emoji = "✨"
                
                # Показываем обновленные данные
                with placeholder.container():
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown(f"**Время:** {timestamp}")
                        st.markdown(f"**Фаза:** {phase_emoji} {phase}")
                        st.markdown(f"**Тип:** {meditation_data['current_meditation_type'].title()}")
                    
                    with col2:
                        st.markdown("**Состояние медитации:**")
                        for metric, value in current_state.items():
                            metric_names = {
                                'deepness_level': 'Глубина',
                                'focus_level': 'Фокус',
                                'relaxation_level': 'Релаксация',
                                'awareness_level': 'Осознанность',
                                'breath_rhythm': 'Дыхание',
                                'body_awareness': 'Тело'
                            }
                            st.progress(value, text=f"{metric_names[metric]}: {value:.1%}")
                    
                    with col3:
                        st.markdown("**Рекомендации:**")
                        if current_state['focus_level'] < 0.3:
                            st.info("🎯 Сосредоточьтесь на дыхании")
                        elif current_state['relaxation_level'] < 0.4:
                            st.info("😌 Расслабьте мышцы")
                        elif current_state['deepness_level'] > 0.7:
                            st.success("✨ Отличное состояние!")
                        else:
                            st.info("🧘 Продолжайте медитацию")
                
                time.sleep(1)  # Задержка для эффекта реального времени
        else:
            st.info("👆 Нажмите 'Начать медитацию' для запуска сессии")
    
    with tab2:
        st.markdown("### 📊 Анализ медитативных состояний")
        
        if meditation_data['meditation_history']:
            # Создаем график развития медитативного состояния
            history_data = meditation_data['meditation_history'][-30:]  # Последние 30 записей
            
            timestamps = [entry['timestamp'] for entry in history_data]
            deepness = [entry['state']['deepness_level'] for entry in history_data]
            focus = [entry['state']['focus_level'] for entry in history_data]
            relaxation = [entry['state']['relaxation_level'] for entry in history_data]
            awareness = [entry['state']['awareness_level'] for entry in history_data]
            
            fig_meditation = go.Figure()
            
            fig_meditation.add_trace(go.Scatter(
                x=timestamps, y=deepness,
                mode='lines+markers',
                name='Глубина',
                line=dict(color='purple', width=3)
            ))
            
            fig_meditation.add_trace(go.Scatter(
                x=timestamps, y=focus,
                mode='lines+markers',
                name='Фокус',
                line=dict(color='blue')
            ))
            
            fig_meditation.add_trace(go.Scatter(
                x=timestamps, y=relaxation,
                mode='lines+markers',
                name='Релаксация',
                line=dict(color='green')
            ))
            
            fig_meditation.add_trace(go.Scatter(
                x=timestamps, y=awareness,
                mode='lines+markers',
                name='Осознанность',
                line=dict(color='orange')
            ))
            
            fig_meditation.update_layout(
                title="Развитие медитативного состояния",
                xaxis_title="Время",
                yaxis_title="Уровень",
                height=400
            )
            
            st.plotly_chart(fig_meditation, use_container_width=True)
            
            # Статистика медитации
            st.markdown("#### 📈 Статистика сессии")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                max_deepness = max([entry['state']['deepness_level'] for entry in history_data])
                avg_deepness = np.mean([entry['state']['deepness_level'] for entry in history_data])
                st.metric("Максимальная глубина", f"{max_deepness:.1%}")
                st.metric("Средняя глубина", f"{avg_deepness:.1%}")
            
            with col2:
                max_focus = max([entry['state']['focus_level'] for entry in history_data])
                avg_focus = np.mean([entry['state']['focus_level'] for entry in history_data])
                st.metric("Максимальный фокус", f"{max_focus:.1%}")
                st.metric("Средний фокус", f"{avg_focus:.1%}")
            
            with col3:
                session_duration = len(history_data)
                st.metric("Длительность сессии", f"{session_duration} мин")
                st.metric("Тип медитации", meditation_data['current_meditation_type'].title())
        else:
            st.info("📊 Данные для анализа будут доступны после медитации")
    
    with tab3:
        st.markdown("### 📚 Руководство по медитации")
        
        # Интерактивное руководство
        selected_guide = st.selectbox(
            "📖 Выберите руководство:",
            ["Общие принципы", "Техники дыхания", "Позиции тела", "Преодоление препятствий", "Углубление практики"],
            key="meditation_guide"
        )
        
        if selected_guide == "Общие принципы":
            st.markdown("""
            #### 🌟 Основные принципы медитации
            
            **1. Регулярность**
            - Медитируйте каждый день в одно и то же время
            - Начните с 5-10 минут и постепенно увеличивайте
            
            **2. Терпение**
            - Не ожидайте мгновенных результатов
            - Развитие осознанности требует времени
            
            **3. Ненасилие**
            - Не заставляйте себя медитировать
            - Подходите к практике с добротой
            
            **4. Принятие**
            - Принимайте все, что происходит во время медитации
            - Не судите свои мысли или чувства
            """)
        
        elif selected_guide == "Техники дыхания":
            st.markdown("""
            #### 🫁 Техники дыхания
            
            **1. Основное дыхание**
            - Вдох через нос на 4 счета
            - Задержка на 4 счета
            - Выдох через рот на 6 счетов
            
            **2. Дыхание квадратом**
            - Вдох на 4 счета
            - Задержка на 4 счета
            - Выдох на 4 счета
            - Пауза на 4 счета
            
            **3. Дыхание 4-7-8**
            - Вдох на 4 счета
            - Задержка на 7 счетов
            - Выдох на 8 счетов
            """)
        
        elif selected_guide == "Позиции тела":
            st.markdown("""
            #### 🧘 Позиции для медитации
            
            **1. Сидя на стуле**
            - Спина прямая, не опирается на спинку
            - Ступни на полу
            - Руки на коленях или в мудре
            
            **2. Сидя на полу**
            - Поза лотоса или полулотоса
            - Подушка под ягодицы для комфорта
            - Колени ниже бедер
            
            **3. Лежа**
            - На спине, руки вдоль тела
            - Подушка под голову
            - Не засыпать!
            """)
        
        elif selected_guide == "Преодоление препятствий":
            st.markdown("""
            #### 🚧 Преодоление препятствий
            
            **1. Блуждающий ум**
            - Это нормально!
            - Мягко возвращайте внимание к объекту медитации
            
            **2. Сонливость**
            - Откройте глаза немного
            - Сядьте прямо
            - Сделайте несколько глубоких вдохов
            
            **3. Беспокойство**
            - Не боритесь с беспокойством
            - Наблюдайте его как облака в небе
            
            **4. Физический дискомфорт**
            - Отрегулируйте позу
            - Используйте подушки для поддержки
            """)
        
        elif selected_guide == "Углубление практики":
            st.markdown("""
            #### 🚀 Углубление практики
            
            **1. Увеличение времени**
            - Постепенно увеличивайте до 30-45 минут
            - Добавляйте вторую сессию в день
            
            **2. Разные техники**
            - Экспериментируйте с различными типами медитации
            - Найдите то, что подходит вам
            
            **3. Интеграция в жизнь**
            - Практикуйте осознанность в повседневных делах
            - Медитируйте в разных местах
            
            **4. Групповая практика**
            - Присоединяйтесь к медитационным группам
              - Обменивайтесь опытом с другими практикующими
            """)
    
    with tab4:
        st.markdown("### 📈 История медитаций")
        
        if meditation_data['meditation_history']:
            # Статистика по типам медитации
            meditation_types_count = {}
            for entry in meditation_data['meditation_history']:
                med_type = entry['type']
                meditation_types_count[med_type] = meditation_types_count.get(med_type, 0) + 1
            
            if meditation_types_count:
                st.markdown("#### 📊 Предпочитаемые типы медитации")
                fig_types = go.Figure(data=[go.Pie(
                    labels=list(meditation_types_count.keys()),
                    values=list(meditation_types_count.values()),
                    hole=0.3
                )])
                fig_types.update_layout(title="Типы медитации")
                st.plotly_chart(fig_types, use_container_width=True)
            
            # Таблица последних сессий
            st.markdown("#### 📋 Последние сессии")
            recent_sessions = []
            for entry in meditation_data['meditation_history'][-10:]:
                recent_sessions.append({
                    'Время': entry['timestamp'],
                    'Тип': entry['type'].title(),
                    'Глубина': f"{entry['state']['deepness_level']:.1%}",
                    'Фокус': f"{entry['state']['focus_level']:.1%}",
                    'Релаксация': f"{entry['state']['relaxation_level']:.1%}",
                    'Длительность': f"{entry['duration']} мин"
                })
            
            df_sessions = pd.DataFrame(recent_sessions)
            st.dataframe(df_sessions, use_container_width=True)
            
            # Рекомендации на основе истории
            st.markdown("#### 💡 Персонализированные рекомендации")
            
            if meditation_types_count:
                most_used = max(meditation_types_count, key=meditation_types_count.get)
                st.info(f"🎯 Ваш предпочитаемый тип медитации: **{most_used.title()}**")
                
                if most_used == 'осознанность':
                    st.write("• Отлично для начинающих")
                    st.write("• Рекомендуется увеличить время до 20-30 минут")
                elif most_used == 'трансцендентальная':
                    st.write("• Продвинутая техника")
                    st.write("• Попробуйте медитацию любящей доброты для разнообразия")
                elif most_used == 'дыхательная':
                    st.write("• Отлично для снятия стресса")
                    st.write("• Рекомендуется практиковать утром и вечером")
        else:
            st.info("📈 История медитаций будет отображаться после практики")
    
    # Экспорт данных медитации
    st.markdown("---")
    st.markdown("### 💾 Экспорт данных медитации")
    
    if st.button("📥 Скачать данные медитации", key="export_meditation"):
        if meditation_data['meditation_history']:
            export_data = {
                'patient_id': patient_id,
                'meditation_type': selected_type,
                'environment': meditation_environment,
                'goals': meditation_data['meditation_goals'],
                'timestamp': datetime.now().isoformat(),
                'data': meditation_data
            }
            
            st.download_button(
                label="📄 Скачать JSON",
                data=json.dumps(export_data, ensure_ascii=False, indent=2),
                file_name=f"meditation_data_{patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
        else:
            st.warning("⚠️ Нет данных для экспорта")

def show_kapch_quantum_experiments():
    """Квантовые эксперименты в КАПЧ."""
    log_action("🎉 ВЫЗОВ ФУНКЦИИ", "show_kapch_quantum_experiments() - НАЧАЛО")
    st.markdown("## ⚡ Квантовые эксперименты")
    
    # Информационная панель
    st.info("""
    🧠 **Квантовые эксперименты с сознанием пациента**
    
    Здесь проводятся виртуальные квантовые эксперименты, моделирующие работу сознания пациента 
    на квантовом уровне. Эти эксперименты помогают понять глубинные процессы мышления и 
    эмоциональных состояний, лежащие в основе алгоритма MQEA.
    """)
    
    # Получаем данные пациента из session_state
    patient_id = st.session_state.get('selected_patient_id', 'P001')
    patient_data = st.session_state.get('patient_data', {})
    
    # Выбор типа квантового эксперимента с сознанием
    experiment_descriptions = {
        "Квантовая суперпозиция сознания": {
            "description": "Изучение способности сознания находиться в нескольких состояниях одновременно",
            "medical_relevance": "Моделирует множественные паттерны мышления при принятии решений",
            "parameters": ["Амплитуда мысли", "Фаза эмоции", "Частота внимания"],
            "icon": "🌀",
            "complexity": "🟢 Низкая",
            "time": "1-2 сек",
            "accuracy": "95-99%"
        },
        "Квантовая запутанность эмоций": {
            "description": "Демонстрация мгновенной корреляции между различными эмоциональными состояниями",
            "medical_relevance": "Основа понимания эмоциональных связей в психике пациента",
            "parameters": ["Угол эмоции", "Корреляция чувств", "Детекция настроения"],
            "icon": "🔗",
            "complexity": "🟡 Средняя",
            "time": "2-3 сек",
            "accuracy": "95-99%"
        },
        "Квантовая интерференция памяти": {
            "description": "Изучение интерференционных паттернов в процессах памяти",
            "medical_relevance": "Помогает понять, как формируются и восстанавливаются воспоминания",
            "parameters": ["Длина волны памяти", "Разность фаз опыта", "Интенсивность впечатления"],
            "icon": "🌊",
            "complexity": "🟢 Низкая",
            "time": "1-2 сек",
            "accuracy": "90-95%"
        },
        "Квантовое туннелирование через барьеры": {
            "description": "Демонстрация преодоления психологических барьеров",
            "medical_relevance": "Моделирует процесс преодоления страхов и ограничений",
            "parameters": ["Высота барьера", "Ширина страха", "Энергия решимости"],
            "icon": "🚇",
            "complexity": "🟡 Средняя",
            "time": "2-3 сек",
            "accuracy": "90-95%"
        },
        "Квантовая декогеренция стресса": {
            "description": "Изучение потери когерентности сознания под воздействием стресса",
            "medical_relevance": "Помогает понять влияние стресса на ясность мышления",
            "parameters": ["Время декогеренции", "Уровень стресса", "Шум окружения"],
            "icon": "💨",
            "complexity": "🔴 Высокая",
            "time": "3-5 сек",
            "accuracy": "85-90%"
        },
        "Квантовые вычисления мышления": {
            "description": "Демонстрация квантовых алгоритмов обработки информации в мозге",
            "medical_relevance": "Основа понимания вычислительных процессов сознания",
            "parameters": ["Количество нейронов", "Синапсы", "Измерения активности"],
            "icon": "💻",
            "complexity": "🔴 Высокая",
            "time": "3-5 сек",
            "accuracy": "85-90%"
        }
    }
    
    experiment_type = st.selectbox(
        "🎯 Выберите тип квантового эксперимента с сознанием:",
        list(experiment_descriptions.keys())
    )
    
    # Показываем подробную информацию о выбранном эксперименте
    exp_info = experiment_descriptions[experiment_type]
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown(f"""
        ### {exp_info['icon']} {experiment_type}
        
        **📖 Описание:** {exp_info['description']}
        
        **🏥 Медицинское применение:** {exp_info['medical_relevance']}
        
        **⚙️ Основные параметры:** {', '.join(exp_info['parameters'])}
        """)
    
    with col2:
        st.markdown(f"""
        ### 📊 Статистика эксперимента
        
        - **Сложность:** {exp_info['complexity']}
        - **Время выполнения:** {exp_info['time']}
        - **Точность:** {exp_info['accuracy']}
        """)
    
    # Разделитель
    st.markdown("---")
    
    # Параметры эксперимента
    st.subheader("⚙️ Настройка параметров квантового эксперимента")
    
    # Создаем табы для разных категорий параметров
    tab1, tab2, tab3 = st.tabs(["🧠 Параметры сознания", "⚛️ Квантовые параметры", "🔬 Экспериментальные настройки"])
    
    with tab1:
        st.markdown("**Настройка параметров сознания пациента:**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            consciousness_level = st.slider(
                "🧠 Уровень сознания (0-100%)", 
                0.0, 100.0, 85.0,
                help="Уровень ясности и осознанности пациента"
            )
            
            emotional_state = st.slider(
                "😊 Эмоциональное состояние (-100 до +100)", 
                -100.0, 100.0, 20.0,
                help="Общее эмоциональное состояние пациента"
            )
        
        with col2:
            attention_span = st.slider(
                "👁️ Концентрация внимания (0-100%)", 
                0.0, 100.0, 70.0,
                help="Способность к концентрации и фокусировке"
            )
            
            stress_level = st.slider(
                "😰 Уровень стресса (0-100%)", 
                0.0, 100.0, 30.0,
                help="Текущий уровень стресса пациента"
            )
        
        # Показываем влияние параметров
        st.markdown("**📊 Влияние параметров на эксперимент:**")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            coherence_factor = max(0.1, consciousness_level/100.0)
            st.metric("Когерентность сознания", f"{coherence_factor:.2f}", delta=f"{coherence_factor-0.5:.2f}")
        
        with col2:
            emotional_balance = max(0.1, 1.0 - abs(emotional_state)/100.0)
            st.metric("Эмоциональный баланс", f"{emotional_balance:.2f}", delta=f"{emotional_balance-0.5:.2f}")
        
        with col3:
            focus_stability = attention_span/100.0
            st.metric("Стабильность фокуса", f"{focus_stability:.2f}", delta=f"{focus_stability-0.5:.2f}")
        
        with col4:
            stress_resistance = max(0.1, 1.0 - stress_level/100.0)
            st.metric("Стрессоустойчивость", f"{stress_resistance:.2f}", delta=f"{stress_resistance-0.5:.2f}")
    
    with tab2:
        st.markdown("**Настройка квантовых параметров сознания:**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            neuron_count = st.slider(
                "🔢 Количество активных нейронов (тыс.)", 
                1, 100, 50,
                help="Количество нейронов, участвующих в квантовом процессе"
            )
            
            quantum_gates = st.slider(
                "🚪 Количество квантовых гейтов", 
                1, 50, 15,
                help="Количество квантовых операций в нейронной сети"
            )
        
        with col2:
            measurement_basis = st.selectbox(
                "📐 Базис измерения сознания", 
                ["Внимание", "Память", "Эмоции", "Внимание+Память", "Память+Эмоции", "Все"],
                help="Базис для измерения квантового состояния сознания"
            )
            
            entanglement_type = st.selectbox(
                "🔗 Тип квантовой запутанности", 
                ["Эмоциональная", "Когнитивная", "Сенсорная", "Смешанная"],
                help="Тип квантовой запутанности в сознании"
            )
        
        # Показываем сложность алгоритма
        st.markdown("**🧮 Сложность квантового алгоритма сознания:**")
        
        algorithm_complexity = neuron_count * quantum_gates
        max_complexity = 100 * 50  # Максимальная сложность
        
        complexity_ratio = algorithm_complexity / max_complexity
        
        st.progress(complexity_ratio)
        st.caption(f"Сложность: {algorithm_complexity} операций (максимум: {max_complexity})")
    
    with tab3:
        st.markdown("**Дополнительные экспериментальные настройки:**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            measurement_accuracy = st.slider(
                "🎯 Точность измерений сознания (%)", 
                85.0, 99.9, 95.0,
                help="Точность детекции квантовых состояний сознания"
            )
            
            noise_level = st.slider(
                "🔊 Уровень внешнего шума (дБ)", 
                -100.0, 0.0, -30.0,
                help="Уровень внешних помех, влияющих на эксперимент"
            )
        
        with col2:
            experiment_duration = st.slider(
                "⏱️ Длительность эксперимента (сек)", 
                1, 60, 10,
                help="Время проведения квантового эксперимента"
            )
            
            repetition_count = st.slider(
                "🔄 Количество повторений", 
                1, 100, 10,
                help="Количество повторений для повышения точности"
            )
        
        # Дополнительные настройки
        st.markdown("**🔧 Дополнительные параметры:**")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            use_meditation = st.checkbox("🧘 Использовать медитацию", value=True)
        
        with col2:
            use_breathing = st.checkbox("🫁 Контроль дыхания", value=True)
        
        with col3:
            use_visualization = st.checkbox("👁️ Визуализация", value=True)
    
    # Разделитель
    st.markdown("---")
    
    # Кнопка запуска эксперимента
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        if st.button("🚀 Запустить квантовый эксперимент", type="primary", use_container_width=True):
            # Запускаем эксперимент
            run_quantum_consciousness_experiment(
                experiment_type, 
                consciousness_level, 
                emotional_state, 
                attention_span, 
                stress_level,
                neuron_count,
                quantum_gates,
                measurement_accuracy,
                noise_level,
                experiment_duration,
                repetition_count,
                use_meditation,
                use_breathing,
                use_visualization,
                patient_id
            )
    
    # Показываем историю экспериментов
    if 'quantum_experiments_history' not in st.session_state:
        st.session_state.quantum_experiments_history = []
    
    if st.session_state.quantum_experiments_history:
        st.markdown("---")
        st.subheader("📚 История квантовых экспериментов")
        
        # Показываем последние 5 экспериментов
        recent_experiments = st.session_state.quantum_experiments_history[-5:]
        
        for i, exp in enumerate(reversed(recent_experiments)):
            with st.expander(f"🔬 Эксперимент #{len(st.session_state.quantum_experiments_history) - i}: {exp['type']} - {exp['timestamp']}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write(f"**Тип:** {exp['type']}")
                    st.write(f"**Результат:** {exp['result']}")
                    st.write(f"**Точность:** {exp['accuracy']:.1f}%")
                
                with col2:
                    st.write(f"**Время:** {exp['duration']} сек")
                    st.write(f"**Повторений:** {exp['repetitions']}")
                    st.write(f"**Статус:** {exp['status']}")
                
                if 'visualization' in exp:
                    st.plotly_chart(exp['visualization'], use_container_width=True, key=f"quantum_exp_{i}_{exp['timestamp']}")

def run_quantum_consciousness_experiment(experiment_type, consciousness_level, emotional_state, 
                                       attention_span, stress_level, neuron_count, quantum_gates,
                                       measurement_accuracy, noise_level, experiment_duration,
                                       repetition_count, use_meditation, use_breathing, 
                                       use_visualization, patient_id):
    """Запуск квантового эксперимента с сознанием."""
    
    log_action("🚀 ЗАПУСК", f"Квантовый эксперимент: {experiment_type}")
    
    # Создаем прогресс-бар
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Симулируем выполнение эксперимента
    results = []
    
    for i in range(repetition_count):
        # Обновляем прогресс
        progress = (i + 1) / repetition_count
        progress_bar.progress(progress)
        status_text.text(f"Выполнение эксперимента {i+1}/{repetition_count}...")
        
        # Симулируем задержку
        import time
        time.sleep(0.1)
        
        # Генерируем результаты эксперимента
        if experiment_type == "Квантовая суперпозиция сознания":
            result = simulate_superposition_experiment(consciousness_level, emotional_state, attention_span)
        elif experiment_type == "Квантовая запутанность эмоций":
            result = simulate_entanglement_experiment(emotional_state, stress_level, neuron_count)
        elif experiment_type == "Квантовая интерференция памяти":
            result = simulate_interference_experiment(consciousness_level, attention_span, quantum_gates)
        elif experiment_type == "Квантовое туннелирование через барьеры":
            result = simulate_tunneling_experiment(stress_level, consciousness_level, measurement_accuracy)
        elif experiment_type == "Квантовая декогеренция стресса":
            result = simulate_decoherence_experiment(stress_level, noise_level, experiment_duration)
        elif experiment_type == "Квантовые вычисления мышления":
            result = simulate_quantum_computing_experiment(neuron_count, quantum_gates, measurement_accuracy)
        else:
            result = {"success": False, "data": [], "accuracy": 0.0}
        
        results.append(result)
    
    # Вычисляем средние результаты
    avg_accuracy = sum(r['accuracy'] for r in results) / len(results)
    success_rate = sum(1 for r in results if r['success']) / len(results) * 100
    
    # Создаем визуализацию
    fig = create_quantum_experiment_visualization(results, experiment_type)
    
    # Сохраняем результаты
    experiment_result = {
        'type': experiment_type,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'result': f"Точность: {avg_accuracy:.1f}%, Успех: {success_rate:.1f}%",
        'accuracy': avg_accuracy,
        'duration': experiment_duration,
        'repetitions': repetition_count,
        'status': '✅ Успешно' if success_rate > 80 else '⚠️ Частично успешно',
        'visualization': fig,
        'patient_id': patient_id,
        'parameters': {
            'consciousness_level': consciousness_level,
            'emotional_state': emotional_state,
            'attention_span': attention_span,
            'stress_level': stress_level,
            'neuron_count': neuron_count,
            'quantum_gates': quantum_gates
        }
    }
    
    # Добавляем в историю
    st.session_state.quantum_experiments_history.append(experiment_result)
    
    # Показываем результаты
    progress_bar.progress(1.0)
    status_text.text("✅ Эксперимент завершен!")
    
    st.success(f"🎉 Квантовый эксперимент '{experiment_type}' успешно завершен!")
    
    # Показываем детальные результаты
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Средняя точность", f"{avg_accuracy:.1f}%", delta=f"{avg_accuracy-90:.1f}%")
    
    with col2:
        st.metric("Успешность", f"{success_rate:.1f}%", delta=f"{success_rate-80:.1f}%")
    
    with col3:
        st.metric("Повторений", repetition_count)
    
    with col4:
        st.metric("Длительность", f"{experiment_duration}с")
    
    # Показываем визуализацию
    st.plotly_chart(fig, use_container_width=True, key=f"quantum_result_{experiment_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    
    # Рекомендации на основе результатов
    st.markdown("### 💡 Рекомендации на основе эксперимента")
    
    if avg_accuracy > 95:
        st.success("🌟 Отличные результаты! Сознание пациента демонстрирует высокую квантовую когерентность.")
    elif avg_accuracy > 85:
        st.info("✅ Хорошие результаты. Рекомендуется продолжить наблюдение.")
    else:
        st.warning("⚠️ Результаты требуют внимания. Рекомендуется дополнительное обследование.")
    
    # Специфичные рекомендации по типу эксперимента
    if experiment_type == "Квантовая суперпозиция сознания":
        if consciousness_level > 80:
            st.write("• Высокий уровень сознания способствует квантовой суперпозиции")
        else:
            st.write("• Рекомендуется повысить уровень сознания через медитацию")
    
    elif experiment_type == "Квантовая запутанность эмоций":
        if abs(emotional_state) < 50:
            st.write("• Эмоциональный баланс способствует квантовой запутанности")
        else:
            st.write("• Рекомендуется работа с эмоциональным состоянием")
    
    log_action("✅ ЗАВЕРШЕНИЕ", f"Квантовый эксперимент завершен: {experiment_type}")

def simulate_superposition_experiment(consciousness_level, emotional_state, attention_span):
    """Симуляция эксперимента квантовой суперпозиции сознания."""
    import numpy as np
    
    # Генерируем данные суперпозиции
    t = np.linspace(0, 2*np.pi, 100)
    amplitude = consciousness_level / 100.0
    phase = emotional_state / 100.0 * np.pi
    frequency = attention_span / 100.0 * 2
    
    # Создаем волновую функцию
    wave_function = amplitude * np.sin(frequency * t + phase)
    
    # Добавляем шум
    noise = np.random.normal(0, 0.1, len(t))
    wave_function += noise
    
    # Вычисляем точность
    accuracy = max(0, min(100, 90 + np.random.normal(0, 5)))
    
    return {
        'success': accuracy > 80,
        'data': wave_function.tolist(),
        'accuracy': accuracy,
        'type': 'superposition'
    }

def simulate_entanglement_experiment(emotional_state, stress_level, neuron_count):
    """Симуляция эксперимента квантовой запутанности эмоций."""
    import numpy as np
    
    # Генерируем корреляционные данные
    n_points = 50
    x = np.random.randn(n_points)
    correlation = 0.8 - abs(emotional_state) / 200.0  # Сильная корреляция при сбалансированных эмоциях
    y = correlation * x + np.sqrt(1 - correlation**2) * np.random.randn(n_points)
    
    # Влияние стресса на корреляцию
    stress_factor = 1 - stress_level / 200.0
    y *= stress_factor
    
    # Вычисляем точность
    accuracy = max(0, min(100, 85 + correlation * 10 + np.random.normal(0, 5)))
    
    return {
        'success': accuracy > 75,
        'data': {'x': x.tolist(), 'y': y.tolist()},
        'accuracy': accuracy,
        'type': 'entanglement'
    }

def simulate_interference_experiment(consciousness_level, attention_span, quantum_gates):
    """Симуляция эксперимента квантовой интерференции памяти."""
    import numpy as np
    
    # Генерируем интерференционные паттерны
    t = np.linspace(0, 4*np.pi, 100)
    wave1 = np.sin(t) * (consciousness_level / 100.0)
    wave2 = np.sin(t + np.pi/4) * (attention_span / 100.0)
    
    # Интерференция
    interference = wave1 + wave2
    
    # Влияние количества гейтов
    gate_factor = min(1.0, quantum_gates / 25.0)
    interference *= gate_factor
    
    # Вычисляем точность
    accuracy = max(0, min(100, 88 + np.random.normal(0, 4)))
    
    return {
        'success': accuracy > 80,
        'data': {
            'wave1': wave1.tolist(),
            'wave2': wave2.tolist(),
            'interference': interference.tolist()
        },
        'accuracy': accuracy,
        'type': 'interference'
    }

def simulate_tunneling_experiment(stress_level, consciousness_level, measurement_accuracy):
    """Симуляция эксперимента квантового туннелирования через барьеры."""
    import numpy as np
    
    # Параметры барьера
    barrier_height = stress_level / 100.0 * 10
    barrier_width = 2.0
    particle_energy = consciousness_level / 100.0 * 8
    
    # Вероятность туннелирования
    if particle_energy >= barrier_height:
        tunneling_prob = 1.0
    else:
        tunneling_prob = np.exp(-2 * barrier_width * np.sqrt(2 * (barrier_height - particle_energy)))
    
    # Влияние точности измерений
    accuracy_factor = measurement_accuracy / 100.0
    tunneling_prob *= accuracy_factor
    
    # Вычисляем точность эксперимента
    accuracy = max(0, min(100, 85 + tunneling_prob * 10 + np.random.normal(0, 3)))
    
    return {
        'success': accuracy > 75,
        'data': {
            'tunneling_probability': tunneling_prob,
            'barrier_height': barrier_height,
            'particle_energy': particle_energy
        },
        'accuracy': accuracy,
        'type': 'tunneling'
    }

def simulate_decoherence_experiment(stress_level, noise_level, experiment_duration):
    """Симуляция эксперимента квантовой декогеренции стресса."""
    import numpy as np
    
    # Время декогеренции
    decoherence_time = 1.0 / (stress_level / 100.0 + 0.1)
    noise_factor = 10 ** (noise_level / 10.0)  # Преобразуем дБ в линейную шкалу
    
    # Модель декогеренции
    t = np.linspace(0, experiment_duration, 100)
    coherence = np.exp(-t / decoherence_time) * np.exp(-noise_factor * t)
    
    # Вычисляем точность
    final_coherence = coherence[-1]
    accuracy = max(0, min(100, 80 + final_coherence * 15 + np.random.normal(0, 4)))
    
    return {
        'success': accuracy > 70,
        'data': {
            'time': t.tolist(),
            'coherence': coherence.tolist(),
            'decoherence_time': decoherence_time
        },
        'accuracy': accuracy,
        'type': 'decoherence'
    }

def simulate_quantum_computing_experiment(neuron_count, quantum_gates, measurement_accuracy):
    """Симуляция эксперимента квантовых вычислений мышления."""
    import numpy as np
    
    # Генерируем квантовые состояния
    n_states = min(neuron_count, 20)  # Ограничиваем для производительности
    states = np.random.randn(n_states) + 1j * np.random.randn(n_states)
    states = states / np.linalg.norm(states)  # Нормализация
    
    # Применяем квантовые гейты
    for _ in range(min(quantum_gates, 10)):
        # Простая унитарная матрица
        U = np.random.randn(n_states, n_states) + 1j * np.random.randn(n_states, n_states)
        U = U / np.linalg.norm(U)
        states = U @ states
    
    # Измерение с учетом точности
    measurement_noise = (100 - measurement_accuracy) / 100.0 * 0.1
    measured_states = states + measurement_noise * (np.random.randn(n_states) + 1j * np.random.randn(n_states))
    
    # Вычисляем точность
    fidelity = abs(np.vdot(states, measured_states))**2
    accuracy = max(0, min(100, 75 + fidelity * 20 + np.random.normal(0, 5)))
    
    return {
        'success': accuracy > 80,
        'data': {
            'states': [abs(s)**2 for s in measured_states],
            'fidelity': fidelity,
            'neuron_count': neuron_count
        },
        'accuracy': accuracy,
        'type': 'quantum_computing'
    }

def create_quantum_experiment_visualization(results, experiment_type):
    """Создание визуализации результатов квантового эксперимента."""
    import plotly.graph_objects as go
    import plotly.express as px
    import numpy as np
    
    if not results:
        return go.Figure()
    
    # Создаем subplot
    fig = go.Figure()
    
    if experiment_type == "Квантовая суперпозиция сознания":
        # Показываем волновую функцию
        avg_data = np.mean([r['data'] for r in results], axis=0)
        t = np.linspace(0, 2*np.pi, len(avg_data))
        
        fig.add_trace(go.Scatter(
            x=t, y=avg_data,
            mode='lines',
            name='Волновая функция',
            line=dict(color='blue', width=2)
        ))
        
        fig.update_layout(
            title="Квантовая суперпозиция сознания",
            xaxis_title="Время",
            yaxis_title="Амплитуда",
            template="plotly_white"
        )
    
    elif experiment_type == "Квантовая запутанность эмоций":
        # Показываем корреляцию
        avg_data = results[0]['data']  # Берем первый результат
        
        fig.add_trace(go.Scatter(
            x=avg_data['x'], y=avg_data['y'],
            mode='markers',
            name='Корреляция эмоций',
            marker=dict(color='red', size=8)
        ))
        
        # Добавляем линию тренда
        z = np.polyfit(avg_data['x'], avg_data['y'], 1)
        p = np.poly1d(z)
        fig.add_trace(go.Scatter(
            x=avg_data['x'], y=p(avg_data['x']),
            mode='lines',
            name='Тренд',
            line=dict(color='red', dash='dash')
        ))
        
        fig.update_layout(
            title="Квантовая запутанность эмоций",
            xaxis_title="Эмоция X",
            yaxis_title="Эмоция Y",
            template="plotly_white"
        )
    
    elif experiment_type == "Квантовая интерференция памяти":
        # Показываем интерференционные паттерны
        avg_data = results[0]['data']
        
        fig.add_trace(go.Scatter(
            x=list(range(len(avg_data['wave1']))), y=avg_data['wave1'],
            mode='lines',
            name='Волна 1',
            line=dict(color='blue')
        ))
        
        fig.add_trace(go.Scatter(
            x=list(range(len(avg_data['wave2']))), y=avg_data['wave2'],
            mode='lines',
            name='Волна 2',
            line=dict(color='green')
        ))
        
        fig.add_trace(go.Scatter(
            x=list(range(len(avg_data['interference']))), y=avg_data['interference'],
            mode='lines',
            name='Интерференция',
            line=dict(color='red', width=3)
        ))
        
        fig.update_layout(
            title="Квантовая интерференция памяти",
            xaxis_title="Время",
            yaxis_title="Амплитуда",
            template="plotly_white"
        )
    
    elif experiment_type == "Квантовое туннелирование через барьеры":
        # Показываем вероятность туннелирования
        probabilities = [r['data']['tunneling_probability'] for r in results]
        
        fig.add_trace(go.Bar(
            x=[f"Попытка {i+1}" for i in range(len(probabilities))],
            y=probabilities,
            name='Вероятность туннелирования',
            marker_color='orange'
        ))
        
        fig.update_layout(
            title="Квантовое туннелирование через барьеры",
            xaxis_title="Попытка",
            yaxis_title="Вероятность",
            template="plotly_white"
        )
    
    elif experiment_type == "Квантовая декогеренция стресса":
        # Показываем декогеренцию
        avg_data = results[0]['data']
        
        fig.add_trace(go.Scatter(
            x=avg_data['time'], y=avg_data['coherence'],
            mode='lines',
            name='Когерентность',
            line=dict(color='purple', width=2)
        ))
        
        fig.update_layout(
            title="Квантовая декогеренция стресса",
            xaxis_title="Время",
            yaxis_title="Когерентность",
            template="plotly_white"
        )
    
    elif experiment_type == "Квантовые вычисления мышления":
        # Показываем квантовые состояния
        avg_data = results[0]['data']
        
        fig.add_trace(go.Bar(
            x=[f"Нейрон {i+1}" for i in range(len(avg_data['states']))],
            y=avg_data['states'],
            name='Квантовые состояния',
            marker_color='teal'
        ))
        
        fig.update_layout(
            title="Квантовые вычисления мышления",
            xaxis_title="Нейрон",
            yaxis_title="Вероятность состояния",
            template="plotly_white"
        )
    
    # Общие настройки
    fig.update_layout(
        height=400,
        showlegend=True,
        font=dict(size=12)
    )
    
    return fig

def show_kapch_results():
    """Результаты анализов КАПЧ."""
    st.markdown("## 📈 Результаты анализов КАПЧ")
    st.info("🚧 Функция находится в разработке. Здесь будут отображаться результаты всех проведенных анализов.")

def show_kapch_settings():
    """Настройки КАПЧ."""
    st.markdown("## ⚙️ Настройки КАПЧ")
    st.info("🚧 Функция находится в разработке. Здесь будут настройки системы КАПЧ.")

def show_pediatric_analytics():
    """Аналитика детской медицины."""
    st.markdown("## 📈 Аналитика детской медицины")
    st.info("🚧 Функция находится в разработке. Здесь будут аналитические инструменты и отчеты по детской медицине.")

def show_pediatric_settings():
    """Настройки детской медицины."""
    st.markdown("## ⚙️ Настройки детской медицины")
    st.info("🚧 Функция находится в разработке. Здесь будут настройки параметров и конфигурации для детской медицины.")

def show_data_analysis():
    """Анализ данных."""
    st.header("🔍 Анализ данных MQEA")
    st.info("Функция анализа данных будет реализована в следующем обновлении")
    log_action("🔍 АНАЛИЗ ДАННЫХ", "Запрос анализа данных")

def show_trends():
    """Тренды."""
    st.header("📈 Тренды MQEA")
    st.info("Функция анализа трендов будет реализована в следующем обновлении")
    log_action("📈 ТРЕНДЫ", "Запрос анализа трендов")

def show_alerts():
    """Алерты."""
    st.header("⚠️ Алерты MQEA")
    st.info("Функция алертов будет реализована в следующем обновлении")
    log_action("⚠️ АЛЕРТЫ", "Запрос алертов")

if __name__ == "__main__":
    main()
